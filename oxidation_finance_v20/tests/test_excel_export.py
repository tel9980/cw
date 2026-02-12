#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
测试Excel导出功能
"""

import pytest
from decimal import Decimal
from datetime import date
from pathlib import Path
import tempfile
import os
from openpyxl import load_workbook

from oxidation_finance_v20.reports.excel_exporter import ExcelExporter
from oxidation_finance_v20.reports.report_templates import ReportTemplates
from oxidation_finance_v20.reports.report_manager import ReportManager
from oxidation_finance_v20.database.db_manager import DatabaseManager


class TestExcelExporter:
    """测试Excel导出器"""
    
    def test_export_balance_sheet(self):
        """测试导出资产负债表"""
        exporter = ExcelExporter()
        sample_data = ReportTemplates.get_sample_balance_sheet()
        
        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = os.path.join(tmpdir, "balance_sheet.xlsx")
            result_path = exporter.export_balance_sheet(sample_data, output_file)
            
            # 验证文件已创建
            assert os.path.exists(result_path)
            assert result_path == output_file
            
            # 验证Excel文件可以打开
            wb = load_workbook(result_path)
            assert "资产负债表" in wb.sheetnames
            ws = wb["资产负债表"]
            
            # 验证标题
            assert ws['A1'].value == "资产负债表"
            
            # 验证数据存在
            assert ws['B5'].value is not None  # 货币资金
            assert ws['D5'].value is not None  # 应付账款
    
    def test_export_income_statement(self):
        """测试导出利润表"""
        exporter = ExcelExporter()
        sample_data = ReportTemplates.get_sample_income_statement()
        
        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = os.path.join(tmpdir, "income_statement.xlsx")
            result_path = exporter.export_income_statement(sample_data, output_file)
            
            # 验证文件已创建
            assert os.path.exists(result_path)
            
            # 验证Excel文件可以打开
            wb = load_workbook(result_path)
            assert "利润表" in wb.sheetnames
            ws = wb["利润表"]
            
            # 验证标题
            assert ws['A1'].value == "利润表"
            
            # 验证数据存在
            assert ws['B5'].value is not None  # 营业收入

    def test_export_cash_flow_statement(self):
        """测试导出现金流量表"""
        exporter = ExcelExporter()
        sample_data = ReportTemplates.get_sample_cash_flow_statement()
        
        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = os.path.join(tmpdir, "cash_flow.xlsx")
            result_path = exporter.export_cash_flow_statement(sample_data, output_file)
            
            # 验证文件已创建
            assert os.path.exists(result_path)
            
            # 验证Excel文件可以打开
            wb = load_workbook(result_path)
            assert "现金流量表" in wb.sheetnames
            ws = wb["现金流量表"]
            
            # 验证标题
            assert ws['A1'].value == "现金流量表"
    
    def test_export_business_analysis(self):
        """测试导出业务分析报告"""
        exporter = ExcelExporter()
        sample_data = ReportTemplates.get_sample_business_analysis()
        
        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = os.path.join(tmpdir, "business_analysis.xlsx")
            result_path = exporter.export_business_analysis(sample_data, output_file)
            
            # 验证文件已创建
            assert os.path.exists(result_path)
            
            # 验证Excel文件可以打开
            wb = load_workbook(result_path)
            
            # 验证包含多个工作表
            assert "关键指标" in wb.sheetnames
            assert "客户分析" in wb.sheetnames
            assert "计价方式分析" in wb.sheetnames
            assert "成本分析" in wb.sheetnames
    
    def test_export_report_generic(self):
        """测试通用导出方法"""
        exporter = ExcelExporter()
        
        with tempfile.TemporaryDirectory() as tmpdir:
            # 测试资产负债表
            output_file = os.path.join(tmpdir, "test_balance.xlsx")
            result = exporter.export_report(
                'balance_sheet',
                ReportTemplates.get_sample_balance_sheet(),
                output_file
            )
            assert os.path.exists(result)
            
            # 测试不支持的报表类型
            with pytest.raises(ValueError, match="不支持的报表类型"):
                exporter.export_report('invalid_type', {}, output_file)


class TestReportTemplates:
    """测试报表模板"""
    
    def test_get_sample_balance_sheet(self):
        """测试获取资产负债表示例"""
        sample = ReportTemplates.get_sample_balance_sheet()
        
        assert sample['report_name'] == "资产负债表"
        assert 'assets' in sample
        assert 'liabilities' in sample
        assert 'equity' in sample
        assert sample['balance_check']['is_balanced'] is True
    
    def test_get_sample_income_statement(self):
        """测试获取利润表示例"""
        sample = ReportTemplates.get_sample_income_statement()
        
        assert sample['report_name'] == "利润表"
        assert 'revenue' in sample
        assert 'cost_of_goods_sold' in sample
        assert 'gross_profit' in sample
        assert 'net_profit' in sample
    
    def test_get_sample_cash_flow_statement(self):
        """测试获取现金流量表示例"""
        sample = ReportTemplates.get_sample_cash_flow_statement()
        
        assert sample['report_name'] == "现金流量表"
        assert 'operating_activities' in sample
        assert 'investing_activities' in sample
        assert 'financing_activities' in sample
        assert 'cash_balance' in sample
    
    def test_get_sample_business_analysis(self):
        """测试获取业务分析报告示例"""
        sample = ReportTemplates.get_sample_business_analysis()
        
        assert sample['report_name'] == "业务分析综合报告"
        assert 'key_metrics' in sample
        assert 'customer_analysis' in sample
        assert 'pricing_analysis' in sample
        assert 'cost_analysis' in sample
    
    def test_generate_sample_excel_reports(self):
        """测试生成示例Excel报表文件"""
        with tempfile.TemporaryDirectory() as tmpdir:
            file_paths = ReportTemplates.generate_sample_excel_reports(tmpdir)
            
            # 验证所有文件都已创建
            assert 'balance_sheet' in file_paths
            assert 'income_statement' in file_paths
            assert 'cash_flow_statement' in file_paths
            assert 'business_analysis' in file_paths
            
            for file_path in file_paths.values():
                assert os.path.exists(file_path)
                # 验证是有效的Excel文件
                wb = load_workbook(file_path)
                assert len(wb.sheetnames) > 0


class TestReportManagerExport:
    """测试ReportManager的导出功能"""
    
    def test_export_to_excel(self):
        """测试导出报表到Excel"""
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = os.path.join(tmpdir, "test.db")
            
            with DatabaseManager(db_path) as db:
                report_manager = ReportManager(db)
                
                # 使用示例数据测试导出
                sample_data = ReportTemplates.get_sample_balance_sheet()
                output_file = os.path.join(tmpdir, "exported_balance.xlsx")
                
                result = report_manager.export_to_excel(
                    'balance_sheet',
                    sample_data,
                    output_file
                )
                
                assert os.path.exists(result)
                assert result == output_file
    
    def test_export_monthly_report_to_excel(self):
        """测试导出月度报表到Excel"""
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = os.path.join(tmpdir, "test.db")
            
            with DatabaseManager(db_path) as db:
                report_manager = ReportManager(db)
                
                # 导出月度报表
                file_paths = report_manager.export_monthly_report_to_excel(
                    2024, 1, tmpdir
                )
                
                # 验证三个报表文件都已创建
                assert 'balance_sheet' in file_paths
                assert 'income_statement' in file_paths
                assert 'cash_flow_statement' in file_paths
                
                for file_path in file_paths.values():
                    assert os.path.exists(file_path)


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
