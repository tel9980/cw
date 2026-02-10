"""
创建Excel模板文件

生成带公式、格式、下拉菜单的Excel模板
方便用户直接使用

使用方法：
    python create_excel_templates.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date
from pathlib import Path


def create_transaction_template():
    """创建交易记录模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "交易记录"
    
    # 设置列宽
    ws.column_dimensions['A'].width = 12  # 日期
    ws.column_dimensions['B'].width = 10  # 类型
    ws.column_dimensions['C'].width = 12  # 金额
    ws.column_dimensions['D'].width = 20  # 往来单位
    ws.column_dimensions['E'].width = 25  # 摘要
    ws.column_dimensions['F'].width = 15  # 类别
    ws.column_dimensions['G'].width = 10  # 状态
    
    # 标题行
    headers = ['日期', '类型', '金额', '往来单位', '摘要', '类别', '状态']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # 数据验证 - 类型
    type_validation = DataValidation(type="list", formula1='"收入,支出,订单"', allow_blank=False)
    type_validation.error = '请选择：收入、支出或订单'
    type_validation.errorTitle = '输入错误'
    ws.add_data_validation(type_validation)
    type_validation.add(f'B2:B1000')
    
    # 数据验证 - 状态
    status_validation = DataValidation(type="list", formula1='"已完成,待处理,已取消"', allow_blank=False)
    status_validation.error = '请选择：已完成、待处理或已取消'
    status_validation.errorTitle = '输入错误'
    ws.add_data_validation(status_validation)
    status_validation.add(f'G2:G1000')
    
    # 示例数据
    examples = [
        [date.today().strftime('%Y-%m-%d'), '收入', 5000, '张三公司', '销售货款', '产品销售', '已完成'],
        [date.today().strftime('%Y-%m-%d'), '支出', 1200, '李四供应商', '采购原料', '原材料采购', '已完成'],
        [date.today().strftime('%Y-%m-%d'), '收入', 3000, '王五客户', '服务费', '服务收入', '已完成'],
    ]
    
    for row_idx, example in enumerate(examples, 2):
        for col_idx, value in enumerate(example, 1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            # 金额列右对齐
            if col_idx == 3:
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.number_format = '#,##0.00'
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    # 保存
    output_dir = Path(__file__).parent / 'Excel模板'
    output_dir.mkdir(exist_ok=True)
    output_path = output_dir / '01_交易记录模板.xlsx'
    wb.save(output_path)
    print(f"✓ 已创建：{output_path}")


def create_counterparty_template():
    """创建往来单位模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "往来单位"
    
    # 设置列宽
    ws.column_dimensions['A'].width = 20  # 单位名称
    ws.column_dimensions['B'].width = 10  # 类型
    ws.column_dimensions['C'].width = 12  # 联系人
    ws.column_dimensions['D'].width = 15  # 电话
    ws.column_dimensions['E'].width = 25  # 邮箱
    ws.column_dimensions['F'].width = 30  # 地址
    ws.column_dimensions['G'].width = 20  # 税号
    
    # 标题行
    headers = ['单位名称', '类型', '联系人', '电话', '邮箱', '地址', '税号']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # 数据验证 - 类型
    type_validation = DataValidation(type="list", formula1='"客户,供应商"', allow_blank=False)
    type_validation.error = '请选择：客户或供应商'
    type_validation.errorTitle = '输入错误'
    ws.add_data_validation(type_validation)
    type_validation.add(f'B2:B1000')
    
    # 示例数据
    examples = [
        ['张三公司', '客户', '张经理', '13800138001', 'zhang@example.com', '北京市朝阳区xxx路xxx号', '91110000XXXXXXXXXX'],
        ['李四供应商', '供应商', '李总', '13900139002', 'li@example.com', '上海市浦东区xxx路xxx号', '91310000XXXXXXXXXX'],
        ['王五客户', '客户', '王主任', '13700137003', 'wang@example.com', '广州市天河区xxx路xxx号', '91440000XXXXXXXXXX'],
    ]
    
    for row_idx, example in enumerate(examples, 2):
        for col_idx, value in enumerate(example, 1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    # 保存
    output_dir = Path(__file__).parent / 'Excel模板'
    output_dir.mkdir(exist_ok=True)
    output_path = output_dir / '02_往来单位模板.xlsx'
    wb.save(output_path)
    print(f"✓ 已创建：{output_path}")


def create_bank_statement_template():
    """创建银行流水模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "银行流水"
    
    # 设置列宽
    ws.column_dimensions['A'].width = 12  # 交易日期
    ws.column_dimensions['B'].width = 25  # 摘要
    ws.column_dimensions['C'].width = 12  # 收入金额
    ws.column_dimensions['D'].width = 12  # 支出金额
    ws.column_dimensions['E'].width = 12  # 余额
    ws.column_dimensions['F'].width = 20  # 对方户名
    
    # 标题行
    headers = ['交易日期', '摘要', '收入金额', '支出金额', '余额', '对方户名']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # 示例数据
    examples = [
        [date.today().strftime('%Y-%m-%d'), '转账', 5000, '', 105000, '张三公司'],
        [date.today().strftime('%Y-%m-%d'), '转账', '', 1200, 103800, '李四供应商'],
        [date.today().strftime('%Y-%m-%d'), '转账', 3000, '', 106800, '王五客户'],
    ]
    
    for row_idx, example in enumerate(examples, 2):
        for col_idx, value in enumerate(example, 1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            # 金额列右对齐
            if col_idx in [3, 4, 5]:
                cell.alignment = Alignment(horizontal='right', vertical='center')
                if value:
                    cell.number_format = '#,##0.00'
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    # 保存
    output_dir = Path(__file__).parent / 'Excel模板'
    output_dir.mkdir(exist_ok=True)
    output_path = output_dir / '03_银行流水模板.xlsx'
    wb.save(output_path)
    print(f"✓ 已创建：{output_path}")


def create_payroll_template():
    """创建工资表模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "工资表"
    
    # 设置列宽
    ws.column_dimensions['A'].width = 12  # 姓名
    ws.column_dimensions['B'].width = 12  # 基本工资
    ws.column_dimensions['C'].width = 12  # 绩效奖金
    ws.column_dimensions['D'].width = 12  # 补贴
    ws.column_dimensions['E'].width = 12  # 应发工资
    ws.column_dimensions['F'].width = 12  # 社保
    ws.column_dimensions['G'].width = 12  # 公积金
    ws.column_dimensions['H'].width = 12  # 个人所得税
    ws.column_dimensions['I'].width = 12  # 实发工资
    
    # 标题行
    headers = ['姓名', '基本工资', '绩效奖金', '补贴', '应发工资', '社保', '公积金', '个人所得税', '实发工资']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # 示例数据（带公式）
    examples = [
        ['张三', 5000, 1000, 500, '=B2+C2+D2', 800, 600, 150, '=E2-F2-G2-H2'],
        ['李四', 6000, 1500, 500, '=B3+C3+D3', 900, 700, 200, '=E3-F3-G3-H3'],
        ['王五', 4000, 800, 300, '=B4+C4+D4', 700, 500, 100, '=E4-F4-G4-H4'],
    ]
    
    for row_idx, example in enumerate(examples, 2):
        for col_idx, value in enumerate(example, 1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            # 数字列格式化
            if col_idx > 1:
                cell.number_format = '#,##0.00'
    
    # 合计行
    total_row = len(examples) + 2
    ws.cell(total_row, 1, '合计').font = Font(bold=True)
    ws.cell(total_row, 1).alignment = Alignment(horizontal='center', vertical='center')
    
    for col in range(2, 10):
        cell = ws.cell(total_row, col)
        if col in [2, 3, 4, 6, 7, 8]:  # 可以求和的列
            cell.value = f'=SUM({get_column_letter(col)}2:{get_column_letter(col)}{total_row-1})'
        cell.font = Font(bold=True)
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    # 保存
    output_dir = Path(__file__).parent / 'Excel模板'
    output_dir.mkdir(exist_ok=True)
    output_path = output_dir / '04_工资表模板.xlsx'
    wb.save(output_path)
    print(f"✓ 已创建：{output_path}")


def create_fixed_assets_template():
    """创建固定资产台账模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "固定资产台账"
    
    # 设置列宽
    ws.column_dimensions['A'].width = 15  # 资产名称
    ws.column_dimensions['B'].width = 12  # 购入日期
    ws.column_dimensions['C'].width = 12  # 原值
    ws.column_dimensions['D'].width = 10  # 折旧年限
    ws.column_dimensions['E'].width = 12  # 月折旧额
    ws.column_dimensions['F'].width = 12  # 累计折旧
    ws.column_dimensions['G'].width = 12  # 净值
    ws.column_dimensions['H'].width = 10  # 状态
    
    # 标题行
    headers = ['资产名称', '购入日期', '原值', '折旧年限(年)', '月折旧额', '累计折旧', '净值', '状态']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # 数据验证 - 状态
    status_validation = DataValidation(type="list", formula1='"在用,闲置,报废"', allow_blank=False)
    status_validation.error = '请选择：在用、闲置或报废'
    status_validation.errorTitle = '输入错误'
    ws.add_data_validation(status_validation)
    status_validation.add(f'H2:H1000')
    
    # 示例数据（带公式）
    examples = [
        ['办公电脑', '2024-01-01', 5000, 5, '=C2/D2/12', 0, '=C2-F2', '在用'],
        ['打印机', '2024-03-01', 3000, 5, '=C3/D3/12', 0, '=C3-F3', '在用'],
        ['办公桌椅', '2024-01-01', 2000, 5, '=C4/D4/12', 0, '=C4-F4', '在用'],
    ]
    
    for row_idx, example in enumerate(examples, 2):
        for col_idx, value in enumerate(example, 1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            # 金额列右对齐
            if col_idx in [3, 5, 6, 7]:
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.number_format = '#,##0.00'
            elif col_idx == 4:
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    # 保存
    output_dir = Path(__file__).parent / 'Excel模板'
    output_dir.mkdir(exist_ok=True)
    output_path = output_dir / '05_固定资产台账模板.xlsx'
    wb.save(output_path)
    print(f"✓ 已创建：{output_path}")


def create_readme():
    """创建使用说明"""
    content = """# Excel模板使用说明

## 📋 模板列表

### 1. 交易记录模板（01_交易记录模板.xlsx）
**用途**: 记录日常收入、支出、订单
**列说明**:
- 日期：交易发生日期（格式：YYYY-MM-DD）
- 类型：收入/支出/订单（下拉选择）
- 金额：交易金额（自动格式化为货币）
- 往来单位：客户或供应商名称
- 摘要：交易说明（如：销售货款、采购原料）
- 类别：交易分类（如：产品销售、原材料采购）
- 状态：已完成/待处理/已取消（下拉选择）

**使用方法**:
1. 删除示例数据（保留标题行）
2. 从第2行开始录入数据
3. 类型和状态从下拉菜单选择
4. 保存后导入小会计系统

---

### 2. 往来单位模板（02_往来单位模板.xlsx）
**用途**: 管理客户和供应商信息
**列说明**:
- 单位名称：公司全称
- 类型：客户/供应商（下拉选择）
- 联系人：主要联系人姓名
- 电话：联系电话
- 邮箱：电子邮箱
- 地址：详细地址
- 税号：纳税人识别号（开票用）

**使用方法**:
1. 删除示例数据（保留标题行）
2. 录入客户和供应商信息
3. 类型从下拉菜单选择
4. 税号务必填写正确（开票必需）

---

### 3. 银行流水模板（03_银行流水模板.xlsx）
**用途**: 导入银行流水进行对账
**列说明**:
- 交易日期：银行交易日期
- 摘要：银行流水摘要
- 收入金额：入账金额
- 支出金额：出账金额
- 余额：账户余额
- 对方户名：对方账户名称

**使用方法**:
1. 从网银下载流水（Excel格式）
2. 复制数据到本模板
3. 确保列名匹配
4. 导入小会计系统进行对账

---

### 4. 工资表模板（04_工资表模板.xlsx）
**用途**: 计算员工工资
**列说明**:
- 姓名：员工姓名
- 基本工资：固定工资
- 绩效奖金：绩效部分
- 补贴：各类补贴
- 应发工资：自动计算（=基本+绩效+补贴）
- 社保：社保扣款
- 公积金：公积金扣款
- 个人所得税：个税
- 实发工资：自动计算（=应发-社保-公积金-个税）

**使用方法**:
1. 录入员工基本信息
2. 应发工资和实发工资自动计算
3. 合计行自动汇总
4. 可直接打印或导出

---

### 5. 固定资产台账模板（05_固定资产台账模板.xlsx）
**用途**: 管理固定资产和折旧
**列说明**:
- 资产名称：资产名称
- 购入日期：购买日期
- 原值：购买价格
- 折旧年限：折旧年限（年）
- 月折旧额：自动计算（=原值/年限/12）
- 累计折旧：手动更新
- 净值：自动计算（=原值-累计折旧）
- 状态：在用/闲置/报废（下拉选择）

**使用方法**:
1. 录入资产信息
2. 月折旧额和净值自动计算
3. 每月更新累计折旧
4. 状态从下拉菜单选择

---

## 💡 使用技巧

### 1. 数据验证
- 带下拉菜单的列，只能从菜单选择
- 输入其他值会提示错误
- 确保数据规范统一

### 2. 自动计算
- 带公式的列会自动计算
- 不要手动修改公式列
- 复制行时公式会自动调整

### 3. 格式化
- 金额列自动格式化为货币
- 日期列使用标准格式（YYYY-MM-DD）
- 冻结首行方便查看

### 4. 导入小会计
- 保存为.xlsx格式
- 不要修改列名
- 删除示例数据后再导入

---

## 🆘 常见问题

### Q1: 下拉菜单不显示？
**A**: 确保单元格在数据验证范围内（第2行到第1000行）

### Q2: 公式不计算？
**A**: 检查Excel是否启用了自动计算（公式 → 计算选项 → 自动）

### Q3: 导入失败？
**A**: 检查列名是否与模板一致，不要修改标题行

### Q4: 如何添加更多行？
**A**: 直接在最后一行下方添加，公式和格式会自动应用

### Q5: 可以修改模板吗？
**A**: 可以，但不要修改列名，否则导入时无法识别

---

## 📞 需要帮助？

1. 查看文档：`docs/` 目录
2. 查看示例：`examples/` 目录
3. 常见问题：`docs/常见问题解答.md`

---

**开始使用吧！** 💼
"""
    
    output_dir = Path(__file__).parent / 'Excel模板'
    output_dir.mkdir(exist_ok=True)
    output_path = output_dir / '使用说明.txt'
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(content)
    print(f"✓ 已创建：{output_path}")


def main():
    """主函数"""
    print("=" * 60)
    print("  创建Excel模板")
    print("=" * 60)
    print()
    
    try:
        create_transaction_template()
        create_counterparty_template()
        create_bank_statement_template()
        create_payroll_template()
        create_fixed_assets_template()
        create_readme()
        
        print()
        print("=" * 60)
        print("✅ 所有模板创建完成！")
        print("=" * 60)
        print()
        print("模板位置：Excel模板/")
        print()
        print("包含模板：")
        print("  1. 交易记录模板")
        print("  2. 往来单位模板")
        print("  3. 银行流水模板")
        print("  4. 工资表模板")
        print("  5. 固定资产台账模板")
        print("  6. 使用说明")
        print()
        
    except Exception as e:
        print(f"❌ 创建失败：{e}")
        return 1
    
    return 0


if __name__ == "__main__":
    exit(main())
