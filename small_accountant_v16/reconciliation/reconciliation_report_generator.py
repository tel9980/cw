"""
Reconciliation Report Generator for V1.6 Small Accountant Practical Enhancement

This module implements the ReconciliationReportGenerator class that generates
professional Excel reports for bank reconciliation discrepancies and customer statements.
"""

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from typing import List, Optional
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from small_accountant_v16.models.core_models import (
    Discrepancy,
    DiscrepancyType,
    BankRecord,
    TransactionRecord,
    Counterparty,
    TransactionType
)


@dataclass
class CustomerAccountData:
    """客户对账数据"""
    customer: Counterparty
    transactions: List[TransactionRecord]
    start_date: date
    end_date: date
    opening_balance: Decimal
    closing_balance: Decimal


class ReconciliationReportGenerator:
    """
    对账报告生成器
    
    负责生成差异报告和客户对账单Excel文件，提供专业格式化的输出。
    """
    
    def __init__(self, output_dir: Optional[str] = None):
        """
        初始化报告生成器
        
        Args:
            output_dir: 输出目录路径，如果为None则使用当前目录
        """
        self.output_dir = Path(output_dir) if output_dir else Path(".")
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def generate_discrepancy_report(
        self,
        discrepancies: List[Discrepancy],
        report_date: Optional[datetime] = None
    ) -> Workbook:
        """
        生成差异报告
        
        Args:
            discrepancies: 差异列表
            report_date: 报告日期，如果为None则使用当前时间
        
        Returns:
            Workbook: Excel工作簿对象
        """
        if report_date is None:
            report_date = datetime.now()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "对账差异报告"
        
        # 设置列宽
        ws.column_dimensions['A'].width = 15  # 差异ID
        ws.column_dimensions['B'].width = 15  # 差异类型
        ws.column_dimensions['C'].width = 12  # 日期
        ws.column_dimensions['D'].width = 25  # 往来单位
        ws.column_dimensions['E'].width = 15  # 银行金额
        ws.column_dimensions['F'].width = 15  # 系统金额
        ws.column_dimensions['G'].width = 15  # 差异金额
        ws.column_dimensions['H'].width = 40  # 描述
        
        # 标题样式
        title_font = Font(name='微软雅黑', size=16, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        title_alignment = Alignment(horizontal='center', vertical='center')
        
        # 表头样式
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 单元格边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 数据样式
        data_alignment = Alignment(horizontal='left', vertical='center')
        number_alignment = Alignment(horizontal='right', vertical='center')
        
        # 写入标题
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = '银行对账差异报告'
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = title_alignment
        title_cell.border = thin_border
        
        # 写入报告信息
        ws.merge_cells('A2:B2')
        ws['A2'] = '报告日期：'
        ws['A2'].font = Font(name='微软雅黑', size=10, bold=True)
        ws.merge_cells('C2:D2')
        ws['C2'] = report_date.strftime('%Y年%m月%d日 %H:%M')
        ws['C2'].font = Font(name='微软雅黑', size=10)
        
        ws.merge_cells('E2:F2')
        ws['E2'] = '差异总数：'
        ws['E2'].font = Font(name='微软雅黑', size=10, bold=True)
        ws.merge_cells('G2:H2')
        ws['G2'] = f'{len(discrepancies)} 条'
        ws['G2'].font = Font(name='微软雅黑', size=10)
        
        # 空行
        ws.row_dimensions[3].height = 5
        
        # 写入表头
        headers = ['差异ID', '差异类型', '日期', '往来单位', '银行金额', '系统金额', '差异金额', '详细描述']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 写入差异数据
        row_num = 5
        for disc in discrepancies:
            # 差异ID
            cell = ws.cell(row=row_num, column=1)
            cell.value = disc.id
            cell.alignment = data_alignment
            cell.border = thin_border
            
            # 差异类型
            cell = ws.cell(row=row_num, column=2)
            cell.value = self._format_discrepancy_type(disc.type)
            cell.alignment = data_alignment
            cell.border = thin_border
            
            # 根据差异类型设置颜色
            if disc.type == DiscrepancyType.AMOUNT_DIFF:
                cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
            elif disc.type == DiscrepancyType.MISSING_SYSTEM:
                cell.fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
            elif disc.type == DiscrepancyType.MISSING_BANK:
                cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
            
            # 日期
            cell = ws.cell(row=row_num, column=3)
            if disc.bank_record:
                cell.value = disc.bank_record.transaction_date.strftime('%Y-%m-%d')
            elif disc.system_record:
                cell.value = disc.system_record.date.strftime('%Y-%m-%d')
            cell.alignment = data_alignment
            cell.border = thin_border
            
            # 往来单位
            cell = ws.cell(row=row_num, column=4)
            if disc.bank_record:
                cell.value = disc.bank_record.counterparty
            elif disc.system_record:
                cell.value = disc.system_record.description
            cell.alignment = data_alignment
            cell.border = thin_border
            
            # 银行金额
            cell = ws.cell(row=row_num, column=5)
            if disc.bank_record:
                cell.value = float(disc.bank_record.amount)
                cell.number_format = '#,##0.00'
            else:
                cell.value = '-'
            cell.alignment = number_alignment
            cell.border = thin_border
            
            # 系统金额
            cell = ws.cell(row=row_num, column=6)
            if disc.system_record:
                cell.value = float(disc.system_record.amount)
                cell.number_format = '#,##0.00'
            else:
                cell.value = '-'
            cell.alignment = number_alignment
            cell.border = thin_border
            
            # 差异金额
            cell = ws.cell(row=row_num, column=7)
            cell.value = float(disc.difference_amount)
            cell.number_format = '#,##0.00'
            cell.alignment = number_alignment
            cell.border = thin_border
            cell.font = Font(name='微软雅黑', size=10, bold=True, color='FF0000')
            
            # 详细描述
            cell = ws.cell(row=row_num, column=8)
            cell.value = disc.description
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
            
            row_num += 1
        
        # 汇总统计
        row_num += 1
        ws.merge_cells(f'A{row_num}:D{row_num}')
        summary_cell = ws[f'A{row_num}']
        summary_cell.value = '差异汇总统计'
        summary_cell.font = Font(name='微软雅黑', size=11, bold=True)
        summary_cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        summary_cell.alignment = Alignment(horizontal='center', vertical='center')
        summary_cell.border = thin_border
        
        # 统计各类差异数量
        amount_diff_count = sum(1 for d in discrepancies if d.type == DiscrepancyType.AMOUNT_DIFF)
        missing_system_count = sum(1 for d in discrepancies if d.type == DiscrepancyType.MISSING_SYSTEM)
        missing_bank_count = sum(1 for d in discrepancies if d.type == DiscrepancyType.MISSING_BANK)
        
        row_num += 1
        ws[f'A{row_num}'] = '金额差异：'
        ws[f'A{row_num}'].font = Font(name='微软雅黑', size=10)
        ws[f'B{row_num}'] = f'{amount_diff_count} 条'
        ws[f'B{row_num}'].font = Font(name='微软雅黑', size=10, bold=True)
        
        ws[f'C{row_num}'] = '系统记录缺失：'
        ws[f'C{row_num}'].font = Font(name='微软雅黑', size=10)
        ws[f'D{row_num}'] = f'{missing_system_count} 条'
        ws[f'D{row_num}'].font = Font(name='微软雅黑', size=10, bold=True)
        
        ws[f'E{row_num}'] = '银行流水缺失：'
        ws[f'E{row_num}'].font = Font(name='微软雅黑', size=10)
        ws[f'F{row_num}'] = f'{missing_bank_count} 条'
        ws[f'F{row_num}'].font = Font(name='微软雅黑', size=10, bold=True)
        
        return wb
    
    def generate_customer_statement_excel(
        self,
        customer_data: CustomerAccountData
    ) -> Workbook:
        """
        生成客户对账单Excel
        
        Args:
            customer_data: 客户对账数据
        
        Returns:
            Workbook: Excel工作簿对象
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "客户对账单"
        
        # 设置列宽
        ws.column_dimensions['A'].width = 12  # 序号
        ws.column_dimensions['B'].width = 15  # 日期
        ws.column_dimensions['C'].width = 15  # 交易类型
        ws.column_dimensions['D'].width = 40  # 摘要
        ws.column_dimensions['E'].width = 15  # 收入金额
        ws.column_dimensions['F'].width = 15  # 支出金额
        ws.column_dimensions['G'].width = 15  # 余额
        
        # 标题样式
        title_font = Font(name='微软雅黑', size=18, bold=True)
        title_alignment = Alignment(horizontal='center', vertical='center')
        
        # 表头样式
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 单元格边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 数据样式
        data_alignment = Alignment(horizontal='left', vertical='center')
        number_alignment = Alignment(horizontal='right', vertical='center')
        
        # 写入标题
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = '客户对账单'
        title_cell.font = title_font
        title_cell.alignment = title_alignment
        ws.row_dimensions[1].height = 30
        
        # 写入客户信息
        row_num = 3
        info_font = Font(name='微软雅黑', size=10)
        info_bold_font = Font(name='微软雅黑', size=10, bold=True)
        
        ws[f'A{row_num}'] = '客户名称：'
        ws[f'A{row_num}'].font = info_bold_font
        ws.merge_cells(f'B{row_num}:D{row_num}')
        ws[f'B{row_num}'] = customer_data.customer.name
        ws[f'B{row_num}'].font = info_font
        
        ws[f'E{row_num}'] = '客户编号：'
        ws[f'E{row_num}'].font = info_bold_font
        ws.merge_cells(f'F{row_num}:G{row_num}')
        ws[f'F{row_num}'] = customer_data.customer.id
        ws[f'F{row_num}'].font = info_font
        
        row_num += 1
        ws[f'A{row_num}'] = '联系人：'
        ws[f'A{row_num}'].font = info_bold_font
        ws.merge_cells(f'B{row_num}:D{row_num}')
        ws[f'B{row_num}'] = customer_data.customer.contact_person
        ws[f'B{row_num}'].font = info_font
        
        ws[f'E{row_num}'] = '联系电话：'
        ws[f'E{row_num}'].font = info_bold_font
        ws.merge_cells(f'F{row_num}:G{row_num}')
        ws[f'F{row_num}'] = customer_data.customer.phone
        ws[f'F{row_num}'].font = info_font
        
        row_num += 1
        ws[f'A{row_num}'] = '对账期间：'
        ws[f'A{row_num}'].font = info_bold_font
        ws.merge_cells(f'B{row_num}:D{row_num}')
        period_str = f'{customer_data.start_date.strftime("%Y年%m月%d日")} 至 {customer_data.end_date.strftime("%Y年%m月%d日")}'
        ws[f'B{row_num}'] = period_str
        ws[f'B{row_num}'].font = info_font
        
        ws[f'E{row_num}'] = '生成日期：'
        ws[f'E{row_num}'].font = info_bold_font
        ws.merge_cells(f'F{row_num}:G{row_num}')
        ws[f'F{row_num}'] = datetime.now().strftime('%Y年%m月%d日')
        ws[f'F{row_num}'].font = info_font
        
        # 空行
        row_num += 1
        ws.row_dimensions[row_num].height = 5
        
        # 写入表头
        row_num += 1
        headers = ['序号', '日期', '交易类型', '摘要', '收入金额', '支出金额', '余额']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 写入期初余额
        row_num += 1
        ws.cell(row=row_num, column=1).value = '-'
        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row_num, column=1).border = thin_border
        
        ws.merge_cells(f'B{row_num}:D{row_num}')
        opening_cell = ws.cell(row=row_num, column=2)
        opening_cell.value = '期初余额'
        opening_cell.font = Font(name='微软雅黑', size=10, bold=True)
        opening_cell.alignment = data_alignment
        opening_cell.border = thin_border
        
        ws.cell(row=row_num, column=5).value = ''
        ws.cell(row=row_num, column=5).border = thin_border
        ws.cell(row=row_num, column=6).value = ''
        ws.cell(row=row_num, column=6).border = thin_border
        
        balance_cell = ws.cell(row=row_num, column=7)
        balance_cell.value = float(customer_data.opening_balance)
        balance_cell.number_format = '#,##0.00'
        balance_cell.alignment = number_alignment
        balance_cell.border = thin_border
        balance_cell.font = Font(name='微软雅黑', size=10, bold=True)
        
        # 写入交易记录
        running_balance = customer_data.opening_balance
        for idx, transaction in enumerate(customer_data.transactions, 1):
            row_num += 1
            
            # 序号
            cell = ws.cell(row=row_num, column=1)
            cell.value = idx
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            
            # 日期
            cell = ws.cell(row=row_num, column=2)
            cell.value = transaction.date.strftime('%Y-%m-%d')
            cell.alignment = data_alignment
            cell.border = thin_border
            
            # 交易类型
            cell = ws.cell(row=row_num, column=3)
            cell.value = self._format_transaction_type(transaction.type)
            cell.alignment = data_alignment
            cell.border = thin_border
            
            # 摘要
            cell = ws.cell(row=row_num, column=4)
            cell.value = f"{transaction.description} - {transaction.category}"
            cell.alignment = data_alignment
            cell.border = thin_border
            
            # 收入金额和支出金额
            income_cell = ws.cell(row=row_num, column=5)
            expense_cell = ws.cell(row=row_num, column=6)
            
            if transaction.type == TransactionType.INCOME:
                income_cell.value = float(transaction.amount)
                income_cell.number_format = '#,##0.00'
                expense_cell.value = ''
                running_balance += transaction.amount
            else:  # EXPENSE
                income_cell.value = ''
                expense_cell.value = float(transaction.amount)
                expense_cell.number_format = '#,##0.00'
                running_balance -= transaction.amount
            
            income_cell.alignment = number_alignment
            income_cell.border = thin_border
            expense_cell.alignment = number_alignment
            expense_cell.border = thin_border
            
            # 余额
            balance_cell = ws.cell(row=row_num, column=7)
            balance_cell.value = float(running_balance)
            balance_cell.number_format = '#,##0.00'
            balance_cell.alignment = number_alignment
            balance_cell.border = thin_border
        
        # 写入期末余额
        row_num += 1
        ws.cell(row=row_num, column=1).value = '-'
        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row_num, column=1).border = thin_border
        
        ws.merge_cells(f'B{row_num}:D{row_num}')
        closing_cell = ws.cell(row=row_num, column=2)
        closing_cell.value = '期末余额'
        closing_cell.font = Font(name='微软雅黑', size=10, bold=True)
        closing_cell.alignment = data_alignment
        closing_cell.border = thin_border
        closing_cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        
        ws.cell(row=row_num, column=5).value = ''
        ws.cell(row=row_num, column=5).border = thin_border
        ws.cell(row=row_num, column=5).fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        
        ws.cell(row=row_num, column=6).value = ''
        ws.cell(row=row_num, column=6).border = thin_border
        ws.cell(row=row_num, column=6).fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        
        final_balance_cell = ws.cell(row=row_num, column=7)
        final_balance_cell.value = float(customer_data.closing_balance)
        final_balance_cell.number_format = '#,##0.00'
        final_balance_cell.alignment = number_alignment
        final_balance_cell.border = thin_border
        final_balance_cell.font = Font(name='微软雅黑', size=11, bold=True, color='FF0000')
        final_balance_cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        
        # 汇总统计
        row_num += 2
        ws.merge_cells(f'A{row_num}:G{row_num}')
        summary_cell = ws[f'A{row_num}']
        summary_cell.value = '对账汇总'
        summary_cell.font = Font(name='微软雅黑', size=12, bold=True)
        summary_cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        summary_cell.alignment = Alignment(horizontal='center', vertical='center')
        summary_cell.border = thin_border
        
        # 计算汇总数据
        total_income = sum(t.amount for t in customer_data.transactions if t.type.value == 'income')
        total_expense = sum(t.amount for t in customer_data.transactions if t.type.value == 'expense')
        transaction_count = len(customer_data.transactions)
        
        row_num += 1
        ws[f'A{row_num}'] = '交易笔数：'
        ws[f'A{row_num}'].font = Font(name='微软雅黑', size=10, bold=True)
        ws[f'B{row_num}'] = f'{transaction_count} 笔'
        ws[f'B{row_num}'].font = Font(name='微软雅黑', size=10)
        
        ws[f'C{row_num}'] = '总收入：'
        ws[f'C{row_num}'].font = Font(name='微软雅黑', size=10, bold=True)
        ws[f'D{row_num}'] = float(total_income)
        ws[f'D{row_num}'].number_format = '#,##0.00'
        ws[f'D{row_num}'].font = Font(name='微软雅黑', size=10)
        
        ws[f'E{row_num}'] = '总支出：'
        ws[f'E{row_num}'].font = Font(name='微软雅黑', size=10, bold=True)
        ws[f'F{row_num}'] = float(total_expense)
        ws[f'F{row_num}'].number_format = '#,##0.00'
        ws[f'F{row_num}'].font = Font(name='微软雅黑', size=10)
        
        # 签字栏
        row_num += 3
        ws[f'A{row_num}'] = '客户确认签字：_______________'
        ws[f'A{row_num}'].font = Font(name='微软雅黑', size=10)
        ws[f'E{row_num}'] = '日期：_______________'
        ws[f'E{row_num}'].font = Font(name='微软雅黑', size=10)
        
        return wb
    
    def save_discrepancy_report(
        self,
        discrepancies: List[Discrepancy],
        filename: Optional[str] = None,
        report_date: Optional[datetime] = None
    ) -> str:
        """
        保存差异报告到文件
        
        Args:
            discrepancies: 差异列表
            filename: 文件名，如果为None则自动生成
            report_date: 报告日期
        
        Returns:
            str: 保存的文件路径
        """
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'对账差异报告_{timestamp}.xlsx'
        
        wb = self.generate_discrepancy_report(discrepancies, report_date)
        filepath = self.output_dir / filename
        wb.save(filepath)
        
        return str(filepath)
    
    def save_customer_statement(
        self,
        customer_data: CustomerAccountData,
        filename: Optional[str] = None
    ) -> str:
        """
        保存客户对账单到文件
        
        Args:
            customer_data: 客户对账数据
            filename: 文件名，如果为None则自动生成
        
        Returns:
            str: 保存的文件路径
        """
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d')
            customer_name = customer_data.customer.name
            filename = f'客户对账单_{customer_name}_{timestamp}.xlsx'
        
        wb = self.generate_customer_statement_excel(customer_data)
        filepath = self.output_dir / filename
        wb.save(filepath)
        
        return str(filepath)
    
    def _format_discrepancy_type(self, disc_type: DiscrepancyType) -> str:
        """格式化差异类型为中文"""
        type_map = {
            DiscrepancyType.AMOUNT_DIFF: '金额差异',
            DiscrepancyType.MISSING_SYSTEM: '系统记录缺失',
            DiscrepancyType.MISSING_BANK: '银行流水缺失'
        }
        return type_map.get(disc_type, str(disc_type.value))
    
    def _format_transaction_type(self, trans_type) -> str:
        """格式化交易类型为中文"""
        type_map = {
            TransactionType.INCOME: '收入',
            TransactionType.EXPENSE: '支出',
            TransactionType.ORDER: '订单'
        }
        return type_map.get(trans_type, str(trans_type.value))
