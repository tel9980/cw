"""
Unit tests for ReportTemplate module

Tests template loading, application, and formatting functionality.
"""

import pytest
import pandas as pd
import openpyxl
from pathlib import Path
from datetime import date
from decimal import Decimal

from small_accountant_v16.reports.report_template import (
    ReportTemplate,
    TemplateType,
    Template,
    TemplateColumn,
    TemplateSection
)


class TestReportTemplate:
    """测试报表模板管理器"""
    
    @pytest.fixture
    def template_manager(self):
        """创建模板管理器实例"""
        return ReportTemplate()
    
    @pytest.fixture
    def temp_output_dir(self, tmp_path):
        """创建临时输出目录"""
        output_dir = tmp_path / "reports"
        output_dir.mkdir()
        return output_dir
    
    # ========== 模板加载测试 ==========
    
    def test_load_revenue_comparison_template(self, template_manager):
        """测试加载收支对比模板"""
        template = template_manager.load_template(TemplateType.REVENUE_COMPARISON)
        
        assert template is not None
        assert template.name == "收支对比表"
        assert template.template_type == TemplateType.REVENUE_COMPARISON
        assert len(template.sections) > 0
        assert template.title == "收支对比表"
    
    def test_load_profit_trend_template(self, template_manager):
        """测试加载利润趋势模板"""
        template = template_manager.load_template(TemplateType.PROFIT_TREND)
        
        assert template is not None
        assert template.name == "利润趋势表"
        assert template.template_type == TemplateType.PROFIT_TREND
        assert len(template.sections) > 0
    
    def test_load_customer_ranking_template(self, template_manager):
        """测试加载客户排名模板"""
        template = template_manager.load_template(TemplateType.CUSTOMER_RANKING)
        
        assert template is not None
        assert template.name == "客户排名表"
        assert template.template_type == TemplateType.CUSTOMER_RANKING
        assert len(template.sections) > 0
    
    def test_load_vat_declaration_template(self, template_manager):
        """测试加载增值税申报表模板"""
        template = template_manager.load_template(TemplateType.VAT_DECLARATION)
        
        assert template is not None
        assert template.name == "增值税申报表"
        assert template.template_type == TemplateType.VAT_DECLARATION
        # 增值税申报表应该有3个区域：销项、进项、应纳税额
        assert len(template.sections) == 3
    
    def test_load_income_tax_declaration_template(self, template_manager):
        """测试加载所得税申报表模板"""
        template = template_manager.load_template(TemplateType.INCOME_TAX_DECLARATION)
        
        assert template is not None
        assert template.name == "所得税申报表"
        assert template.template_type == TemplateType.INCOME_TAX_DECLARATION
        # 所得税申报表应该有3个区域：收入、扣除、应纳税额
        assert len(template.sections) == 3
    
    def test_load_balance_sheet_template(self, template_manager):
        """测试加载资产负债表模板"""
        template = template_manager.load_template(TemplateType.BALANCE_SHEET)
        
        assert template is not None
        assert template.name == "资产负债表"
        assert template.template_type == TemplateType.BALANCE_SHEET
        # 资产负债表应该有3个区域：资产、负债、所有者权益
        assert len(template.sections) == 3
    
    def test_load_income_statement_template(self, template_manager):
        """测试加载利润表模板"""
        template = template_manager.load_template(TemplateType.INCOME_STATEMENT)
        
        assert template is not None
        assert template.name == "利润表"
        assert template.template_type == TemplateType.INCOME_STATEMENT
        assert len(template.sections) > 0
    
    def test_load_cash_flow_statement_template(self, template_manager):
        """测试加载现金流量表模板"""
        template = template_manager.load_template(TemplateType.CASH_FLOW_STATEMENT)
        
        assert template is not None
        assert template.name == "现金流量表"
        assert template.template_type == TemplateType.CASH_FLOW_STATEMENT
        # 现金流量表应该有3个区域：经营、投资、筹资
        assert len(template.sections) == 3
    
    def test_load_invalid_template_type(self, template_manager):
        """测试加载无效的模板类型"""
        # 创建一个不存在的模板类型
        with pytest.raises(ValueError, match="未找到模板类型"):
            # 使用一个不在预定义模板中的枚举值
            template_manager._templates.clear()
            template_manager.load_template(TemplateType.REVENUE_COMPARISON)
    
    # ========== 模板应用测试 ==========
    
    def test_apply_revenue_comparison_template(self, template_manager, temp_output_dir):
        """测试应用收支对比模板"""
        template = template_manager.load_template(TemplateType.REVENUE_COMPARISON)
        
        # 准备测试数据
        data = pd.DataFrame([
            {
                'period': '2024年1月',
                'income': 100000.00,
                'expense': 60000.00,
                'profit': 40000.00,
                'profit_rate': 0.40
            },
            {
                'period': '2024年2月',
                'income': 120000.00,
                'expense': 70000.00,
                'profit': 50000.00,
                'profit_rate': 0.4167
            }
        ])
        
        output_file = temp_output_dir / "revenue_comparison.xlsx"
        
        # 应用模板
        wb = template_manager.apply_template(
            template,
            data,
            str(output_file),
            company_name="测试公司",
            period="2024年1-2月"
        )
        
        assert wb is not None
        assert output_file.exists()
        
        # 验证文件内容
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        
        # 检查标题
        assert "测试公司" in ws['A1'].value
        assert "收支对比表" in ws['A1'].value
        
        # 检查期间
        assert "2024年1-2月" in ws['A2'].value
    
    def test_apply_customer_ranking_template(self, template_manager, temp_output_dir):
        """测试应用客户排名模板"""
        template = template_manager.load_template(TemplateType.CUSTOMER_RANKING)
        
        # 准备测试数据
        data = pd.DataFrame([
            {
                'rank': 1,
                'customer_name': '客户A',
                'sales_amount': 500000.00,
                'transaction_count': 25,
                'percentage': 0.35,
                'customer_type': '重点客户'
            },
            {
                'rank': 2,
                'customer_name': '客户B',
                'sales_amount': 300000.00,
                'transaction_count': 18,
                'percentage': 0.21,
                'customer_type': '普通客户'
            }
        ])
        
        output_file = temp_output_dir / "customer_ranking.xlsx"
        
        # 应用模板
        wb = template_manager.apply_template(
            template,
            data,
            str(output_file),
            company_name="测试公司",
            period="2024年"
        )
        
        assert wb is not None
        assert output_file.exists()
        
        # 验证文件内容
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        
        # 检查数据行数（标题+期间+空行+区域标题+列标题+2行数据+合计行）
        assert ws.max_row >= 8
    
    def test_apply_vat_declaration_template(self, template_manager, temp_output_dir):
        """测试应用增值税申报表模板"""
        template = template_manager.load_template(TemplateType.VAT_DECLARATION)
        
        # 准备测试数据（空数据也应该能生成模板）
        data = pd.DataFrame()
        
        output_file = temp_output_dir / "vat_declaration.xlsx"
        
        # 应用模板
        wb = template_manager.apply_template(
            template,
            data,
            str(output_file),
            company_name="测试公司",
            period="2024年第一季度"
        )
        
        assert wb is not None
        assert output_file.exists()
        
        # 验证文件内容
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        
        # 检查标题
        assert "增值税" in ws['A1'].value
    
    def test_apply_balance_sheet_template(self, template_manager, temp_output_dir):
        """测试应用资产负债表模板"""
        template = template_manager.load_template(TemplateType.BALANCE_SHEET)
        
        # 准备测试数据（空数据）
        data = pd.DataFrame()
        
        output_file = temp_output_dir / "balance_sheet.xlsx"
        
        # 应用模板
        wb = template_manager.apply_template(
            template,
            data,
            str(output_file),
            company_name="测试公司",
            period="2024年12月31日"
        )
        
        assert wb is not None
        assert output_file.exists()
        
        # 验证文件内容
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        
        # 检查标题
        assert "资产负债表" in ws['A1'].value
    
    def test_apply_template_with_empty_data(self, template_manager, temp_output_dir):
        """测试使用空数据应用模板"""
        template = template_manager.load_template(TemplateType.PROFIT_TREND)
        
        # 空数据
        data = pd.DataFrame()
        
        output_file = temp_output_dir / "empty_report.xlsx"
        
        # 应用模板（应该成功，只是没有数据行）
        wb = template_manager.apply_template(
            template,
            data,
            str(output_file),
            company_name="测试公司",
            period="2024年"
        )
        
        assert wb is not None
        assert output_file.exists()
    
    # ========== 模板结构测试 ==========
    
    def test_template_sections_have_columns(self, template_manager):
        """测试所有模板的区域都有列定义"""
        for template_type in TemplateType:
            template = template_manager.load_template(template_type)
            
            for section in template.sections:
                assert len(section.columns) > 0, \
                    f"模板 {template.name} 的区域 {section.title} 没有列定义"
    
    def test_template_columns_have_names(self, template_manager):
        """测试所有列都有名称"""
        for template_type in TemplateType:
            template = template_manager.load_template(template_type)
            
            for section in template.sections:
                for column in section.columns:
                    assert column.name, \
                        f"模板 {template.name} 的区域 {section.title} 有未命名的列"
    
    def test_management_report_templates_exist(self, template_manager):
        """测试管理报表模板都存在"""
        management_templates = [
            TemplateType.REVENUE_COMPARISON,
            TemplateType.PROFIT_TREND,
            TemplateType.CUSTOMER_RANKING
        ]
        
        for template_type in management_templates:
            template = template_manager.load_template(template_type)
            assert template is not None
            assert template.name
            assert template.description
    
    def test_tax_report_templates_exist(self, template_manager):
        """测试税务报表模板都存在"""
        tax_templates = [
            TemplateType.VAT_DECLARATION,
            TemplateType.INCOME_TAX_DECLARATION
        ]
        
        for template_type in tax_templates:
            template = template_manager.load_template(template_type)
            assert template is not None
            assert template.name
            assert template.description
    
    def test_bank_loan_report_templates_exist(self, template_manager):
        """测试银行贷款报表模板都存在"""
        bank_templates = [
            TemplateType.BALANCE_SHEET,
            TemplateType.INCOME_STATEMENT,
            TemplateType.CASH_FLOW_STATEMENT
        ]
        
        for template_type in bank_templates:
            template = template_manager.load_template(template_type)
            assert template is not None
            assert template.name
            assert template.description
    
    # ========== 边界情况测试 ==========
    
    def test_apply_template_with_single_row(self, template_manager, temp_output_dir):
        """测试单行数据"""
        template = template_manager.load_template(TemplateType.REVENUE_COMPARISON)
        
        data = pd.DataFrame([
            {
                'period': '2024年1月',
                'income': 100000.00,
                'expense': 60000.00,
                'profit': 40000.00,
                'profit_rate': 0.40
            }
        ])
        
        output_file = temp_output_dir / "single_row.xlsx"
        
        wb = template_manager.apply_template(
            template,
            data,
            str(output_file)
        )
        
        assert wb is not None
        assert output_file.exists()
    
    def test_apply_template_with_large_numbers(self, template_manager, temp_output_dir):
        """测试大数值"""
        template = template_manager.load_template(TemplateType.CUSTOMER_RANKING)
        
        data = pd.DataFrame([
            {
                'rank': 1,
                'customer_name': '大客户',
                'sales_amount': 99999999.99,
                'transaction_count': 1000,
                'percentage': 0.95,
                'customer_type': '重点客户'
            }
        ])
        
        output_file = temp_output_dir / "large_numbers.xlsx"
        
        wb = template_manager.apply_template(
            template,
            data,
            str(output_file)
        )
        
        assert wb is not None
        assert output_file.exists()
    
    def test_template_with_chinese_characters(self, template_manager, temp_output_dir):
        """测试中文字符处理"""
        template = template_manager.load_template(TemplateType.CUSTOMER_RANKING)
        
        data = pd.DataFrame([
            {
                'rank': 1,
                'customer_name': '北京某某科技有限公司',
                'sales_amount': 100000.00,
                'transaction_count': 10,
                'percentage': 0.50,
                'customer_type': '重点客户'
            }
        ])
        
        output_file = temp_output_dir / "chinese_test.xlsx"
        
        wb = template_manager.apply_template(
            template,
            data,
            str(output_file),
            company_name="测试公司（中文）",
            period="2024年第一季度"
        )
        
        assert wb is not None
        assert output_file.exists()
        
        # 验证中文正确保存
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        assert "测试公司" in ws['A1'].value
        assert "中文" in ws['A1'].value


class TestTemplateDataStructures:
    """测试模板数据结构"""
    
    def test_template_column_creation(self):
        """测试列定义创建"""
        column = TemplateColumn(
            name="测试列",
            width=20,
            data_field="test_field",
            format="#,##0.00"
        )
        
        assert column.name == "测试列"
        assert column.width == 20
        assert column.data_field == "test_field"
        assert column.format == "#,##0.00"
    
    def test_template_section_creation(self):
        """测试区域定义创建"""
        columns = [
            TemplateColumn(name="列1", width=15),
            TemplateColumn(name="列2", width=20)
        ]
        
        section = TemplateSection(
            title="测试区域",
            start_row=5,
            columns=columns,
            has_total=True,
            total_columns=["列1"]
        )
        
        assert section.title == "测试区域"
        assert section.start_row == 5
        assert len(section.columns) == 2
        assert section.has_total is True
        assert "列1" in section.total_columns
    
    def test_template_creation(self):
        """测试模板创建"""
        columns = [TemplateColumn(name="列1", width=15)]
        section = TemplateSection(
            title="区域1",
            start_row=4,
            columns=columns
        )
        
        template = Template(
            template_type=TemplateType.REVENUE_COMPARISON,
            name="测试模板",
            description="测试描述",
            sections=[section],
            title="测试标题"
        )
        
        assert template.name == "测试模板"
        assert template.description == "测试描述"
        assert len(template.sections) == 1
        assert template.title == "测试标题"
