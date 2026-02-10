"""
Unit tests for ReconciliationReportGenerator

Tests the reconciliation report generation functionality including discrepancy reports
and customer statement generation with proper Excel formatting.
"""

import pytest
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
import tempfile
import shutil

from openpyxl import load_workbook

from small_accountant_v16.models.core_models import (
    Discrepancy,
    DiscrepancyType,
    BankRecord,
    TransactionRecord,
    TransactionType,
    TransactionStatus,
    Counterparty,
    CounterpartyType
)
from small_accountant_v16.reconciliation.reconciliation_report_generator import (
    ReconciliationReportGenerator,
    CustomerAccountData
)


class TestReconciliationReportGeneratorDiscrepancyReport:
    """测试差异报告生成功能"""
    
    @pytest.fixture
    def temp_output_dir(self):
        """创建临时输出目录"""
        temp_dir = tempfile.mkdtemp()
        yield temp_dir
        shutil.rmtree(temp_dir)
    
    @pytest.fixture
    def sample_discrepancies(self):
        """创建示例差异数据"""
        bank_record1 = BankRecord(
            id="B001",
            transaction_date=date(2024, 1, 15),
            description="收款",
            amount=Decimal("1000.00"),
            balance=Decimal("5000.00"),
            transaction_type="CREDIT",
            counterparty="ABC公司"
        )
        
        system_record1 = TransactionRecord(
            id="S001",
            date=date(2024, 1, 15),
            type=TransactionType.INCOME,
            amount=Decimal("1010.00"),
            counterparty_id="C001",
            description="ABC公司",
            category="销售收入",
            status=TransactionStatus.COMPLETED,
            created_at=datetime.now(),
            updated_at=datetime.now()
        )
        
        bank_record2 = BankRecord(
            id="B002",
            transaction_date=date(2024, 1, 16),
            description="收款",
            amount=Decimal("2000.00"),
            balance=Decimal("7000.00"),
            transaction_type="CREDIT",
            counterparty="XYZ公司"
        )
        
        system_record2 = TransactionRecord(
            id="S002",
            date=date(2024, 1, 17),
            type=TransactionType.INCOME,
            amount=Decimal("3000.00"),
            counterparty_id="C002",
            description="DEF公司",
            category="销售收入",
            status=TransactionStatus.COMPLETED,
            created_at=datetime.now(),
            updated_at=datetime.now()
        )
        
        return [
            Discrepancy(
                id="DISC-0001",
                type=DiscrepancyType.AMOUNT_DIFF,
                bank_record=bank_record1,
                system_record=system_record1,
                difference_amount=Decimal("10.00"),
                description="金额差异：银行流水 1000.00，系统记录 1010.00，差额 10.00"
            ),
            Discrepancy(
                id="DISC-0002",
                type=DiscrepancyType.MISSING_SYSTEM,
                bank_record=bank_record2,
                system_record=None,
                difference_amount=Decimal("2000.00"),
                description="系统记录缺失：银行流水显示 2024-01-16 XYZ公司 2000.00，但系统中未找到对应记录"
            ),
            Discrepancy(
                id="DISC-0003",
                type=DiscrepancyType.MISSING_BANK,
                bank_record=None,
                system_record=system_record2,
                difference_amount=Decimal("3000.00"),
                description="银行流水缺失：系统记录显示 2024-01-17 DEF公司 3000.00，但银行流水中未找到对应记录"
            )
        ]
    
    def test_generate_discrepancy_report_creates_workbook(self, sample_discrepancies):
        """测试生成差异报告创建工作簿"""
        generator = ReconciliationReportGenerator()
        wb = generator.generate_discrepancy_report(sample_discrepancies)
        
        assert wb is not None
        assert wb.active is not None
        assert wb.active.title == "对账差异报告"
    
    def test_generate_discrepancy_report_has_correct_headers(self, sample_discrepancies):
        """测试差异报告包含正确的表头"""
        generator = ReconciliationReportGenerator()
        wb = generator.generate_discrepancy_report(sample_discrepancies)
        ws = wb.active
        
        # 检查表头（第4行）
        expected_headers = ['差异ID', '差异类型', '日期', '往来单位', '银行金额', '系统金额', '差异金额', '详细描述']
        for col_num, expected_header in enumerate(expected_headers, 1):
            actual_header = ws.cell(row=4, column=col_num).value
            assert actual_header == expected_header
    
    def test_generate_discrepancy_report_contains_all_discrepancies(self, sample_discrepancies):
        """测试差异报告包含所有差异记录"""
        generator = ReconciliationReportGenerator()
        wb = generator.generate_discrepancy_report(sample_discrepancies)
        ws = wb.active
        
        # 数据从第5行开始
        for i, disc in enumerate(sample_discrepancies, 1):
            row_num = 4 + i
            disc_id = ws.cell(row=row_num, column=1).value
            assert disc_id == disc.id
    
    def test_generate_discrepancy_report_formats_amount_diff_correctly(self, sample_discrepancies):
        """测试金额差异格式化正确"""
        generator = ReconciliationReportGenerator()
        wb = generator.generate_discrepancy_report(sample_discrepancies)
        ws = wb.active
        
        # 第一条差异是金额差异
        row_num = 5
        disc_type = ws.cell(row=row_num, column=2).value
        assert disc_type == '金额差异'
        
        bank_amount = ws.cell(row=row_num, column=5).value
        assert bank_amount == 1000.00
        
        system_amount = ws.cell(row=row_num, column=6).value
        assert system_amount == 1010.00
        
        diff_amount = ws.cell(row=row_num, column=7).value
        assert diff_amount == 10.00
    
    def test_generate_discrepancy_report_formats_missing_system_correctly(self, sample_discrepancies):
        """测试系统记录缺失格式化正确"""
        generator = ReconciliationReportGenerator()
        wb = generator.generate_discrepancy_report(sample_discrepancies)
        ws = wb.active
        
        # 第二条差异是系统记录缺失
        row_num = 6
        disc_type = ws.cell(row=row_num, column=2).value
        assert disc_type == '系统记录缺失'
        
        bank_amount = ws.cell(row=row_num, column=5).value
        assert bank_amount == 2000.00
        
        system_amount = ws.cell(row=row_num, column=6).value
        assert system_amount == '-'
    
    def test_generate_discrepancy_report_formats_missing_bank_correctly(self, sample_discrepancies):
        """测试银行流水缺失格式化正确"""
        generator = ReconciliationReportGenerator()
        wb = generator.generate_discrepancy_report(sample_discrepancies)
        ws = wb.active
        
        # 第三条差异是银行流水缺失
        row_num = 7
        disc_type = ws.cell(row=row_num, column=2).value
        assert disc_type == '银行流水缺失'
        
        bank_amount = ws.cell(row=row_num, column=5).value
        assert bank_amount == '-'
        
        system_amount = ws.cell(row=row_num, column=6).value
        assert system_amount == 3000.00
    
    def test_generate_discrepancy_report_includes_summary_statistics(self, sample_discrepancies):
        """测试差异报告包含汇总统计"""
        generator = ReconciliationReportGenerator()
        wb = generator.generate_discrepancy_report(sample_discrepancies)
        ws = wb.active
        
        # 查找汇总统计行（在数据行之后）
        # 数据从第5行开始，3条差异，所以汇总在第9行
        summary_row = 9
        
        # 检查汇总标题
        summary_title = ws.cell(row=summary_row, column=1).value
        assert '差异汇总统计' in summary_title
        
        # 检查统计数据（在下一行）
        stats_row = summary_row + 1
        amount_diff_label = ws.cell(row=stats_row, column=1).value
        assert '金额差异' in amount_diff_label
        
        amount_diff_count = ws.cell(row=stats_row, column=2).value
        assert '1 条' in amount_diff_count
    
    def test_generate_discrepancy_report_with_empty_list(self):
        """测试生成空差异列表的报告"""
        generator = ReconciliationReportGenerator()
        wb = generator.generate_discrepancy_report([])
        ws = wb.active
        
        # 应该仍然有标题和表头
        assert ws['A1'].value == '银行对账差异报告'
        assert ws.cell(row=4, column=1).value == '差异ID'
        
        # 差异总数应该是0
        assert '0 条' in ws['G2'].value
    
    def test_save_discrepancy_report_creates_file(self, temp_output_dir, sample_discrepancies):
        """测试保存差异报告创建文件"""
        generator = ReconciliationReportGenerator(output_dir=temp_output_dir)
        filepath = generator.save_discrepancy_report(
            sample_discrepancies,
            filename='test_discrepancy_report.xlsx'
        )
        
        assert Path(filepath).exists()
        assert Path(filepath).suffix == '.xlsx'
        
        # 验证文件可以被打开
        wb = load_workbook(filepath)
        assert wb.active.title == "对账差异报告"
    
    def test_save_discrepancy_report_auto_generates_filename(self, temp_output_dir, sample_discrepancies):
        """测试保存差异报告自动生成文件名"""
        generator = ReconciliationReportGenerator(output_dir=temp_output_dir)
        filepath = generator.save_discrepancy_report(sample_discrepancies)
        
        assert Path(filepath).exists()
        assert '对账差异报告_' in Path(filepath).name
        assert Path(filepath).suffix == '.xlsx'


class TestReconciliationReportGeneratorCustomerStatement:
    """测试客户对账单生成功能"""
    
    @pytest.fixture
    def temp_output_dir(self):
        """创建临时输出目录"""
        temp_dir = tempfile.mkdtemp()
        yield temp_dir
        shutil.rmtree(temp_dir)
    
    @pytest.fixture
    def sample_customer_data(self):
        """创建示例客户对账数据"""
        customer = Counterparty(
            id="C001",
            name="北京ABC科技有限公司",
            type=CounterpartyType.CUSTOMER,
            contact_person="张三",
            phone="13800138000",
            email="zhangsan@abc.com",
            address="北京市朝阳区",
            tax_id="91110000XXXXXXXX",
            created_at=datetime.now(),
            updated_at=datetime.now()
        )
        
        transactions = [
            TransactionRecord(
                id="T001",
                date=date(2024, 1, 5),
                type=TransactionType.INCOME,
                amount=Decimal("10000.00"),
                counterparty_id="C001",
                description="销售商品",
                category="销售收入",
                status=TransactionStatus.COMPLETED,
                created_at=datetime.now(),
                updated_at=datetime.now()
            ),
            TransactionRecord(
                id="T002",
                date=date(2024, 1, 15),
                type=TransactionType.INCOME,
                amount=Decimal("5000.00"),
                counterparty_id="C001",
                description="销售服务",
                category="服务收入",
                status=TransactionStatus.COMPLETED,
                created_at=datetime.now(),
                updated_at=datetime.now()
            ),
            TransactionRecord(
                id="T003",
                date=date(2024, 1, 20),
                type=TransactionType.EXPENSE,
                amount=Decimal("3000.00"),
                counterparty_id="C001",
                description="退款",
                category="销售退款",
                status=TransactionStatus.COMPLETED,
                created_at=datetime.now(),
                updated_at=datetime.now()
            )
        ]
        
        return CustomerAccountData(
            customer=customer,
            transactions=transactions,
            start_date=date(2024, 1, 1),
            end_date=date(2024, 1, 31),
            opening_balance=Decimal("0.00"),
            closing_balance=Decimal("12000.00")
        )
    
    def test_generate_customer_statement_creates_workbook(self, sample_customer_data):
        """测试生成客户对账单创建工作簿"""
        generator = ReconciliationReportGenerator()
        wb = generator.generate_customer_statement_excel(sample_customer_data)
        
        assert wb is not None
        assert wb.active is not None
        assert wb.active.title == "客户对账单"
    
    def test_generate_customer_statement_has_correct_title(self, sample_customer_data):
        """测试客户对账单包含正确的标题"""
        generator = ReconciliationReportGenerator()
        wb = generator.generate_customer_statement_excel(sample_customer_data)
        ws = wb.active
        
        assert ws['A1'].value == '客户对账单'
    
    def test_generate_customer_statement_includes_customer_info(self, sample_customer_data):
        """测试客户对账单包含客户信息"""
        generator = ReconciliationReportGenerator()
        wb = generator.generate_customer_statement_excel(sample_customer_data)
        ws = wb.active
        
        # 客户名称在第3行
        assert ws['B3'].value == "北京ABC科技有限公司"
        
        # 客户编号
        assert ws['F3'].value == "C001"
        
        # 联系人
        assert ws['B4'].value == "张三"
        
        # 联系电话
        assert ws['F4'].value == "13800138000"
    
    def test_generate_customer_statement_includes_period_info(self, sample_customer_data):
        """测试客户对账单包含对账期间信息"""
        generator = ReconciliationReportGenerator()
        wb = generator.generate_customer_statement_excel(sample_customer_data)
        ws = wb.active
        
        # 对账期间在第5行
        period_value = ws['B5'].value
        assert '2024年01月01日' in period_value
        assert '2024年01月31日' in period_value
    
    def test_generate_customer_statement_has_correct_headers(self, sample_customer_data):
        """测试客户对账单包含正确的表头"""
        generator = ReconciliationReportGenerator()
        wb = generator.generate_customer_statement_excel(sample_customer_data)
        ws = wb.active
        
        # 表头在第7行
        expected_headers = ['序号', '日期', '交易类型', '摘要', '收入金额', '支出金额', '余额']
        for col_num, expected_header in enumerate(expected_headers, 1):
            actual_header = ws.cell(row=7, column=col_num).value
            assert actual_header == expected_header
    
    def test_generate_customer_statement_includes_opening_balance(self, sample_customer_data):
        """测试客户对账单包含期初余额"""
        generator = ReconciliationReportGenerator()
        wb = generator.generate_customer_statement_excel(sample_customer_data)
        ws = wb.active
        
        # 期初余额在第8行
        opening_label = ws.cell(row=8, column=2).value
        assert opening_label == '期初余额'
        
        opening_balance = ws.cell(row=8, column=7).value
        assert opening_balance == 0.00
    
    def test_generate_customer_statement_includes_all_transactions(self, sample_customer_data):
        """测试客户对账单包含所有交易记录"""
        generator = ReconciliationReportGenerator()
        wb = generator.generate_customer_statement_excel(sample_customer_data)
        ws = wb.active
        
        # 交易记录从第9行开始
        transaction_count = len(sample_customer_data.transactions)
        
        for i in range(transaction_count):
            row_num = 9 + i
            seq_num = ws.cell(row=row_num, column=1).value
            assert seq_num == i + 1
    
    def test_generate_customer_statement_calculates_running_balance(self, sample_customer_data):
        """测试客户对账单正确计算余额"""
        generator = ReconciliationReportGenerator()
        wb = generator.generate_customer_statement_excel(sample_customer_data)
        ws = wb.active
        
        # 第一笔交易（收入10000）
        balance1 = ws.cell(row=9, column=7).value
        assert balance1 == 10000.00
        
        # 第二笔交易（收入5000，余额应为15000）
        balance2 = ws.cell(row=10, column=7).value
        assert balance2 == 15000.00
        
        # 第三笔交易（支出3000，余额应为12000）
        balance3 = ws.cell(row=11, column=7).value
        assert balance3 == 12000.00
    
    def test_generate_customer_statement_separates_income_and_expense(self, sample_customer_data):
        """测试客户对账单正确分离收入和支出"""
        generator = ReconciliationReportGenerator()
        wb = generator.generate_customer_statement_excel(sample_customer_data)
        ws = wb.active
        
        # 第一笔交易是收入
        income1 = ws.cell(row=9, column=5).value
        expense1 = ws.cell(row=9, column=6).value
        assert income1 == 10000.00
        assert expense1 == ''
        
        # 第三笔交易是支出
        income3 = ws.cell(row=11, column=5).value
        expense3 = ws.cell(row=11, column=6).value
        assert income3 == ''
        assert expense3 == 3000.00
    
    def test_generate_customer_statement_includes_closing_balance(self, sample_customer_data):
        """测试客户对账单包含期末余额"""
        generator = ReconciliationReportGenerator()
        wb = generator.generate_customer_statement_excel(sample_customer_data)
        ws = wb.active
        
        # 期末余额在交易记录之后（第12行）
        closing_label = ws.cell(row=12, column=2).value
        assert closing_label == '期末余额'
        
        closing_balance = ws.cell(row=12, column=7).value
        assert closing_balance == 12000.00
    
    def test_generate_customer_statement_includes_summary(self, sample_customer_data):
        """测试客户对账单包含汇总统计"""
        generator = ReconciliationReportGenerator()
        wb = generator.generate_customer_statement_excel(sample_customer_data)
        ws = wb.active
        
        # 汇总在第14行
        summary_title = ws.cell(row=14, column=1).value
        assert summary_title == '对账汇总'
        
        # 统计数据在第15行
        transaction_count = ws.cell(row=15, column=2).value
        assert '3 笔' in transaction_count
        
        total_income = ws.cell(row=15, column=4).value
        assert total_income == 15000.00
        
        total_expense = ws.cell(row=15, column=6).value
        assert total_expense == 3000.00
    
    def test_generate_customer_statement_includes_signature_section(self, sample_customer_data):
        """测试客户对账单包含签字栏"""
        generator = ReconciliationReportGenerator()
        wb = generator.generate_customer_statement_excel(sample_customer_data)
        ws = wb.active
        
        # 签字栏在第18行
        signature_text = ws.cell(row=18, column=1).value
        assert '客户确认签字' in signature_text
        
        date_text = ws.cell(row=18, column=5).value
        assert '日期' in date_text
    
    def test_save_customer_statement_creates_file(self, temp_output_dir, sample_customer_data):
        """测试保存客户对账单创建文件"""
        generator = ReconciliationReportGenerator(output_dir=temp_output_dir)
        filepath = generator.save_customer_statement(
            sample_customer_data,
            filename='test_customer_statement.xlsx'
        )
        
        assert Path(filepath).exists()
        assert Path(filepath).suffix == '.xlsx'
        
        # 验证文件可以被打开
        wb = load_workbook(filepath)
        assert wb.active.title == "客户对账单"
    
    def test_save_customer_statement_auto_generates_filename(self, temp_output_dir, sample_customer_data):
        """测试保存客户对账单自动生成文件名"""
        generator = ReconciliationReportGenerator(output_dir=temp_output_dir)
        filepath = generator.save_customer_statement(sample_customer_data)
        
        assert Path(filepath).exists()
        assert '客户对账单_' in Path(filepath).name
        assert '北京ABC科技有限公司' in Path(filepath).name
        assert Path(filepath).suffix == '.xlsx'
    
    def test_generate_customer_statement_with_no_transactions(self):
        """测试生成无交易记录的客户对账单"""
        customer = Counterparty(
            id="C002",
            name="测试客户",
            type=CounterpartyType.CUSTOMER,
            contact_person="李四",
            phone="13900139000",
            email="lisi@test.com",
            address="上海市浦东新区",
            tax_id="91310000XXXXXXXX",
            created_at=datetime.now(),
            updated_at=datetime.now()
        )
        
        customer_data = CustomerAccountData(
            customer=customer,
            transactions=[],
            start_date=date(2024, 1, 1),
            end_date=date(2024, 1, 31),
            opening_balance=Decimal("1000.00"),
            closing_balance=Decimal("1000.00")
        )
        
        generator = ReconciliationReportGenerator()
        wb = generator.generate_customer_statement_excel(customer_data)
        ws = wb.active
        
        # 应该仍然有标题和客户信息
        assert ws['A1'].value == '客户对账单'
        assert ws['B3'].value == "测试客户"
        
        # 期初余额和期末余额应该相同
        opening_balance = ws.cell(row=8, column=7).value
        closing_balance = ws.cell(row=9, column=7).value
        assert opening_balance == 1000.00
        assert closing_balance == 1000.00


class TestReconciliationReportGeneratorEdgeCases:
    """测试边界情况"""
    
    def test_generate_discrepancy_report_with_large_amounts(self):
        """测试生成包含大金额的差异报告"""
        bank_record = BankRecord(
            id="B001",
            transaction_date=date(2024, 1, 15),
            description="大额收款",
            amount=Decimal("9999999.99"),
            balance=Decimal("10000000.00"),
            transaction_type="CREDIT",
            counterparty="大客户"
        )
        
        discrepancy = Discrepancy(
            id="DISC-0001",
            type=DiscrepancyType.MISSING_SYSTEM,
            bank_record=bank_record,
            system_record=None,
            difference_amount=Decimal("9999999.99"),
            description="大额差异"
        )
        
        generator = ReconciliationReportGenerator()
        wb = generator.generate_discrepancy_report([discrepancy])
        ws = wb.active
        
        # 验证大金额正确显示
        bank_amount = ws.cell(row=5, column=5).value
        assert bank_amount == 9999999.99
    
    def test_generate_customer_statement_with_negative_balance(self):
        """测试生成包含负余额的客户对账单"""
        customer = Counterparty(
            id="C001",
            name="测试客户",
            type=CounterpartyType.CUSTOMER,
            contact_person="王五",
            phone="13700137000",
            email="wangwu@test.com",
            address="广州市天河区",
            tax_id="91440000XXXXXXXX",
            created_at=datetime.now(),
            updated_at=datetime.now()
        )
        
        transactions = [
            TransactionRecord(
                id="T001",
                date=date(2024, 1, 5),
                type=TransactionType.EXPENSE,
                amount=Decimal("5000.00"),
                counterparty_id="C001",
                description="预付款",
                category="预付账款",
                status=TransactionStatus.COMPLETED,
                created_at=datetime.now(),
                updated_at=datetime.now()
            )
        ]
        
        customer_data = CustomerAccountData(
            customer=customer,
            transactions=transactions,
            start_date=date(2024, 1, 1),
            end_date=date(2024, 1, 31),
            opening_balance=Decimal("0.00"),
            closing_balance=Decimal("-5000.00")
        )
        
        generator = ReconciliationReportGenerator()
        wb = generator.generate_customer_statement_excel(customer_data)
        ws = wb.active
        
        # 验证负余额正确显示
        closing_balance = ws.cell(row=10, column=7).value
        assert closing_balance == -5000.00
    
    def test_output_directory_creation(self):
        """测试输出目录自动创建"""
        temp_dir = tempfile.mkdtemp()
        try:
            output_dir = Path(temp_dir) / "reports" / "reconciliation"
            
            # 目录不存在
            assert not output_dir.exists()
            
            # 创建生成器应该自动创建目录
            generator = ReconciliationReportGenerator(output_dir=str(output_dir))
            
            # 目录应该被创建
            assert output_dir.exists()
        finally:
            shutil.rmtree(temp_dir)
    
    def test_format_discrepancy_type_all_types(self):
        """测试所有差异类型的格式化"""
        generator = ReconciliationReportGenerator()
        
        assert generator._format_discrepancy_type(DiscrepancyType.AMOUNT_DIFF) == '金额差异'
        assert generator._format_discrepancy_type(DiscrepancyType.MISSING_SYSTEM) == '系统记录缺失'
        assert generator._format_discrepancy_type(DiscrepancyType.MISSING_BANK) == '银行流水缺失'
    
    def test_format_transaction_type_all_types(self):
        """测试所有交易类型的格式化"""
        generator = ReconciliationReportGenerator()
        
        assert generator._format_transaction_type(TransactionType.INCOME) == '收入'
        assert generator._format_transaction_type(TransactionType.EXPENSE) == '支出'
        assert generator._format_transaction_type(TransactionType.ORDER) == '订单'
