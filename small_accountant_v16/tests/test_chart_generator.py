"""
Unit tests for ChartGenerator module

Tests chart generation functionality for revenue comparison, profit trends,
and customer ranking charts.

Requirements: 1.1, 1.4
"""

import pytest
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib.figure import Figure
from openpyxl import Workbook

from small_accountant_v16.reports.chart_generator import ChartGenerator


class TestChartGenerator:
    """测试图表生成器"""
    
    @pytest.fixture
    def chart_generator(self):
        """创建图表生成器实例"""
        return ChartGenerator()
    
    @pytest.fixture
    def revenue_comparison_data(self):
        """收支对比测试数据"""
        return pd.DataFrame([
            {'period': '2024年1月', 'income': 100000.00, 'expense': 60000.00},
            {'period': '2024年2月', 'income': 120000.00, 'expense': 70000.00},
            {'period': '2024年3月', 'income': 110000.00, 'expense': 65000.00},
        ])
    
    @pytest.fixture
    def profit_trend_data(self):
        """利润趋势测试数据"""
        return pd.DataFrame([
            {
                'month': '2024-01',
                'revenue': 100000.00,
                'cost': 50000.00,
                'gross_profit': 50000.00,
                'net_profit': 40000.00
            },
            {
                'month': '2024-02',
                'revenue': 120000.00,
                'cost': 60000.00,
                'gross_profit': 60000.00,
                'net_profit': 48000.00
            },
        ])
    
    @pytest.fixture
    def customer_ranking_data(self):
        """客户排名测试数据"""
        return pd.DataFrame([
            {'customer_name': '客户A', 'sales_amount': 500000.00},
            {'customer_name': '客户B', 'sales_amount': 450000.00},
            {'customer_name': '客户C', 'sales_amount': 400000.00},
            {'customer_name': '客户D', 'sales_amount': 350000.00},
            {'customer_name': '客户E', 'sales_amount': 300000.00},
        ])
    
    def test_create_revenue_comparison_chart_returns_figure(
        self, chart_generator, revenue_comparison_data
    ):
        """测试创建收支对比图返回Figure对象"""
        fig = chart_generator.create_revenue_comparison_chart(revenue_comparison_data)
        assert isinstance(fig, Figure)
        assert len(fig.axes) == 1
        plt.close(fig)
    
    def test_create_revenue_comparison_chart_with_custom_title(
        self, chart_generator, revenue_comparison_data
    ):
        """测试创建收支对比图使用自定义标题"""
        custom_title = "2024年第一季度收支对比"
        fig = chart_generator.create_revenue_comparison_chart(
            revenue_comparison_data, 
            title=custom_title
        )
        ax = fig.axes[0]
        assert ax.get_title() == custom_title
        plt.close(fig)
    
    def test_create_revenue_comparison_chart_data_accuracy(
        self, chart_generator, revenue_comparison_data
    ):
        """测试收支对比图数据正确性"""
        fig = chart_generator.create_revenue_comparison_chart(revenue_comparison_data)
        ax = fig.axes[0]
        bars = ax.patches
        income_bars = bars[:3]
        expense_bars = bars[3:6]
        expected_income = revenue_comparison_data['income'].tolist()
        for i, bar in enumerate(income_bars):
            assert abs(bar.get_height() - expected_income[i]) < 0.01
        expected_expense = revenue_comparison_data['expense'].tolist()
        for i, bar in enumerate(expense_bars):
            assert abs(bar.get_height() - expected_expense[i]) < 0.01
        plt.close(fig)
    
    def test_create_revenue_comparison_chart_with_empty_data(self, chart_generator):
        """测试空数据情况"""
        empty_data = pd.DataFrame(columns=['period', 'income', 'expense'])
        fig = chart_generator.create_revenue_comparison_chart(empty_data)
        assert isinstance(fig, Figure)
        plt.close(fig)
    
    def test_create_profit_trend_chart_returns_figure(
        self, chart_generator, profit_trend_data
    ):
        """测试创建利润趋势图返回Figure对象"""
        fig = chart_generator.create_profit_trend_chart(profit_trend_data)
        assert isinstance(fig, Figure)
        assert len(fig.axes) == 1
        plt.close(fig)
    
    def test_create_profit_trend_chart_data_accuracy(
        self, chart_generator, profit_trend_data
    ):
        """测试利润趋势图数据正确性"""
        fig = chart_generator.create_profit_trend_chart(profit_trend_data)
        ax = fig.axes[0]
        lines = ax.get_lines()
        assert len(lines) == 4
        revenue_line = lines[0]
        revenue_ydata = revenue_line.get_ydata()
        expected_revenue = profit_trend_data['revenue'].tolist()
        for i, value in enumerate(expected_revenue):
            assert abs(revenue_ydata[i] - value) < 0.01
        plt.close(fig)
    
    def test_create_customer_ranking_chart_returns_figure(
        self, chart_generator, customer_ranking_data
    ):
        """测试创建客户排名图返回Figure对象"""
        fig = chart_generator.create_customer_ranking_chart(customer_ranking_data)
        assert isinstance(fig, Figure)
        assert len(fig.axes) == 1
        plt.close(fig)
    
    def test_create_customer_ranking_chart_top_n(
        self, chart_generator, customer_ranking_data
    ):
        """测试客户排名图显示前N名"""
        top_n = 3
        fig = chart_generator.create_customer_ranking_chart(
            customer_ranking_data,
            top_n=top_n
        )
        ax = fig.axes[0]
        bars = ax.patches
        assert len(bars) == top_n
        plt.close(fig)
    
    def test_save_chart_to_excel(
        self, chart_generator, revenue_comparison_data, tmp_path
    ):
        """测试将图表保存到Excel"""
        wb = Workbook()
        fig = chart_generator.create_revenue_comparison_chart(revenue_comparison_data)
        chart_generator.save_chart_to_excel(wb, "收支对比", fig, cell='A1')
        assert "收支对比" in wb.sheetnames
        ws = wb["收支对比"]
        assert len(ws._images) == 1
        output_file = tmp_path / "test_chart.xlsx"
        wb.save(output_file)
        assert output_file.exists()
    
    def test_create_and_embed_revenue_comparison(
        self, chart_generator, revenue_comparison_data, tmp_path
    ):
        """测试创建并嵌入收支对比图"""
        wb = Workbook()
        chart_generator.create_and_embed_revenue_comparison(
            wb, "管理报表", revenue_comparison_data, cell='H4'
        )
        assert "管理报表" in wb.sheetnames
        ws = wb["管理报表"]
        assert len(ws._images) == 1
        output_file = tmp_path / "revenue_comparison.xlsx"
        wb.save(output_file)
        assert output_file.exists()
    
    def test_chart_with_zero_values(self, chart_generator):
        """测试包含零值的数据"""
        data = pd.DataFrame([
            {'period': '2024年1月', 'income': 0.00, 'expense': 0.00},
            {'period': '2024年2月', 'income': 100000.00, 'expense': 50000.00},
        ])
        fig = chart_generator.create_revenue_comparison_chart(data)
        assert isinstance(fig, Figure)
        plt.close(fig)
    
    def test_chart_with_large_numbers(self, chart_generator):
        """测试大数值数据"""
        data = pd.DataFrame([
            {'period': '2024年1月', 'income': 10000000.00, 'expense': 8000000.00},
            {'period': '2024年2月', 'income': 12000000.00, 'expense': 9000000.00},
        ])
        fig = chart_generator.create_revenue_comparison_chart(data)
        assert isinstance(fig, Figure)
        plt.close(fig)
