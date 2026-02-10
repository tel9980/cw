"""
Unit tests for ReportGenerator

Tests the report generation functionality including:
- Management reports (revenue comparison, profit trend, customer ranking)
- Tax reports (VAT and income tax)
- Bank loan reports (balance sheet, income statement, cash flow)
"""

import pytest
from datetime import date, datetime, timedelta
from decimal import Decimal
import os
import tempfile
import shutil
from pathlib import Path

from small_accountant_v16.models.core_models import (
    TransactionRecord, TransactionType, TransactionStatus,
    Counterparty, CounterpartyType, ReportType
)
from small_accountant_v16.storage.transaction_storage import TransactionStorage
from small_accountant_v16.storage.counterparty_storage import CounterpartyStorage
from small_accountant_v16.reports.report_generator import ReportGenerator, TaxReportType


@pytest.fixture
def temp_storage_dir():
    """Create a temporary directory for storage"""
    temp_dir = tempfile.mkdtemp()
    yield temp_dir
    shutil.rmtree(temp_dir)


@pytest.fixture
def temp_output_dir():
    """Create a temporary directory for report output"""
    temp_dir = tempfile.mkdtemp()
    yield temp_dir
    shutil.rmtree(temp_dir)


@pytest.fixture
def transaction_storage(temp_storage_dir):
    """Create transaction storage with test data"""
    storage = TransactionStorage(temp_storage_dir)
    return storage


@pytest.fixture
def counterparty_storage(temp_storage_dir):
    """Create counterparty storage with test data"""
    storage = CounterpartyStorage(temp_storage_dir)
    return storage


@pytest.fixture
def sample_customers(counterparty_storage):
    """Create sample customer data"""
    customers = [
        Counterparty(
            id="C001",
            name="客户A",
            type=CounterpartyType.CUSTOMER,
            contact_person="张三",
            phone="13800138000",
            email="zhangsan@example.com",
            address="北京市朝阳区",
            tax_id="110000000000001",
            created_at=datetime.now(),
            updated_at=datetime.now()
        ),
        Counterparty(
            id="C002",
            name="客户B",
            type=CounterpartyType.CUSTOMER,
            contact_person="李四",
            phone="13800138001",
            email="lisi@example.com",
            address="上海市浦东新区",
            tax_id="310000000000002",
            created_at=datetime.now(),
            updated_at=datetime.now()
        ),
        Counterparty(
            id="C003",
            name="客户C",
            type=CounterpartyType.CUSTOMER,
            contact_person="王五",
            phone="13800138002",
            email="wangwu@example.com",
            address="广州市天河区",
            tax_id="440000000000003",
            created_at=datetime.now(),
            updated_at=datetime.now()
        )
    ]
    
    for customer in customers:
        counterparty_storage.add(customer)
    
    return customers


@pytest.fixture
def sample_suppliers(counterparty_storage):
    """Create sample supplier data"""
    suppliers = [
        Counterparty(
            id="S001",
            name="供应商A",
            type=CounterpartyType.SUPPLIER,
            contact_person="赵六",
            phone="13800138003",
            email="zhaoliu@example.com",
            address="深圳市南山区",
            tax_id="440300000000004",
            created_at=datetime.now(),
            updated_at=datetime.now()
        )
    ]
    
    for supplier in suppliers:
        counterparty_storage.add(supplier)
    
    return suppliers


@pytest.fixture
def sample_transactions(transaction_storage, sample_customers, sample_suppliers):
    """Create sample transaction data"""
    base_date = date(2024, 1, 1)
    transactions = []
    
    # Create transactions for 3 months
    for month in range(3):
        month_start = date(2024, month + 1, 1)
        
        # Income transactions
        for i in range(5):
            trans = TransactionRecord(
                id=f"T{month:02d}{i:02d}I",
                date=month_start + timedelta(days=i * 5),
                type=TransactionType.INCOME,
                amount=Decimal(str(10000 + i * 1000 + month * 5000)),
                counterparty_id=sample_customers[i % len(sample_customers)].id,
                description=f"销售收入 {month + 1}月",
                category="销售",
                status=TransactionStatus.COMPLETED,
                created_at=datetime.now(),
                updated_at=datetime.now()
            )
            transactions.append(trans)
            transaction_storage.add(trans)
        
        # Expense transactions
        for i in range(3):
            trans = TransactionRecord(
                id=f"T{month:02d}{i:02d}E",
                date=month_start + timedelta(days=i * 7),
                type=TransactionType.EXPENSE,
                amount=Decimal(str(5000 + i * 500 + month * 2000)),
                counterparty_id=sample_suppliers[0].id,
                description=f"采购支出 {month + 1}月",
                category="采购",
                status=TransactionStatus.COMPLETED,
                created_at=datetime.now(),
                updated_at=datetime.now()
            )
            transactions.append(trans)
            transaction_storage.add(trans)
    
    return transactions


@pytest.fixture
def report_generator(transaction_storage, counterparty_storage, temp_output_dir):
    """Create report generator instance"""
    return ReportGenerator(
        transaction_storage=transaction_storage,
        counterparty_storage=counterparty_storage,
        output_dir=temp_output_dir
    )



class TestReportGeneratorInitialization:
    """Test report generator initialization"""
    
    def test_initialization(self, report_generator, temp_output_dir):
        """Test that report generator initializes correctly"""
        assert report_generator.transaction_storage is not None
        assert report_generator.counterparty_storage is not None
        assert report_generator.output_dir == temp_output_dir
        assert report_generator.template_manager is not None
        assert report_generator.chart_generator is not None
        assert os.path.exists(temp_output_dir)


class TestManagementReportGeneration:
    """Test management report generation"""
    
    def test_generate_management_report_success(
        self, 
        report_generator, 
        sample_transactions,
        temp_output_dir
    ):
        """Test successful management report generation"""
        start_date = date(2024, 1, 1)
        end_date = date(2024, 3, 31)
        
        result = report_generator.generate_management_report(
            start_date=start_date,
            end_date=end_date,
            company_name="测试公司"
        )
        
        # Verify result
        assert result.success is True
        assert result.report_type == ReportType.MANAGEMENT
        assert result.error_message is None
        assert result.file_path != ""
        assert os.path.exists(result.file_path)
        assert result.file_path.endswith('.xlsx')
        
        # Verify file was created
        assert Path(result.file_path).stat().st_size > 0
    
    def test_generate_management_report_no_data(
        self, 
        report_generator
    ):
        """Test management report generation with no data"""
        start_date = date(2025, 1, 1)
        end_date = date(2025, 3, 31)
        
        result = report_generator.generate_management_report(
            start_date=start_date,
            end_date=end_date
        )
        
        # Should fail gracefully
        assert result.success is False
        assert result.error_message is not None
        assert "没有交易数据" in result.error_message
    
    def test_management_report_contains_all_sections(
        self,
        report_generator,
        sample_transactions
    ):
        """Test that management report contains all required sections"""
        start_date = date(2024, 1, 1)
        end_date = date(2024, 3, 31)
        
        result = report_generator.generate_management_report(
            start_date=start_date,
            end_date=end_date
        )
        
        assert result.success is True
        
        # Verify the Excel file has the expected sheets
        import openpyxl
        wb = openpyxl.load_workbook(result.file_path)
        sheet_names = wb.sheetnames
        
        # Should contain revenue comparison, profit trend, and customer ranking
        assert "收支对比表" in sheet_names
        assert "利润趋势表" in sheet_names
        assert "客户排名表" in sheet_names


class TestTaxReportGeneration:
    """Test tax report generation"""
    
    def test_generate_vat_report_success(
        self,
        report_generator,
        sample_transactions
    ):
        """Test successful VAT report generation"""
        result = report_generator.generate_tax_report(
            report_type=TaxReportType.VAT,
            period="2024年第一季度",
            company_name="测试公司"
        )
        
        # Verify result
        assert result.success is True
        assert result.report_type == ReportType.TAX_VAT
        assert result.error_message is None
        assert result.file_path != ""
        assert os.path.exists(result.file_path)
        assert "增值税申报表" in result.file_path
    
    def test_generate_income_tax_report_success(
        self,
        report_generator,
        sample_transactions
    ):
        """Test successful income tax report generation"""
        result = report_generator.generate_tax_report(
            report_type=TaxReportType.INCOME_TAX,
            period="2024年第一季度",
            company_name="测试公司"
        )
        
        # Verify result
        assert result.success is True
        assert result.report_type == ReportType.TAX_INCOME
        assert result.error_message is None
        assert result.file_path != ""
        assert os.path.exists(result.file_path)
        assert "所得税申报表" in result.file_path
    
    def test_generate_tax_report_invalid_type(
        self,
        report_generator,
        sample_transactions
    ):
        """Test tax report generation with invalid type"""
        result = report_generator.generate_tax_report(
            report_type="invalid_type",
            period="2024年第一季度"
        )
        
        # Should fail gracefully
        assert result.success is False
        assert result.error_message is not None
    
    def test_parse_period_quarterly(self, report_generator):
        """Test period parsing for quarterly reports"""
        start, end = report_generator._parse_period("2024年第一季度")
        assert start == date(2024, 1, 1)
        assert end == date(2024, 3, 31)
        
        start, end = report_generator._parse_period("2024年第二季度")
        assert start == date(2024, 4, 1)
        assert end == date(2024, 6, 30)
    
    def test_parse_period_monthly(self, report_generator):
        """Test period parsing for monthly reports"""
        start, end = report_generator._parse_period("2024年1月")
        assert start == date(2024, 1, 1)
        assert end == date(2024, 1, 31)
        
        start, end = report_generator._parse_period("2024年2月")
        assert start == date(2024, 2, 1)
        assert end == date(2024, 2, 29)  # 2024 is a leap year


class TestBankLoanReportGeneration:
    """Test bank loan report generation"""
    
    def test_generate_bank_loan_report_success(
        self,
        report_generator,
        sample_transactions
    ):
        """Test successful bank loan report generation"""
        report_date = date(2024, 3, 31)
        
        result = report_generator.generate_bank_loan_report(
            report_date=report_date,
            company_name="测试公司"
        )
        
        # Verify result
        assert result.success is True
        assert result.report_type == ReportType.BANK_LOAN
        assert result.error_message is None
        assert result.file_path != ""
        assert os.path.exists(result.file_path)
        assert "银行贷款报表" in result.file_path
    
    def test_bank_loan_report_contains_all_statements(
        self,
        report_generator,
        sample_transactions
    ):
        """Test that bank loan report contains all three financial statements"""
        report_date = date(2024, 3, 31)
        
        result = report_generator.generate_bank_loan_report(
            report_date=report_date
        )
        
        assert result.success is True
        
        # Verify the Excel file has the expected sheets
        import openpyxl
        wb = openpyxl.load_workbook(result.file_path)
        sheet_names = wb.sheetnames
        
        # Should contain balance sheet, income statement, and cash flow statement
        assert "资产负债表" in sheet_names
        assert "利润表" in sheet_names
        assert "现金流量表" in sheet_names


class TestDataPreparation:
    """Test data preparation methods"""
    
    def test_prepare_revenue_comparison_data(
        self,
        report_generator,
        sample_transactions
    ):
        """Test revenue comparison data preparation"""
        data = report_generator._prepare_revenue_comparison_data(
            sample_transactions,
            date(2024, 1, 1),
            date(2024, 3, 31)
        )
        
        # Should have data for 3 months
        assert len(data) == 3
        assert 'period' in data.columns
        assert 'income' in data.columns
        assert 'expense' in data.columns
        assert 'profit' in data.columns
        assert 'profit_rate' in data.columns
        
        # All values should be numeric
        assert data['income'].dtype in ['float64', 'int64']
        assert data['expense'].dtype in ['float64', 'int64']
    
    def test_prepare_profit_trend_data(
        self,
        report_generator,
        sample_transactions
    ):
        """Test profit trend data preparation"""
        data = report_generator._prepare_profit_trend_data(
            sample_transactions,
            date(2024, 1, 1),
            date(2024, 3, 31)
        )
        
        # Should have data for 3 months
        assert len(data) == 3
        assert 'month' in data.columns
        assert 'revenue' in data.columns
        assert 'cost' in data.columns
        assert 'gross_profit' in data.columns
        assert 'net_profit' in data.columns
        assert 'growth_rate' in data.columns
    
    def test_prepare_customer_ranking_data(
        self,
        report_generator,
        sample_transactions,
        sample_customers
    ):
        """Test customer ranking data preparation"""
        data = report_generator._prepare_customer_ranking_data(
            sample_transactions
        )
        
        # Should have customer data
        assert len(data) > 0
        assert 'rank' in data.columns
        assert 'customer_name' in data.columns
        assert 'sales_amount' in data.columns
        assert 'transaction_count' in data.columns
        assert 'percentage' in data.columns
        
        # Rankings should be in order
        assert data['rank'].tolist() == list(range(1, len(data) + 1))
        
        # Sales amounts should be in descending order
        sales = data['sales_amount'].tolist()
        assert sales == sorted(sales, reverse=True)
    
    def test_prepare_vat_data(
        self,
        report_generator,
        sample_transactions
    ):
        """Test VAT data preparation"""
        data = report_generator._prepare_vat_data(sample_transactions)
        
        # Should have VAT calculation data
        assert len(data) > 0
        
        # Check that data contains expected fields
        # (structure may vary based on implementation)
        assert data is not None
    
    def test_prepare_income_tax_data(
        self,
        report_generator,
        sample_transactions
    ):
        """Test income tax data preparation"""
        data = report_generator._prepare_income_tax_data(sample_transactions)
        
        # Should have income tax calculation data
        assert len(data) > 0
        assert 'item' in data.columns
        assert 'amount' in data.columns


class TestEdgeCases:
    """Test edge cases and error handling"""
    
    def test_empty_transactions(
        self,
        transaction_storage,
        counterparty_storage,
        temp_output_dir
    ):
        """Test report generation with no transactions"""
        generator = ReportGenerator(
            transaction_storage=transaction_storage,
            counterparty_storage=counterparty_storage,
            output_dir=temp_output_dir
        )
        
        result = generator.generate_management_report(
            start_date=date(2024, 1, 1),
            end_date=date(2024, 3, 31)
        )
        
        assert result.success is False
        assert result.error_message is not None
    
    def test_single_transaction(
        self,
        transaction_storage,
        counterparty_storage,
        sample_customers,
        temp_output_dir
    ):
        """Test report generation with a single transaction"""
        # Add one transaction
        trans = TransactionRecord(
            id="T001",
            date=date(2024, 1, 15),
            type=TransactionType.INCOME,
            amount=Decimal("10000.00"),
            counterparty_id=sample_customers[0].id,
            description="单笔收入",
            category="销售",
            status=TransactionStatus.COMPLETED,
            created_at=datetime.now(),
            updated_at=datetime.now()
        )
        transaction_storage.add(trans)
        
        generator = ReportGenerator(
            transaction_storage=transaction_storage,
            counterparty_storage=counterparty_storage,
            output_dir=temp_output_dir
        )
        
        result = generator.generate_management_report(
            start_date=date(2024, 1, 1),
            end_date=date(2024, 1, 31)
        )
        
        # Should succeed with single transaction
        assert result.success is True
        assert os.path.exists(result.file_path)
    
    def test_date_range_boundary(
        self,
        report_generator,
        sample_transactions
    ):
        """Test report generation with exact date boundaries"""
        # Test with exact transaction dates
        start_date = date(2024, 1, 1)
        end_date = date(2024, 1, 31)
        
        result = report_generator.generate_management_report(
            start_date=start_date,
            end_date=end_date
        )
        
        assert result.success is True
        assert result.data_period.start_date == start_date
        assert result.data_period.end_date == end_date


class TestReportFileNaming:
    """Test report file naming conventions"""
    
    def test_management_report_filename(
        self,
        report_generator,
        sample_transactions
    ):
        """Test management report filename format"""
        result = report_generator.generate_management_report(
            start_date=date(2024, 1, 1),
            end_date=date(2024, 3, 31)
        )
        
        filename = os.path.basename(result.file_path)
        assert filename.startswith("管理报表_")
        assert filename.endswith(".xlsx")
        assert "20240101" in filename
        assert "20240331" in filename
    
    def test_vat_report_filename(
        self,
        report_generator,
        sample_transactions
    ):
        """Test VAT report filename format"""
        result = report_generator.generate_tax_report(
            report_type=TaxReportType.VAT,
            period="2024年第一季度"
        )
        
        filename = os.path.basename(result.file_path)
        assert filename.startswith("增值税申报表_")
        assert filename.endswith(".xlsx")
