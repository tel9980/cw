"""
报表生成器模块

提供智能报表生成功能，包括管理报表、税务报表和银行贷款报表。
"""

from datetime import date, datetime
from decimal import Decimal
from typing import List, Optional, Dict
import pandas as pd
import os
import logging

from ..models.core_models import (
    ReportResult, ReportType, DateRange, TransactionType,
    TransactionRecord, Counterparty, CounterpartyType, TransactionStatus
)
from ..storage.transaction_storage import TransactionStorage
from ..storage.counterparty_storage import CounterpartyStorage
from .report_template import ReportTemplate, TemplateType
from .chart_generator import ChartGenerator

logger = logging.getLogger(__name__)


class TaxReportType:
    """税务报表类型"""
    VAT = "vat"  # 增值税
    INCOME_TAX = "income_tax"  # 所得税


class ReportGenerator:
    """智能报表生成器
    
    生成各类财务报表，包括：
    - 管理报表（收支对比、利润趋势、客户排名）
    - 税务报表（增值税、所得税申报表）
    - 银行贷款报表（资产负债表、利润表、现金流量表）
    """
    
    def __init__(
        self,
        transaction_storage: TransactionStorage,
        counterparty_storage: CounterpartyStorage,
        output_dir: str = "reports_output"
    ):
        """初始化报表生成器
        
        Args:
            transaction_storage: 交易记录存储
            counterparty_storage: 往来单位存储
            output_dir: 报表输出目录
        """
        self.transaction_storage = transaction_storage
        self.counterparty_storage = counterparty_storage
        self.output_dir = output_dir
        self.template_manager = ReportTemplate()
        self.chart_generator = ChartGenerator()
        
        # 确保输出目录存在
        os.makedirs(output_dir, exist_ok=True)
        
        logger.info(f"报表生成器初始化完成，输出目录: {output_dir}")
    
    def generate_management_report(
        self,
        start_date: date,
        end_date: date,
        company_name: str = "公司"
    ) -> ReportResult:
        """生成管理报表（收支对比、利润趋势、客户排名）
        
        Args:
            start_date: 开始日期
            end_date: 结束日期
            company_name: 公司名称
            
        Returns:
            ReportResult: 报表生成结果
        """
        try:
            logger.info(f"开始生成管理报表: {start_date} 至 {end_date}")
            
            # 获取交易数据
            transactions = self.transaction_storage.get_by_date_range(start_date, end_date)
            
            if not transactions:
                logger.warning("没有找到交易数据")
                return ReportResult(
                    report_type=ReportType.MANAGEMENT,
                    file_path="",
                    generation_date=datetime.now(),
                    data_period=DateRange(start_date, end_date),
                    success=False,
                    error_message="指定期间内没有交易数据"
                )
            
            # 生成输出文件名
            filename = f"管理报表_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
            output_file = os.path.join(self.output_dir, filename)
            
            # 准备收支对比数据
            revenue_data = self._prepare_revenue_comparison_data(transactions, start_date, end_date)
            
            # 准备利润趋势数据
            profit_data = self._prepare_profit_trend_data(transactions, start_date, end_date)
            
            # 准备客户排名数据
            customer_data = self._prepare_customer_ranking_data(transactions)
            
            # 创建工作簿
            import openpyxl
            wb = openpyxl.Workbook()
            
            # 删除默认的Sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            period_str = f"{start_date.strftime('%Y-%m-%d')} 至 {end_date.strftime('%Y-%m-%d')}"
            
            # 1. 收支对比表
            template = self.template_manager.load_template(TemplateType.REVENUE_COMPARISON)
            ws = wb.create_sheet("收支对比表")
            wb.active = ws
            self._apply_template_to_sheet(
                ws, template, revenue_data,
                company_name=company_name,
                period=period_str
            )
            
            # 嵌入收支对比图
            self.chart_generator.create_and_embed_revenue_comparison(
                wb, "收支对比表", revenue_data, cell='H4'
            )
            
            # 2. 利润趋势表
            template = self.template_manager.load_template(TemplateType.PROFIT_TREND)
            ws = wb.create_sheet("利润趋势表")
            self._apply_template_to_sheet(
                ws, template, profit_data,
                company_name=company_name,
                period=period_str
            )
            
            # 嵌入利润趋势图
            self.chart_generator.create_and_embed_profit_trend(
                wb, "利润趋势表", profit_data, cell='H4'
            )
            
            # 3. 客户排名表
            template = self.template_manager.load_template(TemplateType.CUSTOMER_RANKING)
            ws = wb.create_sheet("客户排名表")
            self._apply_template_to_sheet(
                ws, template, customer_data,
                company_name=company_name,
                period=period_str
            )
            
            # 嵌入客户排名图
            self.chart_generator.create_and_embed_customer_ranking(
                wb, "客户排名表", customer_data, cell='H4', top_n=10
            )
            
            # 保存工作簿
            wb.save(output_file)
            
            logger.info(f"管理报表生成成功: {output_file}")
            
            return ReportResult(
                report_type=ReportType.MANAGEMENT,
                file_path=output_file,
                generation_date=datetime.now(),
                data_period=DateRange(start_date, end_date),
                success=True
            )
            
        except Exception as e:
            logger.error(f"生成管理报表失败: {str(e)}", exc_info=True)
            return ReportResult(
                report_type=ReportType.MANAGEMENT,
                file_path="",
                generation_date=datetime.now(),
                data_period=DateRange(start_date, end_date),
                success=False,
                error_message=f"生成报表时发生错误: {str(e)}"
            )
    
    def generate_tax_report(
        self,
        report_type: str,
        period: str,
        company_name: str = "公司"
    ) -> ReportResult:
        """生成税务报表（增值税、所得税申报表）
        
        Args:
            report_type: 税务报表类型 ("vat" 或 "income_tax")
            period: 报表期间（如"2024年第一季度"）
            company_name: 公司名称
            
        Returns:
            ReportResult: 报表生成结果
        """
        try:
            logger.info(f"开始生成税务报表: {report_type}, 期间: {period}")
            
            # 解析期间（简单实现，假设period格式为"2024年第一季度"或"2024年1月"）
            # 实际应用中需要更复杂的日期解析
            start_date, end_date = self._parse_period(period)
            
            # 获取交易数据
            transactions = self.transaction_storage.get_by_date_range(start_date, end_date)
            
            if report_type == TaxReportType.VAT:
                return self._generate_vat_report(transactions, period, company_name, start_date, end_date)
            elif report_type == TaxReportType.INCOME_TAX:
                return self._generate_income_tax_report(transactions, period, company_name, start_date, end_date)
            else:
                raise ValueError(f"不支持的税务报表类型: {report_type}")
                
        except Exception as e:
            logger.error(f"生成税务报表失败: {str(e)}", exc_info=True)
            return ReportResult(
                report_type=ReportType.TAX_VAT if report_type == TaxReportType.VAT else ReportType.TAX_INCOME,
                file_path="",
                generation_date=datetime.now(),
                data_period=DateRange(start_date, end_date),
                success=False,
                error_message=f"生成税务报表时发生错误: {str(e)}"
            )
    
    def generate_bank_loan_report(
        self,
        report_date: date,
        company_name: str = "公司"
    ) -> ReportResult:
        """生成银行贷款报表（资产负债表、利润表、现金流量表）
        
        Args:
            report_date: 报表日期
            company_name: 公司名称
            
        Returns:
            ReportResult: 报表生成结果
        """
        try:
            logger.info(f"开始生成银行贷款报表: {report_date}")
            
            # 获取所有交易数据（用于计算财务状况）
            all_transactions = self.transaction_storage.get_all()
            
            # 生成输出文件名
            filename = f"银行贷款报表_{report_date.strftime('%Y%m%d')}.xlsx"
            output_file = os.path.join(self.output_dir, filename)
            
            # 准备资产负债表数据
            balance_sheet_data = self._prepare_balance_sheet_data(all_transactions, report_date)
            
            # 准备利润表数据
            income_statement_data = self._prepare_income_statement_data(all_transactions, report_date)
            
            # 准备现金流量表数据
            cash_flow_data = self._prepare_cash_flow_statement_data(all_transactions, report_date)
            
            # 创建工作簿
            import openpyxl
            wb = openpyxl.Workbook()
            
            # 删除默认的Sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            period_str = f"截至 {report_date.strftime('%Y年%m月%d日')}"
            
            # 1. 资产负债表
            template = self.template_manager.load_template(TemplateType.BALANCE_SHEET)
            ws = wb.create_sheet("资产负债表")
            wb.active = ws
            self._apply_template_to_sheet(
                ws, template, balance_sheet_data,
                company_name=company_name,
                period=period_str
            )
            
            # 2. 利润表
            template = self.template_manager.load_template(TemplateType.INCOME_STATEMENT)
            ws = wb.create_sheet("利润表")
            self._apply_template_to_sheet(
                ws, template, income_statement_data,
                company_name=company_name,
                period=period_str
            )
            
            # 3. 现金流量表
            template = self.template_manager.load_template(TemplateType.CASH_FLOW_STATEMENT)
            ws = wb.create_sheet("现金流量表")
            self._apply_template_to_sheet(
                ws, template, cash_flow_data,
                company_name=company_name,
                period=period_str
            )
            
            # 保存工作簿
            wb.save(output_file)
            
            logger.info(f"银行贷款报表生成成功: {output_file}")
            
            return ReportResult(
                report_type=ReportType.BANK_LOAN,
                file_path=output_file,
                generation_date=datetime.now(),
                data_period=DateRange(report_date, report_date),
                success=True
            )
            
        except Exception as e:
            logger.error(f"生成银行贷款报表失败: {str(e)}", exc_info=True)
            return ReportResult(
                report_type=ReportType.BANK_LOAN,
                file_path="",
                generation_date=datetime.now(),
                data_period=DateRange(report_date, report_date),
                success=False,
                error_message=f"生成银行贷款报表时发生错误: {str(e)}"
            )
    
    # ========== 数据准备辅助方法 ==========
    
    def _apply_template_to_sheet(
        self,
        ws,
        template,
        data: pd.DataFrame,
        **kwargs
    ):
        """将模板应用到现有工作表
        
        Args:
            ws: 工作表对象
            template: 报表模板
            data: 数据DataFrame
            **kwargs: 额外参数
        """
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        # 设置工作表标题
        ws.title = template.name
        
        # 写入表头
        current_row = 1
        
        # 标题
        if template.title:
            title = template.title
            if 'company_name' in kwargs:
                title = f"{kwargs['company_name']} - {title}"
            
            ws.merge_cells(f'A{current_row}:F{current_row}')
            cell = ws[f'A{current_row}']
            cell.value = title
            cell.font = Font(size=16, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
        
        # 副标题（报表期间）
        if 'period' in kwargs:
            ws.merge_cells(f'A{current_row}:F{current_row}')
            cell = ws[f'A{current_row}']
            cell.value = f"报表期间: {kwargs['period']}"
            cell.font = Font(size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
        
        current_row += 1  # 空行
        
        # 写入各个区域
        for section in template.sections:
            current_row = self._write_section_to_sheet(
                ws, section, data, current_row
            )
            current_row += 2  # 区域间隔
        
        # 应用样式
        self._apply_styles_to_sheet(ws)
    
    def _write_section_to_sheet(self, ws, section, data, start_row):
        """写入报表区域到工作表"""
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        
        current_row = start_row
        
        # 区域标题
        if section.title:
            ws.merge_cells(f'A{current_row}:F{current_row}')
            cell = ws[f'A{current_row}']
            cell.value = section.title
            cell.font = Font(size=12, bold=True)
            cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
            current_row += 1
        
        # 列标题
        for col_idx, column in enumerate(section.columns, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = column.name
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 设置列宽
            ws.column_dimensions[get_column_letter(col_idx)].width = column.width
        
        current_row += 1
        
        # 数据行
        data_start_row = current_row
        for _, row_data in data.iterrows():
            for col_idx, column in enumerate(section.columns, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                
                # 获取数据值
                if column.data_field and column.data_field in row_data:
                    value = row_data[column.data_field]
                    cell.value = value
                    
                    # 应用格式
                    if column.format:
                        cell.number_format = column.format
                
                cell.alignment = Alignment(horizontal='left', vertical='center')
            
            current_row += 1
        
        # 合计行
        if section.has_total and len(data) > 0:
            for col_idx, column in enumerate(section.columns, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                
                if col_idx == 1:
                    cell.value = "合计"
                    cell.font = Font(bold=True)
                elif column.data_field in section.total_columns:
                    # 计算合计
                    col_letter = get_column_letter(col_idx)
                    formula = f"=SUM({col_letter}{data_start_row}:{col_letter}{current_row-1})"
                    cell.value = formula
                    cell.font = Font(bold=True)
                    
                    if column.format:
                        cell.number_format = column.format
                
                cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
            
            current_row += 1
        
        return current_row
    
    def _apply_styles_to_sheet(self, ws):
        """应用样式到工作表"""
        from openpyxl.styles import Border, Side
        
        # 添加边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.border = thin_border
    
    # ========== 数据准备辅助方法 ==========
    
    def _prepare_revenue_comparison_data(
        self,
        transactions: List[TransactionRecord],
        start_date: date,
        end_date: date
    ) -> pd.DataFrame:
        """准备收支对比数据
        
        Args:
            transactions: 交易记录列表
            start_date: 开始日期
            end_date: 结束日期
            
        Returns:
            DataFrame with columns: period, income, expense, profit, profit_rate
        """
        # 按月分组统计
        monthly_data = {}
        
        for trans in transactions:
            month_key = trans.date.strftime('%Y年%m月')
            
            if month_key not in monthly_data:
                monthly_data[month_key] = {
                    'income': Decimal('0'),
                    'expense': Decimal('0')
                }
            
            if trans.type == TransactionType.INCOME:
                monthly_data[month_key]['income'] += trans.amount
            elif trans.type == TransactionType.EXPENSE:
                monthly_data[month_key]['expense'] += trans.amount
        
        # 转换为DataFrame
        data_list = []
        for period, amounts in sorted(monthly_data.items()):
            income = float(amounts['income'])
            expense = float(amounts['expense'])
            profit = income - expense
            profit_rate = (profit / income * 100) if income > 0 else 0
            
            data_list.append({
                'period': period,
                'income': income,
                'expense': expense,
                'profit': profit,
                'profit_rate': profit_rate / 100  # 转换为小数形式
            })
        
        return pd.DataFrame(data_list)
    
    def _prepare_profit_trend_data(
        self,
        transactions: List[TransactionRecord],
        start_date: date,
        end_date: date
    ) -> pd.DataFrame:
        """准备利润趋势数据
        
        Args:
            transactions: 交易记录列表
            start_date: 开始日期
            end_date: 结束日期
            
        Returns:
            DataFrame with columns: month, revenue, cost, gross_profit, net_profit, growth_rate
        """
        # 按月分组统计
        monthly_data = {}
        
        for trans in transactions:
            month_key = trans.date.strftime('%Y-%m')
            
            if month_key not in monthly_data:
                monthly_data[month_key] = {
                    'revenue': Decimal('0'),
                    'cost': Decimal('0')
                }
            
            if trans.type == TransactionType.INCOME:
                monthly_data[month_key]['revenue'] += trans.amount
            elif trans.type == TransactionType.EXPENSE:
                monthly_data[month_key]['cost'] += trans.amount
        
        # 转换为DataFrame并计算利润
        data_list = []
        prev_profit = None
        
        for month, amounts in sorted(monthly_data.items()):
            revenue = float(amounts['revenue'])
            cost = float(amounts['cost'])
            gross_profit = revenue - cost
            net_profit = gross_profit  # 简化处理，实际应考虑税费等
            
            # 计算环比增长率
            if prev_profit is not None and prev_profit != 0:
                growth_rate = (net_profit - prev_profit) / prev_profit
            else:
                growth_rate = 0
            
            data_list.append({
                'month': month,
                'revenue': revenue,
                'cost': cost,
                'gross_profit': gross_profit,
                'net_profit': net_profit,
                'growth_rate': growth_rate
            })
            
            prev_profit = net_profit
        
        return pd.DataFrame(data_list)
    
    def _prepare_customer_ranking_data(
        self,
        transactions: List[TransactionRecord]
    ) -> pd.DataFrame:
        """准备客户排名数据
        
        Args:
            transactions: 交易记录列表
            
        Returns:
            DataFrame with columns: rank, customer_name, sales_amount, transaction_count, percentage, customer_type
        """
        # 统计每个客户的销售额
        customer_sales = {}
        
        for trans in transactions:
            if trans.type == TransactionType.INCOME:
                counterparty = self.counterparty_storage.get(trans.counterparty_id)
                if counterparty and counterparty.type == CounterpartyType.CUSTOMER:
                    if counterparty.id not in customer_sales:
                        customer_sales[counterparty.id] = {
                            'name': counterparty.name,
                            'amount': Decimal('0'),
                            'count': 0
                        }
                    customer_sales[counterparty.id]['amount'] += trans.amount
                    customer_sales[counterparty.id]['count'] += 1
        
        # 计算总销售额
        total_sales = sum(data['amount'] for data in customer_sales.values())
        
        # 转换为列表并排序
        customer_list = []
        for customer_id, data in customer_sales.items():
            amount = float(data['amount'])
            percentage = (amount / float(total_sales)) if total_sales > 0 else 0
            
            customer_list.append({
                'customer_name': data['name'],
                'sales_amount': amount,
                'transaction_count': data['count'],
                'percentage': percentage,
                'customer_type': '客户'
            })
        
        # 按销售额排序
        customer_list.sort(key=lambda x: x['sales_amount'], reverse=True)
        
        # 添加排名
        for i, customer in enumerate(customer_list, start=1):
            customer['rank'] = i
        
        return pd.DataFrame(customer_list)
    
    def _generate_vat_report(
        self,
        transactions: List[TransactionRecord],
        period: str,
        company_name: str,
        start_date: date,
        end_date: date
    ) -> ReportResult:
        """生成增值税申报表
        
        Args:
            transactions: 交易记录列表
            period: 报表期间
            company_name: 公司名称
            start_date: 开始日期
            end_date: 结束日期
            
        Returns:
            ReportResult: 报表生成结果
        """
        # 生成输出文件名
        filename = f"增值税申报表_{period.replace('年', '').replace('月', '').replace('第', '').replace('季度', 'Q')}.xlsx"
        output_file = os.path.join(self.output_dir, filename)
        
        # 准备增值税数据
        vat_data = self._prepare_vat_data(transactions)
        
        # 加载模板并生成报表
        template = self.template_manager.load_template(TemplateType.VAT_DECLARATION)
        wb = self.template_manager.apply_template(
            template,
            vat_data,
            output_file,
            company_name=company_name,
            period=period
        )
        
        # 保存工作簿
        wb.save(output_file)
        
        logger.info(f"增值税申报表生成成功: {output_file}")
        
        return ReportResult(
            report_type=ReportType.TAX_VAT,
            file_path=output_file,
            generation_date=datetime.now(),
            data_period=DateRange(start_date, end_date),
            success=True
        )
    
    def _generate_income_tax_report(
        self,
        transactions: List[TransactionRecord],
        period: str,
        company_name: str,
        start_date: date,
        end_date: date
    ) -> ReportResult:
        """生成所得税申报表
        
        Args:
            transactions: 交易记录列表
            period: 报表期间
            company_name: 公司名称
            start_date: 开始日期
            end_date: 结束日期
            
        Returns:
            ReportResult: 报表生成结果
        """
        # 生成输出文件名
        filename = f"所得税申报表_{period.replace('年', '').replace('月', '').replace('第', '').replace('季度', 'Q')}.xlsx"
        output_file = os.path.join(self.output_dir, filename)
        
        # 准备所得税数据
        income_tax_data = self._prepare_income_tax_data(transactions)
        
        # 加载模板并生成报表
        template = self.template_manager.load_template(TemplateType.INCOME_TAX_DECLARATION)
        wb = self.template_manager.apply_template(
            template,
            income_tax_data,
            output_file,
            company_name=company_name,
            period=period
        )
        
        # 保存工作簿
        wb.save(output_file)
        
        logger.info(f"所得税申报表生成成功: {output_file}")
        
        return ReportResult(
            report_type=ReportType.TAX_INCOME,
            file_path=output_file,
            generation_date=datetime.now(),
            data_period=DateRange(start_date, end_date),
            success=True
        )
    
    def _prepare_vat_data(self, transactions: List[TransactionRecord]) -> pd.DataFrame:
        """准备增值税数据
        
        Args:
            transactions: 交易记录列表
            
        Returns:
            DataFrame: 增值税数据
        """
        # 计算销项税额和进项税额
        # 假设增值税率为13%（一般纳税人）
        vat_rate = Decimal('0.13')
        
        total_sales = sum(
            (t.amount for t in transactions if t.type == TransactionType.INCOME),
            Decimal('0')
        )
        total_purchases = sum(
            (t.amount for t in transactions if t.type == TransactionType.EXPENSE),
            Decimal('0')
        )
        
        # 计算不含税金额和税额
        sales_amount = total_sales / (Decimal('1') + vat_rate)
        sales_tax = sales_amount * vat_rate
        
        purchase_amount = total_purchases / (Decimal('1') + vat_rate)
        purchase_tax = purchase_amount * vat_rate
        
        # 应纳税额
        tax_payable = sales_tax - purchase_tax
        
        # 构造数据（简化版本）
        data = []
        
        # 销项税额部分
        data.append({
            'item': '一般项目销售',
            'sales_amount': float(sales_amount),
            'tax_rate': 0.13,
            'tax_amount': float(sales_tax)
        })
        
        # 进项税额部分
        data.append({
            'item': '采购货物及服务',
            'purchase_amount': float(purchase_amount),
            'tax_rate': 0.13,
            'tax_amount': float(purchase_tax)
        })
        
        # 应纳税额部分
        data.append({
            'item': '应纳税额',
            'amount': float(tax_payable)
        })
        
        return pd.DataFrame(data)
    
    def _prepare_income_tax_data(self, transactions: List[TransactionRecord]) -> pd.DataFrame:
        """准备所得税数据
        
        Args:
            transactions: 交易记录列表
            
        Returns:
            DataFrame: 所得税数据
        """
        # 计算收入总额和扣除项目
        total_income = sum(
            (t.amount for t in transactions if t.type == TransactionType.INCOME),
            Decimal('0')
        )
        total_expense = sum(
            (t.amount for t in transactions if t.type == TransactionType.EXPENSE),
            Decimal('0')
        )
        
        # 应纳税所得额
        taxable_income = total_income - total_expense
        
        # 企业所得税率（假设为25%）
        tax_rate = Decimal('0.25')
        tax_amount = taxable_income * tax_rate if taxable_income > 0 else Decimal('0')
        
        # 构造数据（简化版本）
        data = []
        
        # 收入总额部分
        data.append({
            'item': '营业收入',
            'amount': float(total_income)
        })
        
        # 扣除项目部分
        data.append({
            'item': '营业成本及费用',
            'amount': float(total_expense)
        })
        
        # 应纳税所得额及应纳税额
        data.append({
            'item': '应纳税所得额',
            'amount': float(taxable_income)
        })
        data.append({
            'item': '应纳所得税额',
            'amount': float(tax_amount)
        })
        
        return pd.DataFrame(data)
    
    def _prepare_balance_sheet_data(
        self,
        transactions: List[TransactionRecord],
        report_date: date
    ) -> pd.DataFrame:
        """准备资产负债表数据
        
        Args:
            transactions: 交易记录列表
            report_date: 报表日期
            
        Returns:
            DataFrame: 资产负债表数据
        """
        # 计算截至报表日期的财务状况
        # 这是一个简化版本，实际应用需要更复杂的会计处理
        
        # 计算现金（收入-支出）
        cash = sum(
            (t.amount if t.type == TransactionType.INCOME else -t.amount
             for t in transactions if t.date <= report_date),
            Decimal('0')
        )
        
        # 应收账款（简化处理）
        receivables = sum(
            (t.amount for t in transactions 
             if t.type == TransactionType.INCOME and t.status != TransactionStatus.COMPLETED
             and t.date <= report_date),
            Decimal('0')
        )
        
        # 应付账款（简化处理）
        payables = sum(
            (t.amount for t in transactions 
             if t.type == TransactionType.EXPENSE and t.status != TransactionStatus.COMPLETED
             and t.date <= report_date),
            Decimal('0')
        )
        
        # 所有者权益（资产-负债）
        total_assets = cash + receivables
        total_liabilities = payables
        equity = total_assets - total_liabilities
        
        # 构造数据
        data = []
        
        # 资产部分
        data.append({'item': '货币资金', 'ending_balance': float(cash), 'beginning_balance': 0})
        data.append({'item': '应收账款', 'ending_balance': float(receivables), 'beginning_balance': 0})
        
        # 负债部分
        data.append({'item': '应付账款', 'ending_balance': float(payables), 'beginning_balance': 0})
        
        # 所有者权益部分
        data.append({'item': '实收资本', 'ending_balance': float(equity), 'beginning_balance': 0})
        
        return pd.DataFrame(data)
    
    def _prepare_income_statement_data(
        self,
        transactions: List[TransactionRecord],
        report_date: date
    ) -> pd.DataFrame:
        """准备利润表数据
        
        Args:
            transactions: 交易记录列表
            report_date: 报表日期
            
        Returns:
            DataFrame: 利润表数据
        """
        # 计算本期和上期的收入和支出
        # 本期：截至报表日期
        current_income = sum(
            (t.amount for t in transactions 
             if t.type == TransactionType.INCOME and t.date <= report_date),
            Decimal('0')
        )
        current_expense = sum(
            (t.amount for t in transactions 
             if t.type == TransactionType.EXPENSE and t.date <= report_date),
            Decimal('0')
        )
        
        # 计算利润
        gross_profit = current_income - current_expense
        net_profit = gross_profit  # 简化处理
        
        # 构造数据
        data = []
        data.append({'item': '一、营业收入', 'current_amount': float(current_income), 'previous_amount': 0})
        data.append({'item': '二、营业成本', 'current_amount': float(current_expense), 'previous_amount': 0})
        data.append({'item': '三、营业利润', 'current_amount': float(gross_profit), 'previous_amount': 0})
        data.append({'item': '四、利润总额', 'current_amount': float(gross_profit), 'previous_amount': 0})
        data.append({'item': '五、净利润', 'current_amount': float(net_profit), 'previous_amount': 0})
        
        return pd.DataFrame(data)
    
    def _prepare_cash_flow_statement_data(
        self,
        transactions: List[TransactionRecord],
        report_date: date
    ) -> pd.DataFrame:
        """准备现金流量表数据
        
        Args:
            transactions: 交易记录列表
            report_date: 报表日期
            
        Returns:
            DataFrame: 现金流量表数据
        """
        # 计算经营活动现金流
        operating_inflow = sum(
            (t.amount for t in transactions 
             if t.type == TransactionType.INCOME and t.date <= report_date),
            Decimal('0')
        )
        operating_outflow = sum(
            (t.amount for t in transactions 
             if t.type == TransactionType.EXPENSE and t.date <= report_date),
            Decimal('0')
        )
        operating_net = operating_inflow - operating_outflow
        
        # 构造数据（简化版本）
        data = []
        
        # 经营活动现金流
        data.append({'item': '销售商品、提供劳务收到的现金', 'current_amount': float(operating_inflow), 'previous_amount': 0})
        data.append({'item': '购买商品、接受劳务支付的现金', 'current_amount': float(operating_outflow), 'previous_amount': 0})
        data.append({'item': '经营活动现金流量净额', 'current_amount': float(operating_net), 'previous_amount': 0})
        
        # 投资活动现金流（简化为0）
        data.append({'item': '投资活动现金流量净额', 'current_amount': 0, 'previous_amount': 0})
        
        # 筹资活动现金流（简化为0）
        data.append({'item': '筹资活动现金流量净额', 'current_amount': 0, 'previous_amount': 0})
        
        # 现金净增加额
        data.append({'item': '现金及现金等价物净增加额', 'current_amount': float(operating_net), 'previous_amount': 0})
        
        return pd.DataFrame(data)
    
    def _parse_period(self, period: str) -> tuple[date, date]:
        """解析期间字符串为开始和结束日期
        
        Args:
            period: 期间字符串（如"2024年第一季度"、"2024年1月"）
            
        Returns:
            (start_date, end_date): 开始和结束日期
        """
        import re
        from datetime import timedelta
        from calendar import monthrange
        
        # 提取年份
        year_match = re.search(r'(\d{4})年', period)
        if not year_match:
            raise ValueError(f"无法从期间字符串中提取年份: {period}")
        year = int(year_match.group(1))
        
        # 检查是否为季度
        quarter_match = re.search(r'第([一二三四])季度', period)
        if quarter_match:
            quarter_map = {'一': 1, '二': 2, '三': 3, '四': 4}
            quarter = quarter_map[quarter_match.group(1)]
            
            # 计算季度的开始和结束月份
            start_month = (quarter - 1) * 3 + 1
            end_month = quarter * 3
            
            start_date = date(year, start_month, 1)
            _, last_day = monthrange(year, end_month)
            end_date = date(year, end_month, last_day)
            
            return start_date, end_date
        
        # 检查是否为月份
        month_match = re.search(r'(\d{1,2})月', period)
        if month_match:
            month = int(month_match.group(1))
            start_date = date(year, month, 1)
            _, last_day = monthrange(year, month)
            end_date = date(year, month, last_day)
            
            return start_date, end_date
        
        # 默认返回整年
        start_date = date(year, 1, 1)
        end_date = date(year, 12, 31)
        return start_date, end_date
