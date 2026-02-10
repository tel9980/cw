"""
报表模板管理模块

提供各类报表模板的定义、加载和应用功能。
"""

from dataclasses import dataclass, field
from typing import Dict, List, Any, Optional
from enum import Enum
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd


class TemplateType(Enum):
    """模板类型"""
    # 管理报表
    REVENUE_COMPARISON = "revenue_comparison"  # 收支对比
    PROFIT_TREND = "profit_trend"  # 利润趋势
    CUSTOMER_RANKING = "customer_ranking"  # 客户排名
    
    # 税务报表
    VAT_DECLARATION = "vat_declaration"  # 增值税申报表
    INCOME_TAX_DECLARATION = "income_tax_declaration"  # 所得税申报表
    
    # 银行贷款报表
    BALANCE_SHEET = "balance_sheet"  # 资产负债表
    INCOME_STATEMENT = "income_statement"  # 利润表
    CASH_FLOW_STATEMENT = "cash_flow_statement"  # 现金流量表


@dataclass
class TemplateColumn:
    """模板列定义"""
    name: str  # 列名
    width: int = 15  # 列宽
    data_field: Optional[str] = None  # 数据字段名
    format: Optional[str] = None  # 格式化字符串（如 "0.00" 表示两位小数）


@dataclass
class TemplateSection:
    """模板区域定义"""
    title: str  # 区域标题
    start_row: int  # 起始行
    columns: List[TemplateColumn]  # 列定义
    has_total: bool = False  # 是否有合计行
    total_columns: List[str] = field(default_factory=list)  # 需要合计的列


@dataclass
class Template:
    """报表模板"""
    template_type: TemplateType
    name: str  # 模板名称
    description: str  # 模板描述
    sections: List[TemplateSection]  # 模板区域列表
    header_rows: int = 3  # 表头行数
    title: str = ""  # 报表标题
    subtitle: str = ""  # 报表副标题


class ReportTemplate:
    """报表模板管理器
    
    提供报表模板的加载、应用和格式化功能。
    """
    
    def __init__(self):
        """初始化模板管理器"""
        self._templates: Dict[TemplateType, Template] = {}
        self._initialize_templates()
    
    def _initialize_templates(self):
        """初始化所有预定义模板"""
        # 管理报表模板
        self._templates[TemplateType.REVENUE_COMPARISON] = self._create_revenue_comparison_template()
        self._templates[TemplateType.PROFIT_TREND] = self._create_profit_trend_template()
        self._templates[TemplateType.CUSTOMER_RANKING] = self._create_customer_ranking_template()
        
        # 税务报表模板
        self._templates[TemplateType.VAT_DECLARATION] = self._create_vat_declaration_template()
        self._templates[TemplateType.INCOME_TAX_DECLARATION] = self._create_income_tax_declaration_template()
        
        # 银行贷款报表模板
        self._templates[TemplateType.BALANCE_SHEET] = self._create_balance_sheet_template()
        self._templates[TemplateType.INCOME_STATEMENT] = self._create_income_statement_template()
        self._templates[TemplateType.CASH_FLOW_STATEMENT] = self._create_cash_flow_statement_template()
    
    def load_template(self, template_type: TemplateType) -> Template:
        """加载预定义报表模板
        
        Args:
            template_type: 模板类型
            
        Returns:
            Template: 报表模板
            
        Raises:
            ValueError: 如果模板类型不存在
        """
        if template_type not in self._templates:
            raise ValueError(f"未找到模板类型: {template_type}")
        
        return self._templates[template_type]
    
    def apply_template(
        self,
        template: Template,
        data: pd.DataFrame,
        output_file: str,
        **kwargs
    ) -> openpyxl.Workbook:
        """应用模板生成Excel报表
        
        Args:
            template: 报表模板
            data: 数据DataFrame
            output_file: 输出文件路径
            **kwargs: 额外参数（如报表期间、公司名称等）
            
        Returns:
            openpyxl.Workbook: 生成的工作簿
        """
        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = template.name
        
        # 写入表头
        current_row = self._write_header(ws, template, **kwargs)
        
        # 写入各个区域
        for section in template.sections:
            current_row = self._write_section(
                ws, section, data, current_row, **kwargs
            )
            current_row += 2  # 区域间隔
        
        # 应用样式
        self._apply_styles(ws, template)
        
        # 保存文件
        wb.save(output_file)
        
        return wb
    
    def _write_header(
        self,
        ws: Worksheet,
        template: Template,
        **kwargs
    ) -> int:
        """写入报表表头
        
        Args:
            ws: 工作表
            template: 模板
            **kwargs: 额外参数
            
        Returns:
            下一行行号
        """
        current_row = 1
        
        # 标题
        if template.title:
            title = template.title
            if 'company_name' in kwargs:
                title = f"{kwargs['company_name']} - {title}"
            
            ws.merge_cells(f'A{current_row}:F{current_row}')
            cell = ws[f'A{current_row}']
            cell.value = title
            cell.font = Font(size=16, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
        
        # 副标题（报表期间）
        if 'period' in kwargs:
            ws.merge_cells(f'A{current_row}:F{current_row}')
            cell = ws[f'A{current_row}']
            cell.value = f"报表期间: {kwargs['period']}"
            cell.font = Font(size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
        
        current_row += 1  # 空行
        
        return current_row
    
    def _write_section(
        self,
        ws: Worksheet,
        section: TemplateSection,
        data: pd.DataFrame,
        start_row: int,
        **kwargs
    ) -> int:
        """写入报表区域
        
        Args:
            ws: 工作表
            section: 区域定义
            data: 数据
            start_row: 起始行
            **kwargs: 额外参数
            
        Returns:
            下一行行号
        """
        current_row = start_row
        
        # 区域标题
        if section.title:
            ws.merge_cells(f'A{current_row}:F{current_row}')
            cell = ws[f'A{current_row}']
            cell.value = section.title
            cell.font = Font(size=12, bold=True)
            cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
            current_row += 1
        
        # 列标题
        for col_idx, column in enumerate(section.columns, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = column.name
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 设置列宽
            ws.column_dimensions[get_column_letter(col_idx)].width = column.width
        
        current_row += 1
        
        # 数据行
        data_start_row = current_row
        for _, row_data in data.iterrows():
            for col_idx, column in enumerate(section.columns, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                
                # 获取数据值
                if column.data_field and column.data_field in row_data:
                    value = row_data[column.data_field]
                    cell.value = value
                    
                    # 应用格式
                    if column.format:
                        cell.number_format = column.format
                
                cell.alignment = Alignment(horizontal='left', vertical='center')
            
            current_row += 1
        
        # 合计行
        if section.has_total and len(data) > 0:
            for col_idx, column in enumerate(section.columns, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                
                if col_idx == 1:
                    cell.value = "合计"
                    cell.font = Font(bold=True)
                elif column.data_field in section.total_columns:
                    # 计算合计
                    col_letter = get_column_letter(col_idx)
                    formula = f"=SUM({col_letter}{data_start_row}:{col_letter}{current_row-1})"
                    cell.value = formula
                    cell.font = Font(bold=True)
                    
                    if column.format:
                        cell.number_format = column.format
                
                cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
            
            current_row += 1
        
        return current_row
    
    def _apply_styles(self, ws: Worksheet, template: Template):
        """应用样式到工作表
        
        Args:
            ws: 工作表
            template: 模板
        """
        # 添加边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.border = thin_border
    
    # ========== 管理报表模板 ==========
    
    def _create_revenue_comparison_template(self) -> Template:
        """创建收支对比报表模板"""
        return Template(
            template_type=TemplateType.REVENUE_COMPARISON,
            name="收支对比表",
            description="对比分析收入和支出情况",
            title="收支对比表",
            sections=[
                TemplateSection(
                    title="收支明细",
                    start_row=4,
                    columns=[
                        TemplateColumn(name="期间", width=15, data_field="period"),
                        TemplateColumn(name="收入金额", width=15, data_field="income", format="#,##0.00"),
                        TemplateColumn(name="支出金额", width=15, data_field="expense", format="#,##0.00"),
                        TemplateColumn(name="净利润", width=15, data_field="profit", format="#,##0.00"),
                        TemplateColumn(name="利润率", width=12, data_field="profit_rate", format="0.00%"),
                    ],
                    has_total=True,
                    total_columns=["income", "expense", "profit"]
                )
            ]
        )
    
    def _create_profit_trend_template(self) -> Template:
        """创建利润趋势报表模板"""
        return Template(
            template_type=TemplateType.PROFIT_TREND,
            name="利润趋势表",
            description="分析利润变化趋势",
            title="利润趋势表",
            sections=[
                TemplateSection(
                    title="月度利润趋势",
                    start_row=4,
                    columns=[
                        TemplateColumn(name="月份", width=12, data_field="month"),
                        TemplateColumn(name="营业收入", width=15, data_field="revenue", format="#,##0.00"),
                        TemplateColumn(name="营业成本", width=15, data_field="cost", format="#,##0.00"),
                        TemplateColumn(name="毛利润", width=15, data_field="gross_profit", format="#,##0.00"),
                        TemplateColumn(name="净利润", width=15, data_field="net_profit", format="#,##0.00"),
                        TemplateColumn(name="环比增长", width=12, data_field="growth_rate", format="0.00%"),
                    ],
                    has_total=True,
                    total_columns=["revenue", "cost", "gross_profit", "net_profit"]
                )
            ]
        )
    
    def _create_customer_ranking_template(self) -> Template:
        """创建客户排名报表模板"""
        return Template(
            template_type=TemplateType.CUSTOMER_RANKING,
            name="客户排名表",
            description="按销售额排名客户",
            title="客户排名表",
            sections=[
                TemplateSection(
                    title="客户销售排名",
                    start_row=4,
                    columns=[
                        TemplateColumn(name="排名", width=8, data_field="rank"),
                        TemplateColumn(name="客户名称", width=25, data_field="customer_name"),
                        TemplateColumn(name="销售金额", width=15, data_field="sales_amount", format="#,##0.00"),
                        TemplateColumn(name="交易次数", width=12, data_field="transaction_count"),
                        TemplateColumn(name="占比", width=12, data_field="percentage", format="0.00%"),
                        TemplateColumn(name="客户类型", width=12, data_field="customer_type"),
                    ],
                    has_total=True,
                    total_columns=["sales_amount", "transaction_count"]
                )
            ]
        )
    
    # ========== 税务报表模板 ==========
    
    def _create_vat_declaration_template(self) -> Template:
        """创建增值税申报表模板"""
        return Template(
            template_type=TemplateType.VAT_DECLARATION,
            name="增值税申报表",
            description="增值税纳税申报表",
            title="增值税纳税申报表",
            sections=[
                TemplateSection(
                    title="一、销项税额",
                    start_row=4,
                    columns=[
                        TemplateColumn(name="项目", width=30, data_field="item"),
                        TemplateColumn(name="销售额", width=18, data_field="sales_amount", format="#,##0.00"),
                        TemplateColumn(name="税率", width=10, data_field="tax_rate", format="0.00%"),
                        TemplateColumn(name="税额", width=18, data_field="tax_amount", format="#,##0.00"),
                    ],
                    has_total=True,
                    total_columns=["sales_amount", "tax_amount"]
                ),
                TemplateSection(
                    title="二、进项税额",
                    start_row=10,
                    columns=[
                        TemplateColumn(name="项目", width=30, data_field="item"),
                        TemplateColumn(name="采购额", width=18, data_field="purchase_amount", format="#,##0.00"),
                        TemplateColumn(name="税率", width=10, data_field="tax_rate", format="0.00%"),
                        TemplateColumn(name="税额", width=18, data_field="tax_amount", format="#,##0.00"),
                    ],
                    has_total=True,
                    total_columns=["purchase_amount", "tax_amount"]
                ),
                TemplateSection(
                    title="三、应纳税额",
                    start_row=16,
                    columns=[
                        TemplateColumn(name="项目", width=30, data_field="item"),
                        TemplateColumn(name="金额", width=18, data_field="amount", format="#,##0.00"),
                    ],
                    has_total=False
                )
            ]
        )
    
    def _create_income_tax_declaration_template(self) -> Template:
        """创建所得税申报表模板"""
        return Template(
            template_type=TemplateType.INCOME_TAX_DECLARATION,
            name="所得税申报表",
            description="企业所得税纳税申报表",
            title="企业所得税纳税申报表",
            sections=[
                TemplateSection(
                    title="一、收入总额",
                    start_row=4,
                    columns=[
                        TemplateColumn(name="项目", width=30, data_field="item"),
                        TemplateColumn(name="金额", width=20, data_field="amount", format="#,##0.00"),
                    ],
                    has_total=True,
                    total_columns=["amount"]
                ),
                TemplateSection(
                    title="二、扣除项目",
                    start_row=10,
                    columns=[
                        TemplateColumn(name="项目", width=30, data_field="item"),
                        TemplateColumn(name="金额", width=20, data_field="amount", format="#,##0.00"),
                    ],
                    has_total=True,
                    total_columns=["amount"]
                ),
                TemplateSection(
                    title="三、应纳税所得额及应纳税额",
                    start_row=16,
                    columns=[
                        TemplateColumn(name="项目", width=30, data_field="item"),
                        TemplateColumn(name="金额", width=20, data_field="amount", format="#,##0.00"),
                    ],
                    has_total=False
                )
            ]
        )
    
    # ========== 银行贷款报表模板 ==========
    
    def _create_balance_sheet_template(self) -> Template:
        """创建资产负债表模板"""
        return Template(
            template_type=TemplateType.BALANCE_SHEET,
            name="资产负债表",
            description="企业资产负债表",
            title="资产负债表",
            sections=[
                TemplateSection(
                    title="资产",
                    start_row=4,
                    columns=[
                        TemplateColumn(name="项目", width=25, data_field="item"),
                        TemplateColumn(name="期末余额", width=18, data_field="ending_balance", format="#,##0.00"),
                        TemplateColumn(name="期初余额", width=18, data_field="beginning_balance", format="#,##0.00"),
                    ],
                    has_total=True,
                    total_columns=["ending_balance", "beginning_balance"]
                ),
                TemplateSection(
                    title="负债",
                    start_row=15,
                    columns=[
                        TemplateColumn(name="项目", width=25, data_field="item"),
                        TemplateColumn(name="期末余额", width=18, data_field="ending_balance", format="#,##0.00"),
                        TemplateColumn(name="期初余额", width=18, data_field="beginning_balance", format="#,##0.00"),
                    ],
                    has_total=True,
                    total_columns=["ending_balance", "beginning_balance"]
                ),
                TemplateSection(
                    title="所有者权益",
                    start_row=25,
                    columns=[
                        TemplateColumn(name="项目", width=25, data_field="item"),
                        TemplateColumn(name="期末余额", width=18, data_field="ending_balance", format="#,##0.00"),
                        TemplateColumn(name="期初余额", width=18, data_field="beginning_balance", format="#,##0.00"),
                    ],
                    has_total=True,
                    total_columns=["ending_balance", "beginning_balance"]
                )
            ]
        )
    
    def _create_income_statement_template(self) -> Template:
        """创建利润表模板"""
        return Template(
            template_type=TemplateType.INCOME_STATEMENT,
            name="利润表",
            description="企业利润表",
            title="利润表",
            sections=[
                TemplateSection(
                    title="利润表",
                    start_row=4,
                    columns=[
                        TemplateColumn(name="项目", width=30, data_field="item"),
                        TemplateColumn(name="本期金额", width=18, data_field="current_amount", format="#,##0.00"),
                        TemplateColumn(name="上期金额", width=18, data_field="previous_amount", format="#,##0.00"),
                    ],
                    has_total=False
                )
            ]
        )
    
    def _create_cash_flow_statement_template(self) -> Template:
        """创建现金流量表模板"""
        return Template(
            template_type=TemplateType.CASH_FLOW_STATEMENT,
            name="现金流量表",
            description="企业现金流量表",
            title="现金流量表",
            sections=[
                TemplateSection(
                    title="一、经营活动产生的现金流量",
                    start_row=4,
                    columns=[
                        TemplateColumn(name="项目", width=35, data_field="item"),
                        TemplateColumn(name="本期金额", width=18, data_field="current_amount", format="#,##0.00"),
                        TemplateColumn(name="上期金额", width=18, data_field="previous_amount", format="#,##0.00"),
                    ],
                    has_total=True,
                    total_columns=["current_amount", "previous_amount"]
                ),
                TemplateSection(
                    title="二、投资活动产生的现金流量",
                    start_row=15,
                    columns=[
                        TemplateColumn(name="项目", width=35, data_field="item"),
                        TemplateColumn(name="本期金额", width=18, data_field="current_amount", format="#,##0.00"),
                        TemplateColumn(name="上期金额", width=18, data_field="previous_amount", format="#,##0.00"),
                    ],
                    has_total=True,
                    total_columns=["current_amount", "previous_amount"]
                ),
                TemplateSection(
                    title="三、筹资活动产生的现金流量",
                    start_row=23,
                    columns=[
                        TemplateColumn(name="项目", width=35, data_field="item"),
                        TemplateColumn(name="本期金额", width=18, data_field="current_amount", format="#,##0.00"),
                        TemplateColumn(name="上期金额", width=18, data_field="previous_amount", format="#,##0.00"),
                    ],
                    has_total=True,
                    total_columns=["current_amount", "previous_amount"]
                )
            ]
        )
