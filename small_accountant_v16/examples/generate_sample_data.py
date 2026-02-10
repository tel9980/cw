"""
示例数据生成器

生成用于测试和演示的示例数据，包括：
1. 交易记录（收入、支出、订单）
2. 往来单位（客户、供应商）
3. Excel导入文件

使用方法：
    python examples/generate_sample_data.py
"""

import sys
import os
from pathlib import Path
from datetime import date, datetime, timedelta
from decimal import Decimal
import random
import uuid

# 添加项目根目录到路径
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))

from small_accountant_v16.models.core_models import (
    TransactionRecord, TransactionType, TransactionStatus,
    Counterparty, CounterpartyType
)
from small_accountant_v16.storage.transaction_storage import TransactionStorage
from small_accountant_v16.storage.counterparty_storage import CounterpartyStorage

# 导入openpyxl用于生成Excel文件
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("错误: 需要安装openpyxl库")
    print("请运行: pip install openpyxl")
    sys.exit(1)


# 示例数据配置
SAMPLE_CUSTOMERS = [
    {
        "name": "优质客户有限公司",
        "contact_person": "张经理",
        "phone": "13800138001",
        "email": "zhang@youzhikehu.com",
        "address": "北京市朝阳区建国路88号",
        "tax_id": "91110000MA001ABC01"
    },
    {
        "name": "新兴科技股份有限公司",
        "contact_person": "李总",
        "phone": "13900139002",
        "email": "li@xinxingkeji.com",
        "address": "上海市浦东新区世纪大道100号",
        "tax_id": "91310000MA002DEF02"
    },
    {
        "name": "长期合作伙伴公司",
        "contact_person": "王主任",
        "phone": "13700137003",
        "email": "wang@changqihezuo.com",
        "address": "广州市天河区珠江新城200号",
        "tax_id": "91440000MA003GHI03"
    },
    {
        "name": "稳定收入客户",
        "contact_person": "赵经理",
        "phone": "13600136004",
        "email": "zhao@wendingshouru.com",
        "address": "深圳市南山区科技园300号",
        "tax_id": "91440300MA004JKL04"
    },
    {
        "name": "大型企业集团",
        "contact_person": "刘总监",
        "phone": "13500135005",
        "email": "liu@daxingqiye.com",
        "address": "杭州市西湖区文三路400号",
        "tax_id": "91330000MA005MNO05"
    }
]

SAMPLE_SUPPLIERS = [
    {
        "name": "原材料供应商A",
        "contact_person": "陈经理",
        "phone": "13400134006",
        "email": "chen@yuancailiaoA.com",
        "address": "苏州市工业园区星湖街500号",
        "tax_id": "91320500MA006PQR06"
    },
    {
        "name": "设备租赁公司",
        "contact_person": "周主管",
        "phone": "13300133007",
        "email": "zhou@shebeizulin.com",
        "address": "南京市江宁区科学园600号",
        "tax_id": "91320100MA007STU07"
    },
    {
        "name": "物流运输服务商",
        "contact_person": "吴经理",
        "phone": "13200132008",
        "email": "wu@wuliuyunshu.com",
        "address": "武汉市东湖高新区光谷大道700号",
        "tax_id": "91420100MA008VWX08"
    },
    {
        "name": "办公用品供应商",
        "contact_person": "郑主任",
        "phone": "13100131009",
        "email": "zheng@bangongyongpin.com",
        "address": "成都市高新区天府大道800号",
        "tax_id": "91510100MA009YZA09"
    },
    {
        "name": "技术服务提供商",
        "contact_person": "孙工程师",
        "phone": "13000130010",
        "email": "sun@jishufuwu.com",
        "address": "西安市高新区科技路900号",
        "tax_id": "91610100MA010BCD10"
    }
]

# 交易类别
INCOME_CATEGORIES = ["产品销售", "服务收入", "咨询费", "技术服务费", "租金收入"]
EXPENSE_CATEGORIES = ["原材料采购", "设备租赁", "物流费用", "办公费用", "技术服务费", "工资支出", "水电费"]
ORDER_CATEGORIES = ["加工订单", "定制服务", "批量采购"]


def generate_counterparties():
    """生成示例往来单位"""
    print("正在生成示例往来单位...")
    
    counterparties = []
    now = datetime.now()
    
    # 生成客户
    for i, customer_data in enumerate(SAMPLE_CUSTOMERS, 1):
        counterparty = Counterparty(
            id=f"CUST{i:03d}",
            name=customer_data["name"],
            type=CounterpartyType.CUSTOMER,
            contact_person=customer_data["contact_person"],
            phone=customer_data["phone"],
            email=customer_data["email"],
            address=customer_data["address"],
            tax_id=customer_data["tax_id"],
            created_at=now,
            updated_at=now
        )
        counterparties.append(counterparty)
    
    # 生成供应商
    for i, supplier_data in enumerate(SAMPLE_SUPPLIERS, 1):
        counterparty = Counterparty(
            id=f"SUPP{i:03d}",
            name=supplier_data["name"],
            type=CounterpartyType.SUPPLIER,
            contact_person=supplier_data["contact_person"],
            phone=supplier_data["phone"],
            email=supplier_data["email"],
            address=supplier_data["address"],
            tax_id=supplier_data["tax_id"],
            created_at=now,
            updated_at=now
        )
        counterparties.append(counterparty)
    
    print(f"✓ 生成了 {len(counterparties)} 个往来单位 ({len(SAMPLE_CUSTOMERS)} 个客户, {len(SAMPLE_SUPPLIERS)} 个供应商)")
    return counterparties


def generate_transactions(counterparties):
    """生成示例交易记录"""
    print("正在生成示例交易记录...")
    
    transactions = []
    now = datetime.now()
    start_date = date.today() - timedelta(days=90)  # 最近3个月
    
    # 获取客户和供应商列表
    customers = [c for c in counterparties if c.type == CounterpartyType.CUSTOMER]
    suppliers = [c for c in counterparties if c.type == CounterpartyType.SUPPLIER]
    
    transaction_id = 1
    
    # 生成收入交易（30笔）
    for _ in range(30):
        transaction_date = start_date + timedelta(days=random.randint(0, 90))
        customer = random.choice(customers)
        amount = Decimal(str(random.randint(5000, 50000)))
        category = random.choice(INCOME_CATEGORIES)
        
        transaction = TransactionRecord(
            id=f"TXN{transaction_id:05d}",
            date=transaction_date,
            type=TransactionType.INCOME,
            amount=amount,
            counterparty_id=customer.id,
            description=f"{category} - {customer.name}",
            category=category,
            status=TransactionStatus.COMPLETED,
            created_at=now,
            updated_at=now
        )
        transactions.append(transaction)
        transaction_id += 1
    
    # 生成支出交易（40笔）
    for _ in range(40):
        transaction_date = start_date + timedelta(days=random.randint(0, 90))
        supplier = random.choice(suppliers)
        amount = Decimal(str(random.randint(1000, 30000)))
        category = random.choice(EXPENSE_CATEGORIES)
        
        transaction = TransactionRecord(
            id=f"TXN{transaction_id:05d}",
            date=transaction_date,
            type=TransactionType.EXPENSE,
            amount=amount,
            counterparty_id=supplier.id,
            description=f"{category} - {supplier.name}",
            category=category,
            status=TransactionStatus.COMPLETED,
            created_at=now,
            updated_at=now
        )
        transactions.append(transaction)
        transaction_id += 1
    
    # 生成订单交易（20笔）
    for _ in range(20):
        transaction_date = start_date + timedelta(days=random.randint(0, 90))
        customer = random.choice(customers)
        amount = Decimal(str(random.randint(10000, 100000)))
        category = random.choice(ORDER_CATEGORIES)
        status = random.choice([TransactionStatus.PENDING, TransactionStatus.COMPLETED])
        
        transaction = TransactionRecord(
            id=f"TXN{transaction_id:05d}",
            date=transaction_date,
            type=TransactionType.ORDER,
            amount=amount,
            counterparty_id=customer.id,
            description=f"{category} - {customer.name}",
            category=category,
            status=status,
            created_at=now,
            updated_at=now
        )
        transactions.append(transaction)
        transaction_id += 1
    
    # 按日期排序
    transactions.sort(key=lambda t: t.date)
    
    print(f"✓ 生成了 {len(transactions)} 笔交易记录 (30笔收入, 40笔支出, 20笔订单)")
    return transactions


def save_to_storage(counterparties, transactions):
    """保存到存储系统"""
    print("正在保存到存储系统...")
    
    # 创建examples目录下的data子目录
    examples_dir = Path(__file__).parent
    data_dir = examples_dir / "data"
    data_dir.mkdir(exist_ok=True)
    
    # 初始化存储
    counterparty_storage = CounterpartyStorage(str(data_dir))
    transaction_storage = TransactionStorage(str(data_dir))
    
    # 保存往来单位
    for counterparty in counterparties:
        counterparty_storage.add(counterparty)
    
    # 保存交易记录
    for transaction in transactions:
        transaction_storage.add(transaction)
    
    print(f"✓ 数据已保存到: {data_dir}")


def create_excel_files(counterparties, transactions):
    """创建示例Excel文件"""
    print("正在创建示例Excel文件...")
    
    examples_dir = Path(__file__).parent
    
    # 1. 创建往来单位导入模板
    create_counterparty_excel(counterparties, examples_dir / "示例往来单位.xlsx")
    
    # 2. 创建交易记录导入模板
    create_transaction_excel(transactions, examples_dir / "示例交易记录.xlsx")
    
    # 3. 创建银行流水模板
    create_bank_statement_excel(transactions, examples_dir / "示例银行流水.xlsx")
    
    print("✓ Excel文件创建完成")


def create_counterparty_excel(counterparties, filepath):
    """创建往来单位Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "往来单位"
    
    # 设置表头
    headers = ["单位名称", "类型", "联系人", "电话", "邮箱", "地址", "税号"]
    ws.append(headers)
    
    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 添加数据
    for counterparty in counterparties:
        ws.append([
            counterparty.name,
            "客户" if counterparty.type == CounterpartyType.CUSTOMER else "供应商",
            counterparty.contact_person,
            counterparty.phone,
            counterparty.email,
            counterparty.address,
            counterparty.tax_id
        ])
    
    # 调整列宽
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 20
    
    wb.save(filepath)
    print(f"  - {filepath.name}")


def create_transaction_excel(transactions, filepath):
    """创建交易记录Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "交易记录"
    
    # 设置表头
    headers = ["日期", "类型", "金额", "往来单位", "摘要", "类别", "状态"]
    ws.append(headers)
    
    # 设置表头样式
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 添加数据（只取前50笔）
    for transaction in transactions[:50]:
        type_name = {
            TransactionType.INCOME: "收入",
            TransactionType.EXPENSE: "支出",
            TransactionType.ORDER: "订单"
        }[transaction.type]
        
        status_name = {
            TransactionStatus.PENDING: "待处理",
            TransactionStatus.COMPLETED: "已完成",
            TransactionStatus.CANCELLED: "已取消"
        }[transaction.status]
        
        ws.append([
            transaction.date.strftime("%Y-%m-%d"),
            type_name,
            float(transaction.amount),
            transaction.counterparty_id,
            transaction.description,
            transaction.category,
            status_name
        ])
    
    # 调整列宽
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 10
    
    wb.save(filepath)
    print(f"  - {filepath.name}")


def create_bank_statement_excel(transactions, filepath):
    """创建银行流水Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "银行流水"
    
    # 设置表头
    headers = ["交易日期", "摘要", "收入金额", "支出金额", "余额", "对方户名"]
    ws.append(headers)
    
    # 设置表头样式
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 模拟银行流水（基于交易记录）
    balance = Decimal("500000.00")  # 初始余额
    
    # 只取收入和支出交易，按日期排序
    bank_transactions = [
        t for t in transactions 
        if t.type in [TransactionType.INCOME, TransactionType.EXPENSE]
    ]
    bank_transactions.sort(key=lambda t: t.date)
    
    for transaction in bank_transactions[:40]:  # 只取前40笔
        if transaction.type == TransactionType.INCOME:
            income_amount = float(transaction.amount)
            expense_amount = ""
            balance += transaction.amount
        else:
            income_amount = ""
            expense_amount = float(transaction.amount)
            balance -= transaction.amount
        
        ws.append([
            transaction.date.strftime("%Y-%m-%d"),
            transaction.description,
            income_amount,
            expense_amount,
            float(balance),
            transaction.counterparty_id
        ])
    
    # 调整列宽
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20
    
    wb.save(filepath)
    print(f"  - {filepath.name}")


def create_readme():
    """创建README文件"""
    examples_dir = Path(__file__).parent
    readme_path = examples_dir / "README.md"
    
    content = """# 示例数据说明

本目录包含用于测试和演示的示例数据。

## 文件说明

### 1. 数据文件

- `data/counterparties.json` - 往来单位数据（5个客户 + 5个供应商）
- `data/transactions.json` - 交易记录数据（90笔交易，涵盖最近3个月）

### 2. Excel导入文件

- `示例往来单位.xlsx` - 往来单位导入模板，包含10个示例单位
- `示例交易记录.xlsx` - 交易记录导入模板，包含50笔示例交易
- `示例银行流水.xlsx` - 银行流水导入模板，包含40笔银行流水

## 使用方法

### 生成示例数据

```bash
# 在项目根目录运行
python examples/generate_sample_data.py
```

这将生成：
- JSON格式的数据文件（保存在 `examples/data/` 目录）
- Excel格式的导入文件（保存在 `examples/` 目录）

### 使用示例数据测试系统

1. **测试Excel导入功能**
   - 使用 `示例往来单位.xlsx` 测试往来单位导入
   - 使用 `示例交易记录.xlsx` 测试交易记录导入
   - 使用 `示例银行流水.xlsx` 测试银行流水导入

2. **测试报表生成功能**
   - 导入示例数据后，生成管理报表查看收支对比
   - 生成客户对账单查看往来明细

3. **测试对账功能**
   - 使用银行流水文件进行银行对账
   - 查看匹配结果和差异报告

## 数据特点

### 往来单位
- 5个客户：涵盖不同规模和行业
- 5个供应商：包括原材料、设备、物流、办公用品、技术服务

### 交易记录
- 30笔收入：产品销售、服务收入、咨询费等
- 40笔支出：采购、租赁、物流、办公费用等
- 20笔订单：加工订单、定制服务、批量采购
- 时间跨度：最近3个月
- 金额范围：1,000 - 100,000元

### 银行流水
- 40笔银行交易记录
- 包含收入和支出
- 初始余额：500,000元
- 与交易记录部分匹配（用于测试对账功能）

## 注意事项

1. 示例数据仅用于测试和演示，不代表真实业务数据
2. 所有联系方式、税号等信息均为虚构
3. 重新运行生成脚本会覆盖现有数据
4. 建议在测试环境中使用，避免与生产数据混淆

## 自定义示例数据

如需自定义示例数据，可以修改 `generate_sample_data.py` 中的配置：

- `SAMPLE_CUSTOMERS` - 客户列表
- `SAMPLE_SUPPLIERS` - 供应商列表
- `INCOME_CATEGORIES` - 收入类别
- `EXPENSE_CATEGORIES` - 支出类别
- `ORDER_CATEGORIES` - 订单类别

修改后重新运行生成脚本即可。
"""
    
    with open(readme_path, 'w', encoding='utf-8') as f:
        f.write(content)
    
    print(f"✓ 创建了README文件: {readme_path.name}")


def main():
    """主函数"""
    print("=" * 60)
    print("V1.6 小会计示例数据生成器")
    print("=" * 60)
    print()
    
    # 1. 生成往来单位
    counterparties = generate_counterparties()
    print()
    
    # 2. 生成交易记录
    transactions = generate_transactions(counterparties)
    print()
    
    # 3. 保存到存储系统
    save_to_storage(counterparties, transactions)
    print()
    
    # 4. 创建Excel文件
    create_excel_files(counterparties, transactions)
    print()
    
    # 5. 创建README
    create_readme()
    print()
    
    print("=" * 60)
    print("✓ 示例数据生成完成！")
    print("=" * 60)
    print()
    print("生成的文件：")
    print("  - examples/data/counterparties.json")
    print("  - examples/data/transactions.json")
    print("  - examples/示例往来单位.xlsx")
    print("  - examples/示例交易记录.xlsx")
    print("  - examples/示例银行流水.xlsx")
    print("  - examples/README.md")
    print()
    print("您现在可以使用这些示例数据测试系统功能了！")


if __name__ == "__main__":
    main()
