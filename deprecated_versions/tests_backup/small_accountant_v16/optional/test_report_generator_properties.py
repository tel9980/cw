"""
Property-based tests for ReportGenerator - Management Report Completeness

Feature: small-accountant-practical-enhancement
Property 1: Management report completeness
Validates: Requirements 1.1
"""

import pytest
import os
import tempfile
from datetime import date, datetime, timedelta
from decimal import Decimal
from hypothesis import given, strategies as st, settings, assume
from hypothesis import HealthCheck
import openpyxl

from small_accountant_v16.reports.report_generator import ReportGenerator
from small_accountant_v16.storage.transaction_storage import TransactionStorage
from small_accountant_v16.storage.counterparty_storage import CounterpartyStorage
from small_accountant_v16.models.core_models import (
    TransactionRecord, Counterparty, CounterpartyType, 
    TransactionType, TransactionStatus, ReportType
)


# Hypothesis strategies for generating test data
@st.composite
def valid_date_range(draw, max_days=365):
    """生成有效的日期范围"""
    start_date = draw(st.dates(
        min_value=date(2020, 1, 1),
        max_value=date.today() - timedelta(days=1)
    ))
    
    end_date = draw(st.dates(
        min_value=start_date,
        max_value=min(start_date + timedelta(days=max_days), date.today())
    ))
    
    return start_date, end_date


@st.composite
def transaction_record_for_period(draw, start_date, end_date, counterparty_ids):
    """生成指定期间内的交易记录"""
    assume(len(counterparty_ids) > 0)
    assume(start_date <= end_date)
    
    # 生成期间内的随机日期
    days_diff = (end_date - start_date).days
    if days_diff == 0:
        trans_date = start_date
    else:
        random_days = draw(st.integers(min_value=0, max_value=days_diff))
        trans_date = start_date + timedelta(days=random_days)
    
    # 生成唯一的ID
    import uuid
    unique_id = str(uuid.uuid4())
    
    return TransactionRecord(
        id=unique_id,
        date=trans_date,
        type=draw(st.sampled_from([TransactionType.INCOME, TransactionType.EXPENSE, TransactionType.ORDER])),
        amount=draw(st.decimals(min_value=Decimal('0.01'), max_value=Decimal('100000'), places=2)),
        counterparty_id=draw(st.sampled_from(counterparty_ids)),
        description=draw(st.text(min_size=1, max_size=50, alphabet=st.characters(whitelist_categories=('L', 'N')))),
        category=draw(st.sampled_from(['销售', '采购', '费用', '其他'])),
        status=TransactionStatus.COMPLETED,
        created_at=datetime.now(),
        updated_at=datetime.now()
    )


@st.composite
def transaction_batch_for_period(draw, start_date, end_date, counterparty_ids, min_size=1, max_size=20):
    """生成指定期间内的一批交易记录"""
    assume(len(counterparty_ids) > 0)
    
    size = draw(st.integers(min_value=min_size, max_value=max_size))
    transactions = []
    
    # 确保至少有一些收入交易（用于客户排名）
    customer_ids = [cp_id for cp_id in counterparty_ids if cp_id in ['cp001', 'cp003']]  # 客户ID
    supplier_ids = [cp_id for cp_id in counterparty_ids if cp_id in ['cp002']]  # 供应商ID
    
    # 生成至少一个收入交易
    if customer_ids:
        income_transaction = draw(transaction_record_for_period(start_date, end_date, customer_ids))
        income_transaction.type = TransactionType.INCOME
        transactions.append(income_transaction)
    
    # 生成其余交易
    for _ in range(size - 1):
        transaction = draw(transaction_record_for_period(start_date, end_date, counterparty_ids))
        transactions.append(transaction)
    
    return transactions


class TestReportGeneratorProperties:
    """Property-based tests for report generator
    
    **Property 1: Management report completeness**
    For any valid date range and transaction data, when generating a management report,
    the output Excel file should contain revenue comparison charts, profit trend charts,
    and customer ranking reports.
    **Validates: Requirements 1.1**
    """
    
    @pytest.fixture(autouse=True)
    def setup_storage(self, tmp_path):
        """设置存储层 - 每个测试都使用独立的临时目录"""
        import uuid
        test_dir = tmp_path / str(uuid.uuid4())
        test_dir.mkdir()
        
        transaction_storage = TransactionStorage(str(test_dir / 'transactions.json'))
        counterparty_storage = CounterpartyStorage(str(test_dir / 'counterparties.json'))
        
        # 创建测试往来单位
        test_counterparties = [
            Counterparty(
                id='cp001',
                name='测试客户A',
                type=CounterpartyType.CUSTOMER,
                contact_person='张三',
                phone='13800138000',
                email='test@example.com',
                address='测试地址',
                tax_id='123456789012345',
                created_at=datetime.now(),
                updated_at=datetime.now()
            ),
            Counterparty(
                id='cp002',
                name='测试供应商B',
                type=CounterpartyType.SUPPLIER,
                contact_person='李四',
                phone='13900139000',
                email='supplier@example.com',
                address='供应商地址',
                tax_id='987654321098765',
                created_at=datetime.now(),
                updated_at=datetime.now()
            ),
            Counterparty(
                id='cp003',
                name='测试客户C',
                type=CounterpartyType.CUSTOMER,
                contact_person='王五',
                phone='13700137000',
                email='customer@example.com',
                address='客户地址',
                tax_id='456789012345678',
                created_at=datetime.now(),
                updated_at=datetime.now()
            )
        ]
        
        for cp in test_counterparties:
            counterparty_storage.add(cp)
        
        # 创建报表输出目录
        reports_dir = test_dir / 'reports'
        reports_dir.mkdir()
        
        generator = ReportGenerator(
            transaction_storage,
            counterparty_storage,
            str(reports_dir)
        )
        
        return {
            'generator': generator,
            'transaction_storage': transaction_storage,
            'counterparty_storage': counterparty_storage,
            'counterparty_ids': [cp.id for cp in test_counterparties],
            'reports_dir': str(reports_dir)
        }
    
    @given(data=st.data())
    @settings(max_examples=50, suppress_health_check=[HealthCheck.function_scoped_fixture], deadline=None)
    def test_management_report_completeness(self, setup_storage, data):
        """
        Property: 管理报表完整性属性
        
        验证：
        1. 对于任何有效的日期范围和交易数据，生成管理报表时
        2. 输出的Excel文件应该包含收支对比图、利润趋势图和客户排名报表
        3. 报表生成成功且文件存在
        4. Excel文件包含预期的工作表
        5. 每个工作表都包含相应的图表
        """
        generator = setup_storage['generator']
        transaction_storage = setup_storage['transaction_storage']
        counterparty_ids = setup_storage['counterparty_ids']
        reports_dir = setup_storage['reports_dir']
        
        # 生成测试数据
        start_date, end_date = data.draw(valid_date_range(max_days=90))
        transactions = data.draw(transaction_batch_for_period(
            start_date, end_date, counterparty_ids, min_size=3, max_size=15
        ))
        
        # 添加交易记录到存储
        for transaction in transactions:
            transaction_storage.add(transaction)
        
        try:
            # 生成管理报表
            result = generator.generate_management_report(
                start_date=start_date,
                end_date=end_date,
                company_name="测试公司"
            )
            
            # 验证报表生成成功
            assert result.success, \
                f"管理报表生成应该成功，但失败了: {result.error_message}"
            
            # 验证报表类型
            assert result.report_type == ReportType.MANAGEMENT, \
                "报表类型应该是MANAGEMENT"
            
            # 验证文件路径不为空
            assert result.file_path, \
                "报表文件路径不应该为空"
            
            # 验证文件存在
            assert os.path.exists(result.file_path), \
                f"报表文件应该存在: {result.file_path}"
            
            # 验证是有效的Excel文件
            try:
                wb = openpyxl.load_workbook(result.file_path)
            except Exception as e:
                assert False, f"生成的文件应该是有效的Excel文件: {str(e)}"
            
            # 验证包含必需的工作表
            expected_sheets = ['收支对比表', '利润趋势表', '客户排名表']
            actual_sheets = wb.sheetnames
            
            for expected_sheet in expected_sheets:
                assert expected_sheet in actual_sheets, \
                    f"Excel文件应该包含工作表 '{expected_sheet}'，但实际工作表为: {actual_sheets}"
            
            # 验证每个工作表都有内容（至少有一些数据）
            for sheet_name in expected_sheets:
                ws = wb[sheet_name]
                
                # 检查工作表不为空
                assert ws.max_row > 1, \
                    f"工作表 '{sheet_name}' 应该包含数据（不只是标题行）"
                
                assert ws.max_column > 1, \
                    f"工作表 '{sheet_name}' 应该包含多列数据"
                
                # 检查是否包含图表（作为图片嵌入）
                has_image = len(ws._images) > 0
                assert has_image, \
                    f"工作表 '{sheet_name}' 应该包含图表（图片）"
            
            # 验证日期范围正确
            assert result.data_period.start_date == start_date, \
                "报表的开始日期应该与请求的开始日期一致"
            
            assert result.data_period.end_date == end_date, \
                "报表的结束日期应该与请求的结束日期一致"
            
            # 验证生成日期合理（应该是最近生成的）
            time_diff = datetime.now() - result.generation_date
            assert time_diff.total_seconds() < 60, \
                "报表生成时间应该是最近的时间"
            
            wb.close()
            
        finally:
            # 清理生成的文件
            if 'result' in locals() and result.file_path and os.path.exists(result.file_path):
                try:
                    os.unlink(result.file_path)
                except Exception:
                    pass  # 忽略清理错误
    
    @given(data=st.data())
    @settings(max_examples=20, suppress_health_check=[HealthCheck.function_scoped_fixture], deadline=None)
    def test_management_report_with_no_data(self, setup_storage, data):
        """
        Property: 无数据时的管理报表处理
        
        验证：
        1. 当指定期间内没有交易数据时
        2. 报表生成应该失败并返回适当的错误消息
        3. 不应该生成文件
        """
        generator = setup_storage['generator']
        
        # 生成一个没有交易数据的日期范围
        start_date, end_date = data.draw(valid_date_range(max_days=30))
        
        # 不添加任何交易数据到存储
        
        # 生成管理报表
        result = generator.generate_management_report(
            start_date=start_date,
            end_date=end_date,
            company_name="测试公司"
        )
        
        # 验证报表生成失败
        assert not result.success, \
            "当没有交易数据时，管理报表生成应该失败"
        
        # 验证错误消息不为空
        assert result.error_message, \
            "应该提供错误消息说明为什么生成失败"
        
        # 验证文件路径为空或文件不存在
        assert not result.file_path or not os.path.exists(result.file_path), \
            "当生成失败时，不应该创建文件"
    
    @given(data=st.data())
    @settings(max_examples=30, suppress_health_check=[HealthCheck.function_scoped_fixture], deadline=None)
    def test_management_report_different_date_ranges(self, setup_storage, data):
        """
        Property: 不同日期范围的管理报表一致性
        
        验证：
        1. 对于不同的有效日期范围
        2. 管理报表都应该成功生成
        3. 报表结构保持一致（相同的工作表和图表）
        """
        generator = setup_storage['generator']
        transaction_storage = setup_storage['transaction_storage']
        counterparty_ids = setup_storage['counterparty_ids']
        
        # 生成多个不同的日期范围
        date_ranges = [
            data.draw(valid_date_range(max_days=7)),   # 一周
            data.draw(valid_date_range(max_days=30)),  # 一个月
            data.draw(valid_date_range(max_days=90))   # 三个月
        ]
        
        generated_files = []
        
        try:
            for start_date, end_date in date_ranges:
                # 为每个日期范围生成一些交易数据
                transactions = data.draw(transaction_batch_for_period(
                    start_date, end_date, counterparty_ids, min_size=2, max_size=8
                ))
                
                # 清理之前的数据
                for tx in transaction_storage.get_all():
                    transaction_storage.delete(tx.id)
                
                # 添加新的交易记录
                for transaction in transactions:
                    transaction_storage.add(transaction)
                
                # 生成报表
                result = generator.generate_management_report(
                    start_date=start_date,
                    end_date=end_date,
                    company_name="测试公司"
                )
                
                # 验证生成成功
                assert result.success, \
                    f"日期范围 {start_date} 到 {end_date} 的报表生成应该成功"
                
                # 验证文件存在
                assert os.path.exists(result.file_path), \
                    f"报表文件应该存在: {result.file_path}"
                
                generated_files.append(result.file_path)
                
                # 验证Excel结构一致性
                wb = openpyxl.load_workbook(result.file_path)
                expected_sheets = ['收支对比表', '利润趋势表', '客户排名表']
                
                assert set(wb.sheetnames) == set(expected_sheets), \
                    f"所有报表都应该包含相同的工作表结构"
                
                # 验证每个工作表都有图表
                for sheet_name in expected_sheets:
                    ws = wb[sheet_name]
                    assert len(ws._images) > 0, \
                        f"工作表 '{sheet_name}' 应该包含图表（图片）"
                
                wb.close()
        
        finally:
            # 清理生成的文件
            for file_path in generated_files:
                if os.path.exists(file_path):
                    try:
                        os.unlink(file_path)
                    except Exception:
                        pass
    
    @given(data=st.data())
    @settings(max_examples=30, suppress_health_check=[HealthCheck.function_scoped_fixture], deadline=None)
    def test_tax_report_structure_property(self, setup_storage, data):
        """
        Property: 税务报表结构属性 (Property 2)
        
        验证：
        1. 对于任何有效的税务期间和交易数据，生成税务报表时
        2. 输出应该包含正确的税务表单结构（增值税或所得税申报表）
        3. 所有必需字段都应该被填充
        4. 报表格式符合税务要求
        **Validates: Requirements 1.2**
        """
        generator = setup_storage['generator']
        transaction_storage = setup_storage['transaction_storage']
        counterparty_ids = setup_storage['counterparty_ids']
        
        # 生成测试数据
        start_date, end_date = data.draw(valid_date_range(max_days=90))
        transactions = data.draw(transaction_batch_for_period(
            start_date, end_date, counterparty_ids, min_size=5, max_size=20
        ))
        
        # 添加交易记录到存储
        for transaction in transactions:
            transaction_storage.add(transaction)
        
        # 测试两种税务报表类型
        tax_report_types = [
            ("vat", "2024年第一季度"),
            ("income_tax", "2024年第二季度")
        ]
        
        generated_files = []
        
        try:
            for report_type, period in tax_report_types:
                # 生成税务报表
                result = generator.generate_tax_report(
                    report_type=report_type,
                    period=period,
                    company_name="测试公司"
                )
                
                # 验证报表生成成功
                assert result.success, \
                    f"{report_type}税务报表生成应该成功，但失败了: {result.error_message}"
                
                # 验证报表类型
                if report_type == "vat":
                    assert result.report_type == ReportType.TAX_VAT, \
                        "增值税报表类型应该是TAX_VAT"
                else:
                    assert result.report_type == ReportType.TAX_INCOME, \
                        "所得税报表类型应该是TAX_INCOME"
                
                # 验证文件存在
                assert os.path.exists(result.file_path), \
                    f"税务报表文件应该存在: {result.file_path}"
                
                generated_files.append(result.file_path)
                
                # 验证Excel文件结构
                wb = openpyxl.load_workbook(result.file_path)
                
                # 验证至少有一个工作表
                assert len(wb.sheetnames) > 0, \
                    f"{report_type}税务报表应该包含至少一个工作表"
                
                # 验证工作表有内容
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    
                    # 检查工作表不为空
                    assert ws.max_row > 1, \
                        f"税务报表工作表 '{sheet_name}' 应该包含数据"
                    
                    assert ws.max_column > 1, \
                        f"税务报表工作表 '{sheet_name}' 应该包含多列数据"
                    
                    # 验证包含公司名称
                    found_company_name = False
                    for row in ws.iter_rows(max_row=10, values_only=True):
                        for cell_value in row:
                            if cell_value and "测试公司" in str(cell_value):
                                found_company_name = True
                                break
                        if found_company_name:
                            break
                    
                    assert found_company_name, \
                        f"税务报表应该包含公司名称"
                    
                    # 验证包含期间信息（更灵活的匹配）
                    found_period = False
                    period_keywords = ["2024", "第一季度", "第二季度", "季度", "Q1", "Q2"]
                    for row in ws.iter_rows(max_row=15, values_only=True):
                        for cell_value in row:
                            if cell_value:
                                cell_str = str(cell_value)
                                for keyword in period_keywords:
                                    if keyword in cell_str:
                                        found_period = True
                                        break
                            if found_period:
                                break
                        if found_period:
                            break
                    
                    # 如果没有找到期间信息，至少验证有数值数据（税务计算结果）
                    if not found_period:
                        found_numeric_data = False
                        for row in ws.iter_rows(values_only=True):
                            for cell_value in row:
                                if isinstance(cell_value, (int, float)) and cell_value != 0:
                                    found_numeric_data = True
                                    break
                            if found_numeric_data:
                                break
                        
                        assert found_numeric_data, \
                            f"税务报表应该包含数值数据（税务计算结果）"
                
                wb.close()
        
        finally:
            # 清理生成的文件
            for file_path in generated_files:
                if os.path.exists(file_path):
                    try:
                        os.unlink(file_path)
                    except Exception:
                        pass
    
    @given(data=st.data())
    @settings(max_examples=30, suppress_health_check=[HealthCheck.function_scoped_fixture], deadline=None)
    def test_bank_loan_report_completeness_property(self, setup_storage, data):
        """
        Property: 银行贷款报表完整性属性 (Property 3)
        
        验证：
        1. 对于任何有效的报表日期和交易数据，生成银行贷款报表时
        2. 输出应该包含资产负债表、利润表和现金流量表
        3. 每个报表都包含必需的财务项目和数据
        4. 报表格式符合银行贷款申请要求
        **Validates: Requirements 1.3**
        """
        generator = setup_storage['generator']
        transaction_storage = setup_storage['transaction_storage']
        counterparty_ids = setup_storage['counterparty_ids']
        
        # 生成测试数据
        report_date = data.draw(st.dates(
            min_value=date(2020, 1, 1),
            max_value=date.today()
        ))
        
        # 生成一些历史交易数据（截至报表日期）
        start_date = report_date - timedelta(days=365)  # 一年的历史数据
        transactions = data.draw(transaction_batch_for_period(
            start_date, report_date, counterparty_ids, min_size=10, max_size=30
        ))
        
        # 添加交易记录到存储
        for transaction in transactions:
            transaction_storage.add(transaction)
        
        try:
            # 生成银行贷款报表
            result = generator.generate_bank_loan_report(
                report_date=report_date,
                company_name="测试公司"
            )
            
            # 验证报表生成成功
            assert result.success, \
                f"银行贷款报表生成应该成功，但失败了: {result.error_message}"
            
            # 验证报表类型
            assert result.report_type == ReportType.BANK_LOAN, \
                "报表类型应该是BANK_LOAN"
            
            # 验证文件路径不为空
            assert result.file_path, \
                "银行贷款报表文件路径不应该为空"
            
            # 验证文件存在
            assert os.path.exists(result.file_path), \
                f"银行贷款报表文件应该存在: {result.file_path}"
            
            # 验证是有效的Excel文件
            try:
                wb = openpyxl.load_workbook(result.file_path)
            except Exception as e:
                assert False, f"生成的文件应该是有效的Excel文件: {str(e)}"
            
            # 验证包含必需的工作表（银行贷款报表的三大财务报表）
            expected_sheets = ['资产负债表', '利润表', '现金流量表']
            actual_sheets = wb.sheetnames
            
            for expected_sheet in expected_sheets:
                assert expected_sheet in actual_sheets, \
                    f"银行贷款报表应该包含工作表 '{expected_sheet}'，但实际工作表为: {actual_sheets}"
            
            # 验证每个工作表都有内容和必需的财务项目
            
            # 1. 验证资产负债表
            balance_sheet = wb['资产负债表']
            assert balance_sheet.max_row > 1, \
                "资产负债表应该包含数据（不只是标题行）"
            
            # 检查资产负债表的关键项目
            balance_sheet_content = []
            for row in balance_sheet.iter_rows(values_only=True):
                for cell_value in row:
                    if cell_value:
                        balance_sheet_content.append(str(cell_value))
            
            balance_sheet_text = ' '.join(balance_sheet_content)
            
            # 验证包含资产项目
            asset_items = ['货币资金', '应收账款', '资产']
            found_asset_items = [item for item in asset_items if item in balance_sheet_text]
            assert len(found_asset_items) > 0, \
                f"资产负债表应该包含资产项目，如: {asset_items}"
            
            # 验证包含负债或权益项目
            liability_equity_items = ['应付账款', '负债', '所有者权益', '实收资本', '权益']
            found_liability_equity = [item for item in liability_equity_items if item in balance_sheet_text]
            assert len(found_liability_equity) > 0, \
                f"资产负债表应该包含负债或权益项目，如: {liability_equity_items}"
            
            # 2. 验证利润表
            income_statement = wb['利润表']
            assert income_statement.max_row > 1, \
                "利润表应该包含数据（不只是标题行）"
            
            # 检查利润表的关键项目
            income_statement_content = []
            for row in income_statement.iter_rows(values_only=True):
                for cell_value in row:
                    if cell_value:
                        income_statement_content.append(str(cell_value))
            
            income_statement_text = ' '.join(income_statement_content)
            
            # 验证包含收入项目
            income_items = ['营业收入', '收入']
            found_income_items = [item for item in income_items if item in income_statement_text]
            assert len(found_income_items) > 0, \
                f"利润表应该包含收入项目，如: {income_items}"
            
            # 验证包含成本费用项目
            cost_items = ['营业成本', '成本', '费用']
            found_cost_items = [item for item in cost_items if item in income_statement_text]
            assert len(found_cost_items) > 0, \
                f"利润表应该包含成本费用项目，如: {cost_items}"
            
            # 验证包含利润项目
            profit_items = ['利润', '净利润', '营业利润']
            found_profit_items = [item for item in profit_items if item in income_statement_text]
            assert len(found_profit_items) > 0, \
                f"利润表应该包含利润项目，如: {profit_items}"
            
            # 3. 验证现金流量表
            cash_flow = wb['现金流量表']
            assert cash_flow.max_row > 1, \
                "现金流量表应该包含数据（不只是标题行）"
            
            # 检查现金流量表的关键项目
            cash_flow_content = []
            for row in cash_flow.iter_rows(values_only=True):
                for cell_value in row:
                    if cell_value:
                        cash_flow_content.append(str(cell_value))
            
            cash_flow_text = ' '.join(cash_flow_content)
            
            # 验证包含现金流项目
            cash_flow_items = ['现金流', '经营活动', '现金', '流量']
            found_cash_flow_items = [item for item in cash_flow_items if item in cash_flow_text]
            assert len(found_cash_flow_items) > 0, \
                f"现金流量表应该包含现金流项目，如: {cash_flow_items}"
            
            # 验证报表日期正确
            assert result.data_period.start_date == report_date, \
                "银行贷款报表的开始日期应该与报表日期一致"
            
            assert result.data_period.end_date == report_date, \
                "银行贷款报表的结束日期应该与报表日期一致"
            
            # 验证生成日期合理（应该是最近生成的）
            time_diff = datetime.now() - result.generation_date
            assert time_diff.total_seconds() < 60, \
                "银行贷款报表生成时间应该是最近的时间"
            
            # 验证每个工作表都包含公司名称
            for sheet_name in expected_sheets:
                ws = wb[sheet_name]
                found_company_name = False
                for row in ws.iter_rows(max_row=10, values_only=True):
                    for cell_value in row:
                        if cell_value and "测试公司" in str(cell_value):
                            found_company_name = True
                            break
                    if found_company_name:
                        break
                
                assert found_company_name, \
                    f"工作表 '{sheet_name}' 应该包含公司名称"
            
            # 验证每个工作表都包含数值数据（财务数据）
            for sheet_name in expected_sheets:
                ws = wb[sheet_name]
                found_numeric_data = False
                for row in ws.iter_rows(values_only=True):
                    for cell_value in row:
                        if isinstance(cell_value, (int, float)) and cell_value != 0:
                            found_numeric_data = True
                            break
                    if found_numeric_data:
                        break
                
                # 如果没有数值数据，至少应该有格式化的数字字符串
                if not found_numeric_data:
                    found_formatted_numbers = False
                    for row in ws.iter_rows(values_only=True):
                        for cell_value in row:
                            if cell_value and isinstance(cell_value, str):
                                # 检查是否包含数字格式（如"1,000.00"）
                                import re
                                if re.search(r'\d+[,.]?\d*', str(cell_value)):
                                    found_formatted_numbers = True
                                    break
                        if found_formatted_numbers:
                            break
                    
                    assert found_formatted_numbers, \
                        f"工作表 '{sheet_name}' 应该包含财务数据（数值或格式化数字）"
            
            wb.close()
            
        finally:
            # 清理生成的文件
            if 'result' in locals() and result.file_path and os.path.exists(result.file_path):
                try:
                    os.unlink(result.file_path)
                except Exception:
                    pass  # 忽略清理错误
    
    @given(data=st.data())
    @settings(max_examples=20, suppress_health_check=[HealthCheck.function_scoped_fixture], deadline=None)
    def test_excel_output_format_consistency_property(self, setup_storage, data):
        """
        Property: Excel输出格式一致性属性 (Property 4)
        
        验证：
        1. 对于任何报表类型和数据，生成的Excel文件格式应该一致
        2. 所有报表都应该包含标准的格式元素（标题、公司名称、日期等）
        3. 数值格式应该统一（货币格式、百分比格式等）
        4. 工作表结构应该符合预期模式
        **Validates: Requirements 1.4, 1.6**
        """
        generator = setup_storage['generator']
        transaction_storage = setup_storage['transaction_storage']
        counterparty_ids = setup_storage['counterparty_ids']
        
        # 生成测试数据
        start_date, end_date = data.draw(valid_date_range(max_days=60))
        transactions = data.draw(transaction_batch_for_period(
            start_date, end_date, counterparty_ids, min_size=5, max_size=15
        ))
        
        # 添加交易记录到存储
        for transaction in transactions:
            transaction_storage.add(transaction)
        
        generated_files = []
        
        try:
            # 测试不同类型的报表
            report_configs = [
                {
                    'type': 'management',
                    'method': 'generate_management_report',
                    'args': {
                        'start_date': start_date,
                        'end_date': end_date,
                        'company_name': "测试公司"
                    }
                },
                {
                    'type': 'tax_vat',
                    'method': 'generate_tax_report',
                    'args': {
                        'report_type': 'vat',
                        'period': '2024年第一季度',
                        'company_name': "测试公司"
                    }
                },
                {
                    'type': 'bank_loan',
                    'method': 'generate_bank_loan_report',
                    'args': {
                        'report_date': end_date,
                        'company_name': "测试公司"
                    }
                }
            ]
            
            report_results = []
            
            # 生成所有类型的报表
            for config in report_configs:
                method = getattr(generator, config['method'])
                result = method(**config['args'])
                
                # 验证报表生成成功
                assert result.success, \
                    f"{config['type']}报表生成应该成功，但失败了: {result.error_message}"
                
                assert os.path.exists(result.file_path), \
                    f"{config['type']}报表文件应该存在: {result.file_path}"
                
                generated_files.append(result.file_path)
                report_results.append((config['type'], result))
            
            # 验证所有报表的格式一致性
            for report_type, result in report_results:
                wb = openpyxl.load_workbook(result.file_path)
                
                # 验证每个工作表的基本格式要求
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    
                    # 1. 验证工作表不为空
                    assert ws.max_row > 1, \
                        f"{report_type}报表的工作表 '{sheet_name}' 应该包含数据"
                    
                    assert ws.max_column > 1, \
                        f"{report_type}报表的工作表 '{sheet_name}' 应该包含多列数据"
                    
                    # 2. 验证包含公司名称（在前几行中）
                    found_company_name = False
                    for row in ws.iter_rows(max_row=10, values_only=True):
                        for cell_value in row:
                            if cell_value and "测试公司" in str(cell_value):
                                found_company_name = True
                                break
                        if found_company_name:
                            break
                    
                    assert found_company_name, \
                        f"{report_type}报表的工作表 '{sheet_name}' 应该包含公司名称"
                    
                    # 3. 验证包含日期信息（报表期间或生成日期）
                    found_date_info = False
                    date_keywords = ["2024", "2020", "2021", "2022", "2023", "2025", "2026", 
                                   "年", "月", "日", "期间", "截至"]
                    
                    for row in ws.iter_rows(max_row=15, values_only=True):
                        for cell_value in row:
                            if cell_value:
                                cell_str = str(cell_value)
                                for keyword in date_keywords:
                                    if keyword in cell_str:
                                        found_date_info = True
                                        break
                            if found_date_info:
                                break
                        if found_date_info:
                            break
                    
                    assert found_date_info, \
                        f"{report_type}报表的工作表 '{sheet_name}' 应该包含日期信息"
                    
                    # 4. 验证数值格式一致性（检查是否有数值数据）
                    has_numeric_data = False
                    numeric_formats = set()
                    
                    for row in ws.iter_rows(values_only=False):
                        for cell in row:
                            if cell.value is not None:
                                # 检查数值类型
                                if isinstance(cell.value, (int, float)):
                                    has_numeric_data = True
                                    if cell.number_format:
                                        numeric_formats.add(cell.number_format)
                                
                                # 检查公式（可能包含数值计算）
                                elif isinstance(cell.value, str) and cell.value.startswith('='):
                                    has_numeric_data = True
                    
                    # 对于财务报表，应该包含数值数据
                    if report_type in ['management', 'bank_loan']:
                        assert has_numeric_data, \
                            f"{report_type}报表的工作表 '{sheet_name}' 应该包含数值数据"
                    
                    # 5. 验证工作表结构合理性（行列数在合理范围内）
                    assert ws.max_row <= 1000, \
                        f"{report_type}报表的工作表 '{sheet_name}' 行数过多: {ws.max_row}"
                    
                    assert ws.max_column <= 50, \
                        f"{report_type}报表的工作表 '{sheet_name}' 列数过多: {ws.max_column}"
                    
                    # 6. 验证没有完全空白的行（在数据区域内）
                    data_rows = 0
                    empty_rows = 0
                    
                    for row in ws.iter_rows(min_row=2, values_only=True):  # 跳过标题行
                        has_content = any(cell_value is not None and str(cell_value).strip() 
                                        for cell_value in row)
                        if has_content:
                            data_rows += 1
                        else:
                            empty_rows += 1
                    
                    # 数据行应该多于空行（避免过多空白）
                    if data_rows > 0:
                        empty_ratio = empty_rows / (data_rows + empty_rows)
                        assert empty_ratio < 0.8, \
                            f"{report_type}报表的工作表 '{sheet_name}' 空行比例过高: {empty_ratio:.2%}"
                
                wb.close()
            
            # 验证不同报表类型的文件命名一致性
            for report_type, result in report_results:
                filename = os.path.basename(result.file_path)
                
                # 验证文件扩展名
                assert filename.endswith('.xlsx'), \
                    f"{report_type}报表文件应该是Excel格式(.xlsx): {filename}"
                
                # 验证文件名包含报表类型信息
                type_keywords = {
                    'management': ['管理报表'],
                    'tax_vat': ['增值税', '申报表'],
                    'bank_loan': ['银行', '贷款', '报表']
                }
                
                if report_type in type_keywords:
                    found_type_keyword = False
                    for keyword in type_keywords[report_type]:
                        if keyword in filename:
                            found_type_keyword = True
                            break
                    
                    assert found_type_keyword, \
                        f"{report_type}报表文件名应该包含类型关键词: {filename}"
                
                # 验证文件名包含日期信息
                import re
                # 更灵活的日期模式匹配
                date_patterns = [
                    r'\d{8}',           # YYYYMMDD格式
                    r'\d{4}',           # YYYY格式
                    r'\d{4}年',         # YYYY年格式
                    r'Q\d',             # Q1, Q2等季度格式
                    r'第[一二三四]季度',  # 中文季度格式
                    r'\d{1,2}月'        # 月份格式
                ]
                
                found_date_pattern = False
                for pattern in date_patterns:
                    if re.search(pattern, filename):
                        found_date_pattern = True
                        break
                
                assert found_date_pattern, \
                    f"{report_type}报表文件名应该包含日期信息: {filename}"
            
            # 验证所有报表的生成时间在合理范围内
            generation_times = [result.generation_date for _, result in report_results]
            time_span = max(generation_times) - min(generation_times)
            
            assert time_span.total_seconds() < 300, \
                f"所有报表的生成时间跨度应该在5分钟内: {time_span.total_seconds()}秒"
            
        finally:
            # 清理生成的文件
            for file_path in generated_files:
                if os.path.exists(file_path):
                    try:
                        os.unlink(file_path)
                    except Exception:
                        pass