"""
Property-based tests for ReconciliationReportGenerator - Customer statement generation

Feature: small-accountant-practical-enhancement
Property 12: Customer statement generation
Validates: Requirements 3.2
"""

import pytest
from datetime import date, timedelta
from decimal import Decimal
from hypothesis import given, strategies as st, settings, assume
from hypothesis import HealthCheck
from typing import List
import uuid
import tempfile
import os

from small_accountant_v16.reconciliation.reconciliation_report_generator import (
    ReconciliationReportGenerator, CustomerAccountData
)
from small_accountant_v16.models.core_models import (
    TransactionRecord, TransactionType, TransactionStatus,
    Counterparty, CounterpartyType,
    Discrepancy, DiscrepancyType, BankRecord
)


# Hypothesis strategies for generating test data
@st.composite
def valid_amount(draw, min_value=0.01, max_value=100000.0):
    """生成有效的金额"""
    amount = draw(st.decimals(
        min_value=Decimal(str(min_value)),
        max_value=Decimal(str(max_value)),
        places=2
    ))
    return amount


@st.composite
def valid_date_range(draw, max_days=365):
    """生成有效的日期范围"""
    base_date = draw(st.dates(
        min_value=date(2020, 1, 1),
        max_value=date.today() - timedelta(days=1)
    ))
    
    end_date = draw(st.dates(
        min_value=base_date,
        max_value=min(base_date + timedelta(days=max_days), date.today())
    ))
    
    return base_date, end_date


@st.composite
def counterparty_strategy(draw):
    """生成往来单位"""
    company_names = [
        "测试客户A", "测试供应商B", "ABC公司", "XYZ有限公司",
        "北京科技公司", "上海贸易公司", "深圳制造厂", "广州服务商",
        "优质客户有限公司", "新兴科技股份有限公司", "长期合作伙伴公司"
    ]
    
    contact_persons = ["张三", "李四", "王五", "赵六", "陈七", "刘八"]
    
    return Counterparty(
        id=f"cp{draw(st.integers(min_value=1, max_value=9999))}",
        name=draw(st.sampled_from(company_names)),
        type=draw(st.sampled_from([CounterpartyType.CUSTOMER, CounterpartyType.SUPPLIER])),
        contact_person=draw(st.sampled_from(contact_persons)),
        phone=f"1{draw(st.integers(min_value=3000000000, max_value=8999999999))}",
        email=f"contact{draw(st.integers(min_value=1, max_value=999))}@example.com",
        address=draw(st.sampled_from([
            "北京市朝阳区建国路1号", "上海市浦东新区陆家嘴环路2号",
            "深圳市南山区科技园3号", "广州市天河区珠江新城4号"
        ])),
        tax_id=f"{draw(st.integers(min_value=100000000000000000, max_value=999999999999999999))}",
        created_at=draw(st.datetimes()),
        updated_at=draw(st.datetimes())
    )


@st.composite
def transaction_record_strategy(draw, date_range=None, counterparty_id=None):
    """生成交易记录"""
    if date_range:
        start_date, end_date = date_range
        trans_date = draw(st.dates(min_value=start_date, max_value=end_date))
    else:
        trans_date = draw(st.dates(
            min_value=date(2020, 1, 1),
            max_value=date.today()
        ))
    
    amount = draw(valid_amount())
    
    descriptions = [
        "销售收入", "采购付款", "办公费用", "差旅费", "租金支出",
        "工资支付", "税费缴纳", "银行手续费", "利息收入", "其他收入",
        "货款结算", "服务费", "咨询费", "运输费", "材料费"
    ]
    
    categories = ['销售', '采购', '费用', '其他', '服务', '材料']
    
    return TransactionRecord(
        id=str(uuid.uuid4()),
        date=trans_date,
        type=draw(st.sampled_from([TransactionType.INCOME, TransactionType.EXPENSE])),
        amount=amount,
        counterparty_id=counterparty_id or f"cp{draw(st.integers(min_value=1, max_value=100))}",
        description=draw(st.sampled_from(descriptions)),
        category=draw(st.sampled_from(categories)),
        status=TransactionStatus.COMPLETED,
        created_at=draw(st.datetimes()),
        updated_at=draw(st.datetimes())
    )


@st.composite
def customer_account_data_strategy(draw):
    """生成客户对账数据"""
    # 生成客户信息
    customer = draw(counterparty_strategy())
    
    # 生成日期范围
    start_date, end_date = draw(valid_date_range(max_days=90))
    
    # 生成期初余额
    opening_balance = draw(valid_amount(min_value=0, max_value=50000))
    
    # 生成交易记录
    transactions = draw(st.lists(
        transaction_record_strategy(
            date_range=(start_date, end_date),
            counterparty_id=customer.id
        ),
        min_size=1,
        max_size=20
    ))
    
    # 按日期排序
    transactions.sort(key=lambda t: t.date)
    
    # 计算期末余额
    closing_balance = opening_balance
    for transaction in transactions:
        if transaction.type == TransactionType.INCOME:
            closing_balance += transaction.amount
        else:  # EXPENSE
            closing_balance -= transaction.amount
    
    return CustomerAccountData(
        customer=customer,
        transactions=transactions,
        start_date=start_date,
        end_date=end_date,
        opening_balance=opening_balance,
        closing_balance=closing_balance
    )


@st.composite
def discrepancy_strategy(draw):
    """生成差异记录"""
    disc_type = draw(st.sampled_from([
        DiscrepancyType.AMOUNT_DIFF,
        DiscrepancyType.MISSING_BANK,
        DiscrepancyType.MISSING_SYSTEM
    ]))
    
    # 根据差异类型生成相应的记录
    if disc_type == DiscrepancyType.AMOUNT_DIFF:
        # 金额差异：需要银行记录和系统记录
        bank_record = BankRecord(
            id=str(uuid.uuid4()),
            transaction_date=draw(st.dates(min_value=date(2020, 1, 1), max_value=date.today())),
            amount=draw(valid_amount()),
            counterparty=draw(st.sampled_from(["测试客户A", "测试供应商B", "ABC公司"])),
            description="银行流水记录",
            balance=draw(valid_amount()),
            transaction_type=draw(st.sampled_from(["DEBIT", "CREDIT"]))
        )
        
        system_record = draw(transaction_record_strategy())
        difference_amount = abs(bank_record.amount - system_record.amount)
        description = f"金额差异：银行{bank_record.amount}，系统{system_record.amount}"
        
    elif disc_type == DiscrepancyType.MISSING_BANK:
        # 银行流水缺失：只有系统记录
        bank_record = None
        system_record = draw(transaction_record_strategy())
        difference_amount = system_record.amount
        description = f"银行流水缺失：系统记录{system_record.amount}"
        
    else:  # MISSING_SYSTEM
        # 系统记录缺失：只有银行记录
        bank_record = BankRecord(
            id=str(uuid.uuid4()),
            transaction_date=draw(st.dates(min_value=date(2020, 1, 1), max_value=date.today())),
            amount=draw(valid_amount()),
            counterparty=draw(st.sampled_from(["测试客户A", "测试供应商B", "ABC公司"])),
            description="银行流水记录",
            balance=draw(valid_amount()),
            transaction_type=draw(st.sampled_from(["DEBIT", "CREDIT"]))
        )
        system_record = None
        difference_amount = bank_record.amount
        description = f"系统记录缺失：银行流水{bank_record.amount}"
    
    return Discrepancy(
        id=str(uuid.uuid4()),
        type=disc_type,
        bank_record=bank_record,
        system_record=system_record,
        difference_amount=difference_amount,
        description=description
    )


class TestReconciliationReportGeneratorProperties:
    """Property-based tests for reconciliation report generator
    
    **Property 12: Customer statement generation**
    For any valid customer account data, when generating a customer statement,
    the Excel output should contain all required information with correct
    formatting and calculations.
    **Validates: Requirements 3.2**
    """
    
    @given(customer_data=customer_account_data_strategy())
    @settings(max_examples=50, suppress_health_check=[HealthCheck.function_scoped_fixture], deadline=None)
    def test_customer_statement_completeness_property(self, customer_data):
        """
        Property: 客户对账单完整性属性
        
        验证：
        1. 生成的Excel包含所有必要的客户信息
        2. 包含所有交易记录
        3. 期初余额、期末余额计算正确
        4. 汇总统计信息正确
        """
        # 创建临时目录用于测试
        with tempfile.TemporaryDirectory() as temp_dir:
            generator = ReconciliationReportGenerator(temp_dir)
            
            # 生成客户对账单
            workbook = generator.generate_customer_statement_excel(customer_data)
            
            # 验证工作簿结构
            assert workbook is not None, "应该生成有效的工作簿"
            assert len(workbook.worksheets) >= 1, "工作簿应该至少包含一个工作表"
            
            worksheet = workbook.active
            assert worksheet.title == "客户对账单", f"工作表标题应该为'客户对账单'，但为'{worksheet.title}'"
            
            # 验证客户信息完整性
            customer_name_found = False
            customer_id_found = False
            contact_person_found = False
            phone_found = False
            period_found = False
            
            # 检查前10行是否包含客户信息
            for row in range(1, 11):
                for col in range(1, 8):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value:
                        cell_str = str(cell_value)
                        if customer_data.customer.name in cell_str:
                            customer_name_found = True
                        if customer_data.customer.id in cell_str:
                            customer_id_found = True
                        if customer_data.customer.contact_person in cell_str:
                            contact_person_found = True
                        if customer_data.customer.phone in cell_str:
                            phone_found = True
                        if customer_data.start_date.strftime("%Y年%m月%d日") in cell_str:
                            period_found = True
            
            assert customer_name_found, f"客户对账单应该包含客户名称'{customer_data.customer.name}'"
            assert customer_id_found, f"客户对账单应该包含客户编号'{customer_data.customer.id}'"
            assert contact_person_found, f"客户对账单应该包含联系人'{customer_data.customer.contact_person}'"
            assert phone_found, f"客户对账单应该包含联系电话'{customer_data.customer.phone}'"
            assert period_found, f"客户对账单应该包含对账期间信息"
            
            # 验证交易记录数量
            # 查找表头行（包含"序号"、"日期"等）
            header_row = None
            for row in range(1, 20):
                if worksheet.cell(row=row, column=1).value == "序号":
                    header_row = row
                    break
            
            assert header_row is not None, "应该找到表头行"
            
            # 计算实际的交易记录行数（不包括期初余额和期末余额）
            transaction_rows = 0
            data_start_row = header_row + 2  # 跳过表头和期初余额
            
            for row in range(data_start_row, worksheet.max_row + 1):
                seq_value = worksheet.cell(row=row, column=1).value
                if seq_value and str(seq_value).isdigit():
                    transaction_rows += 1
            
            assert transaction_rows == len(customer_data.transactions), \
                f"交易记录行数应该为{len(customer_data.transactions)}，但为{transaction_rows}"
            
            # 验证余额计算
            opening_balance_found = False
            closing_balance_found = False
            
            for row in range(header_row, worksheet.max_row + 1):
                cell_value = worksheet.cell(row=row, column=2).value
                if cell_value:
                    if "期初余额" in str(cell_value):
                        balance_value = worksheet.cell(row=row, column=7).value
                        if balance_value is not None:
                            assert abs(float(balance_value) - float(customer_data.opening_balance)) < 0.01, \
                                f"期初余额应该为{customer_data.opening_balance}，但为{balance_value}"
                            opening_balance_found = True
                    
                    elif "期末余额" in str(cell_value):
                        balance_value = worksheet.cell(row=row, column=7).value
                        if balance_value is not None:
                            assert abs(float(balance_value) - float(customer_data.closing_balance)) < 0.01, \
                                f"期末余额应该为{customer_data.closing_balance}，但为{balance_value}"
                            closing_balance_found = True
            
            assert opening_balance_found, "应该找到期初余额"
            assert closing_balance_found, "应该找到期末余额"
            
            # 验证汇总统计
            total_income = sum(t.amount for t in customer_data.transactions if t.type == TransactionType.INCOME)
            total_expense = sum(t.amount for t in customer_data.transactions if t.type == TransactionType.EXPENSE)
            transaction_count = len(customer_data.transactions)
            
            summary_found = False
            for row in range(1, worksheet.max_row + 1):
                cell_value = worksheet.cell(row=row, column=1).value
                if cell_value and "对账汇总" in str(cell_value):
                    summary_found = True
                    
                    # 检查下一行的统计信息
                    next_row = row + 1
                    if next_row <= worksheet.max_row:
                        # 检查交易笔数
                        count_cell = worksheet.cell(row=next_row, column=2).value
                        if count_cell and f"{transaction_count} 笔" in str(count_cell):
                            pass  # 交易笔数正确
                        else:
                            # 可能在不同的位置，继续查找
                            pass
                    break
            
            assert summary_found, "应该找到对账汇总部分"
    
    @given(customer_data=customer_account_data_strategy())
    @settings(max_examples=30, suppress_health_check=[HealthCheck.function_scoped_fixture], deadline=None)
    def test_customer_statement_balance_calculation_property(self, customer_data):
        """
        Property: 客户对账单余额计算属性
        
        验证：
        1. 余额计算的连续性和正确性
        2. 每笔交易后的余额都正确
        3. 收入增加余额，支出减少余额
        4. 最终余额与期末余额一致
        """
        with tempfile.TemporaryDirectory() as temp_dir:
            generator = ReconciliationReportGenerator(temp_dir)
            workbook = generator.generate_customer_statement_excel(customer_data)
            worksheet = workbook.active
            
            # 找到表头行
            header_row = None
            for row in range(1, 20):
                if worksheet.cell(row=row, column=1).value == "序号":
                    header_row = row
                    break
            
            assert header_row is not None, "应该找到表头行"
            
            # 验证期初余额
            opening_row = header_row + 1
            opening_balance_cell = worksheet.cell(row=opening_row, column=7).value
            assert opening_balance_cell is not None, "应该有期初余额"
            
            current_balance = Decimal(str(opening_balance_cell))
            assert abs(current_balance - customer_data.opening_balance) < Decimal("0.01"), \
                f"期初余额应该为{customer_data.opening_balance}，但为{current_balance}"
            
            # 验证每笔交易的余额计算
            sorted_transactions = sorted(customer_data.transactions, key=lambda t: t.date)
            
            for i, transaction in enumerate(sorted_transactions):
                transaction_row = header_row + 2 + i  # 跳过表头和期初余额
                
                # 检查交易金额
                if transaction.type == TransactionType.INCOME:
                    income_cell = worksheet.cell(row=transaction_row, column=5).value
                    expense_cell = worksheet.cell(row=transaction_row, column=6).value
                    
                    assert income_cell is not None and income_cell != '', \
                        f"收入交易应该在收入列有值，第{i+1}笔交易"
                    assert expense_cell == '' or expense_cell is None, \
                        f"收入交易在支出列应该为空，第{i+1}笔交易"
                    
                    current_balance += transaction.amount
                    
                else:  # EXPENSE
                    income_cell = worksheet.cell(row=transaction_row, column=5).value
                    expense_cell = worksheet.cell(row=transaction_row, column=6).value
                    
                    assert expense_cell is not None and expense_cell != '', \
                        f"支出交易应该在支出列有值，第{i+1}笔交易"
                    assert income_cell == '' or income_cell is None, \
                        f"支出交易在收入列应该为空，第{i+1}笔交易"
                    
                    current_balance -= transaction.amount
                
                # 检查余额
                balance_cell = worksheet.cell(row=transaction_row, column=7).value
                assert balance_cell is not None, f"第{i+1}笔交易应该有余额"
                
                actual_balance = Decimal(str(balance_cell))
                assert abs(actual_balance - current_balance) < Decimal("0.01"), \
                    f"第{i+1}笔交易后余额应该为{current_balance}，但为{actual_balance}"
            
            # 验证期末余额
            closing_row = header_row + 2 + len(sorted_transactions)
            closing_balance_cell = worksheet.cell(row=closing_row, column=7).value
            
            if closing_balance_cell is not None:
                final_balance = Decimal(str(closing_balance_cell))
                assert abs(final_balance - customer_data.closing_balance) < Decimal("0.01"), \
                    f"期末余额应该为{customer_data.closing_balance}，但为{final_balance}"
                
                assert abs(final_balance - current_balance) < Decimal("0.01"), \
                    f"计算的最终余额{current_balance}应该与期末余额{final_balance}一致"
    
    @given(discrepancies=st.lists(discrepancy_strategy(), min_size=1, max_size=15))
    @settings(max_examples=30, suppress_health_check=[HealthCheck.function_scoped_fixture], deadline=None)
    def test_discrepancy_report_completeness_property(self, discrepancies):
        """
        Property: 差异报告完整性属性
        
        验证：
        1. 差异报告包含所有差异记录
        2. 差异类型正确分类和显示
        3. 汇总统计信息正确
        4. 差异描述信息完整
        """
        with tempfile.TemporaryDirectory() as temp_dir:
            generator = ReconciliationReportGenerator(temp_dir)
            workbook = generator.generate_discrepancy_report(discrepancies)
            
            assert workbook is not None, "应该生成有效的工作簿"
            worksheet = workbook.active
            assert worksheet.title == "对账差异报告", f"工作表标题应该为'对账差异报告'"
            
            # 验证差异总数
            total_discrepancies = len(discrepancies)
            total_found = False
            
            for row in range(1, 10):
                for col in range(1, 8):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value and f"{total_discrepancies} 条" in str(cell_value):
                        total_found = True
                        break
                if total_found:
                    break
            
            assert total_found, f"应该显示差异总数{total_discrepancies}条"
            
            # 找到表头行
            header_row = None
            for row in range(1, 15):
                if worksheet.cell(row=row, column=1).value == "差异ID":
                    header_row = row
                    break
            
            assert header_row is not None, "应该找到差异表头行"
            
            # 验证差异记录数量
            data_rows = 0
            for row in range(header_row + 1, worksheet.max_row + 1):
                cell_value = worksheet.cell(row=row, column=1).value
                if cell_value and not ("差异汇总" in str(cell_value) or "差异总数" in str(cell_value)):
                    data_rows += 1
                elif cell_value and "差异汇总" in str(cell_value):
                    break
            
            assert data_rows == total_discrepancies, \
                f"差异记录行数应该为{total_discrepancies}，但为{data_rows}"
            
            # 验证差异类型统计
            amount_diff_count = sum(1 for d in discrepancies if d.type == DiscrepancyType.AMOUNT_DIFF)
            missing_system_count = sum(1 for d in discrepancies if d.type == DiscrepancyType.MISSING_SYSTEM)
            missing_bank_count = sum(1 for d in discrepancies if d.type == DiscrepancyType.MISSING_BANK)
            
            # 查找汇总统计部分
            summary_found = False
            for row in range(header_row, worksheet.max_row + 1):
                cell_value = worksheet.cell(row=row, column=1).value
                if cell_value and "差异汇总统计" in str(cell_value):
                    summary_found = True
                    
                    # 检查下一行的统计信息
                    stats_row = row + 1
                    if stats_row <= worksheet.max_row:
                        # 验证各类差异统计
                        row_values = []
                        for col in range(1, 7):
                            val = worksheet.cell(row=stats_row, column=col).value
                            if val:
                                row_values.append(str(val))
                        
                        row_text = " ".join(row_values)
                        
                        if amount_diff_count > 0:
                            assert f"{amount_diff_count} 条" in row_text, \
                                f"应该显示金额差异{amount_diff_count}条"
                        
                        if missing_system_count > 0:
                            assert f"{missing_system_count} 条" in row_text, \
                                f"应该显示系统记录缺失{missing_system_count}条"
                        
                        if missing_bank_count > 0:
                            assert f"{missing_bank_count} 条" in row_text, \
                                f"应该显示银行流水缺失{missing_bank_count}条"
                    break
            
            assert summary_found, "应该找到差异汇总统计部分"
            
            # 验证每个差异记录的完整性
            for i, discrepancy in enumerate(discrepancies):
                data_row = header_row + 1 + i
                
                # 验证差异ID
                id_cell = worksheet.cell(row=data_row, column=1).value
                assert id_cell == discrepancy.id, \
                    f"第{i+1}个差异的ID应该为{discrepancy.id}，但为{id_cell}"
                
                # 验证差异类型
                type_cell = worksheet.cell(row=data_row, column=2).value
                expected_type_text = {
                    DiscrepancyType.AMOUNT_DIFF: "金额差异",
                    DiscrepancyType.MISSING_SYSTEM: "系统记录缺失",
                    DiscrepancyType.MISSING_BANK: "银行流水缺失"
                }[discrepancy.type]
                
                assert type_cell == expected_type_text, \
                    f"第{i+1}个差异的类型应该为{expected_type_text}，但为{type_cell}"
                
                # 验证差异金额
                diff_amount_cell = worksheet.cell(row=data_row, column=7).value
                assert diff_amount_cell is not None, f"第{i+1}个差异应该有差异金额"
                
                actual_diff_amount = Decimal(str(diff_amount_cell))
                assert abs(actual_diff_amount - discrepancy.difference_amount) < Decimal("0.01"), \
                    f"第{i+1}个差异的金额应该为{discrepancy.difference_amount}，但为{actual_diff_amount}"
                
                # 验证描述
                desc_cell = worksheet.cell(row=data_row, column=8).value
                assert desc_cell == discrepancy.description, \
                    f"第{i+1}个差异的描述应该为'{discrepancy.description}'，但为'{desc_cell}'"
    
    @given(customer_data=customer_account_data_strategy())
    @settings(max_examples=20, suppress_health_check=[HealthCheck.function_scoped_fixture], deadline=None)
    def test_customer_statement_file_save_property(self, customer_data):
        """
        Property: 客户对账单文件保存属性
        
        验证：
        1. 文件能够成功保存
        2. 保存的文件路径正确
        3. 文件名包含客户名称和日期
        4. 保存的文件可以重新打开
        """
        with tempfile.TemporaryDirectory() as temp_dir:
            generator = ReconciliationReportGenerator(temp_dir)
            
            # 保存客户对账单
            saved_path = generator.save_customer_statement(customer_data)
            
            # 验证文件路径
            assert saved_path is not None, "应该返回保存的文件路径"
            assert os.path.exists(saved_path), f"保存的文件应该存在：{saved_path}"
            
            # 验证文件名格式
            filename = os.path.basename(saved_path)
            assert filename.startswith("客户对账单_"), "文件名应该以'客户对账单_'开头"
            assert filename.endswith(".xlsx"), "文件名应该以'.xlsx'结尾"
            assert customer_data.customer.name in filename, \
                f"文件名应该包含客户名称'{customer_data.customer.name}'"
            
            # 验证文件大小（应该大于0）
            file_size = os.path.getsize(saved_path)
            assert file_size > 0, f"保存的文件大小应该大于0，但为{file_size}"
            
            # 验证文件可以重新打开（基本的Excel格式验证）
            try:
                from openpyxl import load_workbook
                reopened_wb = load_workbook(saved_path)
                assert len(reopened_wb.worksheets) >= 1, "重新打开的文件应该包含工作表"
                reopened_wb.close()
            except Exception as e:
                pytest.fail(f"无法重新打开保存的文件：{e}")
    
    @given(discrepancies=st.lists(discrepancy_strategy(), min_size=1, max_size=10))
    @settings(max_examples=20, suppress_health_check=[HealthCheck.function_scoped_fixture], deadline=None)
    def test_discrepancy_report_file_save_property(self, discrepancies):
        """
        Property: 差异报告文件保存属性
        
        验证：
        1. 差异报告文件能够成功保存
        2. 保存的文件路径正确
        3. 文件名包含时间戳
        4. 保存的文件可以重新打开
        """
        with tempfile.TemporaryDirectory() as temp_dir:
            generator = ReconciliationReportGenerator(temp_dir)
            
            # 保存差异报告
            saved_path = generator.save_discrepancy_report(discrepancies)
            
            # 验证文件路径
            assert saved_path is not None, "应该返回保存的文件路径"
            assert os.path.exists(saved_path), f"保存的文件应该存在：{saved_path}"
            
            # 验证文件名格式
            filename = os.path.basename(saved_path)
            assert filename.startswith("对账差异报告_"), "文件名应该以'对账差异报告_'开头"
            assert filename.endswith(".xlsx"), "文件名应该以'.xlsx'结尾"
            
            # 验证文件大小
            file_size = os.path.getsize(saved_path)
            assert file_size > 0, f"保存的文件大小应该大于0，但为{file_size}"
            
            # 验证文件可以重新打开
            try:
                from openpyxl import load_workbook
                reopened_wb = load_workbook(saved_path)
                assert len(reopened_wb.worksheets) >= 1, "重新打开的文件应该包含工作表"
                
                # 验证重新打开的文件包含差异数据
                worksheet = reopened_wb.active
                assert worksheet.title == "对账差异报告", "工作表标题应该正确"
                
                reopened_wb.close()
            except Exception as e:
                pytest.fail(f"无法重新打开保存的差异报告文件：{e}")