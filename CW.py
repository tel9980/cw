# -*- coding: utf-8 -*-
"""
飞书财务小助手V8.8 全场景自动化终极版
✅ 新增：自动对账+税务统计+批量导入导出+异常自愈
✅ 集成：你的Bot + Wiki + 台账
✅ 适配：lark-oapi V2 SDK
"""
import os
import sys
import json
import time
import shutil
import logging
import requests
import itertools
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import filedialog
from concurrent.futures import ThreadPoolExecutor, as_completed
import base64
from io import BytesIO
from PIL import Image, ImageGrab
from datetime import datetime, timedelta
from dotenv import load_dotenv

# Lark OAPI V2 Imports
import lark_oapi as lark
from lark_oapi.api.bitable.v1 import *
from lark_oapi.api.bitable.v1.model import *

# ZhipuAI Import
from zhipuai import ZhipuAI

class Color:
    HEADER = '\033[95m'
    BLUE = '\033[94m'
    OKBLUE = '\033[94m'
    CYAN = '\033[96m'
    GREEN = '\033[92m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

# Compatibility constants for Field Types
class FT:
    TEXT = 1
    NUMBER = 2
    SELECT = 3
    MULTI_SELECT = 4
    DATE = 5
    CHECKBOX = 7
    USER = 11
    PHONE = 13
    URL = 15
    ATTACHMENT = 17
    LINK = 18
    LOOKUP = 19
    FORMULA = 20
    DUPLEX_LINK = 21
    LOCATION = 22
    GROUP_CHAT = 23
    CREATED_TIME = 1001
    MODIFIED_TIME = 1002
    CREATED_USER = 1003
    MODIFIED_USER = 1004

# 加载环境变量
# load_dotenv() # Moved to after path config

# -------------------------------------------------------------------------
# 新增功能：启动引导向导
# -------------------------------------------------------------------------
def setup_wizard():
    """新手引导配置向导"""
    if os.path.exists(".env"):
        # 检查关键变量是否存在
        with open(".env", "r", encoding="utf-8") as f:
            content = f.read()
        if "FEISHU_APP_ID" in content and "FEISHU_APP_SECRET" in content and "FEISHU_APP_TOKEN" in content:
            return # 配置已存在，跳过

    print(f"{Color.HEADER}===============================================")
    print(f"       👋 欢迎使用飞书财务小助手！(初次设置)")
    print(f"==============================================={Color.ENDC}")
    print(f"{Color.CYAN}检测到您是第一次运行或配置文件缺失。{Color.ENDC}")
    print("请按照提示输入飞书开放平台的 App ID 和 Secret。")
    print("如果您还没有这些信息，请先去 open.feishu.cn 创建企业自建应用。")
    print("-" * 50)
    
    app_id = input(f"👉 请输入 {Color.BOLD}App ID{Color.ENDC} (cli_...): ").strip()
    app_secret = input(f"👉 请输入 {Color.BOLD}App Secret{Color.ENDC}: ").strip()
    app_token = input(f"👉 请输入 {Color.BOLD}App Token{Color.ENDC} (多维表格的base_token): ").strip()
    
    if not app_id or not app_secret or not app_token:
        print(f"{Color.FAIL}❌ 输入不完整，无法继续。{Color.ENDC}")
        sys.exit(1)
        
    # 写入 .env
    with open(".env", "w", encoding="utf-8") as f:
        f.write(f"FEISHU_APP_ID={app_id}\n")
        f.write(f"FEISHU_APP_SECRET={app_secret}\n")
        f.write(f"FEISHU_APP_TOKEN={app_token}\n")
        f.write("VAT_RATE=3\n")
        f.write("TOLERANCE_DAYS=2\n")
        
    print(f"\n{Color.GREEN}✅ 配置已保存！正在启动...{Color.ENDC}\n")
    # 重新加载
    load_dotenv()
    global APP_ID, APP_SECRET, APP_TOKEN
    APP_ID = os.getenv("FEISHU_APP_ID")
    APP_SECRET = os.getenv("FEISHU_APP_SECRET")
    APP_TOKEN = os.getenv("FEISHU_APP_TOKEN")

# 辅助：交互式文件选择
def select_file_interactively(pattern="*.xlsx", prompt="请选择文件"):
    """列出当前目录下匹配的文件供用户选择 (优先尝试 GUI)"""
    # 尝试使用 GUI 选择文件
    try:
        import tkinter as tk
        from tkinter import filedialog
        
        print(f"\n📂 正在打开文件选择窗口...")
        root = tk.Tk()
        root.withdraw() # 隐藏主窗口
        root.attributes('-topmost', True) # 置顶
        
        # 优先打开待处理目录
        init_dir = os.getcwd()
        if 'PENDING_DIR' in globals() and os.path.exists(PENDING_DIR):
            init_dir = PENDING_DIR
        
        file_path = filedialog.askopenfilename(
            title=prompt,
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialdir=init_dir
        )
        root.destroy()
        
        if file_path:
            # 转换为相对路径以保持显示整洁 (如果是在当前目录下)
            try:
                rel_path = os.path.relpath(file_path, os.getcwd())
                if not rel_path.startswith(".."):
                    return rel_path
                return file_path
            except:
                return file_path
        else:
            print("   (用户取消了选择)")
            # 如果用户取消，回退到列表模式？或者直接返回None
            # 让我们回退到列表模式，以防万一GUI不好用
    except Exception as e:
        print(f"⚠️ GUI 启动失败 ({e})，切换回列表模式。")

    import glob
    # 搜索当前目录和待处理目录
    files = [f for f in glob.glob(pattern) if not f.startswith("~$")]
    
    if 'PENDING_DIR' in globals() and os.path.exists(PENDING_DIR):
        pending_files = [os.path.join(PENDING_DIR, f) for f in os.listdir(PENDING_DIR) 
                         if f.lower().endswith('.xlsx') and not f.startswith("~$")]
        files.extend(pending_files)
        
    # 去重
    files = list(set(files))
    
    if not files:
        return None
        
    print(f"\n📂 {prompt} (列表模式):")
    for i, f in enumerate(files):
        print(f"  {i+1}. {f}")
    print(f"  0. 手动输入路径")
    
    while True:
        choice = input(f"👉 请选择 (1-{len(files)}, 0): ").strip()
        if choice == '0':
            return None
        if choice.isdigit():
            idx = int(choice) - 1
            if 0 <= idx < len(files):
                return files[idx]
        print(f"{Color.FAIL}❌ 无效选择{Color.ENDC}")

# 辅助：选择文件
def select_file(title="请选择Excel文件"):
    root = tk.Tk()
    root.withdraw() # 隐藏主窗口
    root.attributes('-topmost', True) # 置顶
    
    # 优先打开待处理目录
    init_dir = os.getcwd()
    if 'PENDING_DIR' in globals() and os.path.exists(PENDING_DIR):
        init_dir = PENDING_DIR
            
    file_path = filedialog.askopenfilename(
        title=title,
        filetypes=[("Excel files", "*.xlsx;*.xls")],
        initialdir=init_dir
    )
    root.destroy()
    return file_path

# -------------------------- 路径配置 --------------------------
ROOT_DIR = os.path.dirname(os.path.abspath(__file__)) if getattr(sys, 'frozen', False) else os.getcwd()
DATA_ROOT = os.path.join(ROOT_DIR, "财务数据")
CONFIG_DIR = os.path.join(DATA_ROOT, "配置文件")
REPORT_DIR = os.path.join(DATA_ROOT, "查询报告")
BACKUP_DIR = os.path.join(DATA_ROOT, "自动备份")
TEMPLATE_DIR = os.path.join(DATA_ROOT, "Excel模版")
LOG_DIR = os.path.join(DATA_ROOT, "运行日志")
ARCHIVE_DIR = os.path.join(DATA_ROOT, "已处理归档")
PENDING_DIR = os.path.join(DATA_ROOT, "待处理单据")

# 确保目录存在
for d in [DATA_ROOT, CONFIG_DIR, REPORT_DIR, BACKUP_DIR, TEMPLATE_DIR, LOG_DIR, ARCHIVE_DIR, PENDING_DIR]:
    os.makedirs(d, exist_ok=True)

# 加载环境变量 (优先加载配置文件目录下的，兼容根目录)
env_path_config = os.path.join(CONFIG_DIR, ".env")
env_path_data = os.path.join(DATA_ROOT, ".env") # Support env in Data Root
env_path_root = os.path.join(ROOT_DIR, ".env")

if os.path.exists(env_path_config):
    load_dotenv(env_path_config)
    print(f"🔧 已加载配置: {env_path_config}")
elif os.path.exists(env_path_data):
    load_dotenv(env_path_data)
    print(f"🔧 已加载配置: {env_path_data}")
elif os.path.exists(env_path_root):
    load_dotenv(env_path_root)
    
# 文件路径常量
FILE_CATEGORY_RULES = os.path.join(CONFIG_DIR, "category_rules.json")
FILE_PARTNER_ALIASES = os.path.join(CONFIG_DIR, "partner_aliases.json")
FILE_VOUCHER_TEMPLATES = os.path.join(CONFIG_DIR, "voucher_templates.json")
FILE_AI_CACHE = os.path.join(DATA_ROOT, "ai_category_cache.json")
FILE_DASHBOARD_CACHE = os.path.join(DATA_ROOT, "dashboard_cache.json")

# 自动迁移旧文件
def migrate_legacy_files():
    # 1. Migrate Files
    moves = [
        ("category_rules.json", FILE_CATEGORY_RULES),
        ("partner_aliases.json", FILE_PARTNER_ALIASES),
        ("voucher_templates.json", FILE_VOUCHER_TEMPLATES),
        ("ai_category_cache.json", FILE_AI_CACHE),
        ("dashboard_cache.json", FILE_DASHBOARD_CACHE),
    ]
    for src_name, dst_path in moves:
        src = os.path.join(ROOT_DIR, src_name)
        if os.path.exists(src) and not os.path.exists(dst_path):
            try:
                shutil.move(src, dst_path)
                print(f"📦 已迁移: {src_name} -> {dst_path}")
            except: pass

    # 2. Migrate Directories
    dir_moves = [
        ("待处理单据", PENDING_DIR),
        ("财务数据备份", BACKUP_DIR),
        ("查询报告", REPORT_DIR),
        ("Excel模版", TEMPLATE_DIR),
        ("运行日志", LOG_DIR),
        ("已处理归档", ARCHIVE_DIR)
    ]
    for src_name, dst_path in dir_moves:
        src = os.path.join(ROOT_DIR, src_name)
        # Avoid moving if src is same as dst (e.g. if ROOT_DIR is already DATA_ROOT's parent correctly configured)
        if os.path.exists(src) and os.path.abspath(src) != os.path.abspath(dst_path):
            try:
                if not os.path.exists(dst_path):
                    shutil.move(src, dst_path)
                    print(f"📦 已迁移目录: {src_name} -> {dst_path}")
                else:
                    # Merge contents
                    for item in os.listdir(src):
                        s = os.path.join(src, item)
                        d = os.path.join(dst_path, item)
                        if not os.path.exists(d):
                            shutil.move(s, d)
                    # Try remove empty src dir
                    try:
                        os.rmdir(src)
                    except: pass
            except Exception as e:
                pass

migrate_legacy_files()

# -------------------------- 核心配置 --------------------------
LOG_FILE = os.path.join(LOG_DIR, f"feishu_table_log_{datetime.now().strftime('%Y%m%d')}.log")
TEST_PRODUCT_COUNT = 10
TEST_LEDGER_COUNT = 5
# TABLE_NAME 在此处意为 Base Name (应用名称)
BASE_NAME = "飞书财务台账-2026"
BOT_WEBHOOK = os.getenv("BOT_WEBHOOK", "")
WIKI_LINK = os.getenv("WIKI_LINK", "")
WIKI_EXCEPTION = f"{WIKI_LINK}# 异常排查" if WIKI_LINK else "请联系管理员"
WIKI_TAX = f"{WIKI_LINK}# 税务申报" if WIKI_LINK else "请联系管理员"
LOCAL_FOLDER = BACKUP_DIR
# os.makedirs(LOCAL_FOLDER, exist_ok=True) # Already created in path config

# 业务配置
VAT_RATE = float(os.getenv("VAT_RATE", 3))
TOLERANCE_DAYS = int(os.getenv("TOLERANCE_DAYS", 2))
# -------------------------------------------------------------------------

# 初始化日志
class SolutionFormatter(logging.Formatter):
    def format(self, record):
        if not hasattr(record, 'solution'):
            record.solution = "无"
        return super().format(record)

file_handler = logging.FileHandler(LOG_FILE, encoding="utf-8")
stream_handler = logging.StreamHandler()

# 使用自定义Formatter
formatter = SolutionFormatter("%(asctime)s - %(levelname)s - %(message)s - 解决方案：%(solution)s")
file_handler.setFormatter(formatter)
stream_handler.setFormatter(formatter)

logging.basicConfig(
    level=logging.INFO,
    handlers=[file_handler, stream_handler],
    force=True # Ensure we override any existing config
)
logger = logging.getLogger(__name__)

class SolutionLogAdapter(logging.LoggerAdapter):
    def process(self, msg, kwargs):
        kwargs["extra"] = {"solution": self.extra.get("solution", "无")}
        return msg, kwargs
log = SolutionLogAdapter(logger, {"solution": "无"})

# 加载环境变量 (已在上方加载)
APP_ID = os.getenv("FEISHU_APP_ID")
APP_SECRET = os.getenv("FEISHU_APP_SECRET")
APP_TOKEN = os.getenv("FEISHU_APP_TOKEN") # Base Token
USER_ID = os.getenv("FEISHU_USER_ID", "")
BOSS_ID = os.getenv("BOSS_FEISHU_ID", "")
ZHIPUAI_API_KEY = os.getenv("ZHIPUAI_API_KEY", "")

# 初始化 GLM-4 客户端
zhipu_client = None
if ZHIPUAI_API_KEY:
    try:
        zhipu_client = ZhipuAI(api_key=ZHIPUAI_API_KEY)
        # log.info("🧠 GLM-4 AI 模型已加载", extra={"solution": "无"}) # Avoid logging too early if log not setup, but log is setup at line 95
    except Exception as e:
        pass

# 税务配置
TAX_CONFIG = {
    "is_small": os.getenv("IS_SMALL", "true").lower() == "true",
    "vat_rate": float(os.getenv("VAT_RATE", 3)),
    "corporate_tax_rate": float(os.getenv("CORP_TAX_RATE", 25)),
    "surtax_rates": {"city": 7, "education": 3, "local_education": 2}
}
RECONCILE_THRESHOLD = float(os.getenv("RECONCILE_THRESHOLD", 0.01))

def ai_guess_category(description, partner):
    """使用 AI 猜测交易分类"""
    if not zhipu_client: return None
    
    try:
        # 构造提示词
        prompt = f"""
你是一名资深会计。请根据交易描述判断费用类型。
交易对象: {partner}
交易摘要: {description}
可选分类: [差旅费-交通, 差旅费-住宿, 差旅费-加油, 业务招待费, 办公费, 房租物业, 水电费, 快递费, 营销推广费, 技术服务费, 采购款, 员工工资, 社保公积金, 税费]
如果不确定，请根据经验推断最可能的分类。
只返回分类名称，不要其他废话。
"""
        resp = zhipu_client.chat.completions.create(
            model="glm-4-flash",
            messages=[{"role": "user", "content": prompt}]
        )
        category = resp.choices[0].message.content.strip()
        # 简单清洗
        category = category.replace("分类：", "").replace("。", "").strip()
        return category
    except Exception as e:
        # print(f"AI error: {e}")
        return None

# 重试装饰器
def retry_on_failure(max_retries=3, delay=3):
    def decorator(func):
        def wrapper(*args, **kwargs):
            retries = 0
            while retries < max_retries:
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    retries += 1
                    log.error(f"❌ 函数{func.__name__}执行失败（第{retries}次重试）：{str(e)}",
                              extra={"solution": f"等待{delay}秒后重试"})
                    time.sleep(delay)
            log.error(f"❌ 函数{func.__name__}重试{max_retries}次失败",
                      extra={"solution": f"查看Wiki：{WIKI_EXCEPTION}"})
            return False
        return wrapper
    return decorator

# 发送Bot消息 (支持卡片)
def send_bot_message(content, msg_type="text", card_data=None):
    if not BOT_WEBHOOK:
        log.warning("⚠️ 未配置Bot Webhook，跳过消息推送", extra={"solution": "在.env配置BOT_WEBHOOK"})
        return

    headers = {"Content-Type": "application/json"}
    
    if msg_type == "interactive" and card_data:
        payload = {
            "msg_type": "interactive",
            "card": card_data
        }
    else:
        # 默认文本消息
        payload = {
            "msg_type": "text",
            "content": {"text": content}
        }
        
    try:
        resp = requests.post(BOT_WEBHOOK, json=payload, headers=headers)
        if resp.status_code == 200:
            resp_json = resp.json()
            if resp_json.get("code") == 0:
                log.info("✅ Bot推送成功", extra={"solution": "无"})
                return True
            else:
                log.error(f"❌ Bot推送失败：{resp_json.get('msg')}", extra={"solution": "检查Bot配置"})
                return False
        else:
            log.error(f"❌ Bot网络错误：{resp.status_code}", extra={"solution": "检查网络"})
            return False
    except Exception as e:
        log.error(f"❌ Bot推送异常：{str(e)}", extra={"solution": "检查网络"})
        return False

# 初始化客户端
@retry_on_failure(max_retries=3, delay=3)
def init_clients():
    if not APP_ID or not APP_SECRET:
        log.error("❌ 未配置 APP_ID 或 APP_SECRET", extra={"solution": "请在 .env 文件中配置"})
        return None
        
    client = lark.Client.builder() \
        .app_id(APP_ID) \
        .app_secret(APP_SECRET) \
        .build()
    log.info("✅ 飞书客户端初始化成功", extra={"solution": "无"})
    send_bot_message("飞书财务小助手V8.8已启动 (Lark OAPI V2)", "accountant")
    return client

# 辅助：根据表名获取TableID
def get_table_id_by_name(client, app_token, table_name):
    req = ListAppTableRequest.builder() \
        .app_token(app_token) \
        .page_size(50) \
        .build()
    resp = client.bitable.v1.app_table.list(req)
    if not resp.success():
        log.error(f"❌ 获取表格列表失败: {resp.msg}", extra={"solution": "检查App Token"})
        return None
        
    if resp.data and resp.data.items:
        for table in resp.data.items:
            if table.name == table_name:
                return table.table_id
    return None

# 批量导入Excel
@retry_on_failure(max_retries=2, delay=3)
def import_from_excel(client, app_token, excel_path=None):
    try:
        # 如果没有指定路径，尝试交互式选择或弹窗
        if not excel_path:
            # 优先尝试交互式选择
            excel_path = select_file_interactively("*.xlsx", "请选择要导入的数据文件")
            
            # 如果还是没有，回退到弹窗
            if not excel_path:
                log.info("📂 请选择导入数据的Excel文件...", extra={"solution": "弹窗选择"})
                excel_path = select_file("请选择要导入的Excel文件")
                
            if not excel_path:
                log.warning("⚠️ 未选择文件，操作取消", extra={"solution": "无"})
                return False

        # 导入基础信息表
        with pd.ExcelFile(excel_path) as excel_file:
            if "基础信息表" in excel_file.sheet_names:
                table_id = get_table_id_by_name(client, app_token, "基础信息表")
                if table_id:
                    df = pd.read_excel(excel_file, sheet_name="基础信息表").fillna("")
                    records = []
                    for _, row in df.iterrows():
                        fields = {
                            "产品名称": str(row["产品名称"]),
                            "单位成本": float(row["单位成本"]) if row["单位成本"] != "" else 0,
                            "备注": str(row.get("备注", ""))
                        }
                        records.append(AppTableRecord.builder().fields(fields).build())
                    
                    # 分批写入 (API限制每次100条)
                    batch_size = 100
                    for i in range(0, len(records), batch_size):
                        batch = records[i:i+batch_size]
                        req = BatchCreateAppTableRecordRequest.builder() \
                            .app_token(app_token) \
                            .table_id(table_id) \
                            .request_body(BatchCreateAppTableRecordRequestBody.builder().records(batch).build()) \
                            .build()
                        resp = client.bitable.v1.app_table_record.batch_create(req)
                        if not resp.success():
                            log.error(f"❌ 基础信息表部分导入失败: {resp.msg}", extra={"solution": "检查数据格式"})
                    log.info(f"✅ 基础信息表导入完成: {len(records)}条", extra={"solution": "无"})
                else:
                    log.error("❌ 未找到'基础信息表'", extra={"solution": "请先创建表格"})

        # 导入日常台账表 (优化：支持任意Sheet名，智能识别表头)
        table_id = get_table_id_by_name(client, app_token, "日常台账表")
        if table_id:
            # 1. 智能读取Excel数据
            df = read_excel_smart(excel_path).fillna("")
            if df.empty:
                 log.info("⚠️ Excel中没有有效数据", extra={"solution": "请检查文件内容"})
                 return True

            # 2. 智能银行识别 (基于文件名)
            filename = os.path.basename(excel_path).upper()
            default_bank = "N银行（现金）"
            default_is_cash = "是"
            default_has_ticket = "无票"
            
            if "微信" in filename or "WECHAT" in filename:
                default_bank = "微信"
                default_is_cash = "是"
                default_has_ticket = "无票"
            elif "支付宝" in filename or "ALIPAY" in filename:
                default_bank = "支付宝"
                default_is_cash = "是"
                default_has_ticket = "无票"
            elif "工商" in filename or "ICBC" in filename:
                default_bank = "工商银行"
                default_is_cash = "否"
                default_has_ticket = "有票"
            elif "G银行" in filename:
                default_bank = "G银行基本户"
                default_is_cash = "否"
                default_has_ticket = "有票"

            log.info(f"🤖 智能识别默认银行: {default_bank}", extra={"solution": "如需修改请重命名文件"})

            # 3. 获取日期范围用于过滤查询 (优化)
            min_ts = None
            max_ts = None
            valid_dates = []
            for idx, row in df.iterrows():
                try:
                    dt = pd.to_datetime(row["记账日期"])
                    ts = int(dt.timestamp() * 1000)
                    valid_dates.append(ts)
                except:
                    pass
            
            filter_cmd = None
            if valid_dates:
                min_ts = min(valid_dates) - 24*3600*1000 # 放宽1天
                max_ts = max(valid_dates) + 24*3600*1000
                filter_cmd = f'CurrentValue.[记账日期]>={min_ts}&&CurrentValue.[记账日期]<={max_ts}'
                log.info(f"📅 启用日期范围过滤: {pd.to_datetime(min_ts, unit='ms').date()} 至 {pd.to_datetime(max_ts, unit='ms').date()}", extra={"solution": "无"})

            # 4. 获取现有记录 (仅获取必要字段 + 日期过滤)
            log.info("🔍 正在拉取现有数据进行去重检查...", extra={"solution": "无"})
            required_fields = ["记账日期", "实际收付金额", "业务类型", "备注"]
            existing_records = get_all_records(client, app_token, table_id, filter_info=filter_cmd, field_names=required_fields)
            
            existing_hashes = set()
            existing_meta = [] # 用于模糊查重
            for r in existing_records:
                f = r.fields
                d = f.get("记账日期", 0)
                a = round(float(f.get("实际收付金额", 0)), 2)
                t = f.get("业务类型", "")
                m = str(f.get("备注", ""))[:10]
                existing_hashes.add(f"{d}_{a}_{t}_{m}")
                existing_meta.append({"ts": d, "amt": a, "type": t})
            
            log.info(f"✅ 已索引 {len(existing_hashes)} 条现有记录", extra={"solution": "无"})

            records = []
            skipped_count = 0
            possible_dup_count = 0
            
            for _, row in df.iterrows():
                # 预处理数据以生成Hash
                try:
                    r_date_str = str(row["记账日期"])
                    if not r_date_str: continue
                    
                    # 处理日期格式
                    if isinstance(row["记账日期"], (int, float)):
                        ts = int(pd.to_datetime(row["记账日期"]).timestamp() * 1000)
                    else:
                        ts = int(pd.to_datetime(row["记账日期"]).timestamp() * 1000)
                        
                    r_amt = round(float(row["实际收付金额"]), 2)
                    r_type = str(row["业务类型"])
                    r_memo = str(row.get("备注", ""))[:10]
                    
                    # 1. 严格查重 (完全跳过)
                    row_hash = f"{ts}_{r_amt}_{r_type}_{r_memo}"
                    if row_hash in existing_hashes:
                        skipped_count += 1
                        continue

                    # 2. 智能模糊查重 (仅提醒)
                    # 规则: 金额相同 + 类型相同 + 日期相差在 48小时内
                    for ex in existing_meta:
                        if abs(ex["amt"] - r_amt) < 0.01 and ex["type"] == r_type:
                            if abs(ex["ts"] - ts) <= 48 * 3600 * 1000: # 48小时
                                log.warning(f"⚠️ 发现疑似重复数据: {r_date_str} {r_amt} {r_type} (库中已有相近记录)", extra={"solution": "请人工核对"})
                                possible_dup_count += 1
                                break
                        
                except Exception as e:
                    log.warning(f"⚠️ 数据行解析失败跳过: {e}", extra={"solution": "检查日期/金额格式"})
                    continue

                desc = str(row.get("往来单位费用", ""))
                
                # 优化：解析别名
                resolved_desc = resolve_partner(desc)
                
                # 如果户名列无效，尝试从摘要列匹配别名
                if resolved_desc == desc:
                    memo = str(row.get("备注", ""))
                    memo_resolved = resolve_partner(memo)
                    if memo_resolved != memo:
                        resolved_desc = memo_resolved
                
                desc = resolved_desc
                if not desc or desc == "nan" or desc == "未知" or desc == "":
                    desc = "散户" # 默认

                # 尝试自动分类补全 (费用归类)
                category = str(row.get("费用归类", ""))
                if not category or category == "nan" or category == "未知" or category == "":
                    memo = str(row.get("备注", ""))
                    category = auto_categorize(memo, "其他", partner_name=desc)
                    
                fields = {
                    "记账日期": ts,
                    "凭证号": int(row.get("凭证号", 0)) if str(row.get("凭证号", "")).strip() != "" else 0,
                    "业务类型": r_type,
                    "费用归类": category,
                    "往来单位费用": desc,
                    "账面金额": float(row.get("账面金额", 0)),
                    "实际收付金额": r_amt,
                    "交易银行": str(row.get("交易银行", "")) or default_bank,
                    "是否现金": str(row.get("是否现金", "")) or default_is_cash,
                    "发票流水单号": str(row.get("发票流水单号", "")),
                    "是否有票": str(row.get("是否有票", "")) or default_has_ticket,
                    "待补票标记": str(row.get("待补票标记", "无")),
                    "有票成本": float(row.get("有票成本", 0)),
                    "无票成本": float(row.get("无票成本", 0)),
                    "本次实际利润": float(row.get("本次实际利润", 0)),
                    "手工式分录": str(row.get("手工式分录", "")),
                    "操作人": str(row.get("操作人", USER_ID)),
                    "合同订单号": str(row.get("合同订单号", "")),
                    "备注": str(row.get("备注", ""))
                }
                    
                records.append(AppTableRecord.builder().fields(fields).build())
            
            if skipped_count > 0:
                log.info(f"⏭️ 已自动跳过 {skipped_count} 条重复记录", extra={"solution": "无"})
            
            if not records:
                log.info("✅ 没有新数据需要导入", extra={"solution": "无"})
                return True

            batch_size = 100
            for i in range(0, len(records), batch_size):
                batch = records[i:i+batch_size]
                req = BatchCreateAppTableRecordRequest.builder() \
                    .app_token(app_token) \
                    .table_id(table_id) \
                    .request_body(BatchCreateAppTableRecordRequestBody.builder().records(batch).build()) \
                    .build()
                resp = client.bitable.v1.app_table_record.batch_create(req)
                if not resp.success():
                     log.error(f"❌ 日常台账表部分导入失败: {resp.msg}", extra={"solution": "检查数据"})
            log.info(f"✅ 日常台账表导入完成: {len(records)}条", extra={"solution": "无"})
        else:
             log.error("❌ 未找到'日常台账表'", extra={"solution": "请先创建表格"})
                 
        # 导入成功后，静默刷新仪表盘缓存
        try:
            update_dashboard_cache_silent(client, app_token)
        except:
            pass
        return True
    except Exception as e:
        log.error(f"❌ Excel导入异常：{str(e)}", extra={"solution": "检查文件"})
        return False

# 辅助：获取所有记录 (支持过滤和字段选择，带TTL缓存)
# 缓存结构: {(table_id, filter_str, fields_str): (timestamp, records)}
RECORD_CACHE = {}
CACHE_TTL = 300 # 5分钟

def get_all_records(client, app_token, table_id, filter_info=None, field_names=None, use_cache=False):
    """
    获取所有记录
    use_cache: 是否使用内存缓存 (默认False，对于频繁读取的场景建议开启)
    """
    global RECORD_CACHE
    
    # 构造缓存Key
    cache_key = (table_id, str(filter_info), str(field_names))
    
    # 检查缓存
    if use_cache:
        if cache_key in RECORD_CACHE:
            ts, cached_records = RECORD_CACHE[cache_key]
            if time.time() - ts < CACHE_TTL:
                # 缓存有效
                return cached_records
            else:
                # 缓存过期
                del RECORD_CACHE[cache_key]

    records = []
    page_token = None
    
    # 只有当开启缓存且数据量很大时才显示进度提示
    # 为了简化，暂不加进度条
    
    while True:
        builder = ListAppTableRecordRequest.builder() \
            .app_token(app_token) \
            .table_id(table_id) \
            .page_size(100)
        
        if filter_info:
            builder.filter(filter_info)
            
        if field_names:
            builder.field_names(field_names)
            
        if page_token:
            builder.page_token(page_token)
            
        req = builder.build()
        resp = client.bitable.v1.app_table_record.list(req)
        if not resp.success():
            log.error(f"❌ 获取记录失败: {resp.msg}", extra={"solution": "检查网络或Token"})
            break
        if resp.data.items:
            records.extend(resp.data.items)
        if not resp.data.has_more:
            break
        page_token = resp.data.page_token
        
    # 写入缓存
    if use_cache:
        RECORD_CACHE[cache_key] = (time.time(), records)
        
    return records

# 自动分类规则 (关键词 -> 往来单位/费用类型)
def load_category_rules():
    default_rules = {
        "电费": "水电费",
        "水费": "水电费",
        "燃气": "水电费",
        "中石化": "差旅费-加油",
        "中石油": "差旅费-加油",
        "滴滴": "差旅费-交通",
        "铁路": "差旅费-交通",
        "航空": "差旅费-交通",
        "餐饮": "业务招待费",
        "酒店": "差旅费-住宿",
        "住宿": "差旅费-住宿",
        "工资": "工资薪金",
        "社保": "社保公积金",
        "公积金": "社保公积金",
        "税": "税费",
        "利息": "财务费用-利息",
        "手续费": "财务费用-手续费",
        "租金": "房租",
        "物业": "物业费",
        "推广": "市场推广费",
        "广告": "市场推广费",
        "阿里云": "技术服务费",
        "腾讯云": "技术服务费",
        "采购": "原材料采购",
        "货款": "原材料采购",
        "微信提现": "现金互转"
    }
    
    if os.path.exists(FILE_CATEGORY_RULES):
        try:
            with open(FILE_CATEGORY_RULES, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            log.warning(f"⚠️ 规则文件读取失败: {e}，使用默认规则")
            return default_rules
    else:
        # 创建默认文件方便用户修改
        try:
            with open(FILE_CATEGORY_RULES, "w", encoding="utf-8") as f:
                json.dump(default_rules, f, ensure_ascii=False, indent=4)
        except:
            pass
        return default_rules

AUTO_CATEGORY_RULES = load_category_rules()

def load_partner_aliases():
    """加载往来单位别名映射"""
    default_aliases = {}
    if os.path.exists(FILE_PARTNER_ALIASES):
        try:
            with open(FILE_PARTNER_ALIASES, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            log.warning(f"⚠️ 别名文件读取失败: {e}")
            return default_aliases
    return default_aliases

PARTNER_ALIASES = load_partner_aliases()

def resolve_partner(name):
    """解析往来单位别名 (支持模糊匹配)"""
    if not name: return ""
    name = str(name).strip()
    
    # 1. 优先完全匹配
    if name in PARTNER_ALIASES:
        return PARTNER_ALIASES[name]
        
    # 2. 模糊匹配 (按别名长度倒序，优先匹配更长的别名)
    # 例如：规则 "张三"->A, "张三丰"->B. 输入 "张三丰转账".
    # 应该匹配 "张三丰" 而不是 "张三".
    sorted_aliases = sorted(PARTNER_ALIASES.keys(), key=len, reverse=True)
    
    for alias in sorted_aliases:
        if alias in name:
            # log.info(f"💡 触发别名匹配: '{alias}' in '{name}' -> '{PARTNER_ALIASES[alias]}'")
            return PARTNER_ALIASES[alias]
            
    return name

def clean_description(text):
    """清洗摘要，移除无用前缀"""
    if not text: return ""
    text = str(text).strip()
    # 常见银行摘要垃圾词 (按长度倒序排列，优先匹配长的)
    garbage = [
        "PAYMENT TO", "TRANSFER FROM", "REMITTANCE", 
        "跨行转账", "网银转账", "银企直联", "手机转账", "批量转账",
        "付款给", "收到", "支付", "转账", "汇款", "网转", "电汇", "回单", "记账",
        "用途:", "摘要:", "附言:", "备注:", "说明:",
        "工资:", "报销:", "代发:",
        "用途：", "摘要：", "附言：", "备注：", "说明：",
        "工资：", "报销：", "代发："
    ]
    # 排序：长的在前面，防止误伤
    garbage.sort(key=len, reverse=True)
    
    clean_text = text
    # 循环去除前缀，直到没有匹配项
    while True:
        original = clean_text
        for g in garbage:
            if clean_text.upper().startswith(g):
                clean_text = clean_text[len(g):].strip()
        if clean_text == original:
            break
            
    return clean_text

def read_excel_smart(file_path):
    """
    智能读取 Excel：
    1. 自动寻找表头行 (包含 '日期', '金额' 等关键词)
    2. 自动重命名列为标准字段
    3. 返回标准化的 DataFrame
    """
    try:
        with pd.ExcelFile(file_path) as xl:
            # 优先读 '日常台账表'，否则读第一个 Sheet
            sheet_name = "日常台账表" if "日常台账表" in xl.sheet_names else xl.sheet_names[0]
            
            # 先读前 20 行来找表头
            df_preview = pd.read_excel(xl, sheet_name=sheet_name, header=None, nrows=20)
            
            header_row_idx = -1
            column_map = {}
            
            # 关键词映射表 (可能的列名 -> 标准列名)
            # 增加更多模糊匹配词
            keyword_map = {
                # 日期类
                "日期": "记账日期", "时间": "记账日期", "交易日": "记账日期", "记账日": "记账日期", 
                "入账时间": "记账日期", "交易时间": "记账日期",
                
                # 金额类
                "金额": "实际收付金额", "发生额": "实际收付金额", "收支金额": "实际收付金额",
                "交易金额": "实际收付金额", "收/支": "实际收付金额", "金额(元)": "实际收付金额",
                
                # 备注/摘要类
                "摘要": "备注", "说明": "备注", "用途": "备注", "商品": "备注", "附言": "备注",
                "交易摘要": "备注", "备注说明": "备注", "项目名称": "备注", "内容": "备注",
                
                # 往来单位类
                "对方": "往来单位费用", "户名": "往来单位费用", "单位": "往来单位费用",
                "对方户名": "往来单位费用", "对方账号名称": "往来单位费用", "交易对方": "往来单位费用",
                "收/付款人": "往来单位费用", "商户名称": "往来单位费用",
                
                # 业务类型类 (通常不用，自动推断)
                "借贷": "业务类型", "收付标志": "业务类型"
            }
            
            # 扫描寻找表头
            # 策略优化：只要包含"日期"和("金额"或"发生额"或"支出")的行，就算表头
            for idx, row in df_preview.iterrows():
                row_str = " ".join([str(x) for x in row.values])
                if ("日期" in row_str or "时间" in row_str) and ("金额" in row_str or "发生额" in row_str or "支出" in row_str):
                    header_row_idx = idx
                    # 构建列映射 (精准匹配 -> 包含匹配)
                    for col_idx, val in enumerate(row.values):
                        val_str = str(val).strip()
                        # 1. 精准匹配
                        if val_str in keyword_map:
                            column_map[val_str] = keyword_map[val_str]
                            continue
                        # 2. 包含匹配
                        for k, v in keyword_map.items():
                            if k in val_str:
                                column_map[val_str] = v 
                                break 
                    break
                    
            if header_row_idx == -1:
                # 没找到明显表头，假设第一行就是
                header_row_idx = 0
                log.warning("⚠️ 未找到明显的表头行，尝试默认第一行读取", extra={"solution": "请检查Excel格式"})
                
            # 重新读取数据
            df = pd.read_excel(xl, sheet_name=sheet_name, header=header_row_idx)
        
        # 重命名列
        df.rename(columns=column_map, inplace=True)
        
        # 再次检查关键列是否存在，如果不存在尝试根据位置猜测
        # 假设：如果不包含标准列名，尝试猜：第一列是日期，最后一列是金额？(风险较大，暂不激进)
        
        return df
        
    except Exception as e:
        log.error(f"❌ 智能读取Excel失败: {e}", extra={"solution": "文件可能损坏"})
        return pd.DataFrame()

# 智能分类：历史记忆库
HISTORY_CATEGORY_MAP = {}
AI_CACHE_FILE = FILE_AI_CACHE
AI_CACHE_MAP = {}
AI_CACHE_LOADED = False

def load_ai_cache():
    """加载本地AI分类缓存"""
    global AI_CACHE_MAP, AI_CACHE_LOADED
    if os.path.exists(AI_CACHE_FILE):
        try:
            with open(AI_CACHE_FILE, "r", encoding="utf-8") as f:
                AI_CACHE_MAP = json.load(f)
            log.info(f"🧠 已加载 {len(AI_CACHE_MAP)} 条AI分类缓存", extra={"solution": "无"})
        except Exception as e:
            log.warning(f"⚠️ 加载AI缓存失败: {e}", extra={"solution": "无"})
            AI_CACHE_MAP = {}
    AI_CACHE_LOADED = True

def save_ai_cache():
    """保存AI分类缓存到本地"""
    try:
        with open(AI_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(AI_CACHE_MAP, f, ensure_ascii=False, indent=2)
    except Exception as e:
        log.warning(f"⚠️ 保存AI缓存失败: {e}")

def load_history_knowledge(client, app_token):
    """从飞书加载最近的历史分类习惯 (智能记忆)"""
    global HISTORY_CATEGORY_MAP
    HISTORY_CATEGORY_MAP = {}
    
    # 同时加载AI缓存
    load_ai_cache()
    
    table_id = get_table_id_by_name(client, app_token, "日常台账表")
    if not table_id: return
    
    # 获取最近2000条记录
    log.info("🧠 正在学习历史分类习惯...", extra={"solution": "无"})
    # 使用进度条
    # records = get_all_records(client, app_token, table_id, field_names=["备注", "往来单位费用", "费用归类"])
    # 优化：get_all_records 本身比较慢，但这里无法直接插入进度条，除非修改 get_all_records
    # 暂时保持原样，或者给 get_all_records 加一个 verbose 参数
    records = get_all_records(client, app_token, table_id, field_names=["备注", "往来单位费用", "费用归类"])
    
    # 倒序遍历，越新的越优先
    total = len(records)
    for i, r in enumerate(reversed(records)):
        if i % 200 == 0: # 每200条刷新一次进度
            show_progress_bar(i + 1, total, prefix='学习中', suffix='', length=20)
            
        f = r.fields
        memo = str(f.get("备注") or "").strip()
        partner = str(f.get("往来单位费用") or "").strip()
        cat = str(f.get("费用归类") or "").strip()
        
        if not cat: continue
        
        # 1. 记住 "摘要关键词" -> "分类" (取前10个字作为特征)
        if memo and len(memo) > 1:
            key = memo[:10].lower()
            if key not in HISTORY_CATEGORY_MAP:
                HISTORY_CATEGORY_MAP[key] = cat
                
        # 2. 记住 "往来单位" -> "分类"
        if partner and partner not in ["散户", ""]:
            # 往来单位优先级低一点，加上前缀区分
            p_key = f"PARTNER:{partner}"
            if p_key not in HISTORY_CATEGORY_MAP:
                HISTORY_CATEGORY_MAP[p_key] = cat
                
    show_progress_bar(total, total, prefix='学习完成', suffix='', length=20)
    log.info(f"✅ 已学习 {len(HISTORY_CATEGORY_MAP)} 条历史分类规则", extra={"solution": "无"})

def auto_categorize(description, default_val, partner_name=None):
    if not description and not partner_name:
        return default_val
    
    # 重新加载规则，支持热修改
    global AUTO_CATEGORY_RULES, HISTORY_CATEGORY_MAP, AI_CACHE_LOADED
    
    # 确保AI缓存已加载
    if not AI_CACHE_LOADED:
        load_ai_cache()
    
    desc_str = str(description).lower()
    
    # 1. 优先匹配明确的【规则库】 (category_rules.json)
    for key, value in AUTO_CATEGORY_RULES.items():
        if key.lower() in desc_str:
            return value
            
    # 2. 其次匹配【历史记忆】 (History Knowledge)
    # 2.1 匹配摘要前缀
    if desc_str:
        key = desc_str[:10]
        if key in HISTORY_CATEGORY_MAP:
            return HISTORY_CATEGORY_MAP[key]
            
    # 2.2 匹配往来单位
    if partner_name:
        p_key = f"PARTNER:{partner_name}"
        if p_key in HISTORY_CATEGORY_MAP:
            return HISTORY_CATEGORY_MAP[p_key]
            
    # 2.3 [V9.5新特性] 匹配本地AI缓存 (Smart Cache)
    # 避免重复调用AI接口，节省Token并提升速度
    cache_key = f"{desc_str}|{str(partner_name).lower()}"
    if cache_key in AI_CACHE_MAP:
        # print(f"   🧠 命中本地AI缓存: {cache_key[:20]}... -> {AI_CACHE_MAP[cache_key]}")
        return AI_CACHE_MAP[cache_key]
            
    # 3. [V9.4新特性] 尝试 AI 智能推断
    # 只有当描述足够长(>2)或有明确往来单位时才调用，避免浪费 Token
    if (len(desc_str) > 2 or partner_name) and ZHIPUAI_API_KEY:
        ai_cat = ai_guess_category(description, partner_name)
        if ai_cat:
            print(f"   🧠 AI 智能推断: '{description}' -> [{ai_cat}]")
            # 更新缓存
            AI_CACHE_MAP[cache_key] = ai_cat
            save_ai_cache()
            return ai_cat
            
    return default_val

def parse_smart_text(text):
    """
    智能解析自然语言账目 (V1.0)
    输入: "昨天付给张三货款5000元"
    输出: {"date": "...", "type": "...", "amount": 5000, "partner": "...", ...}
    """
    import re
    text = text.strip()
    if not text: return None
    
    res = {
        "date": datetime.now().strftime("%Y-%m-%d"),
        "type": "费用", # 默认
        "amount": 0.0,
        "partner": "散户",
        "category": "未分类",
        "remark": text,
        "has_invoice": "无票"
    }
    
    # 1. 解析日期
    if "昨天" in text:
        res["date"] = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
    elif "前天" in text:
        res["date"] = (datetime.now() - timedelta(days=2)).strftime("%Y-%m-%d")
    
    # 2. 解析金额
    try:
        amount_found = False
        # 1. 优先找明确单位 (w, k)
        m_w = re.search(r'(\d+(?:\.\d+)?)\s*[wW万]', text)
        if m_w: 
            res["amount"] = float(m_w.group(1)) * 10000
            amount_found = True
        
        if not amount_found:
            m_k = re.search(r'(\d+(?:\.\d+)?)\s*[kK千]', text)
            if m_k: 
                res["amount"] = float(m_k.group(1)) * 1000
                amount_found = True
                
        if not amount_found:
             m_unit = re.search(r'(\d+(?:\.\d+)?)\s*[元块]', text)
             if m_unit:
                 res["amount"] = float(m_unit.group(1))
                 amount_found = True

        if not amount_found:
            # 找独立数字，排除年份(202x年)和手机号
            nums = re.findall(r'\d+(?:\.\d+)?', text)
            valid_nums = []
            for n in nums:
                val = float(n)
                # Check for Year context: "2024年"
                if re.search(str(n) + r"\s*年", text): continue
                # Check for strict year range if 4 digits (e.g. 2024) and no decimal
                if val >= 2000 and val <= 2030 and "." not in n: 
                    # If it's the only number, maybe it is amount? Unlikely for small amounts.
                    # Let's assume 2000-2030 are years unless we have strong evidence otherwise
                    pass 
                
                if len(n) == 11 and n.startswith("1") and "." not in n: continue # Phone-like
                valid_nums.append(val)
            
            if valid_nums:
                res["amount"] = max(valid_nums) # 猜测最大的数字是金额
            
    except: pass
    
    # 3. 解析类型
    if any(k in text for k in ["收入", "收到", "收款", "入账", "转入", "退回"]):
        res["type"] = "收款"
    elif any(k in text for k in ["付", "支", "转给", "消费", "买", "交"]):
        res["type"] = "付款"
    
    # 4. 解析往来单位 (Refined)
    # 强匹配: 给xxx, 收到xxx, 来自xxx
    m_p = re.search(r'(?:给|收到|来自)\s*([^0-9\s元块,，。]+)', text)
    if m_p:
        raw_p = m_p.group(1)
        # 清理后缀
        raw_p = re.sub(r'(货款|款|费|工资|报销|转账)$', '', raw_p)
        if len(raw_p) > 1:
            res["partner"] = raw_p
    
    # 弱匹配: 付xxx (如果还没找到)
    if res["partner"] == "散户":
         m_p_weak = re.search(r'(?:付)\s*([^0-9\s元块,，。]+)', text)
         if m_p_weak:
             raw_p = m_p_weak.group(1)
             # 排除常见非人名
             if raw_p not in ["款", "工资", "货款", "租金", "电费", "水费", "定金", "押金"]:
                  raw_p = re.sub(r'(货款|款|费|工资|报销|转账)$', '', raw_p)
                  if len(raw_p) > 1:
                      res["partner"] = raw_p

    # 5. 自动归类
    # 如果已解析出 partner，传入辅助归类
    cat = auto_categorize(text, "", res["partner"])
    if cat: res["category"] = cat
    
    # 修正类型: 如果归类暗示了类型
    if res["type"] == "付款":
        # 常见费用词
        if any(c in str(res["category"]) for c in ["费", "税", "租金", "薪", "社保"]):
            res["type"] = "费用"
            
    return res

# 导入未匹配流水到飞书
def import_bank_records_to_feishu(client, app_token, records_list):
    """
    将未匹配的银行流水直接导入飞书台账
    """
    if not records_list:
        return
        
    table_id = get_table_id_by_name(client, app_token, "日常台账表")
    if not table_id:
        log.error("❌ 找不到日常台账表，无法导入", extra={"solution": "请先初始化表格"})
        return

    log.info(f"🚀 正在批量导入 {len(records_list)} 条流水...", extra={"solution": "无"})
    
    # 导入未匹配流水到飞书
    # 自动识别逻辑：G银行默认有票，N银行/微信默认现金
    
    feishu_records = []
    for r in records_list:
        # 解析日期字符串 "YYYY-MM-DD" -> timestamp
        try:
            dt = datetime.strptime(r["记账日期"], "%Y-%m-%d")
            ts = int(dt.timestamp() * 1000)
        except:
            ts = int(datetime.now().timestamp() * 1000)

        # 绝对值处理
        amt = abs(float(r["实际收付金额"]))
        
        # 智能判断默认值
        txn_bank = r.get("交易银行", "G银行基本户")
        is_cash = "否"
        has_ticket = "无票"
        
        if "G银行" in txn_bank:
            has_ticket = "有票"
            is_cash = "否"
        elif "N银行" in txn_bank or "微信" in txn_bank or "现金" in txn_bank:
            is_cash = "是"
            has_ticket = "无票" # 现金通常默认无票，除非明确指定

        # 如果原记录已有值，则优先使用
        if r.get("是否有票"): has_ticket = r.get("是否有票")
        if r.get("是否现金"): is_cash = r.get("是否现金")

        # 尝试自动分类补全 (费用归类)
        category = str(r.get("费用归类", ""))
        if not category or category == "nan" or category == "未知" or category == "":
            category = auto_categorize(r.get("备注", ""), "其他", partner_name=r.get("往来单位费用", ""))

        fields = {
            "记账日期": ts,
            "凭证号": 0, # 默认为0
            "业务类型": r["业务类型"],
            "费用归类": category,
            "往来单位费用": r["往来单位费用"],
            "账面金额": amt, # 默认账面=实际 (按实际发生)
            "实际收付金额": amt,
            "交易银行": txn_bank,
            "是否现金": is_cash,
            "是否有票": has_ticket,
            "待补票标记": "否",
            "备注": r["备注"]
        }
        feishu_records.append(AppTableRecord.builder().fields(fields).build())
    
    # 分批提交 (每次100条)
    batch_size = 100
    success_count = 0
    for i in range(0, len(feishu_records), batch_size):
        batch = feishu_records[i:i+batch_size]
        req = BatchCreateAppTableRecordRequest.builder() \
            .app_token(app_token) \
            .table_id(table_id) \
            .request_body(BatchCreateAppTableRecordRequestBody.builder().records(batch).build()) \
            .build()
        
        resp = client.bitable.v1.app_table_record.batch_create(req)
        if resp.success():
            success_count += len(batch)
            log.info(f"✅ 第 {i//batch_size + 1} 批导入成功 ({len(batch)}条)")
        else:
            log.error(f"❌ 第 {i//batch_size + 1} 批导入失败: {resp.msg}")
            
    if success_count > 0:
        send_bot_message(f"✅ 已自动导入 {success_count} 条银行流水到台账！", "reconcile")
        print(f"✅ 成功导入 {success_count} 条记录。")
        # 导入成功后静默刷新仪表盘缓存
        try:
            update_dashboard_cache_silent(client, app_token)
        except:
            pass

def generate_reconciliation_report(matched_count, unmatched_list, ledger_unmatched_list=None):
    """生成对账结果可视化报告 (包含双向差异)"""
    if ledger_unmatched_list is None: ledger_unmatched_list = []
    
    total_bank = matched_count + len(unmatched_list)
    total_ledger_issues = len(ledger_unmatched_list)
    
    if total_bank == 0 and total_ledger_issues == 0: return
    
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>银行对账报告 - {datetime.now().strftime('%Y-%m-%d')}</title>
        <script src="https://cdn.jsdelivr.net/npm/echarts@5.4.3/dist/echarts.min.js"></script>
        <style>
            body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: #f4f6f9; margin: 0; padding: 20px; }}
            .container {{ max-width: 1200px; margin: 0 auto; background: white; padding: 30px; border-radius: 10px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); }}
            h1 {{ color: #2c3e50; text-align: center; border-bottom: 2px solid #3498db; padding-bottom: 10px; }}
            .summary {{ display: flex; justify-content: space-around; margin: 30px 0; }}
            .card {{ text-align: center; padding: 20px; background: #f8f9fa; border-radius: 8px; width: 22%; }}
            .number {{ font-size: 24px; font-weight: bold; color: #2c3e50; }}
            .chart-row {{ display: flex; justify-content: space-between; height: 400px; margin: 20px 0; }}
            .chart-box {{ width: 48%; height: 100%; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 20px; font-size: 14px; }}
            th, td {{ border: 1px solid #ddd; padding: 12px; text-align: left; }}
            th {{ background-color: #3498db; color: white; }}
            tr:nth-child(even) {{ background-color: #f2f2f2; }}
            .badge {{ padding: 5px 10px; border-radius: 4px; font-size: 12px; }}
            .badge-danger {{ background-color: #e74c3c; color: white; }}
            .badge-warning {{ background-color: #f39c12; color: white; }}
            .section-title {{ margin-top: 40px; color: #2c3e50; border-left: 5px solid #3498db; padding-left: 10px; }}
        </style>
    </head>
    <body>
        <div class="container">
            <h1>🏦 银行对账报告</h1>
            <p style="text-align: center; color: #7f8c8d;">生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            
            <div class="summary">
                <div class="card">
                    <div class="number" style="color: #3498db;">{total_bank}</div>
                    <div>银行流水总数</div>
                </div>
                <div class="card">
                    <div class="number" style="color: #27ae60;">{matched_count}</div>
                    <div>✅ 自动匹配成功</div>
                </div>
                <div class="card">
                    <div class="number" style="color: #e74c3c;">{len(unmatched_list)}</div>
                    <div>❌ 银行有而台账无</div>
                </div>
                <div class="card">
                    <div class="number" style="color: #f39c12;">{len(ledger_unmatched_list)}</div>
                    <div>❓ 台账有而银行无</div>
                </div>
            </div>

            <div class="chart-row">
                <div id="pie-chart" class="chart-box"></div>
                <div id="bar-chart" class="chart-box"></div>
            </div>

            <h3 class="section-title">❌ 异常类型一：银行流水有，但台账未记录 ({len(unmatched_list)}条)</h3>
            <p style="color: #7f8c8d; font-size: 14px;">👉 建议：检查是否漏记，可使用"待补录流水.xlsx"直接导入</p>
            <table>
                <thead>
                    <tr>
                        <th>日期</th>
                        <th>摘要</th>
                        <th>金额</th>
                        <th>对象</th>
                        <th>建议分类</th>
                        <th>原因</th>
                    </tr>
                </thead>
                <tbody>
    """
    
    for item in unmatched_list:
        html += f"""
                    <tr>
                        <td>{item.get('记账日期')}</td>
                        <td>{item.get('备注')}</td>
                        <td>{item.get('实际收付金额')}</td>
                        <td>{item.get('往来单位费用')}</td>
                        <td>{item.get('费用归类')}</td>
                        <td><span class="badge badge-danger">{item.get('原因')}</span></td>
                    </tr>
        """
        
    html += f"""
                </tbody>
            </table>

            <h3 class="section-title">❓ 异常类型二：台账已记，但银行流水无 ({len(ledger_unmatched_list)}条)</h3>
            <p style="color: #7f8c8d; font-size: 14px;">👉 建议：检查是否多记、重复记账、日期偏差过大(>2天)或银行选错</p>
            <table>
                <thead>
                    <tr>
                        <th>日期</th>
                        <th>摘要</th>
                        <th>金额</th>
                        <th>往来对象</th>
                        <th>登记银行</th>
                        <th>原因</th>
                    </tr>
                </thead>
                <tbody>
    """

    for item in ledger_unmatched_list:
        html += f"""
                    <tr>
                        <td>{item.get('记账日期')}</td>
                        <td>{item.get('摘要')}</td>
                        <td>{item.get('金额')}</td>
                        <td>{item.get('往来')}</td>
                        <td>{item.get('交易银行')}</td>
                        <td><span class="badge badge-warning">{item.get('原因')}</span></td>
                    </tr>
        """

    html += f"""
                </tbody>
            </table>
            
            <script>
                var chartDom = document.getElementById('pie-chart');
                var myChart = echarts.init(chartDom);
                var option = {{
                    title: {{ text: '银行流水匹配情况', left: 'center' }},
                    tooltip: {{ trigger: 'item' }},
                    legend: {{ orient: 'vertical', left: 'left' }},
                    series: [
                        {{
                            name: '匹配结果',
                            type: 'pie',
                            radius: ['40%', '70%'],
                            avoidLabelOverlap: false,
                            itemStyle: {{
                                borderRadius: 10,
                                borderColor: '#fff',
                                borderWidth: 2
                            }},
                            label: {{
                                show: false,
                                position: 'center'
                            }},
                            emphasis: {{
                                label: {{
                                    show: true,
                                    fontSize: 20,
                                    fontWeight: 'bold'
                                }}
                            }},
                            labelLine: {{ show: false }},
                            data: [
                                {{ value: {matched_count}, name: '匹配成功', itemStyle: {{ color: '#27ae60' }} }},
                                {{ value: {len(unmatched_list)}, name: '银行未入账', itemStyle: {{ color: '#e74c3c' }} }}
                            ]
                        }}
                    ]
                }};
                myChart.setOption(option);
                
                var barDom = document.getElementById('bar-chart');
                var barChart = echarts.init(barDom);
                var barOption = {{
                    title: {{ text: '双向差异概览', left: 'center' }},
                    tooltip: {{ trigger: 'axis', axisPointer: {{ type: 'shadow' }} }},
                    grid: {{ left: '3%', right: '4%', bottom: '3%', containLabel: true }},
                    xAxis: [ {{ type: 'category', data: ['银行有台账无', '台账有银行无'], axisTick: {{ alignWithLabel: true }} }} ],
                    yAxis: [ {{ type: 'value' }} ],
                    series: [
                        {{
                            name: '笔数',
                            type: 'bar',
                            barWidth: '60%',
                            data: [
                                {{ value: {len(unmatched_list)}, itemStyle: {{ color: '#e74c3c' }} }},
                                {{ value: {len(ledger_unmatched_list)}, itemStyle: {{ color: '#f39c12' }} }}
                            ]
                        }}
                    ]
                }};
                barChart.setOption(barOption);
            </script>
        </div>
    </body>
    </html>
    """
    
    report_dir = "财务数据备份"
    if not os.path.exists(report_dir):
        os.makedirs(report_dir)
        
    filename = f"{report_dir}/对账报告_{datetime.now().strftime('%Y%m%d%H%M%S')}.html"
    with open(filename, "w", encoding="utf-8") as f:
        f.write(html)
        
    log.info(f"📊 对账可视化报告已生成: {filename}", extra={"solution": "浏览器打开查看"})
    try:
        os.startfile(filename)
    except:
        pass

# 银行流水对账 (智能模糊匹配 + 性能优化)
@retry_on_failure(max_retries=2, delay=3)
def reconcile_bank_flow(client, app_token, bank_excel_path):
    log.info("📊 开始智能对账...", extra={"solution": "无"})
    
    # 1. 先读取银行流水 (为了获取日期范围，减少飞书数据拉取量)
    try:
        if not bank_excel_path:
            # 优先尝试交互式选择
            bank_excel_path = select_file_interactively("*.xlsx", "请选择银行流水Excel文件")
            
            # 如果还是没有，回退到弹窗
            if not bank_excel_path:
                log.info("📂 请选择银行流水Excel文件...", extra={"solution": "弹窗选择"})
                bank_excel_path = select_file("请选择银行流水Excel文件")

            if not bank_excel_path:
                log.warning("⚠️ 未选择文件，操作取消", extra={"solution": "无"})
                return False

        # 智能识别银行类型 (基于文件名)
        bank_choice = "1" # Default G Bank
        base_name = os.path.basename(bank_excel_path).upper()
        
        if any(k in base_name for k in ["微信", "N银行", "现金", "WECHAT", "ALIPAY", "支付宝"]):
            log.info(f"🤖 检测到文件名包含关键信息，自动识别为【N银行/微信（现金）】模式", extra={"solution": "无需操作"})
            bank_choice = "2"
        elif any(k in base_name for k in ["G银行", "工商", "ICBC", "对公"]):
             log.info(f"🤖 检测到文件名包含关键信息，自动识别为【G银行（对公）】模式", extra={"solution": "无需操作"})
             bank_choice = "1"
        else:
            # 交互式选择银行类型
            print("\n🏦 请选择当前对账的银行类型：")
            print("1. G银行 (对公账户 - 默认有票)")
            print("2. N银行/微信 (现金/私户 - 默认现金)")
            user_input = input(f"请输入数字 (1/2) [默认{bank_choice}]: ").strip()
            if user_input:
                bank_choice = user_input
        
        bank_name = "G银行基本户"
        default_ticket = "有票"
        is_cash = "否"
        
        if bank_choice == "2":
            bank_name = "N银行/微信（现金）"
            default_ticket = "无票"
            is_cash = "是"
        log.info(f"✅ 当前设定: {bank_name}", extra={"solution": "无"})

        # 使用智能读取
        df = read_excel_smart(bank_excel_path)
        if df.empty:
            return False

        # 标准列名
        date_col = "记账日期"
        amount_col = "实际收付金额"
        
        if date_col not in df.columns or amount_col not in df.columns:
            log.error(f"❌ 银行流水Excel缺少必要的列 (需包含 日期/金额 关键词)", extra={"solution": "修改表头或使用智能模式"})
            log.info(f"当前识别到的列: {df.columns.tolist()}", extra={"solution": "无"})
            return False

        # 获取日期范围用于过滤
        try:
            dates = pd.to_datetime(df[date_col])
            min_date = dates.min()
            max_date = dates.max()
            # 扩大范围前后各7天，防止容差漏掉
            filter_start_ts = int((min_date - timedelta(days=7)).timestamp() * 1000)
            filter_end_ts = int((max_date + timedelta(days=7)).timestamp() * 1000)
            log.info(f"📅 提取流水日期范围: {min_date.date()} 至 {max_date.date()}", extra={"solution": "无"})
        except Exception as e:
            log.warning(f"⚠️ 日期解析失败，将拉取全量数据: {e}", extra={"solution": "检查日期格式"})
            filter_start_ts = None
            filter_end_ts = None

    except Exception as e:
        log.error(f"❌ 读取Excel失败: {str(e)}", extra={"solution": "检查文件是否被占用"})
        return False

    # 2. 读取飞书台账 (带过滤器)
    table_id = get_table_id_by_name(client, app_token, "日常台账表")
    if not table_id:
        log.error("❌ 未找到'日常台账表'", extra={"solution": "先创建表格"})
        return False
        
    filter_info = None
    if filter_start_ts and filter_end_ts:
        # 构造组合过滤器
        filter_info = f'AND(CurrentValue.[记账日期]>={filter_start_ts}, CurrentValue.[记账日期]<={filter_end_ts})'
        log.info("🚀 启用云端数据过滤，仅拉取相关日期记录", extra={"solution": "性能优化"})

    feishu_records = get_all_records(client, app_token, table_id, filter_info=filter_info)
    log.info(f"📥 拉取到 {len(feishu_records)} 条相关记录", extra={"solution": "无"})
    
    # 构建飞书数据索引: Amount -> List of Records
    feishu_amount_map = {}
    for record in feishu_records:
        fields = record['fields']
        try:
            # 修正：字段名应为 '实际收付金额'
            f_amount = float(fields.get('实际收付金额', 0))
            f_date = pd.to_datetime(fields.get('记账日期', 0), unit='ms')
            
            key = round(f_amount, 2)
            if key not in feishu_amount_map:
                feishu_amount_map[key] = []
            
            feishu_amount_map[key].append({
                "record_id": record['record_id'],
                "date": f_date,
                "fields": fields,
                "matched": False
            })
        except:
            continue

            
    unmatched = []
    matched_count = 0
    
    # 容差设置
    # TOLERANCE_DAYS = 2 # Moved to global config
    
    # 加载历史分类知识
    load_history_knowledge(client, app_token)

    for idx, row in df.iterrows():
        try:
            b_date = pd.to_datetime(row[date_col])
            b_amount = float(row[amount_col])
            b_amount_key = round(b_amount, 2)
        except:
            continue 
        
        # 查找匹配
        candidates = feishu_amount_map.get(b_amount_key, [])
        match_found = False
        
        for cand in candidates:
            if cand["matched"]:
                continue # 已经被其他流水匹配过了
            
            # 检查日期差
            delta = abs((cand["date"] - b_date).days)
            if delta <= TOLERANCE_DAYS:
                cand["matched"] = True
                match_found = True
                matched_count += 1
                break
        
        if not match_found:
            # 构造符合导入格式的数据
            raw_desc = str(row.get("对方户名", str(row.get("对方账号", row.get("摘要", "未知")))))
            memo = str(row.get('摘要', ''))
            
            # 优化：清洗摘要
            cleaned_memo = clean_description(memo)
            cleaned_desc = clean_description(raw_desc)
            
            # 优化：解析别名 (张三 -> A客户)
            cleaned_desc = resolve_partner(cleaned_desc)

            # 尝试自动分类 (使用历史记忆)
            # 逻辑：
            # 1. 如果有明确规则匹配 cleaned_memo -> 用规则
            # 2. 如果 cleaned_memo 在历史中出现过 -> 用历史
            # 3. 如果 cleaned_desc 在历史中出现过 -> 用历史
            category = auto_categorize(cleaned_memo, "其他", partner_name=cleaned_desc) 
            
            # 如果自动分类返回默认值，尝试单独匹配 cleaned_desc
            if category == "其他":
                 category = auto_categorize(cleaned_desc, "其他", partner_name=cleaned_desc)
            
            unmatched.append({
                "记账日期": b_date.strftime("%Y-%m-%d"),
                "凭证号": "",
                "业务类型": "付款" if b_amount < 0 else "收款",
                "费用归类": category,
                "往来单位费用": cleaned_desc,
                "实际收付金额": b_amount,
                "交易银行": bank_name,
                "是否现金": is_cash,
                "是否有票": default_ticket,
                "待补票标记": "否",
                "备注": f"流水导入: {memo}",
                "原因": "飞书无此金额或日期超2天"
            })
            
    # [新增] 反向对账：检查台账中有，但银行流水中没有的记录 (可能是多记、重复或日期错误)
    ledger_unmatched = []
    
    # 定义当前银行的关键词
    target_bank_keywords = []
    if bank_choice == "1":
        target_bank_keywords = ["G银行", "工行", "ICBC", "对公"]
    elif bank_choice == "2":
        target_bank_keywords = ["N银行", "微信", "现金", "私户"]
        
    for key, records_list in feishu_amount_map.items():
        for r in records_list:
            if not r["matched"]:
                # 检查该记录是否属于当前对账的银行
                r_bank = str(r["fields"].get("交易银行", "")).strip()
                
                # 如果台账里没写银行，默认不报错(避免误报)；或者如果用户希望严查，可以调整策略
                if not r_bank: continue
                
                is_target = False
                for k in target_bank_keywords:
                    if k in r_bank:
                        is_target = True
                        break
                        
                if is_target:
                    # 找到了属于该银行但未匹配流水的数据
                    f = r["fields"]
                    ledger_unmatched.append({
                        "记账日期": datetime.fromtimestamp(f.get("记账日期",0)/1000).strftime("%Y-%m-%d"),
                        "业务类型": f.get("业务类型",""),
                        "金额": f.get("实际收付金额",0),
                        "摘要": f.get("备注",""),
                        "往来": f.get("往来单位费用",""),
                        "交易银行": r_bank,
                        "原因": "台账有但流水无 (可能是多记、日期偏差大或金额不一致)"
                    })

    # 3. 输出结果
    msg = f"智能对账完成！\n✅ 自动匹配：{matched_count}笔\n❌ 银行流水未入账：{len(unmatched)}笔"
    if ledger_unmatched:
        msg += f"\n⚠️ 台账多余记录 (疑似错误)：{len(ledger_unmatched)}笔"
        
    log.info(msg, extra={"solution": "查看导出文件"})
    
    # 生成可视化报告
    generate_reconciliation_report(matched_count, unmatched, ledger_unmatched)

    # 导出 银行流水未入账
    if unmatched:
        res_df = pd.DataFrame(unmatched)
        # 确保列顺序符合导入要求
        cols = ["记账日期", "凭证号", "业务类型", "费用归类", "往来单位费用", "实际收付金额", 
                "交易银行", "是否现金", "是否有票", "待补票标记", "备注", "原因"]
        # 动态调整列，防止KeyError
        final_cols = [c for c in cols if c in res_df.columns]
        res_df = res_df[final_cols]
        
        res_path = f"待补录流水_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        res_df.to_excel(res_path, index=False, sheet_name="日常台账表") 
        log.info(f"📄 待补录清单已导出: {res_path}", extra={"solution": "检查后导入"})
        
    # 导出 台账多余记录
    if ledger_unmatched:
        l_df = pd.DataFrame(ledger_unmatched)
        l_path = f"台账异常记录_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        l_df.to_excel(l_path, index=False)
        log.warning(f"📄 发现台账异常记录 (流水中没有): {l_path}", extra={"solution": "请核对是否多记或日期错误"})

    if unmatched:
        # 新增：询问是否直接导入 (按实际发生)
        print(f"\n💡 发现 {len(unmatched)} 笔未匹配流水 (可能是新发生的收支)。")
        print("💡 小提示: 小企业通常付款/回款不一一对应，建议按'实际发生'直接导入。")
        import_choice = input("👉 是否直接将这些流水作为新账目导入飞书? (y/n) [推荐y]: ").strip().lower()
        if import_choice != 'n': 
            import_bank_records_to_feishu(client, app_token, unmatched)
            
    else:
        send_bot_message(f"{msg}\n🎉 账目完美平衡！", "reconcile")
        
    return True

# 往来对账 (导入外部账单核对)
@retry_on_failure(max_retries=2, delay=3)
def reconcile_partner_flow(client, app_token, partner_excel_path=None):
    log.info("🤝 开始往来对账流程...", extra={"solution": "无"})
    
    # 1. 获取外部账单文件
    if not partner_excel_path:
        partner_excel_path = select_file_interactively("*.xlsx", "请选择客户/供应商的对账单文件")
        
    if not partner_excel_path:
        log.warning("⚠️ 未选择文件，操作取消", extra={"solution": "无"})
        return False
        
    # 2. 读取外部账单
    try:
        df = read_excel_smart(partner_excel_path)
        if df.empty:
            log.error("❌ 文件为空或无法识别", extra={"solution": "检查文件内容"})
            return False
            
        # 尝试从文件名猜测往来单位
        filename = os.path.basename(partner_excel_path)
        guessed_partner = filename.split('.')[0].replace("对账单", "").replace("往来", "").strip()
        
        print(f"\n🏢 识别到的往来单位: {Color.BOLD}{guessed_partner}{Color.ENDC}")
        partner_name = input(f"👉 确认往来单位名称 (回车默认, 或输入新名称): ").strip()
        if not partner_name:
            partner_name = guessed_partner
            
        log.info(f"✅ 当前对账对象: {partner_name}", extra={"solution": "无"})
        
        # 识别关键列
        # 我们需要: 日期, 金额 (正负代表方向), 摘要
        # read_excel_smart 已经尽力标准化了 '记账日期', '实际收付金额', '摘要'
        if "记账日期" not in df.columns or "实际收付金额" not in df.columns:
            log.error("❌ 无法识别必要的列 (日期/金额)", extra={"solution": "请确保表头包含 '日期' 和 '金额' 相关字样"})
            return False
            
    except Exception as e:
        log.error(f"❌ 读取文件失败: {e}", extra={"solution": "检查文件格式"})
        return False

    # 3. 拉取系统内部数据 (按往来单位过滤)
    table_id = get_table_id_by_name(client, app_token, "日常台账表")
    if not table_id: return False
    
    # 构造日期范围过滤器 (为了性能)
    try:
        dates = pd.to_datetime(df["记账日期"])
        min_ts = int((dates.min() - timedelta(days=15)).timestamp() * 1000) # 放宽范围
        max_ts = int((dates.max() + timedelta(days=15)).timestamp() * 1000)
        
        # 组合过滤器: 日期范围 AND 往来单位包含
        # 注意: 飞书公式中字符串包含用 CurrentValue.[往来单位费用].contains("Name") ? 不, API filter syntax
        # 简单起见，先按日期拉取，内存过滤往来单位 (更稳健)
        filter_info = f'AND(CurrentValue.[记账日期]>={min_ts}, CurrentValue.[记账日期]<={max_ts})'
        
        log.info("📥 正在拉取系统内部账目...", extra={"solution": "无"})
        all_records = get_all_records(client, app_token, table_id, filter_info=filter_info)
        
        # 内存过滤往来单位 (支持模糊匹配)
        internal_records = []
        for r in all_records:
            f = r.fields
            p = str(f.get("往来单位费用", "")).strip()
            # 简单的包含关系检查
            if partner_name in p or p in partner_name:
                internal_records.append(r)
                
        log.info(f"✅ 系统内找到相关记录: {len(internal_records)} 条", extra={"solution": "无"})
        
    except Exception as e:
        log.error(f"❌ 数据拉取失败: {e}", extra={"solution": "检查网络"})
        return False
        
    # 4. 执行核心对账逻辑 (升级版: 模糊匹配 + 组合匹配)
    log.info("🔄 正在进行智能比对 (精准+模糊+组合)...", extra={"solution": "无"})
    
    # 构建内部数据池 (Flat List)
    internal_pool = []
    for i, r in enumerate(internal_records):
        f = r.fields
        try:
            amt = round(float(f.get("实际收付金额", 0)), 2)
            ts = f.get("记账日期", 0)
            date_obj = datetime.fromtimestamp(ts/1000)
            internal_pool.append({
                "record": r,
                "amount": amt,
                "date": date_obj,
                "matched": False,
                "match_type": None,
                "id": i
            })
        except:
            pass

    # 构建外部数据池
    external_pool = []
    for idx, row in df.iterrows():
        try:
            e_amt = round(float(row["实际收付金额"]), 2)
            e_date = pd.to_datetime(row["记账日期"])
            e_desc = str(row.get("摘要", "") or row.get("备注", ""))
            external_pool.append({
                "amount": e_amt,
                "date": e_date,
                "desc": e_desc,
                "matched": False,
                "match_type": None,
                "original_idx": idx
            })
        except:
            continue

    matched_count = 0
    fuzzy_count = 0
    combo_count = 0

    # --- Pass 1: 精准匹配 (金额一致 & 日期在容差内) ---
    for e_item in external_pool:
        if e_item['matched']: continue
        
        best_match = None
        min_day_diff = 999
        
        for i_item in internal_pool:
            if i_item['matched']: continue
            
            # 金额严格一致
            if abs(i_item['amount'] - e_item['amount']) < 0.01:
                day_diff = abs((i_item['date'] - e_item['date']).days)
                if day_diff <= TOLERANCE_DAYS:
                    if day_diff < min_day_diff:
                        min_day_diff = day_diff
                        best_match = i_item
        
        if best_match:
            e_item['matched'] = True
            e_item['match_type'] = "精准匹配"
            best_match['matched'] = True
            best_match['match_type'] = "精准匹配"
            matched_count += 1

    # --- Pass 2: 模糊金额匹配 (金额相差<=1元 & 日期在容差内) ---
    for e_item in external_pool:
        if e_item['matched']: continue
        
        best_match = None
        min_day_diff = 999
        
        for i_item in internal_pool:
            if i_item['matched']: continue
            
            diff = abs(i_item['amount'] - e_item['amount'])
            if 0.01 < diff <= 1.0: # 允许1元以内误差
                day_diff = abs((i_item['date'] - e_item['date']).days)
                if day_diff <= TOLERANCE_DAYS:
                    if day_diff < min_day_diff:
                        min_day_diff = day_diff
                        best_match = i_item
        
        if best_match:
            diff_val = best_match['amount'] - e_item['amount']
            e_item['matched'] = True
            e_item['match_type'] = f"模糊匹配 (差{diff_val:+.2f})"
            best_match['matched'] = True
            best_match['match_type'] = f"模糊匹配 (差{-diff_val:+.2f})"
            fuzzy_count += 1

    # --- Pass 3: 组合匹配 (1笔外部 vs 多笔内部) ---
    # 场景: 客户一次转账对应我们多笔发票/订单
    # 限制: 仅尝试未匹配的记录, 且组合数量限制在 2-3 笔以防性能问题
    
    # 筛选候选池 (仅保留日期接近的未匹配内部记录)
    for e_item in external_pool:
        if e_item['matched']: continue
        
        candidates = []
        for i_item in internal_pool:
            if not i_item['matched']:
                # 放宽日期限制给组合匹配
                day_diff = abs((i_item['date'] - e_item['date']).days)
                if day_diff <= TOLERANCE_DAYS + 5: 
                    candidates.append(i_item)
        
        if len(candidates) > 50: continue # 候选太多跳过组合尝试
        
        found_combo = False
        
        # 尝试 2 笔组合
        for i in range(len(candidates)):
            if found_combo: break
            for j in range(i+1, len(candidates)):
                s = candidates[i]['amount'] + candidates[j]['amount']
                if abs(s - e_item['amount']) < 0.05:
                    # 找到组合!
                    e_item['matched'] = True
                    e_item['match_type'] = "组合匹配 (2笔)"
                    candidates[i]['matched'] = True
                    candidates[i]['match_type'] = "组合成员"
                    candidates[j]['matched'] = True
                    candidates[j]['match_type'] = "组合成员"
                    combo_count += 1
                    found_combo = True
                    break
        
        # 尝试 3 笔组合 (仅当候选较少时)
        if not found_combo and len(candidates) < 20:
             for i in range(len(candidates)):
                if found_combo: break
                for j in range(i+1, len(candidates)):
                    for k in range(j+1, len(candidates)):
                        s = candidates[i]['amount'] + candidates[j]['amount'] + candidates[k]['amount']
                        if abs(s - e_item['amount']) < 0.05:
                            e_item['matched'] = True
                            e_item['match_type'] = "组合匹配 (3笔)"
                            candidates[i]['matched'] = True
                            candidates[i]['match_type'] = "组合成员"
                            candidates[j]['matched'] = True
                            candidates[j]['match_type'] = "组合成员"
                            candidates[k]['matched'] = True
                            candidates[k]['match_type'] = "组合成员"
                            combo_count += 1
                            found_combo = True
                            break

    # --- 结果汇总 ---
    internal_missing = [] # 漏记
    external_missing = [] # 多记
    
    for e_item in external_pool:
        if not e_item['matched']:
            internal_missing.append({
                "日期": e_item['date'].strftime("%Y-%m-%d"),
                "金额": e_item['amount'],
                "摘要": e_item['desc'],
                "原因": "我方缺失 (需补录)"
            })
            
    for i_item in internal_pool:
        if not i_item['matched']:
             f = i_item['record'].fields
             external_missing.append({
                "日期": i_item['date'].strftime("%Y-%m-%d"),
                "金额": i_item['amount'],
                "摘要": f.get("备注", "") or f.get("往来单位费用", ""),
                "原因": "我方多出 (对方无此记录)"
             })

    # 5. 输出结果
    total_ok = matched_count + fuzzy_count + combo_count
    print(f"\n{Color.HEADER}📊 对账结果摘要 ({partner_name}){Color.ENDC}")
    print(f"✅ 匹配成功: {total_ok} 笔 (精准: {matched_count}, 模糊: {fuzzy_count}, 组合: {combo_count})")
    print(f"❌ 我方缺失: {len(internal_missing)} 笔 (可能是漏记)")
    print(f"❓ 我方多出: {len(external_missing)} 笔 (可能是多记/对方漏记)")
    
    # 6. 生成差异报告 Excel
    if internal_missing or external_missing:
        report_data = []
        for item in internal_missing:
            item["类型"] = "漏记风险"
            report_data.append(item)
        for item in external_missing:
            item["类型"] = "多记风险"
            report_data.append(item)
            
        df_res = pd.DataFrame(report_data)
        out_file = f"往来对账差异_{partner_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        # 简单美化 Excel (通过 pandas writer)
        try:
            with pd.ExcelWriter(out_file, engine='xlsxwriter') as writer:
                df_res.to_excel(writer, sheet_name='差异明细', index=False)
                workbook = writer.book
                worksheet = writer.sheets['差异明细']
                
                # 定义格式
                red_fmt = workbook.add_format({'font_color': '#9C0006', 'bg_color': '#FFC7CE'})
                yellow_fmt = workbook.add_format({'font_color': '#9C6500', 'bg_color': '#FFEB9C'})
                
                # 设置列宽
                worksheet.set_column('A:A', 12) # 日期
                worksheet.set_column('B:B', 12) # 金额
                worksheet.set_column('C:C', 30) # 摘要
                worksheet.set_column('D:D', 20) # 原因
                worksheet.set_column('E:E', 10) # 类型
                
                # 条件格式
                worksheet.conditional_format('E2:E1000', {'type': 'text',
                                                        'criteria': 'containing',
                                                        'value': '漏记',
                                                        'format': red_fmt})
                worksheet.conditional_format('E2:E1000', {'type': 'text',
                                                        'criteria': 'containing',
                                                        'value': '多记',
                                                        'format': yellow_fmt})
        except:
            # Fallback if xlsxwriter not available
            df_res.to_excel(out_file, index=False)
            
        log.info(f"📄 差异报告已生成: {out_file}", extra={"solution": "请打开Excel查看详情"})
        
        try:
            os.startfile(out_file)
        except:
            pass
            
        # 7. 交互式补录询问
        if internal_missing:
            print(f"\n🔧 发现 {len(internal_missing)} 笔漏记记录。")
            if input("👉 是否将这些记录自动补录到台账? (y/n): ").strip().lower() == 'y':
                # 转换格式适配 import_bank_records_to_feishu 或直接写入
                to_import = []
                for item in internal_missing:
                    # 构造导入所需的字典格式
                    # 需要: 记账日期, 实际收付金额, 往来单位费用, 业务类型, 费用归类
                    
                    # 简单推断业务类型
                    b_type = "付款" if item["金额"] < 0 else "收款"
                    
                    to_import.append({
                        "记账日期": item["日期"],
                        "实际收付金额": item["金额"],
                        "往来单位费用": partner_name,
                        "摘要": item["摘要"],
                        "业务类型": b_type,
                        "费用归类": "待确认", # 暂时设为待确认
                        "交易银行": "未指定",
                        "是否现金": "否",
                        "是否有票": "有票",
                        "备注": f"对账补录: {item['摘要']}"
                    })
                
                # 调用现有的导入逻辑
                batch_add_records(client, app_token, table_id, to_import)

    else:
        print(f"\n{Color.GREEN}🎉 完美匹配！双方账目一致。{Color.ENDC}")
        
    return True

def batch_add_records(client, app_token, table_id, data_list):
    """批量写入记录辅助函数"""
    records = []
    for item in data_list:
        fields = {
            "记账日期": int(pd.to_datetime(item["记账日期"]).timestamp() * 1000),
            "实际收付金额": float(item["实际收付金额"]),
            "往来单位费用": str(item["往来单位费用"]),
            "业务类型": str(item["业务类型"]),
            "费用归类": str(item["费用归类"]),
            "备注": str(item["备注"])
        }
        records.append(AppTableRecord.builder().fields(fields).build())
        
    # 分批写入
    batch_size = 100
    for i in range(0, len(records), batch_size):
        batch = records[i:i+batch_size]
        req = BatchCreateAppTableRecordRequest.builder() \
            .app_token(app_token) \
            .table_id(table_id) \
            .request_body(BatchCreateAppTableRecordRequestBody.builder().records(batch).build()) \
            .build()
            
        resp = client.bitable.v1.app_table_record.batch_create(req)
        if resp.success():
            log.info(f"✅ 已补录 {len(batch)} 条记录", extra={"solution": "无"})
        else:
            log.error(f"❌ 补录失败: {resp.msg}", extra={"solution": "检查权限"})

# 薪酬管理流程 (新)
def manage_salary_flow(client, app_token):
    """薪酬管理流程：导入工资表、生成凭证"""
    while True:
        print(f"\n{Color.HEADER}💰 薪酬管理 (工资/个税/社保){Color.ENDC}")
        print("1. 导入工资表 (Excel)")
        print("2. 查看薪酬列表 (最近10条)")
        print("3. 生成记账凭证 (同步到台账)")
        print("4. 个税计算器 (实用工具) [New]")
        print("0. 返回主菜单")
        
        choice = input(f"{Color.OKBLUE}请选择功能 (0-4): {Color.ENDC}").strip()
        
        if choice == '0': break

        if choice == '4':
            print(f"\n{Color.CYAN}🧮 2024 个人所得税计算器 (月度综合所得){Color.ENDC}")
            try:
                salary = float(input("请输入税前工资: ") or 0)
                social = float(input("请输入社保公积金扣除(个人部分): ") or 0)
                special = float(input("请输入专项附加扣除(如租金/养老等): ") or 0)
                threshold = 5000 # 起征点
                
                taxable = salary - social - special - threshold
                tax = 0
                if taxable <= 0:
                    tax = 0
                elif taxable <= 3000:
                    tax = taxable * 0.03
                elif taxable <= 12000:
                    tax = taxable * 0.1 - 210
                elif taxable <= 25000:
                    tax = taxable * 0.2 - 1410
                elif taxable <= 35000:
                    tax = taxable * 0.25 - 2660
                elif taxable <= 55000:
                    tax = taxable * 0.3 - 4410
                elif taxable <= 80000:
                    tax = taxable * 0.35 - 7160
                else:
                    tax = taxable * 0.45 - 15160
                
                net_salary = salary - social - max(0, tax)
                
                print(f"\n{Color.BOLD}计算结果:{Color.ENDC}")
                print(f"应纳税所得额: {max(0, taxable):.2f} 元")
                print(f"预计个税:     {Color.FAIL}{max(0, tax):.2f} 元{Color.ENDC}")
                print(f"税后实发:     {Color.GREEN}{net_salary:.2f} 元{Color.ENDC}")
                
            except ValueError:
                print("❌ 输入格式错误，请输入数字")
            
            input("\n按回车继续...")
            continue
        
        if choice == '1':
            file_path = input(f"{Color.OKBLUE}请输入工资表Excel路径 (直接回车扫描当前目录): {Color.ENDC}").strip()
            if not file_path:
                candidates = [f for f in os.listdir('.') if '工资' in f and f.endswith('.xlsx')]
                if candidates:
                    file_path = candidates[0]
                    print(f"🔍 自动找到: {file_path}")
                else:
                    print(f"{Color.WARNING}⚠️ 未找到工资表，请手动输入路径{Color.ENDC}")
                    continue
            
            if not os.path.exists(file_path):
                print(f"{Color.FAIL}❌ 文件不存在{Color.ENDC}")
                continue

            try:
                df = pd.read_excel(file_path)
                month_input = input(f"{Color.OKBLUE}请输入归属月份 (YYYY-MM): {Color.ENDC}").strip()
                
                table_id = get_table_id_by_name(client, app_token, "薪酬管理表")
                if not table_id:
                    print(f"{Color.FAIL}❌ 薪酬管理表不存在{Color.ENDC}")
                    continue

                records = []
                for _, row in df.iterrows():
                    fields = {
                        "月份": month_input,
                        "姓名": str(row.get('姓名', '')),
                        "部门": str(row.get('部门', '')),
                        "实发工资": float(row.get('实发工资', 0) or 0),
                        "状态": "已发放"
                    }
                    if '基本工资' in row: fields["基本工资"] = float(row.get('基本工资', 0) or 0)
                    if '绩效工资' in row: fields["绩效工资"] = float(row.get('绩效工资', 0) or 0)
                    if '社保个人' in row: fields["社保扣除"] = float(row.get('社保个人', 0) or 0)
                    if '个税' in row: fields["个税扣除"] = float(row.get('个税', 0) or 0)
                    
                    records.append(AppTableRecord.builder().fields(fields).build())

                # Batch Write
                batch_size = 100
                total_success = 0
                for i in range(0, len(records), batch_size):
                    batch = records[i:i+batch_size]
                    req = BatchCreateAppTableRecordRequest.builder() \
                        .app_token(app_token) \
                        .table_id(table_id) \
                        .request_body(BatchCreateAppTableRecordRequestBody.builder().records(batch).build()) \
                        .build()
                    
                    resp = client.bitable.v1.app_table_record.batch_create(req)
                    if resp.success():
                        total_success += len(batch)
                    else:
                        print(f"{Color.FAIL}❌ 部分写入失败: {resp.msg}{Color.ENDC}")
                
                print(f"{Color.OKGREEN}✅ 成功导入 {total_success} 条薪酬记录 ({month_input}){Color.ENDC}")
                
                # Ask to generate accounting vouchers
                if input(f"{Color.OKBLUE}是否生成记账凭证(写入日常台账)? (y/n): {Color.ENDC}").lower() == 'y':
                    total_net = df["实发工资"].sum()
                    ledger_id = get_table_id_by_name(client, app_token, "日常台账表")
                    
                    print(f"{Color.WARNING}⚠️ 注意: 将生成【实发工资】的支出记录。个税和社保请在缴纳时通过银行流水导入。{Color.ENDC}")
                    
                    record_fields = {
                        "记账日期": int(datetime.now().timestamp() * 1000),
                        "实际收付金额": float(total_net),
                        "业务类型": "费用",
                        "费用归类": "工资薪金",
                        "摘要": f"{month_input} 工资发放 (共{len(df)}人)",
                        "交易银行": "待确认",
                        "是否有票": "无票",
                        "备注": "自动生成自薪酬管理"
                    }
                    
                    rec = AppTableRecord.builder().fields(record_fields).build()
                    req_l = BatchCreateAppTableRecordRequest.builder().app_token(app_token).table_id(ledger_id).request_body(
                        BatchCreateAppTableRecordRequestBody.builder().records([rec]).build()).build()
                    resp_l = client.bitable.v1.app_table_record.batch_create(req_l)
                    if resp_l.success():
                         print(f"{Color.OKGREEN}✅ 已生成支出凭证: {total_net} 元{Color.ENDC}")
                    else:
                         print(f"{Color.FAIL}❌ 生成凭证失败: {resp_l.msg}{Color.ENDC}")

            except Exception as e:
                log.error(f"操作失败: {e}")
                print(f"{Color.FAIL}❌ 操作失败: {e}{Color.ENDC}")
                
        elif choice == '2':
             table_id = get_table_id_by_name(client, app_token, "薪酬管理表")
             if not table_id:
                 print(f"{Color.WARNING}⚠️ 薪酬表不存在{Color.ENDC}")
                 continue
             records = get_all_records(client, app_token, table_id) # Should limit? get_all_records gets all.
             # Assuming get_all_records is efficient enough for now or we just take last 10
             if not records:
                 print("📭 暂无记录")
             else:
                 print(f"\n{Color.UNDERLINE}最近 10 条薪酬记录:{Color.ENDC}")
                 # Sort by creation? The API returns in some order.
                 # Just show last 10
                 for r in records[-10:]:
                     f = r.fields
                     print(f"- {f.get('月份')} | {f.get('姓名')} | 实发: {f.get('实发工资')} | 状态: {f.get('状态')}")
             
        elif choice == '3':
            month_input = input(f"{Color.OKBLUE}请输入月份 (YYYY-MM): {Color.ENDC}").strip()
            if len(month_input) != 7 or '-' not in month_input:
                print(f"{Color.WARNING}⚠️ 格式错误，请使用 YYYY-MM 格式{Color.ENDC}")
                continue

            table_id = get_table_id_by_name(client, app_token, "薪酬管理表")
            if not table_id:
                print(f"{Color.FAIL}❌ 未找到薪酬管理表{Color.ENDC}")
                continue

            # 查询该月记录
            filter_str = f'CurrentValue.[月份]="{month_input}"'
            print(f"🔍 正在查询 {month_input} 的薪酬记录...")
            records = get_all_records(client, app_token, table_id, filter_info=filter_str)

            if not records:
                print(f"{Color.WARNING}📭 {month_input} 暂无薪酬记录{Color.ENDC}")
                continue

            # 统计金额
            total_net = 0.0
            person_count = 0
            details_summary = []

            for r in records:
                f = r.fields
                try:
                    net = float(f.get("实发工资", 0))
                except:
                    net = 0.0
                
                total_net += net
                person_count += 1
                if len(details_summary) < 3:
                    details_summary.append(f"{f.get('姓名', '未知')}")

            print(f"\n{Color.OKGREEN}📊 {month_input} 薪酬统计概览:{Color.ENDC}")
            print(f"--------------------------------")
            print(f"👥 发放人数: {person_count} 人 ({', '.join(details_summary)}...)")
            print(f"💰 实发总额: {total_net:,.2f} 元")
            print(f"--------------------------------")

            if input(f"{Color.OKBLUE}❓ 确认生成记账凭证(同步到日常台账)? (y/n): {Color.ENDC}").lower() == 'y':
                ledger_id = get_table_id_by_name(client, app_token, "日常台账表")
                if not ledger_id:
                    print(f"{Color.FAIL}❌ 未找到日常台账表{Color.ENDC}")
                    continue

                record_fields = {
                    "记账日期": int(datetime.now().timestamp() * 1000),
                    "实际收付金额": float(total_net),
                    "业务类型": "费用",
                    "费用归类": "工资薪金",
                    "摘要": f"{month_input} 工资发放 (共{person_count}人)",
                    "交易银行": "待确认", 
                    "是否有票": "无票",
                    "备注": f"薪酬模块自动生成 - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                }

                rec = AppTableRecord.builder().fields(record_fields).build()
                req = BatchCreateAppTableRecordRequest.builder() \
                    .app_token(app_token) \
                    .table_id(ledger_id) \
                    .request_body(BatchCreateAppTableRecordRequestBody.builder().records([rec]).build()) \
                    .build()
                
                resp = client.bitable.v1.app_table_record.batch_create(req)
                
                if resp.success():
                    print(f"{Color.OKGREEN}✅ 凭证生成成功！已写入日常台账。{Color.ENDC}")
                else:
                    print(f"{Color.FAIL}❌ 生成凭证失败: {resp.msg}{Color.ENDC}")

# AI 财务诊断 (GLM-4-Flash)
def get_ai_insight(data_context):
    if not ZHIPUAI_API_KEY:
        log.info("🤖 AI 未配置，跳过智能分析", extra={"solution": "在.env配置 ZHIPUAI_API_KEY"})
        return ""
        
    log.info("🧠 正在进行AI财务诊断 (GLM-4-Flash)...", extra={"solution": "无"})
    try:
        client = ZhipuAI(api_key=ZHIPUAI_API_KEY)
        
        prompt = f"""
        你是一位经验丰富的CFO（首席财务官）。请根据以下小企业本月财务数据，给出一段简练、专业的经营评价，并提出1条具体的改进建议。
        
        【财务数据】
        - 有票收入：{data_context['income']}
        - 总支出：{data_context['cost']} (其中无票支出：{data_context['no_ticket_cost']})
        - 账面利润：{data_context['profit']} (利润率：{data_context['margin']}%)
        - 预计税负：{data_context['tax']}
        - 风险提示：{data_context['risk_msg']}
        
        【要求】
        1. 语气专业、客观，但也通俗易懂（给老板看）。
        2. 重点关注利润率、合规风险（无票占比）。
        3. 100字以内。
        4. 不要用Markdown格式，直接输出纯文本。
        """
        
        response = client.chat.completions.create(
            model="glm-4-flash",
            messages=[
                {"role": "user", "content": prompt}
            ],
            stream=False
        )
        insight = response.choices[0].message.content.strip()
        log.info("✅ AI诊断完成", extra={"solution": "无"})
        return f"\n🤖 [AI 财务诊断]\n{insight}"
        
    except Exception as e:
        log.error(f"❌ AI分析失败: {str(e)}", extra={"solution": "检查API Key或网络"})
        return ""

# 税务统计 (含风险预警)
@retry_on_failure(max_retries=2, delay=3)
def calculate_tax(client, app_token, target_year=None):
    if target_year:
        log.info(f"🧮 开始 {target_year}年度 税务及风险分析...", extra={"solution": "无"})
        year = target_year
    else:
        log.info("🧮 开始税务及风险分析...", extra={"solution": "无"})
        year = datetime.now().year
        
    table_id = get_table_id_by_name(client, app_token, "日常台账表")
    if not table_id:
        return False
        
    # 优化：只获取指定年度数据
    start_ts = int(datetime(year, 1, 1).timestamp() * 1000)
    end_ts = int(datetime(year + 1, 1, 1).timestamp() * 1000)
    filter_str = f'AND(CurrentValue.[记账日期]>={start_ts}, CurrentValue.[记账日期]<{end_ts})'
    
    log.info(f"🔍 正在拉取 {year} 年度数据...", extra={"solution": "无"})
    records = get_all_records(client, app_token, table_id, filter_info=filter_str)
    
    total_income_ticket = 0.0 # 有票收入
    total_cost_ticket = 0.0   # 有票成本
    total_cost_no_ticket = 0.0 # 无票成本 (总)
    
    # 细分风险统计
    g_bank_no_ticket = 0.0   # G银行(对公)无票支出 -> 高风险
    n_bank_flow = 0.0        # N银行(私户)流水总额 -> 私户避税风险
    
    for r in records:
        fields = r.fields
        has_ticket = fields.get("是否有票") == "有票"
        bank = str(fields.get("交易银行", ""))
        is_g_bank = "G银行" in bank or "对公" in bank
        is_n_bank = "N银行" in bank or "现金" in bank or "私户" in bank
        
        biz_type = fields.get("业务类型")
        amount = float(fields.get("实际收付金额", 0))
        
        if biz_type == "收款":
            if has_ticket:
                total_income_ticket += amount
            if is_n_bank:
                n_bank_flow += amount
        elif biz_type == "付款" or biz_type == "费用":
            if has_ticket:
                total_cost_ticket += amount
            else:
                total_cost_no_ticket += amount
                if is_g_bank:
                    g_bank_no_ticket += amount
            
            if is_n_bank:
                n_bank_flow += abs(amount)
                
    # 计算
    vat_rate = VAT_RATE / 100.0
    estimated_vat = (total_income_ticket / (1 + vat_rate)) * vat_rate
    
    profit = total_income_ticket - total_cost_ticket
    corp_tax = max(0, profit * (TAX_CONFIG["corporate_tax_rate"] / 100.0))
    
    # 风险分析
    total_cost = total_cost_ticket + total_cost_no_ticket
    no_ticket_ratio = (total_cost_no_ticket / total_cost * 100) if total_cost > 0 else 0
    profit_margin = (profit / total_income_ticket * 100) if total_income_ticket > 0 else 0
    
    risk_msg = ""
    if g_bank_no_ticket > 0:
        risk_msg += f"\n⚠️ [严重风险] G银行(对公)存在无票支出: {g_bank_no_ticket:,.2f} (必须补票)"
    if no_ticket_ratio > 30:
        risk_msg += f"\n⚠️ [经营风险] 无票支出整体占比 {no_ticket_ratio:.1f}%"
    if n_bank_flow > 500000: # 假设阈值
        risk_msg += f"\n⚠️ [私户风险] N银行(私户)流水过大: {n_bank_flow:,.2f}"
        
    if profit_margin < 5 and total_income_ticket > 0:
        risk_msg += f"\n⚠️ [异常] 账面利润率仅 {profit_margin:.1f}% (易被稽查)"
    
    # AI 分析
    ai_context = {
        "income": f"{total_income_ticket:,.2f}",
        "cost": f"{total_cost:,.2f}",
        "no_ticket_cost": f"{total_cost_no_ticket:,.2f}",
        "profit": f"{profit:,.2f}",
        "margin": f"{profit_margin:.1f}",
        "tax": f"{estimated_vat + corp_tax:,.2f}",
        "risk_msg": risk_msg.replace("\n", "; ")
    }
    ai_insight = get_ai_insight(ai_context)
    
    msg = (
        f"📊 财务经营月报 (费率: {VAT_RATE}%)\n"
        f"------------------------\n"
        f"💰 有票收入: {total_income_ticket:,.2f}\n"
        f"💸 有票成本: {total_cost_ticket:,.2f}\n"
        f"🚫 无票支出: {total_cost_no_ticket:,.2f} (含G银行: {g_bank_no_ticket:,.2f})\n"
        f"------------------------\n"
        f"🧾 预计增值税: {estimated_vat:,.2f}\n"
        f"🏦 预计所得税: {corp_tax:,.2f}\n"
        f"📈 账面利润率: {profit_margin:.1f}%\n"
        f"------------------------{risk_msg}\n"
        f"{ai_insight}\n"
        f"------------------------\n"
        f"💡 仅供参考，具体以申报为准"
    )
    log.info("✅ 税务统计完成", extra={"solution": "无"})
    print(msg) # 保持原有打印
    return msg
    
    # 构造卡片
    header_color = "green" if profit_margin >= 5 else "red"
    
    elements = [
        {
            "tag": "div",
            "fields": [
                {"is_short": True, "text": {"tag": "lark_md", "content": f"**有票收入**\n¥ {total_income_ticket:,.2f}"}},
                {"is_short": True, "text": {"tag": "lark_md", "content": f"**有票成本**\n¥ {total_cost_ticket:,.2f}"}},
                {"is_short": True, "text": {"tag": "lark_md", "content": f"**预计增值税**\n¥ {estimated_vat:,.2f}"}},
                {"is_short": True, "text": {"tag": "lark_md", "content": f"**预计所得税**\n¥ {corp_tax:,.2f}"}}
            ]
        },
        {"tag": "hr"},
        {
            "tag": "div",
            "text": {
                "tag": "lark_md", 
                "content": f"📈 **利润率**: {profit_margin:.1f}% | 🚫 **无票占比**: {no_ticket_ratio:.1f}%\n{risk_msg}"
            }
        }
    ]
    
    if ai_insight:
        elements.append({
            "tag": "div",
            "text": {"tag": "lark_md", "content": f"**🤖 AI CFO 观点**:\n{ai_insight.replace('🤖 [AI 财务诊断]', '').strip()}"}
        })
        
    elements.append({
        "tag": "action",
        "actions": [
            {
                "tag": "button",
                "text": {"tag": "plain_text", "content": "查看税务指南"},
                "url": WIKI_TAX,
                "type": "default"
            }
        ]
    })
    
    card = {
        "header": {
            "title": {"tag": "plain_text", "content": "📊 财务月度经营分析"},
            "template": header_color
        },
        "elements": elements
    }
    
    send_bot_message("税务分析报告", "interactive", card)
    return True

# 导出待补票清单 (新功能)
@retry_on_failure(max_retries=2, delay=3)
def export_missing_tickets(client, app_token, silent=False):
    log.info("🔍 正在查找待补票记录...", extra={"solution": "无"})
    table_id = get_table_id_by_name(client, app_token, "日常台账表")
    if not table_id:
        return 0
        
    records = get_all_records(client, app_token, table_id)
    missing_list = []
    
    for r in records:
        fields = r.fields
        # 条件: (无票 OR 待补票标记=是) AND (业务类型=付款/费用)
        is_expense = fields.get("业务类型") in ["付款", "费用"]
        no_ticket = fields.get("是否有票") == "无票"
        pending = "是" in str(fields.get("待补票标记", ""))
        
        if is_expense and (no_ticket or pending):
            row = fields.copy()
            # 注入 record_id 以便后续更新
            row['record_id'] = r.record_id
            
            # 日期格式化
            if isinstance(row.get("记账日期"), int):
                row["记账日期"] = datetime.fromtimestamp(row["记账日期"] / 1000).strftime("%Y-%m-%d")
            missing_list.append(row)
            
    if missing_list:
        df = pd.DataFrame(missing_list)
        # 整理列顺序
        cols = ["记账日期", "凭证号", "费用归类", "往来单位费用", "实际收付金额", "是否有票", "待补票标记", "备注", "操作人"]
        # 只保留存在的列
        final_cols = [c for c in cols if c in df.columns]
        df = df[final_cols]
        
        filename = f"待补票清单_{datetime.now().strftime('%Y%m%d')}.xlsx"
        file_path = os.path.join(LOCAL_FOLDER, filename)
        df.to_excel(file_path, index=False)
        
        msg = f"已生成待补票清单: {len(missing_list)}条"
        log.info(f"✅ {msg}", extra={"solution": "发送给业务员补票"})
        if not silent:
            send_bot_message(f"📢 {msg}\n📄 文件位置: {file_path}", "alert")
        
        # 询问是否进入交互式补录模式 (仅在非静默模式下)
        if not silent:
            print(f"\n💡 发现 {len(missing_list)} 笔待补票记录。")
            if input("👉 是否现在开始【交互式补录】(逐条确认收到发票)? (y/n): ").strip().lower() == 'y':
                resolve_missing_tickets(client, app_token, missing_list, table_id)
            
    else:
        log.info("✅ 没有发现待补票记录", extra={"solution": "无"})
        if not silent:
            send_bot_message("👏 只有完美的账单！目前没有待补票记录。", "alert")
            
    return len(missing_list)

def resolve_missing_tickets(client, app_token, missing_list, table_id):
    """交互式补录发票状态"""
    print(f"\n🎫 启动交互式补票模式 ({len(missing_list)}条待处理)...")
    print("-----------------------------------")
    print("说明: 按 'y' 标记为【有票】，按 'n' 或回车跳过，按 'q' 退出。")
    print("-----------------------------------")
    
    count = 0
    for row in missing_list:
        # 显示记录详情
        date_str = row.get("记账日期", "未知日期")
        partner = row.get("往来单位费用", "未知")
        amount = row.get("实际收付金额", 0)
        memo = row.get("备注", "")
        
        print(f"\n📝 [{count+1}/{len(missing_list)}] {date_str} | {partner} | {amount}元 | {memo}")
        choice = input("👉 是否已收到发票? (y/n/q): ").strip().lower()
        
        if choice == 'q':
            break
            
        if choice == 'y':
            # 更新记录
            record_id = row.get("record_id") # 需要确保 get_all_records 返回了 record_id
            if not record_id:
                # 尝试通过原始对象获取 (如果 row 是 dict，可能没有 record_id，除非 get_all_records 特殊处理)
                # 这里假设 get_all_records 返回的 record 对象包含 record_id，但我们之前转换为了 dict
                # 这是一个潜在 bug，我们需要检查 get_all_records 的实现或 missing_list 的构造
                # 修正: 在 export_missing_tickets 中构造 missing_list 时，应该包含 record_id
                print("❌ 无法获取记录ID，跳过")
                continue
                
            try:
                # 更新字段: 是否有票=有票, 待补票标记=""
                fields = {"是否有票": "有票", "待补票标记": ""}
                
                req = UpdateAppTableRecordRequest.builder() \
                    .app_token(app_token) \
                    .table_id(table_id) \
                    .record_id(record_id) \
                    .app_table_record(AppTableRecord.builder().fields(fields).build()) \
                    .build()
                    
                resp = client.bitable.v1.app_table_record.update(req)
                if resp.success():
                    print("✅ 已更新为 [有票]")
                    count += 1
                else:
                    print(f"❌ 更新失败: {resp.msg}")
            except Exception as e:
                print(f"❌ 错误: {e}")
                
    print(f"\n🎉 补录完成！共更新 {count} 条记录。")


# 生成HTML可视化报表
@retry_on_failure(max_retries=2, delay=3)
def generate_html_report(client, app_token, target_year=None):
    if target_year:
        log.info(f"📊 正在生成 {target_year}年度 可视化报表...", extra={"solution": "无"})
        year = target_year
    else:
        log.info("📊 正在生成可视化报表...", extra={"solution": "无"})
        year = datetime.now().year

    table_id = get_table_id_by_name(client, app_token, "日常台账表")
    if not table_id:
        return False
        
    # 获取指定年度数据
    start_ts = int(datetime(year, 1, 1).timestamp() * 1000)
    end_ts = int(datetime(year + 1, 1, 1).timestamp() * 1000)
    filter_str = f'AND(CurrentValue.[记账日期]>={start_ts}, CurrentValue.[记账日期]<{end_ts})'
    records = get_all_records(client, app_token, table_id, filter_info=filter_str)
    
    # 数据处理
    monthly_data = {} # month -> {income, expense, profit}
    
    for r in records:
        fields = r.fields
        r_date = fields.get("记账日期")
        if not r_date: continue
        
        if isinstance(r_date, int):
            date_obj = datetime.fromtimestamp(r_date / 1000)
        else:
            try:
                date_obj = datetime.strptime(str(r_date).split(" ")[0], "%Y-%m-%d")
            except: continue
            
        month_key = date_obj.strftime("%Y-%m")
        if month_key not in monthly_data:
            monthly_data[month_key] = {"income": 0, "expense": 0}
            
        amount = float(fields.get("实际收付金额", 0))
        biz_type = fields.get("业务类型")
        
        if biz_type == "收款":
            monthly_data[month_key]["income"] += amount
        elif biz_type == "付款" or biz_type == "费用":
            monthly_data[month_key]["expense"] += amount
            
    # 生成HTML
    months = sorted(monthly_data.keys())
    incomes = [round(monthly_data[m]["income"], 2) for m in months]
    expenses = [round(monthly_data[m]["expense"], 2) for m in months]
    profits = [round(monthly_data[m]["income"] - monthly_data[m]["expense"], 2) for m in months]
    
    # 风险/合规数据
    total_cost = sum(expenses)
    total_cost_ticket = 0
    total_cost_no_ticket = 0
    
    for r in records:
        fields = r.fields
        biz_type = fields.get("业务类型")
        if biz_type in ["付款", "费用"]:
            amount = float(fields.get("实际收付金额", 0))
            if fields.get("是否有票") == "有票":
                total_cost_ticket += amount
            else:
                total_cost_no_ticket += amount
                
    compliance_data = [
        {"value": round(total_cost_ticket, 2), "name": "有票成本 (合规)"},
        {"value": round(total_cost_no_ticket, 2), "name": "无票成本 (风险)"}
    ]
    
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <title>财务经营分析仪表盘</title>
        <script src="https://cdn.jsdelivr.net/npm/echarts@5.4.3/dist/echarts.min.js"></script>
        <style>
            body {{ font-family: 'Segoe UI', sans-serif; background: #f5f6fa; padding: 20px; }}
            .container {{ max-width: 1200px; margin: 0 auto; }}
            .card {{ background: white; border-radius: 8px; padding: 20px; margin-bottom: 20px; box-shadow: 0 2px 5px rgba(0,0,0,0.05); }}
            h1 {{ color: #2c3e50; text-align: center; }}
            .summary {{ display: flex; justify-content: space-around; margin-bottom: 20px; }}
            .stat-box {{ text-align: center; padding: 15px; background: #fff; border-radius: 8px; flex: 1; margin: 0 10px; box-shadow: 0 2px 5px rgba(0,0,0,0.05); }}
            .stat-value {{ font-size: 24px; font-weight: bold; color: #3498db; }}
            .stat-label {{ color: #7f8c8d; font-size: 14px; }}
            #main-chart {{ width: 100%; height: 500px; }}
            #pie-chart {{ width: 100%; height: 400px; }}
            .row {{ display: flex; gap: 20px; }}
            .col-8 {{ flex: 2; }}
            .col-4 {{ flex: 1; }}
        </style>
    </head>
    <body>
        <div class="container">
            <h1>📊 {year}年度财务经营分析</h1>
            
            <div class="summary">
                <div class="stat-box">
                    <div class="stat-value">¥ {sum(incomes):,.2f}</div>
                    <div class="stat-label">年度总收入</div>
                </div>
                <div class="stat-box">
                    <div class="stat-value" style="color: #e74c3c">¥ {sum(expenses):,.2f}</div>
                    <div class="stat-label">年度总支出</div>
                </div>
                <div class="stat-box">
                    <div class="stat-value" style="color: #27ae60">¥ {sum(profits):,.2f}</div>
                    <div class="stat-label">年度净利润</div>
                </div>
            </div>
            
            <div class="row">
                <div class="card col-8">
                    <div id="main-chart"></div>
                </div>
                <div class="card col-4">
                    <div id="pie-chart"></div>
                </div>
            </div>
            
            <script>
                var chartDom = document.getElementById('main-chart');
                var myChart = echarts.init(chartDom);
                var option;

                option = {{
                    title: {{ text: '月度收支趋势图' }},
                    tooltip: {{ trigger: 'axis' }},
                    legend: {{ data: ['收入', '支出', '利润'] }},
                    grid: {{ left: '3%', right: '4%', bottom: '3%', containLabel: true }},
                    xAxis: {{ type: 'category', boundaryGap: false, data: {months} }},
                    yAxis: {{ type: 'value' }},
                    series: [
                        {{ name: '收入', type: 'line', stack: 'Total', areaStyle: {{}}, data: {incomes}, itemStyle: {{ color: '#3498db' }} }},
                        {{ name: '支出', type: 'line', stack: 'Total', areaStyle: {{}}, data: {expenses}, itemStyle: {{ color: '#e74c3c' }} }},
                        {{ name: '利润', type: 'bar', data: {profits}, itemStyle: {{ color: '#27ae60' }} }}
                    ]
                }};
                option && myChart.setOption(option);
                
                // 饼图
                var pieDom = document.getElementById('pie-chart');
                var pieChart = echarts.init(pieDom);
                var pieOption = {{
                    title: {{ text: '成本合规性分析', subtext: '有票 vs 无票', left: 'center' }},
                    tooltip: {{ trigger: 'item' }},
                    legend: {{ orient: 'vertical', left: 'left' }},
                    series: [
                        {{
                            name: '成本构成',
                            type: 'pie',
                            radius: '50%',
                            data: {compliance_data},
                            emphasis: {{
                                itemStyle: {{
                                    shadowBlur: 10,
                                    shadowOffsetX: 0,
                                    shadowColor: 'rgba(0, 0, 0, 0.5)'
                                }}
                            }},
                            itemStyle: {{
                                color: function(params) {{
                                    var colorList = ['#27ae60', '#e74c3c'];
                                    return colorList[params.dataIndex];
                                }}
                            }}
                        }}
                    ]
                }};
                pieOption && pieChart.setOption(pieOption);
            </script>
        </div>
    </body>
    </html>
    """
    
    filename = f"财务分析报表_{datetime.now().strftime('%Y%m%d')}.html"
    filepath = os.path.join(LOCAL_FOLDER, filename)
    with open(filepath, "w", encoding="utf-8") as f:
        f.write(html_content)
        
    log.info(f"✅ 报表已生成: {filepath}", extra={"solution": "双击打开HTML文件"})
    # 尝试自动打开
    try:
        os.startfile(filepath)
    except:
        pass
        
    return True

# 导出备份
@retry_on_failure(max_retries=2, delay=3)
def export_to_excel(client, app_token, target_path=None):
    """全量备份：导出所有数据表到 Excel"""
    log.info("💾 开始全量云端数据备份...", extra={"solution": "无"})
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    if target_path:
        backup_path = os.path.join(target_path, f"飞书台账全量备份_{timestamp}.xlsx")
    else:
        backup_path = os.path.join(LOCAL_FOLDER, f"飞书台账全量备份_{timestamp}.xlsx")
    
    try:
        # 1. 获取所有数据表
        tables = []
        page_token = None
        while True:
            req = ListAppTableRequest.builder() \
                .app_token(app_token) \
                .page_size(20) \
                .page_token(page_token) \
                .build()
            resp = client.bitable.v1.app_table.list(req)
            if resp.success():
                if resp.data.items:
                    tables.extend(resp.data.items)
                if not resp.data.has_more:
                    break
                page_token = resp.data.page_token
            else:
                log.error(f"无法获取表格列表: {resp.msg}")
                break

        if not tables:
            return False

        # 2. [V9.5新特性] 并行获取数据 (Parallel Backup)
        table_data_map = {}
        
        def fetch_table_data(table):
            t_name = table.name
            t_id = table.table_id
            try:
                # print(f"   ⏳ [并行] 正在拉取: {t_name}...") # 减少刷屏
                records = get_all_records(client, app_token, t_id)
                clean_data = []
                if records:
                    for r in records:
                        row = r.fields.copy()
                        # 转换时间戳
                        for k, v in row.items():
                            if isinstance(v, int) and v > 1000000000000: # 简单判断毫秒时间戳
                                try:
                                    row[k] = datetime.fromtimestamp(v / 1000).strftime("%Y-%m-%d %H:%M:%S")
                                except:
                                    pass
                        clean_data.append(row)
                return t_name, pd.DataFrame(clean_data)
            except Exception as e:
                log.error(f"❌ 获取表 {t_name} 失败: {e}")
                return t_name, pd.DataFrame()

        log.info(f"🚀 启动并行备份，正在同时拉取 {len(tables)} 张表...", extra={"solution": "无"})
        with ThreadPoolExecutor(max_workers=5) as executor:
            future_to_table = {executor.submit(fetch_table_data, t): t for t in tables}
            for future in as_completed(future_to_table):
                t_name, df = future.result()
                table_data_map[t_name] = df
                print(f"   ✅ 已就绪: {t_name} ({len(df)} 条)")

        # 3. 写入 Excel (带美化)
        log.info("💾 正在写入Excel文件...", extra={"solution": "无"})
        # 使用 xlsxwriter 引擎以支持样式
        with pd.ExcelWriter(backup_path, engine='xlsxwriter') as writer:
            for table in tables: # 保持原有顺序
                table_name = table.name
                if table_name in table_data_map:
                    df = table_data_map[table_name]
                    # Excel Sheet 名字不能超过31个字符
                    safe_name = table_name[:30]
                    # 处理重名Sheet (极其罕见)
                    if safe_name in writer.sheets:
                        safe_name = (table_name[:25] + "_1")
                    
                    df.to_excel(writer, sheet_name=safe_name, index=False)
                    
                    # --- 美化开始 ---
                    workbook = writer.book
                    worksheet = writer.sheets[safe_name]
                    
                    # 格式定义
                    header_fmt = workbook.add_format({
                        'bold': True,
                        'text_wrap': False,
                        'valign': 'top',
                        'fg_color': '#D7E4BC', # 浅绿背景
                        'border': 1
                    })
                    data_fmt = workbook.add_format({
                        'border': 1
                    })
                    date_fmt = workbook.add_format({
                        'num_format': 'yyyy-mm-dd hh:mm:ss',
                        'border': 1
                    })
                    num_fmt = workbook.add_format({
                        'num_format': '#,##0.00', # 千分位
                        'border': 1
                    })
                    
                    # 应用表头格式
                    for col_num, value in enumerate(df.columns.values):
                        worksheet.write(0, col_num, value, header_fmt)
                        
                    # 自动调整列宽
                    for i, col in enumerate(df.columns):
                        max_len = 0
                        # 检查列名长度
                        max_len = max(max_len, len(str(col)) * 2) 
                        # 检查数据长度 (取前50行采样，避免太慢)
                        sample_vals = df[col].head(50).astype(str)
                        for v in sample_vals:
                            l = len(v)
                            # 中文占2字符简单估算
                            utf8_len = len(v.encode('utf-8'))
                            display_len = (utf8_len - l)/2 + l
                            max_len = max(max_len, display_len)
                            
                        # 限制最大宽度
                        final_width = min(max_len + 2, 50) 
                        worksheet.set_column(i, i, final_width, data_fmt)
                        
                        # 针对特定列应用特定格式
                        if '金额' in str(col) or '单价' in str(col) or '原值' in str(col):
                             worksheet.set_column(i, i, final_width, num_fmt)
                        elif '日期' in str(col) or '时间' in str(col):
                             worksheet.set_column(i, i, 20, date_fmt) # 日期固定宽一点
                    # --- 美化结束 ---
                    
                else:
                    pd.DataFrame().to_excel(writer, sheet_name=table_name[:30])

        log.info(f"✅ 全量备份成功: {backup_path}", extra={"solution": "妥善保管"})
        if not target_path: # 如果是手动触发，发送通知
            send_bot_message(f"✅ 数据已备份至本地:\n{backup_path}", "accountant")
        return True
    except Exception as e:
        log.error(f"❌ 备份失败: {str(e)}", extra={"solution": "检查磁盘空间"})
        return False

# 创建基础信息表
@retry_on_failure(max_retries=2, delay=3)
def create_basic_info_table(client, app_token):
    # 先检查是否存在
    existing_id = get_table_id_by_name(client, app_token, "基础信息表")
    if existing_id:
        log.info(f"⚠️ 基础信息表已存在 (ID: {existing_id})", extra={"solution": "无"})
        return True, existing_id

    req = CreateAppTableRequest.builder() \
        .app_token(app_token) \
        .request_body(CreateAppTableRequestBody.builder()
            .table(ReqTable.builder()
                .name("基础信息表")
                .fields([
                    AppTableCreateHeader.builder().field_name("产品名称").type(FT.TEXT).build(),
                    AppTableCreateHeader.builder().field_name("单位成本").type(FT.NUMBER).build(), # Number format can be set later or via property
                    AppTableCreateHeader.builder().field_name("备注").type(FT.TEXT).build(),
                    # "最后更新时间" is usually a system field, but we can add a Date field
                    AppTableCreateHeader.builder().field_name("最后更新时间").type(FT.DATE).build() 
                ])
                .build())
            .build()) \
        .build()
    
    resp = client.bitable.v1.app_table.create(req)
    if resp.success():
        log.info("✅ 基础信息表创建成功", extra={"solution": "无"})
        return True, resp.data.table_id
    else:
        log.error(f"❌ 基础信息表创建失败: {resp.msg}", extra={"solution": "检查权限"})
        return False, None

# 创建往来单位表 (新)
@retry_on_failure(max_retries=2, delay=3)
def create_partner_table(client, app_token):
    existing_id = get_table_id_by_name(client, app_token, "往来单位表")
    if existing_id:
        log.info(f"⚠️ 往来单位表已存在 (ID: {existing_id})", extra={"solution": "无"})
        return True, existing_id

    req = CreateAppTableRequest.builder() \
        .app_token(app_token) \
        .request_body(CreateAppTableRequestBody.builder()
            .table(ReqTable.builder()
                .name("往来单位表")
                .fields([
                    AppTableCreateHeader.builder().field_name("单位名称").type(FT.TEXT).build(),
                    AppTableCreateHeader.builder().field_name("类型").type(FT.SELECT).property(AppTableFieldProperty.builder().options([
                        AppTableFieldPropertyOption.builder().name("客户").build(),
                        AppTableFieldPropertyOption.builder().name("供应商").build(),
                        AppTableFieldPropertyOption.builder().name("外发加工").build(),
                        AppTableFieldPropertyOption.builder().name("其他").build()
                    ]).build()).build(),
                    AppTableCreateHeader.builder().field_name("纳税人识别号").type(FT.TEXT).build(),
                    AppTableCreateHeader.builder().field_name("开户行及账号").type(FT.TEXT).build(),
                    AppTableCreateHeader.builder().field_name("地址及电话").type(FT.TEXT).build(),
                    AppTableCreateHeader.builder().field_name("联系人").type(FT.TEXT).build(),
                    AppTableCreateHeader.builder().field_name("联系电话").type(FT.TEXT).build(),
                    AppTableCreateHeader.builder().field_name("备注").type(FT.TEXT).build()
                ])
                .build())
            .build()) \
        .build()
    
    resp = client.bitable.v1.app_table.create(req)
    if resp.success():
        log.info("✅ 往来单位表创建成功", extra={"solution": "无"})
        return True, resp.data.table_id
    else:
        log.error(f"❌ 往来单位表创建失败: {resp.msg}", extra={"solution": "检查权限"})
        return False, None

# 创建发票管理表 (新)
@retry_on_failure(max_retries=2, delay=3)
def create_invoice_table(client, app_token):
    existing_id = get_table_id_by_name(client, app_token, "发票管理表")
    if existing_id:
        log.info(f"⚠️ 发票管理表已存在 (ID: {existing_id})", extra={"solution": "无"})
        return True, existing_id

    req = CreateAppTableRequest.builder() \
        .app_token(app_token) \
        .request_body(CreateAppTableRequestBody.builder()
            .table(ReqTable.builder()
                .name("发票管理表")
                .fields([
                    AppTableCreateHeader.builder().field_name("发票号码").type(FT.TEXT).build(),
                    AppTableCreateHeader.builder().field_name("发票代码").type(FT.TEXT).build(),
                    AppTableCreateHeader.builder().field_name("开票日期").type(FT.DATE).build(),
                    AppTableCreateHeader.builder().field_name("类型").type(FT.SELECT).property(AppTableFieldProperty.builder().options([
                        AppTableFieldPropertyOption.builder().name("进项专票").build(),
                        AppTableFieldPropertyOption.builder().name("进项普票").build(),
                        AppTableFieldPropertyOption.builder().name("销项专票").build(),
                        AppTableFieldPropertyOption.builder().name("销项普票").build()
                    ]).build()).build(),
                    AppTableCreateHeader.builder().field_name("购买方/销售方").type(FT.TEXT).build(),
                    AppTableCreateHeader.builder().field_name("不含税金额").type(FT.NUMBER).build(),
                    AppTableCreateHeader.builder().field_name("税额").type(FT.NUMBER).build(),
                    AppTableCreateHeader.builder().field_name("价税合计").type(FT.NUMBER).build(),
                    AppTableCreateHeader.builder().field_name("状态").type(FT.SELECT).property(AppTableFieldProperty.builder().options([
                        AppTableFieldPropertyOption.builder().name("正常").build(),
                        AppTableFieldPropertyOption.builder().name("作废").build(),
                        AppTableFieldPropertyOption.builder().name("红冲").build()
                    ]).build()).build(),
                    AppTableCreateHeader.builder().field_name("电子发票文件").type(FT.ATTACHMENT).build(),
                    AppTableCreateHeader.builder().field_name("备注").type(FT.TEXT).build()
                ])
                .build())
            .build()) \
        .build()
    
    resp = client.bitable.v1.app_table.create(req)
    if resp.success():
        log.info("✅ 发票管理表创建成功", extra={"solution": "无"})
        return True, resp.data.table_id
    else:
        log.error(f"❌ 发票管理表创建失败: {resp.msg}", extra={"solution": "检查权限"})
        return False, None

# 创建固定资产表 (新)
@retry_on_failure(max_retries=2, delay=3)
def create_asset_table(client, app_token):
    existing_id = get_table_id_by_name(client, app_token, "固定资产表")
    if existing_id:
        log.info(f"⚠️ 固定资产表已存在 (ID: {existing_id})", extra={"solution": "无"})
        return True, existing_id

    req = CreateAppTableRequest.builder() \
        .app_token(app_token) \
        .request_body(CreateAppTableRequestBody.builder()
            .table(ReqTable.builder()
                .name("固定资产表")
                .fields([
                    AppTableCreateHeader.builder().field_name("资产名称").type(FT.TEXT).build(),
                    AppTableCreateHeader.builder().field_name("类别").type(FT.SELECT).property(AppTableFieldProperty.builder().options([
                        AppTableFieldPropertyOption.builder().name("电子设备").build(),
                        AppTableFieldPropertyOption.builder().name("办公家具").build(),
                        AppTableFieldPropertyOption.builder().name("交通工具").build(),
                        AppTableFieldPropertyOption.builder().name("其他").build()
                    ]).build()).build(),
                    AppTableCreateHeader.builder().field_name("购买日期").type(FT.DATE).build(),
                    AppTableCreateHeader.builder().field_name("原值").type(FT.NUMBER).build(),
                    AppTableCreateHeader.builder().field_name("残值率(%)").type(FT.NUMBER).build(), # New
                    AppTableCreateHeader.builder().field_name("使用年限(年)").type(FT.NUMBER).build(),
                    AppTableCreateHeader.builder().field_name("存放地点").type(FT.TEXT).build(),
                    AppTableCreateHeader.builder().field_name("保管人").type(FT.TEXT).build(),
                    AppTableCreateHeader.builder().field_name("状态").type(FT.SELECT).property(AppTableFieldProperty.builder().options([
                        AppTableFieldPropertyOption.builder().name("使用中").build(),
                        AppTableFieldPropertyOption.builder().name("闲置").build(),
                        AppTableFieldPropertyOption.builder().name("报废").build()
                    ]).build()).build(),
                    AppTableCreateHeader.builder().field_name("备注").type(FT.TEXT).build()
                ])
                .build())
            .build()) \
        .build()
    
    resp = client.bitable.v1.app_table.create(req)
    if resp.success():
        log.info("✅ 固定资产表创建成功", extra={"solution": "无"})
        return True, resp.data.table_id
    else:
        log.error(f"❌ 固定资产表创建失败: {resp.msg}", extra={"solution": "检查权限"})
        return False, None

# 创建薪酬表 (新)
@retry_on_failure(max_retries=2, delay=3)
def create_salary_table(client, app_token):
    existing_id = get_table_id_by_name(client, app_token, "薪酬管理表")
    if existing_id:
        log.info(f"⚠️ 薪酬管理表已存在 (ID: {existing_id})", extra={"solution": "无"})
        return True, existing_id

    req = CreateAppTableRequest.builder() \
        .app_token(app_token) \
        .request_body(CreateAppTableRequestBody.builder()
            .table(ReqTable.builder()
                .name("薪酬管理表")
                .fields([
                    AppTableCreateHeader.builder().field_name("月份").type(FT.TEXT).build(), # YYYY-MM
                    AppTableCreateHeader.builder().field_name("姓名").type(FT.TEXT).build(),
                    AppTableCreateHeader.builder().field_name("基本工资").type(FT.NUMBER).build(),
                    AppTableCreateHeader.builder().field_name("绩效奖金").type(FT.NUMBER).build(),
                    AppTableCreateHeader.builder().field_name("社保扣款(个人)").type(FT.NUMBER).build(),
                    AppTableCreateHeader.builder().field_name("公积金扣款(个人)").type(FT.NUMBER).build(),
                    AppTableCreateHeader.builder().field_name("个税").type(FT.NUMBER).build(),
                    AppTableCreateHeader.builder().field_name("实发工资").type(FT.NUMBER).build(),
                    AppTableCreateHeader.builder().field_name("公司社保承担").type(FT.NUMBER).build(),
                    AppTableCreateHeader.builder().field_name("公司公积金承担").type(FT.NUMBER).build(),
                    AppTableCreateHeader.builder().field_name("发放日期").type(FT.DATE).build(),
                    AppTableCreateHeader.builder().field_name("状态").type(FT.SELECT).property(AppTableFieldProperty.builder().options([
                        AppTableFieldPropertyOption.builder().name("草稿").build(),
                        AppTableFieldPropertyOption.builder().name("已发放").build()
                    ]).build()).build(),
                    AppTableCreateHeader.builder().field_name("备注").type(FT.TEXT).build()
                ])
                .build())
            .build()) \
        .build()
    
    resp = client.bitable.v1.app_table.create(req)
    if resp.success():
        log.info("✅ 薪酬管理表创建成功", extra={"solution": "无"})
        return True, resp.data.table_id
    else:
        log.error(f"❌ 薪酬管理表创建失败: {resp.msg}", extra={"solution": "检查权限"})
        return False, None

# 创建日常台账表
@retry_on_failure(max_retries=2, delay=3)
def create_ledger_table(client, app_token):
    existing_id = get_table_id_by_name(client, app_token, "日常台账表")
    if existing_id:
        log.info(f"⚠️ 日常台账表已存在 (ID: {existing_id})", extra={"solution": "无"})
        return True, existing_id

    req = CreateAppTableRequest.builder() \
        .app_token(app_token) \
        .request_body(CreateAppTableRequestBody.builder()
            .table(ReqTable.builder()
                .name("日常台账表")
                .fields([
                    AppTableCreateHeader.builder().field_name("记账日期").type(FT.DATE).build(),
                    AppTableCreateHeader.builder().field_name("凭证号").type(FT.NUMBER).build(),
                    AppTableCreateHeader.builder().field_name("业务类型").type(FT.SELECT).property(AppTableFieldProperty.builder().options([
                        AppTableFieldPropertyOption.builder().name("收款").build(),
                        AppTableFieldPropertyOption.builder().name("付款").build(),
                        AppTableFieldPropertyOption.builder().name("费用").build()
                    ]).build()).build(),
                    AppTableCreateHeader.builder().field_name("费用归类").type(FT.SELECT).property(AppTableFieldProperty.builder().options([
                        AppTableFieldPropertyOption.builder().name("原材料-三酸/片碱/色粉").build(),
                        AppTableFieldPropertyOption.builder().name("辅料-挂具/除油剂").build(),
                        AppTableFieldPropertyOption.builder().name("外协加工费").build(),
                        AppTableFieldPropertyOption.builder().name("房租水电").build(),
                        AppTableFieldPropertyOption.builder().name("人力成本").build(),
                        AppTableFieldPropertyOption.builder().name("日常费用").build(),
                        AppTableFieldPropertyOption.builder().name("税费").build(),
                        AppTableFieldPropertyOption.builder().name("其他").build()
                    ]).build()).build(),
                    AppTableCreateHeader.builder().field_name("关联项目").type(FT.TEXT).build(),
                    AppTableCreateHeader.builder().field_name("往来单位费用").type(FT.TEXT).build(),
                    AppTableCreateHeader.builder().field_name("账面金额").type(FT.NUMBER).build(),
                    AppTableCreateHeader.builder().field_name("实际收付金额").type(FT.NUMBER).build(),
                    AppTableCreateHeader.builder().field_name("交易银行").type(FT.SELECT).property(AppTableFieldProperty.builder().options([
                         AppTableFieldPropertyOption.builder().name("G银行基本户(有票)").build(),
                         AppTableFieldPropertyOption.builder().name("N银行/微信(无票)").build()
                    ]).build()).build(),
                    AppTableCreateHeader.builder().field_name("是否现金").type(FT.SELECT).property(AppTableFieldProperty.builder().options([ # Use Select for Yes/No
                         AppTableFieldPropertyOption.builder().name("是").build(),
                         AppTableFieldPropertyOption.builder().name("否").build()
                    ]).build()).build(),
                    AppTableCreateHeader.builder().field_name("发票流水单号").type(FT.TEXT).build(),
                    AppTableCreateHeader.builder().field_name("是否有票").type(FT.SELECT).property(AppTableFieldProperty.builder().options([
                         AppTableFieldPropertyOption.builder().name("有票").build(),
                         AppTableFieldPropertyOption.builder().name("无票").build()
                    ]).build()).build(),
                    AppTableCreateHeader.builder().field_name("待补票标记").type(FT.SELECT).property(AppTableFieldProperty.builder().options([
                         AppTableFieldPropertyOption.builder().name("是（大额无票）").build(),
                         AppTableFieldPropertyOption.builder().name("无").build()
                    ]).build()).build(),
                    AppTableCreateHeader.builder().field_name("有票成本").type(FT.NUMBER).build(),
                    AppTableCreateHeader.builder().field_name("无票成本").type(FT.NUMBER).build(),
                    AppTableCreateHeader.builder().field_name("本次实际利润").type(FT.NUMBER).build(),
                    AppTableCreateHeader.builder().field_name("手工式分录").type(FT.TEXT).build(),
                    AppTableCreateHeader.builder().field_name("操作人").type(FT.TEXT).build(),
                    AppTableCreateHeader.builder().field_name("合同订单号").type(FT.TEXT).build(),
                    AppTableCreateHeader.builder().field_name("备注").type(FT.TEXT).build()
                ])
                .build())
            .build()) \
        .build()
    
    resp = client.bitable.v1.app_table.create(req)
    if resp.success():
        log.info("✅ 日常台账表创建成功", extra={"solution": "无"})
        return True, resp.data.table_id
    else:
        log.error(f"❌ 日常台账表创建失败: {resp.msg}", extra={"solution": "检查权限"})
        return False, None

# 填充测试数据
def fill_test_data(client, app_token):
    log.info("🚀 正在填充高质量演示数据...", extra={"solution": "请稍候"})
    
    # 1. 填充往来单位表 (先填充这个，方便后面引用)
    table_id = get_table_id_by_name(client, app_token, "往来单位表")
    if table_id:
        records = []
        sample_partners = [
            {"name": "杭州阿里云计算有限公司", "type": "供应商", "tax_id": "913301066739887754", "remark": "云服务器费用"},
            {"name": "滴滴出行科技有限公司", "type": "供应商", "tax_id": "91120116340983320T", "remark": "员工差旅"},
            {"name": "上海xx贸易有限公司", "type": "客户", "tax_id": "91310115MA1H888888", "remark": "核心大客户"},
            {"name": "京东办公用品", "type": "供应商", "tax_id": "91110105MA00C7XE48", "remark": "办公耗材"},
            {"name": "中国移动通信", "type": "供应商", "tax_id": "911100007109250324", "remark": "电话宽带"}
        ]
        for p in sample_partners:
            fields = {
                "单位名称": p["name"],
                "类型": p["type"],
                "纳税人识别号": p["tax_id"],
                "备注": p["remark"]
            }
            records.append(AppTableRecord.builder().fields(fields).build())
        
        req = BatchCreateAppTableRecordRequest.builder() \
            .app_token(app_token) \
            .table_id(table_id) \
            .request_body(BatchCreateAppTableRecordRequestBody.builder().records(records).build()) \
            .build()
        client.bitable.v1.app_table_record.batch_create(req)
        log.info("✅ 往来单位表：已写入 5 条标准数据")

    # 2. 填充基础信息表
    table_id = get_table_id_by_name(client, app_token, "基础信息表")
    if table_id:
        records = []
        products = [
            {"name": "咨询服务费", "cost": 0},
            {"name": "标准SaaS订阅", "cost": 2000},
            {"name": "高级定制开发", "cost": 15000}
        ]
        for p in products:
            fields = {
                "产品名称": p["name"],
                "单位成本": p["cost"],
                "备注": "演示产品"
            }
            records.append(AppTableRecord.builder().fields(fields).build())
        
        req = BatchCreateAppTableRecordRequest.builder() \
            .app_token(app_token) \
            .table_id(table_id) \
            .request_body(BatchCreateAppTableRecordRequestBody.builder().records(records).build()) \
            .build()
        client.bitable.v1.app_table_record.batch_create(req)
        log.info("✅ 基础信息表：已写入 3 条产品数据")

    # 3. 填充日常台账表 (生成本月真实流水的台账)
    table_id = get_table_id_by_name(client, app_token, "日常台账表")
    if table_id:
        records = []
        now = datetime.now()
        # 生成几笔典型交易
        txs = [
            {"day": 1, "type": "收款", "name": "上海xx贸易有限公司", "amt": 50000.0, "bank": "G银行基本户", "ticket": "有票", "cash": "否", "remark": "Q1服务费首款"},
            {"day": 5, "type": "付款", "name": "杭州阿里云计算有限公司", "amt": 2500.0, "bank": "G银行基本户", "ticket": "有票", "cash": "否", "remark": "1月云资源费"},
            {"day": 8, "type": "费用", "name": "滴滴出行科技有限公司", "amt": 356.5, "bank": "G银行基本户", "ticket": "有票", "cash": "否", "remark": "销售部拜访客户打车"},
            {"day": 10, "type": "费用", "name": "京东办公用品", "amt": 899.0, "bank": "G银行基本户", "ticket": "有票", "cash": "否", "remark": "采购A4纸和墨盒"},
            {"day": 12, "type": "收款", "name": "上海xx贸易有限公司", "amt": 30000.0, "bank": "G银行基本户", "ticket": "有票", "cash": "否", "remark": "Q1服务费尾款"},
            {"day": 15, "type": "费用", "name": "中国移动通信", "amt": 199.0, "bank": "G银行基本户", "ticket": "有票", "cash": "否", "remark": "公司宽带费"},
            {"day": 20, "type": "付款", "name": "房租物业", "amt": 12000.0, "bank": "N银行/微信（现金）", "ticket": "无票", "cash": "是", "remark": "临时仓库租金(房东个人)"},
            {"day": 22, "type": "费用", "name": "顺丰快递", "amt": 56.0, "bank": "N银行/微信（现金）", "ticket": "无票", "cash": "是", "remark": "寄送合同快递费(微信支付)"}
        ]
        
        for tx in txs:
            # 构造日期
            tx_date = datetime(now.year, now.month, tx["day"])
            ts = int(tx_date.timestamp() * 1000)
            
            fields = {
                "记账日期": ts,
                "凭证号": int(tx_date.strftime("%Y%m%d")) + hash(tx["name"]) % 100,
                "业务类型": tx["type"],
                "往来单位费用": tx["name"],
                "账面金额": tx["amt"],
                "实际收付金额": tx["amt"],
                "交易银行": tx["bank"],
                "是否现金": tx["cash"],
                "是否有票": tx["ticket"],
                "待补票标记": "是（大额无票）" if tx["ticket"] == "无票" and tx["amt"] > 1000 else "无",
                "操作人": "自动初始化",
                "备注": tx["remark"]
            }
            records.append(AppTableRecord.builder().fields(fields).build())
        
        req = BatchCreateAppTableRecordRequest.builder() \
            .app_token(app_token) \
            .table_id(table_id) \
            .request_body(BatchCreateAppTableRecordRequestBody.builder().records(records).build()) \
            .build()
        client.bitable.v1.app_table_record.batch_create(req)
        log.info(f"✅ 日常台账表：已写入 {len(records)} 条演示流水")

    # 4. 填充发票管理表
    table_id = get_table_id_by_name(client, app_token, "发票管理表")
    if table_id:
        records = []
        fields = {
            "发票号码": "88888888",
            "发票代码": "031001800104",
            "开票日期": int(datetime.now().timestamp() * 1000),
            "类型": "进项专票",
            "购买方/销售方": "杭州阿里云计算有限公司",
            "价税合计": 2500.0,
            "不含税金额": 2358.49,
            "税额": 141.51,
            "状态": "正常",
            "备注": "对应1月云资源费"
        }
        records.append(AppTableRecord.builder().fields(fields).build())
        
        req = BatchCreateAppTableRecordRequest.builder() \
            .app_token(app_token) \
            .table_id(table_id) \
            .request_body(BatchCreateAppTableRecordRequestBody.builder().records(records).build()) \
            .build()
        client.bitable.v1.app_table_record.batch_create(req)
        log.info("✅ 发票管理表：已写入演示发票")

    # 5. 填充固定资产表
    table_id = get_table_id_by_name(client, app_token, "固定资产表")
    if table_id:
        records = []
        assets = [
            {"name": "MacBook Pro M3", "type": "电子设备", "val": 14999, "user": "老板"},
            {"name": "佳能打印机", "type": "电子设备", "val": 2500, "user": "行政"},
            {"name": "人体工学椅", "type": "办公家具", "val": 800, "user": "全体"}
        ]
        for a in assets:
            fields = {
                "资产名称": a["name"],
                "类别": a["type"],
                "购买日期": int((datetime.now() - timedelta(days=30)).timestamp() * 1000),
                "原值": a["val"],
                "使用年限(年)": 3,
                "存放地点": "公司总部",
                "保管人": a["user"],
                "状态": "使用中",
                "备注": "初始资产"
            }
            records.append(AppTableRecord.builder().fields(fields).build())
        
        req = BatchCreateAppTableRecordRequest.builder() \
            .app_token(app_token) \
            .table_id(table_id) \
            .request_body(BatchCreateAppTableRecordRequestBody.builder().records(records).build()) \
            .build()
        client.bitable.v1.app_table_record.batch_create(req)
        log.info("✅ 固定资产表：已写入演示资产")

# 月度结账
@retry_on_failure(max_retries=2, delay=3)
def monthly_close(client, app_token, ym_input=None):
    log.info("📅 开始月度结账流程...", extra={"solution": "无"})
    
    # 确定结账月份 (默认上个月)
    today = datetime.now()
    last_month_date = today.replace(day=1) - timedelta(days=1)
    default_ym = last_month_date.strftime("%Y%m")
    
    if ym_input:
        ym_str = ym_input
    else:
        print(f"\n{Color.YELLOW}💡 默认结账月份为上个月 ({default_ym}){Color.ENDC}")
        ym_str = input(f"请输入结账月份 (格式 YYYYMM, 直接回车使用默认值): ").strip()
    
    target_year = None
    target_month = None
    
    if not ym_str:
        target_year = last_month_date.year
        target_month = last_month_date.month
    else:
        try:
            if len(ym_str) == 4:
                target_year = int(ym_str)
                target_month = None
                print(f"🎯 选定结账年度: {target_year}年 (生成年度报表)")
            elif len(ym_str) == 6:
                target_year = int(ym_str[:4])
                target_month = int(ym_str[4:])
                print(f"🎯 选定结账月份: {target_year}年{target_month}月")
            else:
                raise ValueError("Length mismatch")
        except:
            print(f"❌ 格式错误，将处理当前年度所有数据")
            target_year = today.year

    # 1. 自动修复缺失分类
    print("\n[1/5] 正在检查并修复缺失分类...")
    auto_fix_missing_categories(client, app_token, target_year)
    
    # 2. 导出备份
    print("\n[2/5] 正在执行全量备份...")
    backup_ok = export_to_excel(client, app_token)
    
    # 3. 生成报表
    print("\n[3/5] 正在生成分析报表...")
    report_ok = generate_html_report(client, app_token, target_year)
    
    if backup_ok and report_ok:
        # 生成Excel利润表
        generate_excel_pnl_report(client, app_token, target_year, target_month)
        
        # 4. 税务测算 (一键结转增强)
        print("\n[4/5] 正在进行税务风险测算及财务体检...")
        calculate_tax(client, app_token, target_year)
        financial_health_check(client, app_token, target_year)

        # 5. 导出标准凭证 (一键结转增强)
        print("\n[5/5] 正在导出标准财务凭证...")
        export_standard_voucher(client, app_token, target_year, target_month)
        
        if target_month:
            msg = f"📅 {target_year}年{target_month}月 月度结账完成！\n✅ 数据已备份\n✅ 报表已生成\n✅ 税务已测算\n✅ 凭证已导出\n💡 请务必将本地生成的 Excel 和 HTML 文件打包存档。"
        else:
            msg = f"🏆 {target_year}年度 年结完成！\n✅ 全年数据已备份\n✅ 年度报表已生成\n✅ 年度税务测算完成\n✅ 全年凭证已导出\n💡 请务必将本地生成的 Excel 和 HTML 文件打包存档。"
            
        log.info("✅ 结账流程结束", extra={"solution": "存档"})
        send_bot_message(msg, "accountant")
        return True
    else:
        log.error("❌ 结账部分失败", extra={"solution": "检查日志"})
        return False

def apply_excel_styles(ws, title_row=1):
    """通用 Excel 样式美化函数"""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # 打印设置
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        # 如果列数较多 (>5)，自动横向
        if ws.max_column > 5:
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToHeight = False # 允许无限长
        ws.page_setup.fitToWidth = 1      # 限制一页宽
        
        # 遍历单元格
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')
                
                # 表头
                if row[0].row == title_row:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                # 数据行
                elif isinstance(cell.value, (int, float)):
                    # 模糊匹配金额列 (列名包含 金额/单价/Cost/Price)
                    col_header = ws.cell(row=title_row, column=cell.column).value
                    if col_header and any(k in str(col_header) for k in ["金额", "单价", "Cost", "Price", "费用", "余额"]):
                        cell.number_format = '#,##0.00'
                        
        # 自动列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column].width = min((max_length + 2) * 1.2, 50)
    except Exception as e:
        print(f"⚠️ 样式应用失败: {e}")

def year_end_closing(client, app_token):
    """一键年结：调用月度结账逻辑，但锁定为年度模式"""
    print(f"\n{Color.HEADER}📅 启动年结流程 (Year-End Closing)...{Color.ENDC}")
    print(f"{Color.CYAN}此功能将生成全年的财务报表、税务测算及凭证导出。{Color.ENDC}")
    
    last_year = datetime.now().year - 1
    year_str = input(f"请输入结账年度 (默认: {last_year}): ").strip()
    if not year_str:
        year_str = str(last_year)
        
    # 再次确认
    confirm = input(f"❓ 确认对 {year_str} 年度进行年结吗? (y/n) [y]: ").strip().lower()
    if confirm == 'n':
        print("已取消。")
        return

    # 新增：年结前建议进行折旧计提
    print("-" * 30)
    dep_confirm = input(f"📉 是否先进行 {year_str}年12月 的固定资产折旧计提 (通常作为年度最后调整)? (y/n) [y]: ").strip().lower()
    if dep_confirm != 'n':
        calculate_depreciation(client, app_token, auto_run=True, target_year=int(year_str), target_month=12)

    # 复用 monthly_close 逻辑，它已经包含了年度处理的所有分支
    monthly_close(client, app_token, ym_input=year_str)

# 生成Excel利润表
def generate_excel_pnl_report(client, app_token, target_year=None, target_month=None):
    if target_year and target_month:
        log.info(f"📊 正在生成 {target_year}年{target_month}月 利润表(Excel)...", extra={"solution": "无"})
        filename_prefix = f"利润表_{target_year}{target_month:02d}"
    elif target_year:
        log.info(f"📊 正在生成 {target_year}年度 利润表(Excel)...", extra={"solution": "无"})
        filename_prefix = f"利润表_{target_year}"
    else:
        log.info("📊 正在生成标准利润表(Excel)...", extra={"solution": "无"})
        filename_prefix = f"利润表_{datetime.now().strftime('%Y%m')}"

    table_id = get_table_id_by_name(client, app_token, "日常台账表")
    if not table_id:
        return False

    # 构建过滤条件
    filter_str = None
    if target_year:
        try:
            if target_month:
                start_dt = datetime(target_year, target_month, 1)
                if target_month == 12:
                    end_dt = datetime(target_year + 1, 1, 1)
                else:
                    end_dt = datetime(target_year, target_month + 1, 1)
            else:
                start_dt = datetime(target_year, 1, 1)
                end_dt = datetime(target_year + 1, 1, 1)
            
            start_ts = int(start_dt.timestamp() * 1000)
            end_ts = int(end_dt.timestamp() * 1000)
            filter_str = f'AND(CurrentValue.[记账日期]>={start_ts}, CurrentValue.[记账日期]<{end_ts})'
        except Exception as e:
            log.error(f"日期计算错误: {e}")
            return False

    records = get_all_records(client, app_token, table_id, filter_info=filter_str)
    if not records:
        log.warning("⚠️ 该期间无数据，跳过生成利润表")
        return False
        
    data = []
    for r in records:
        fields = r.fields
        data.append({
            "记账日期": datetime.fromtimestamp(fields.get("记账日期", 0)/1000).strftime('%Y-%m-%d') if fields.get("记账日期") else "",
            "业务类型": fields.get("业务类型", ""),
            "往来单位费用": fields.get("往来单位费用", ""),
            "费用归类": fields.get("费用归类", "其他"),
            "实际收付金额": float(fields.get("实际收付金额", 0)),
            "是否有票": fields.get("是否有票", "无票")
        })
        
    df = pd.DataFrame(data)
    
    # 简单的利润表逻辑
    income = df[df["业务类型"] == "收款"]["实际收付金额"].sum()
    cost = df[df["业务类型"].isin(["付款", "费用"])]["实际收付金额"].sum()
    gross_profit = income - cost
    
    # 按费用分类汇总 (往来单位)
    partner_summary = pd.DataFrame()
    if not df[df["业务类型"].isin(["付款", "费用"])].empty:
        partner_summary = df[df["业务类型"].isin(["付款", "费用"])].groupby("往来单位费用")["实际收付金额"].sum().reset_index()
        partner_summary.columns = ["往来单位", "金额"]
        partner_summary = partner_summary.sort_values(by="金额", ascending=False)
    
    # 按费用分类汇总 (费用归类)
    category_summary = pd.DataFrame()
    if not df[df["业务类型"].isin(["付款", "费用"])].empty:
        category_summary = df[df["业务类型"].isin(["付款", "费用"])].groupby("费用归类")["实际收付金额"].sum().reset_index()
        category_summary.columns = ["费用科目", "金额"]
        category_summary = category_summary.sort_values(by="金额", ascending=False)
    
    # 月度趋势 (仅在年度报表时生成)
    monthly_trend = pd.DataFrame()
    if not target_month and not df.empty:
        try:
            # Extract month from date
            df['Month'] = df['记账日期'].apply(lambda x: x[:7] if x else '') # YYYY-MM
            expense_df = df[df["业务类型"].isin(["付款", "费用"])]
            if not expense_df.empty:
                monthly_trend = expense_df.pivot_table(
                    index='费用归类', 
                    columns='Month', 
                    values='实际收付金额', 
                    aggfunc='sum', 
                    fill_value=0
                )
                
                # 按总金额排序 (降序)
                # 计算每行的总和
                monthly_trend['Total'] = monthly_trend.sum(axis=1)
                # 排序
                monthly_trend = monthly_trend.sort_values(by='Total', ascending=False)
                # 移除 Total 列，避免写入 Excel 时重复
                monthly_trend = monthly_trend.drop(columns=['Total'])
                
                monthly_trend = monthly_trend.reset_index()
        except Exception as e:
            log.warning(f"生成月度趋势失败: {e}")

    # 写入Excel
    filename = f"{filename_prefix}.xlsx"
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # 概览页
        summary_data = [
            ["项目", "金额", "备注"],
            ["营业收入", income, "所有收款汇总"],
            ["营业成本/费用", cost, "所有付款/费用汇总"],
            ["净利润", gross_profit, "收入 - 成本"],
            ["", "", ""],
            ["利润率", f"{(gross_profit/income*100):.2f}%" if income > 0 else "0%", ""]
        ]
        pd.DataFrame(summary_data).to_excel(writer, sheet_name="利润表概览", index=False, header=False)
        
        # 费用明细页 (按科目)
        category_summary.to_excel(writer, sheet_name="费用明细(按科目)", index=False)
        
        # 费用明细页 (按单位)
        partner_summary.to_excel(writer, sheet_name="费用明细(按单位)", index=False)
        
        # 年度趋势页
        if not monthly_trend.empty:
             monthly_trend.to_excel(writer, sheet_name="年度费用趋势", index=False)

        # 原始数据页
        df.to_excel(writer, sheet_name="流水底稿", index=False)
        
    log.info(f"✅ 利润表已生成: {filename}", extra={"solution": "无"})
    return True

def setup_auto_task():
    """设置 Windows 计划任务 (每天18:00自动运行)"""
    print(f"\n⏰ 正在配置每日自动运行任务...")
    
    # 1. 创建自动运行脚本
    cwd = os.getcwd()
    python_exe = sys.executable
    script_path = os.path.join(cwd, "CW.py")
    
    bat_content = f"""@echo off
cd /d "{cwd}"
"{python_exe}" "{script_path}" --auto-run >> auto_run.log 2>&1
"""
    bat_file = os.path.join(cwd, "run_daily_task.bat")
    with open(bat_file, "w", encoding="utf-8") as f:
        f.write(bat_content)
        
    print(f"✅ 已创建执行脚本: {bat_file}")
    
    # 2. 调用 schtasks 创建任务
    # 任务名: FeishuFinanceAuto
    # 触发: 每天 18:00
    cmd = f'schtasks /create /sc daily /tn "FeishuFinanceAuto" /tr "{bat_file}" /st 18:00 /f'
    
    print(f"👉 正在向 Windows 注册任务 (可能需要管理员权限)...")
    print(f"   执行命令: {cmd}")
    
    try:
        ret = os.system(cmd)
        if ret == 0:
            print(f"{Color.GREEN}✅ 成功！系统将在每天 18:00 自动帮您处理账务。{Color.ENDC}")
            print("   (记得电脑那时候要开机哦)")
        else:
            print(f"{Color.FAIL}❌ 创建失败。请尝试以管理员身份运行 run.bat 再试一次。{Color.ENDC}")
    except Exception as e:
        print(f"❌ 发生错误: {e}")

def restore_from_backup():
    """从备份恢复数据"""
    backup_root = "backup"
    if not os.path.exists(backup_root):
        print("❌ 没有找到备份文件夹 (backup)")
        return
        
    # 列出备份目录
    dirs = [d for d in os.listdir(backup_root) if os.path.isdir(os.path.join(backup_root, d))]
    dirs.sort(reverse=True) # 最新在通过
    
    if not dirs:
        print("❌ 没有找到任何备份记录")
        return
        
    print(f"\n💾 [系统恢复] 请选择要恢复的备份点:")
    for i, d in enumerate(dirs[:10]):
        print(f"   {i+1}. {d}")
        
    choice = input("\n👉 请输入序号 (慎重! 会覆盖当前配置): ").strip()
    if not choice.isdigit(): return
    
    idx = int(choice) - 1
    if 0 <= idx < len(dirs):
        target = dirs[idx]
        src_path = os.path.join(backup_root, target)
        
        print(f"{Color.WARNING}⚠️  警告: 将从 {target} 恢复文件。当前目录下的同名文件将被覆盖！{Color.ENDC}")
        if input("确认继续吗? (输入 'yes' 确认): ").strip().lower() != "yes":
            return
            
        print("⏳ 正在恢复...")
        import shutil
        
        # 恢复文件
        for f in os.listdir(src_path):
            full_src = os.path.join(src_path, f)
            if os.path.isfile(full_src) and not f.endswith(".zip"):
                try:
                    shutil.copy(full_src, ".")
                    print(f"   - 已恢复: {f}")
                except Exception as e:
                    print(f"   ❌ 失败: {f} ({e})")
                    
        print(f"{Color.GREEN}✅ 恢复完成！请重启程序。{Color.ENDC}")
        sys.exit(0)

# 每日简报 (老板看板)
@retry_on_failure(max_retries=2, delay=3)
def draw_ascii_bar_chart(data, title="统计图表"):
    """在终端绘制简单的ASCII柱状图"""
    if not data: return
    
    print(f"\n📊 {title}")
    print("-" * 50)
    
    # 过滤掉 0 值，除非全是 0
    valid_data = {k: v for k, v in data.items() if v > 0}
    if not valid_data and any(v > 0 for v in data.values()):
        pass # 只有部分是0
    elif not valid_data:
        valid_data = data # 全是0
        
    if not valid_data:
        print("   (暂无数据)")
        print("-" * 50)
        return

    max_val = max(valid_data.values())
    if max_val == 0: max_val = 1
    
    max_len = 25 # 柱子最大长度
    
    for label, value in data.items():
        bar_len = int((value / max_val) * max_len)
        bar = "█" * bar_len
        # 使用 ANSI 颜色让它更好看
        color = Color.GREEN if "收入" in label else (Color.FAIL if "支出" in label else Color.CYAN)
        print(f"{label:<12} | {color}{bar:<25}{Color.ENDC} {value:,.2f}")
    print("-" * 50)

def daily_briefing(client, app_token):
    log.info("📅 正在生成每日经营简报...", extra={"solution": "无"})
    table_id = get_table_id_by_name(client, app_token, "日常台账表")
    if not table_id: return False

    # 获取本月数据
    now = datetime.now()
    start_of_month = datetime(now.year, now.month, 1)
    start_ts = int(start_of_month.timestamp() * 1000)
    filter_info = f'CurrentValue.[记账日期]>={start_ts}'
    
    records = get_all_records(client, app_token, table_id, filter_info=filter_info)
    
    today_str = now.strftime("%Y-%m-%d")
    
    today_income = 0.0
    today_cost = 0.0
    month_income = 0.0
    month_cost = 0.0
    
    today_tx_count = 0
    latest_txs = []
    
    for r in records:
        fields = r.fields
        # 日期处理
        ts = fields.get("记账日期", 0)
        try:
            r_date = datetime.fromtimestamp(ts/1000)
            r_date_str = r_date.strftime("%Y-%m-%d")
        except:
            continue
            
        amt = float(fields.get("实际收付金额", 0))
        biz_type = fields.get("业务类型", "")
        desc = fields.get("往来单位费用", "未知")
        
        # 本月累计
        if biz_type == "收款":
            month_income += amt
        elif biz_type in ["付款", "费用"]:
            month_cost += amt
            
        # 今日统计
        if r_date_str == today_str:
            today_tx_count += 1
            if biz_type == "收款":
                today_income += amt
            elif biz_type in ["付款", "费用"]:
                today_cost += amt
                
            # 收集今日明细
            latest_txs.append(f"{biz_type}: {desc} ({amt:,.2f})")

    # 构造飞书卡片
    net_cash = month_income - month_cost
    
    # [V9.7] 利润率分析
    profit_margin = 0.0
    if month_income > 0:
        profit_margin = (net_cash / month_income) * 100
        
    margin_color = Color.OKGREEN if profit_margin >= 10 else (Color.WARNING if profit_margin >= 0 else Color.FAIL)
    margin_str = f"{margin_color}{profit_margin:+.1f}%{Color.ENDC}"

    # [新增] 终端显示 ASCII 图表
    chart_data = {
        "今日收入": today_income,
        "今日支出": today_cost,
        "本月收入": month_income,
        "本月支出": month_cost,
        "本月净利": net_cash  # Add Net Profit to chart
    }
    
    # [V9.4] 简单的趋势预测
    days_passed = now.day
    pred_msg = ""
    if days_passed >= 3: # 至少3天才预测
        avg_cost = month_cost / days_passed
        pred_total_cost = avg_cost * 30
        chart_data[f"预测月底支出"] = pred_total_cost
        pred_msg = f" (按当前趋势，月底预计支出: {pred_total_cost:,.0f})"
        
    draw_ascii_bar_chart(chart_data, title=f"今日经营简报 (利润率: {margin_str}){pred_msg}")
    
    if latest_txs:
        print(f"\n📝 今日明细 ({today_tx_count}笔):")
        for t in latest_txs[:5]:
            print(f"  - {t}")
        if len(latest_txs) > 5:
            print(f"  ... (还有 {len(latest_txs)-5} 笔)")
    else:
        print("\n💤 今日暂无收支记录")

    # [V9.7] 保存到仪表盘缓存
    try:
        cache_data = {
            "updated_at": now.strftime("%Y-%m-%d %H:%M"),
            "month": now.strftime("%Y-%m"),
            "income": month_income,
            "expense": month_cost,
            "net": net_cash
        }
        with open(FILE_DASHBOARD_CACHE, "w", encoding="utf-8") as f:
            json.dump(cache_data, f, ensure_ascii=False)
    except: pass

    # [V9.4] 检查待处理单据
    watch_dir = PENDING_DIR
    if os.path.exists(watch_dir):
        pending_files = [f for f in os.listdir(watch_dir) if f.lower().endswith(('.xlsx', '.xls'))]
        if pending_files:
            print(f"\n🔔 【待办提醒】 发现 {len(pending_files)} 个待处理文件在 '{watch_dir}'")
            for pf in pending_files[:3]:
                print(f"   - {pf}")
            if len(pending_files) > 3:
                print(f"   ... 等 {len(pending_files)} 个文件")
            print("   💡 建议运行菜单 [20] 启动自动监听，或手动导入")
    
    elements = [
        {
            "tag": "div",
            "fields": [
                {"is_short": True, "text": {"tag": "lark_md", "content": f"**今日收入**\n¥ {today_income:,.2f}"}},
                {"is_short": True, "text": {"tag": "lark_md", "content": f"**今日支出**\n¥ {today_cost:,.2f}"}},
                {"is_short": True, "text": {"tag": "lark_md", "content": f"**本月累计收入**\n¥ {month_income:,.2f}"}},
                {"is_short": True, "text": {"tag": "lark_md", "content": f"**本月累计支出**\n¥ {month_cost:,.2f}"}}
            ]
        },
        {"tag": "hr"},
        {
             "tag": "div",
             "text": {"tag": "lark_md", "content": f"💰 **本月净现金流**: ¥ {net_cash:,.2f}"}
        }
    ]
    
    # 如果今天有交易，列出前5笔
    if latest_txs:
        tx_list = "\n".join([f"- {t}" for t in latest_txs[:5]])
        if len(latest_txs) > 5:
            tx_list += f"\n... (还有 {len(latest_txs)-5} 笔)"
            
        elements.append({
            "tag": "div",
            "text": {"tag": "lark_md", "content": f"**📝 今日交易明细 ({today_tx_count}笔)**:\n{tx_list}"}
        })
    else:
        elements.append({
            "tag": "div",
            "text": {"tag": "lark_md", "content": "💤 今日暂无收支记录"}
        })
        
    # AI 点评 (如果有Key)
    if ZHIPUAI_API_KEY and today_tx_count > 0:
        try:
            client_ai = ZhipuAI(api_key=ZHIPUAI_API_KEY)
            prompt = f"今日公司收入{today_income}，支出{today_cost}。请用一句话给老板汇报，语气积极。"
            resp = client_ai.chat.completions.create(model="glm-4-flash", messages=[{"role": "user", "content": prompt}])
            ai_msg = resp.choices[0].message.content.strip()
            print(f"\n🤖 AI汇报: {ai_msg}") # 终端也显示
            elements.append({
                "tag": "note",
                "elements": [{"tag": "plain_text", "content": f"🤖 AI汇报: {ai_msg}"}]
            })
        except:
            pass

    card = {
        "header": {
            "title": {"tag": "plain_text", "content": f"📅 每日经营简报 ({today_str})"},
            "template": "blue"
        },
        "elements": elements
    }
    
    send_bot_message("每日简报", "interactive", card)
    log.info("✅ 每日简报已推送", extra={"solution": "查看飞书"})
    return True

def update_dashboard_cache_silent(client, app_token):
    """静默更新仪表盘缓存 (不发送通知，不打印日志)"""
    try:
        table_id = get_table_id_by_name(client, app_token, "日常台账表")
        if not table_id: return

        # 获取本月数据
        now = datetime.now()
        start_of_month = datetime(now.year, now.month, 1)
        start_ts = int(start_of_month.timestamp() * 1000)
        filter_info = f'CurrentValue.[记账日期]>={start_ts}'
        
        records = get_all_records(client, app_token, table_id, filter_info=filter_info)
        
        month_income = 0.0
        month_cost = 0.0
        
        for r in records:
            fields = r.fields
            amt = float(fields.get("实际收付金额", 0))
            biz_type = fields.get("业务类型", "")
            
            if biz_type == "收款":
                month_income += amt
            elif biz_type in ["付款", "费用"]:
                month_cost += amt
                
        net_cash = month_income - month_cost
        
        cache_data = {
            "updated_at": now.strftime("%Y-%m-%d %H:%M"),
            "month": now.strftime("%Y-%m"),
            "income": month_income,
            "expense": month_cost,
            "net": net_cash
        }
        
        with open(FILE_DASHBOARD_CACHE, "w", encoding="utf-8") as f:
            json.dump(cache_data, f, ensure_ascii=False)
            
    except Exception:
        pass

# 显示数据后台链接
def show_cloud_urls(client, app_token):
    print("\n🌐 飞书云端数据后台 (请复制链接在浏览器打开):")
    print(f"🔗 多维表格主页: https://feishu.cn/base/{app_token}")
    
    tables = ["日常台账表", "往来单位表", "基础信息表", "发票管理表", "固定资产表"]
    for t in tables:
        tid = get_table_id_by_name(client, app_token, t)
        if tid:
            print(f"   📂 {t}: https://feishu.cn/base/{app_token}?table={tid}")
            
    print("\n💡 提示:")
    print("1. 往来单位、产品信息、银行账户等**基础档案**，请直接在网页端修改。")
    print(f"2. 自动分类规则，请修改本地的 {FILE_CATEGORY_RULES} 文件。")
    print("3. 税率、容差等参数，请使用 [8. 系统设置] 修改。")
    
    # 尝试自动打开
    try:
        import webbrowser
        print("\n🚀 正在尝试自动打开浏览器...")
        webbrowser.open(f"https://feishu.cn/base/{app_token}")
    except:
        pass

# -------------------------------------------------------------------------
# 新增功能：AI 查数助手 & 财务体检
# -------------------------------------------------------------------------

def ai_data_query(client, app_token):
    """AI 查数助手：允许用户用自然语言查询财务数据"""
    if not zhipu_client:
        log.error("❌ 未配置 GLM-4 API Key，无法使用 AI 功能", extra={"solution": "请在 .env 文件中配置 ZHIPU_API_KEY"})
        return

    print("\n🤖 AI 财务助手已启动 (输入 'q' 退出)")
    print("你可以问：'上个月餐饮费多少？' 或 '最近一笔大额支出是什么？'")
    
    # 获取最近的数据作为上下文 (为了节省 token，只取最近 50 条)
    table_id = get_table_id_by_name(client, app_token, "日常台账表")
    if not table_id:
        log.error("❌ 找不到日常台账表", extra={"solution": "请先初始化表格"})
        return

    records = get_all_records(client, app_token, table_id)
    # 简单的上下文构建
    context_data = []
    # 排序：最新的在后
    records_sorted = sorted(records, key=lambda r: r.fields.get("记账日期", 0))
    
    for r in records_sorted[-50:]: # 取最后50条
        f = r.fields
        # 转换时间戳
        date_str = "未知日期"
        if isinstance(f.get("记账日期"), int):
            date_str = datetime.fromtimestamp(f["记账日期"] / 1000).strftime("%Y-%m-%d")
        
        context_data.append(f"{date_str} | {f.get('业务类型')} | {f.get('往来单位费用')} | {f.get('实际收付金额')} | {f.get('备注') or ''}")
    
    data_context = "\n".join(context_data)
    
    while True:
        user_input = input("\n🗣️ 请提问: ").strip()
        if user_input.lower() in ['q', 'quit', 'exit']:
            break
            
        if not user_input:
            continue
            
        try:
            log.info("🤔 AI 正在思考...", extra={"solution": "请稍候"})
            response = zhipu_client.chat.completions.create(
                model="glm-4-flash",
                messages=[
                    {"role": "system", "content": f"你是一名专业的财务助理。以下是最近的财务流水数据（格式：日期|类型|对象|金额|备注）：\n\n{data_context}\n\n请根据以上数据回答用户的问题。如果数据中没有答案，请如实告知。金额单位为元。"},
                    {"role": "user", "content": user_input}
                ],
                stream=False
            )
            answer = response.choices[0].message.content
            print(f"\n🤖 AI 回答:\n{answer}")
        except Exception as e:
            log.error(f"AI 响应失败: {e}")

def check_duplicate(client, app_token, table_id, amount, date_str, partner, summary):
    """检查是否存在重复记录 (最近7天，金额相同，摘要相似)"""
    try:
        # 获取最近记录
        records = get_all_records(client, app_token, table_id, field_names=["记账日期", "实际收付金额", "备注", "往来单位费用"])
        
        target_date = datetime.strptime(date_str, "%Y-%m-%d")
        target_amount = float(amount)
        
        for r in records:
            f = r.fields
            try:
                r_date = datetime.fromtimestamp(f.get("记账日期", 0) / 1000)
                r_amount = float(f.get("实际收付金额", 0))
                r_partner = str(f.get("往来单位费用", ""))
                r_summary = str(f.get("备注", ""))
                
                # 规则1: 金额必须完全相同
                if abs(r_amount - target_amount) > 0.01:
                    continue
                    
                # 规则2: 日期在前后3天内
                if abs((r_date - target_date).days) > 3:
                    continue
                    
                # 规则3: 往来单位或摘要高度相似
                if partner and partner in r_partner:
                    return True, f"发现相似记录: {r_date.strftime('%Y-%m-%d')} {r_amount} {r_summary}"
                if summary and summary[:5] in r_summary:
                    return True, f"发现相似记录: {r_date.strftime('%Y-%m-%d')} {r_amount} {r_summary}"
                    
            except:
                continue
                
        return False, ""
    except Exception as e:
        log.warning(f"查重失败: {e}")
        return False, ""

def smart_text_entry(client, app_token):
    """智能文本录入：将自然语言/微信消息解析为记账记录"""
    if not zhipu_client:
        log.error("❌ 未配置 GLM-4 API Key，无法使用 AI 功能", extra={"solution": "请在 .env 文件中配置 ZHIPU_API_KEY"})
        return

    table_id = get_table_id_by_name(client, app_token, "日常台账表")
    if not table_id:
        log.error("❌ 找不到日常台账表", extra={"solution": "请先初始化表格"})
        return

    print("\n📝 智能文本记账助手")
    print("-----------------------------------")
    print("👉 请直接粘贴微信/支付宝的文本，或者老板在群里发的消息。")
    print("例如：")
    print("  - '今天收到张三货款5000元'")
    print("  - '支付阿里云服务器续费 1200元'")
    print("  - '微信支付打车费 56元'")
    print("-----------------------------------")
    
    import json
    
    # 预加载历史知识，用于优化AI结果
    load_history_knowledge(client, app_token)

    while True:
        text = input("\n⌨️ 请输入/粘贴文本 (输入 'q' 退出): ").strip()
        if text.lower() in ['q', 'quit', 'exit']:
            break
        if not text:
            continue
            
        log.info("🧠 AI 正在解析...", extra={"solution": "请稍候"})
        
        try:
            # 1. AI 解析
            prompt = f"""
            你是一个专业的会计助手。请从以下文本中提取财务记账所需的关键信息，并以 JSON 格式返回。
            文本："{text}"
            
            JSON 字段要求：
            - date: 记账日期 (格式 YYYY-MM-DD)。如果文本中没有明确日期，默认为今天({datetime.now().strftime('%Y-%m-%d')})。
            - amount: 金额 (数字，保留2位小数)。
            - type: 业务类型 (只能是 "收款" 或 "付款" 或 "费用")。如果是支出但有票，优先选"费用"；如果是纯支出无票，选"付款"。如果无法判断，默认为"费用"。
            - category: 费用类型/资金账户 (例如：主营业务收入, 办公费, 差旅费, 技术服务费, 预收账款, 预付账款)。请根据内容猜测。
            - partner: 往来单位/对象 (例如：张三, 阿里云, 滴滴)。如果没提到，留空。
            - summary: 备注/摘要 (尽量保留原意，去除无关废话)。
            - is_cash: 是否现金/私户 (true/false)。如果提到"微信"、"支付宝"、"现金"，则为 true，否则 false。
            - has_ticket: 是否有票 (有票/无票)。如果是现金/私户，默认为"无票"，否则"有票"。
            
            只返回纯 JSON 字符串，不要包含 Markdown 格式。
            """
            
            response = zhipu_client.chat.completions.create(
                model="glm-4-flash",
                messages=[
                    {"role": "user", "content": prompt}
                ],
                stream=False
            )
            
            content = response.choices[0].message.content.replace("```json", "").replace("```", "").strip()
            data = json.loads(content)
            
            # 优化1：解析别名
            if data.get('partner'):
                data['partner'] = resolve_partner(data['partner'])

            # 优化2：利用历史知识修正分类
            # 如果 AI 猜的分类不在常用列表中，或者我们有更明确的规则
            history_cat = auto_categorize(data.get('summary'), data.get('category'), partner_name=data.get('partner'))
            if history_cat != data.get('category'):
                log.info(f"💡 根据历史习惯，将 '{data.get('category')}' 修正为 '{history_cat}'")
                data['category'] = history_cat

            # 查重检测
            is_dup, dup_msg = check_duplicate(client, app_token, table_id, data.get('amount'), data.get('date'), data.get('partner'), data.get('summary'))
            if is_dup:
                print(f"\n⚠️  警告: {dup_msg}")
                print("    (可能重复录入！)")
            
            # 3. 交互式确认与修改 (复用逻辑)
            while True:
                print("\n🤖 AI 解析结果 (请核对):")
                print(f"  1. 📅 日期: {data.get('date')}")
                print(f"  2. 💰 金额: {data.get('amount')}")
                print(f"  3. 🏷️ 类型: {data.get('type')}")
                print(f"  4. 📂 分类: {data.get('category')}")
                print(f"  5. 👤 对象: {data.get('partner')}")
                print(f"  6. 📝 摘要: {data.get('summary')}")
                print(f"  7. 🏦 账户: {'现金/私户' if data.get('is_cash') else '对公账户'}")
                print(f"  8. 🧾 发票: {data.get('has_ticket')}")
                
                action = input("\n👉 输入 'y' 确认录入，输入数字(1-8)修改对应项，输入 'n' 取消: ").strip().lower()
                
                if action == 'y':
                    break
                elif action == 'n':
                    print("❌ 已取消")
                    return
                elif action.isdigit():
                    idx = int(action)
                    if idx == 1:
                        val = input(f"请输入新日期 ({data.get('date')}): ").strip()
                        if val: data['date'] = val
                    elif idx == 2:
                        val = input(f"请输入新金额 ({data.get('amount')}): ").strip()
                        if val: data['amount'] = val
                    elif idx == 3:
                        val = input(f"请输入新类型 ({data.get('type')}): ").strip()
                        if val: data['type'] = val
                    elif idx == 4:
                        val = input(f"请输入新分类 ({data.get('category')}): ").strip()
                        if val: data['category'] = val
                    elif idx == 5:
                        val = input(f"请输入新对象 ({data.get('partner')}): ").strip()
                        if val: data['partner'] = val
                    elif idx == 6:
                        val = input(f"请输入新摘要 ({data.get('summary')}): ").strip()
                        if val: data['summary'] = val
                    elif idx == 7:
                        data['is_cash'] = not data.get('is_cash') # 切换
                    elif idx == 8:
                        curr = data.get('has_ticket')
                        data['has_ticket'] = "无票" if curr == "有票" else "有票" # 切换
                else:
                    print("❌ 无效指令")

            # 4. 构造 Record 并上传
            fields = {
                "记账日期": int(datetime.strptime(data.get('date'), "%Y-%m-%d").timestamp() * 1000),
                "业务类型": data.get('type'),
                "费用归类": data.get('category'),
                "往来单位费用": data.get('partner') or "散户",
                "实际收付金额": float(data.get('amount')),
                "备注": data.get('summary'),
                "是否现金": "是" if data.get('is_cash') else "否",
                "是否有票": data.get('has_ticket')
            }
            
            req = CreateAppTableRecordRequest.builder() \
                .app_token(app_token) \
                .table_id(table_id) \
                .app_table_record(AppTableRecord.builder().fields(fields).build()) \
                .build()
            
            resp = client.bitable.v1.app_table_record.create(req)
            if resp.success():
                print("✅ 录入成功！")
                send_bot_message(f"✅ AI 文本录入成功: {data.get('summary')} - {data.get('amount')}元", "accountant")
            else:
                log.error(f"❌ 录入失败: {resp.msg}")

        except Exception as e:
            log.error(f"处理失败: {e}", extra={"solution": "请重试"})

def smart_image_entry(client, app_token, file_path=None, auto_confirm=False):
    """智能截图记账：OCR识别+AI解析"""
    if not zhipu_client:
        log.error("❌ 未配置 GLM-4 API Key", extra={"solution": "请在 .env 文件中配置 ZHIPU_API_KEY"})
        return

    table_id = get_table_id_by_name(client, app_token, "日常台账表")
    if not table_id: return

    # 预加载历史知识
    load_history_knowledge(client, app_token)

    image = None
    if file_path:
        # 自动模式：直接使用传入的文件路径
        try:
            image = Image.open(file_path)
            print(f"✅ 已加载图片: {os.path.basename(file_path)}")
        except Exception as e:
            log.error(f"无法打开图片: {e}")
            return
    else:
        # 交互模式：从剪贴板或对话框获取
        print("\n📸 智能截图记账助手")
        print("-----------------------------------")
        print("👉 请先将【微信/支付宝截图】或【银行回单截图】复制到剪贴板。")
        print("   (或者按回车键选择本地图片文件)")
        print("-----------------------------------")
        
        input("📋 复制好图片后，请按回车继续... (输入 q 退出)")
        
        try:
            # 1. 获取图片
            image = ImageGrab.grabclipboard()
            
            # Windows上抓取的文件列表可能不是Image对象
            if isinstance(image, list):
                 # 用户复制了文件，不是图片内容
                 if len(image) > 0:
                     try:
                         image = Image.open(image[0])
                     except:
                         image = None

            if isinstance(image, Image.Image):
                print("✅ 已从剪贴板获取图片")
            else:
                print("⚠️ 剪贴板中没有图片，请选择文件...")
                # 隐藏主窗口
                root = tk.Tk()
                root.withdraw()
                root.attributes('-topmost', True) # 确保弹窗在最前
                
                file_path_dialog = filedialog.askopenfilename(
                    title="选择图片文件",
                    filetypes=[("Images", "*.png;*.jpg;*.jpeg;*.bmp;*.webp")]
                )
                root.destroy()
                
                if not file_path_dialog:
                    print("❌ 未选择文件")
                    return
                try:
                    image = Image.open(file_path_dialog)
                    print(f"✅ 已加载图片: {os.path.basename(file_path_dialog)}")
                except Exception as e:
                    log.error(f"无法打开图片: {e}")
                    return
        except Exception as e:
            log.error(f"获取图片失败: {e}")
            return

    try:

        # 2. 转 base64
        # 压缩图片以避免超出token限制或传输过慢
        image.thumbnail((1024, 1024)) 
        
        buffered = BytesIO()
        image.save(buffered, format="PNG") # 统一转PNG
        img_base64 = base64.b64encode(buffered.getvalue()).decode('utf-8')
        
        log.info("👀 AI 正在“看”图并提取数据...", extra={"solution": "请稍候"})
        
        response = zhipu_client.chat.completions.create(
            model="glm-4v-flash",
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "text",
                            "text": """
                            请分析这张财务单据/聊天截图，提取记账所需的关键信息，并以 JSON 格式返回。
                            
                            JSON 字段要求：
                            - date: 交易日期 (格式 YYYY-MM-DD)。如果图中没有年份，默认为2026年。如果完全没日期，默认为今天。
                            - amount: 金额 (数字，保留2位小数)。
                            - type: 业务类型 (收款/付款/费用)。
                            - category: 费用类型/资金账户 (例如：主营业务收入, 办公费, 差旅费, 技术服务费, 预收账款, 预付账款)。请根据内容猜测。
                            - partner: 往来单位/对象。
                            - summary: 备注/摘要 (简要描述交易内容)。
                            - is_cash: 是否现金/私户 (true/false)。微信/支付宝/私卡截图通常为 true。
                            - has_ticket: 是否有票 (有票/无票)。截图通常默认为"无票"，除非是发票截图。
                            
                            只返回纯 JSON 字符串，不要包含 Markdown 格式。
                            """
                        },
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": img_base64
                            }
                        }
                    ]
                }
            ]
        )
        
        import json
        content = response.choices[0].message.content.replace("```json", "").replace("```", "").strip()
        data = json.loads(content)
        
        # 优化：解析别名
        if data.get('partner'):
            data['partner'] = resolve_partner(data['partner'])

        # 优化：利用历史知识修正分类
        history_cat = auto_categorize(data.get('summary'), data.get('category'), partner_name=data.get('partner'))
        if history_cat != data.get('category'):
            log.info(f"💡 根据历史习惯，将 '{data.get('category')}' 修正为 '{history_cat}'")
            data['category'] = history_cat

        # 查重检测
        is_dup, dup_msg = check_duplicate(client, app_token, table_id, data.get('amount'), data.get('date'), data.get('partner'), data.get('summary'))
        if is_dup:
            print(f"\n⚠️  警告: {dup_msg}")
            print("    (可能重复录入！)")

        # 3. 交互式确认与修改
        while True:
            print("\n🤖 AI 解析结果 (请核对):")
            print(f"  1. 📅 日期: {data.get('date')}")
            print(f"  2. 💰 金额: {data.get('amount')}")
            print(f"  3. 🏷️ 类型: {data.get('type')}")
            print(f"  4. 📂 分类: {data.get('category')}")
            print(f"  5. 👤 对象: {data.get('partner')}")
            print(f"  6. 📝 摘要: {data.get('summary')}")
            print(f"  7. 🏦 账户: {'现金/私户' if data.get('is_cash') else '对公账户'}")
            print(f"  8. 🧾 发票: {data.get('has_ticket')}")
            
            if auto_confirm:
                print("✅ 自动确认模式：直接录入")
                action = 'y'
            else:
                action = input("\n👉 输入 'y' 确认录入，输入数字(1-8)修改对应项，输入 'n' 取消: ").strip().lower()
            
            if action == 'y':
                break
            elif action == 'n':
                print("❌ 已取消")
                return
            elif action.isdigit():
                idx = int(action)
                if idx == 1:
                    val = input(f"请输入新日期 ({data.get('date')}): ").strip()
                    if val: data['date'] = val
                elif idx == 2:
                    val = input(f"请输入新金额 ({data.get('amount')}): ").strip()
                    if val: data['amount'] = val
                elif idx == 3:
                    val = input(f"请输入新类型 ({data.get('type')}): ").strip()
                    if val: data['type'] = val
                elif idx == 4:
                    val = input(f"请输入新分类 ({data.get('category')}): ").strip()
                    if val: data['category'] = val
                elif idx == 5:
                    val = input(f"请输入新对象 ({data.get('partner')}): ").strip()
                    if val: data['partner'] = val
                elif idx == 6:
                    val = input(f"请输入新摘要 ({data.get('summary')}): ").strip()
                    if val: data['summary'] = val
                elif idx == 7:
                    data['is_cash'] = not data.get('is_cash') # 切换
                elif idx == 8:
                    curr = data.get('has_ticket')
                    data['has_ticket'] = "无票" if curr == "有票" else "有票" # 切换
            else:
                print("❌ 无效指令")

        # 4. 构造 Record 并上传
        fields = {
            "记账日期": int(datetime.strptime(data.get('date'), "%Y-%m-%d").timestamp() * 1000),
            "业务类型": data.get('type'),
            "费用归类": data.get('category'),
            "往来单位费用": data.get('partner') or "散户",
            "实际收付金额": float(data.get('amount')),
            "备注": data.get('summary'),
            "是否现金": "是" if data.get('is_cash') else "否",
            "是否有票": data.get('has_ticket')
        }
        
        req = CreateAppTableRecordRequest.builder() \
            .app_token(app_token) \
            .table_id(table_id) \
            .app_table_record(AppTableRecord.builder().fields(fields).build()) \
            .build()
        
        resp = client.bitable.v1.app_table_record.create(req)
        if resp.success():
            print("✅ 录入成功！")
            send_bot_message(f"✅ AI 截图录入成功: {data.get('summary')} - {data.get('amount')}元", "accountant")
        else:
            log.error(f"❌ 录入失败: {resp.msg}")

    except Exception as e:
        log.error(f"处理失败: {e}", extra={"solution": "请重试"})

def learn_category_rules(client, app_token):
    """智能学习：从历史数据中挖掘分类规则"""
    log.info("🧠 正在分析历史数据以优化分类规则...", extra={"solution": "无"})
    
    table_id = get_table_id_by_name(client, app_token, "日常台账表")
    if not table_id:
        return

    records = get_all_records(client, app_token, table_id)
    
    from collections import Counter
    
    # 现有规则
    global AUTO_CATEGORY_RULES
    existing_rules = AUTO_CATEGORY_RULES
    
    # 候选池
    candidates = []
    
    for r in records:
        f = r.fields
        desc = str(f.get("备注", "")).strip()
        cat = f.get("费用归类", "")
        
        if not desc or not cat:
            continue
            
        # 简单清洗：如果全是数字（如订单号），跳过
        if desc.isdigit(): continue
        
        # 检查是否已被规则覆盖
        is_covered = False
        for k, v in existing_rules.items():
            if k in desc:
                is_covered = True
                break
        
        if not is_covered:
            candidates.append((desc, cat))
            
    # 统计频率
    counts = Counter(candidates)
    
    # 筛选出高频项 (出现 >= 2 次)
    print("\n🔍 发现以下潜在的分类规则 (基于您的历史习惯):")
    print(f"{'关键词 (摘要)':<30} | {'建议分类':<15} | {'出现次数'}")
    print("-" * 60)
    
    index = 1
    suggested_map = {}
    
    for (desc, cat), count in counts.most_common(20): # 只看前20个高频
        if count >= 2:
            # 简单启发式：截取前10个字作为关键词
            keyword = desc[:10]
            print(f"{index}. {keyword:<27} -> {cat:<15} ({count}次)")
            suggested_map[index] = (keyword, cat)
            index += 1
            
    if not suggested_map:
        print("✅ 当前没有发现明显的新规则 (现有规则已覆盖大部分场景)")
        return
        
    print("-" * 60)
    print("👉 输入序号添加规则 (例如 '1,3,5')，输入 'all' 全部添加，直接回车跳过")
    choice = input("您的选择: ").strip().lower()
    
    to_add = []
    if choice == 'all':
        to_add = list(suggested_map.values())
    elif choice:
        try:
            indices = [int(x.strip()) for x in choice.replace("，", ",").split(",") if x.strip()]
            for i in indices:
                if i in suggested_map:
                    to_add.append(suggested_map[i])
        except:
            pass
            
    if to_add:
        # 更新规则文件
        import json
        try:
            with open(FILE_CATEGORY_RULES, "r", encoding="utf-8") as f:
                rules = json.load(f)
        except:
            rules = {}
            
        count = 0
        for k, v in to_add:
            rules[k] = v
            AUTO_CATEGORY_RULES[k] = v # 更新内存
            count += 1
            
        with open(FILE_CATEGORY_RULES, "w", encoding="utf-8") as f:
            json.dump(rules, f, ensure_ascii=False, indent=4)
            
        print(f"✅ 已成功添加 {count} 条新规则！下次记账更智能。")
    else:
        print("未添加任何规则。")

def quick_search_records(client, app_token):
    """快速查账功能"""
    print("\n🔍 快速查账助手")
    print("-----------------------------------")
    keyword = input("👉 请输入关键词 (日期/金额/对象/摘要): ").strip()
    if not keyword: return

    table_id = get_table_id_by_name(client, app_token, "日常台账表")
    if not table_id: return

    print("正在搜索...")
    records = get_all_records(client, app_token, table_id)
    
    found = []
    total_in = 0.0
    total_out = 0.0
    
    for r in records:
        f = r.fields
        # 将所有字段拼接成字符串进行搜索
        date_ts = f.get("记账日期", 0)
        date_str = datetime.fromtimestamp(date_ts/1000).strftime('%Y-%m-%d') if date_ts else ""
        
        full_text = f"{date_str} {f.get('业务类型','')} {f.get('费用归类','')} {f.get('往来单位费用','')} {f.get('实际收付金额','')} {f.get('备注','')} {f.get('合同订单号','')}"
        
        if keyword in full_text:
            found.append(r)
            amt = float(f.get("实际收付金额", 0))
            if f.get("业务类型") == "收款":
                total_in += amt
            elif f.get("业务类型") in ["付款", "费用"]:
                total_out += amt
                
    if not found:
        print(f"❌ 未找到包含 '{keyword}' 的记录")
        return
        
    print(f"\n✅ 找到 {len(found)} 条记录:")
    print("-" * 60)
    print(f"{'日期':<12} | {'类型':<6} | {'对象':<15} | {'金额':<10} | {'摘要'}")
    print("-" * 60)
    
    # 按日期倒序
    found.sort(key=lambda x: x.fields.get("记账日期", 0), reverse=True)
    
    for r in found[:20]: # 最多显示20条
        f = r.fields
        date_ts = f.get("记账日期", 0)
        date_str = datetime.fromtimestamp(date_ts/1000).strftime('%Y-%m-%d') if date_ts else ""
        
        # 截断过长字符串
        partner = str(f.get("往来单位费用", ""))[:14]
        memo = str(f.get("备注", ""))[:20]
        
        print(f"{date_str:<12} | {f.get('业务类型',''):<6} | {partner:<15} | {f.get('实际收付金额',0):<10} | {memo}")
        
    if len(found) > 20:
        print(f"... (还有 {len(found)-20} 条未显示)")
        
    print("-" * 60)
    print(f"💰 统计结果: 收款 {total_in:,.2f} | 支出 {total_out:,.2f} | 净额 {total_in - total_out:,.2f}")
    
    # 导出选项
    if input("\n👉 是否导出搜索结果到 Excel? (y/n): ").strip().lower() == 'y':
        data = []
        for r in found:
            f = r.fields
            f['记账日期'] = datetime.fromtimestamp(f.get('记账日期', 0)/1000).strftime('%Y-%m-%d') if f.get('记账日期') else ""
            data.append(f)
            
        filename = f"查账结果_{keyword}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        pd.DataFrame(data).to_excel(filename, index=False)
        print(f"✅ 已导出: {filename}")
        try: os.startfile(filename)
        except: pass

def manage_partners_flow(client, app_token):
    """往来单位综合管理：主数据(云端) + 别名(本地)"""
    global PARTNER_ALIASES
    
    while True:
        print(f"\n{Color.HEADER}🤝 往来单位管理 (客户/供应商/外发){Color.ENDC}")
        print("---------------------------------------")
        print("1. [云端] 查看往来单位列表 (最新20条)")
        print("2. [云端] 新增往来单位 (单个)")
        print("3. [本地] 管理名称别名 (用于对账)")
        print("0. 返回主菜单")
        
        choice = input(f"{Color.OKBLUE}请选择功能 (0-3): {Color.ENDC}").strip()
        
        if choice == '0':
            break
            
        elif choice == '1':
            table_id = get_table_id_by_name(client, app_token, "往来单位表")
            if not table_id:
                print(f"{Color.FAIL}❌ 往来单位表不存在{Color.ENDC}")
                continue
                
            records = get_all_records(client, app_token, table_id)
            if not records:
                print("📭 暂无往来单位数据")
            else:
                print(f"\n{Color.UNDERLINE}往来单位列表 (共 {len(records)} 个):{Color.ENDC}")
                print(f"{'单位名称':<20} | {'类型':<8} | {'联系人'}")
                print("-" * 50)
                # Show last 20
                for r in records[-20:]:
                    f = r.fields
                    print(f"{f.get('单位名称', '')[:18]:<20} | {f.get('类型', ''):<8} | {f.get('联系人', '')}")
                    
        elif choice == '2':
            print(f"\n{Color.CYAN}➕ 新增往来单位{Color.ENDC}")
            name = input("请输入单位名称: ").strip()
            if not name: continue
            
            p_type = input("请输入类型 (1.客户 2.供应商 3.外发加工 4.其他): ").strip()
            type_map = {'1': '客户', '2': '供应商', '3': '外发加工', '4': '其他'}
            type_str = type_map.get(p_type, '其他')
            
            contact = input("联系人 (选填): ").strip()
            phone = input("联系电话 (选填): ").strip()
            
            table_id = get_table_id_by_name(client, app_token, "往来单位表")
            if not table_id:
                create_partner_table(client, app_token)
                table_id = get_table_id_by_name(client, app_token, "往来单位表")
                
            fields = {
                "单位名称": name,
                "类型": type_str,
                "联系人": contact,
                "联系电话": phone,
                "备注": f"CLI添加 {datetime.now().strftime('%Y-%m-%d')}"
            }
            
            req = BatchCreateAppTableRecordRequest.builder() \
                .app_token(app_token) \
                .table_id(table_id) \
                .request_body(BatchCreateAppTableRecordRequestBody.builder().records([
                    AppTableRecord.builder().fields(fields).build()
                ]).build()) \
                .build()
                
            resp = client.bitable.v1.app_table_record.batch_create(req)
            if resp.success():
                print(f"{Color.OKGREEN}✅ 添加成功: {name} ({type_str}){Color.ENDC}")
            else:
                print(f"{Color.FAIL}❌ 添加失败: {resp.msg}{Color.ENDC}")

        elif choice == '3':
            manage_aliases() # Call existing function

def manage_aliases():
    """管理往来单位别名 (增删改查)"""
    global PARTNER_ALIASES
    
    while True:
        print("\n📇 往来单位别名管理")
        print("-------------------")
        print("1. 查看当前别名")
        print("2. 添加新别名")
        print("3. 删除别名")
        print("4. 批量导入别名 (文本粘贴) [新]")
        print("5. 批量导入别名 (Excel文件) [新]")
        print("0. 返回主菜单")
        print("-------------------")
        
        choice = input("请选择 (0-5): ").strip()
        
        if choice == '0':
            break
            
        elif choice == '1':
            print(f"\n📋 当前映射规则 ({len(PARTNER_ALIASES)}条):")
            print(f"{'别名 (对方户名)':<20} -> {'标准名称 (系统)'}")
            print("-" * 50)
            if not PARTNER_ALIASES:
                print("(暂无别名)")
            else:
                for k, v in PARTNER_ALIASES.items():
                    print(f"{k:<25} -> {v}")
            input("\n按回车继续...")
            
        elif choice == '2':
            print("\n➕ 添加新别名")
            print("例如: 银行流水显示'张三'，实际是'A客户公司'")
            alias = input("请输入别名 (对方户名/微信名): ").strip()
            if not alias: continue
            
            real_name = input(f"请输入 '{alias}' 对应的标准单位名称: ").strip()
            if not real_name: continue
            
            PARTNER_ALIASES[alias] = real_name
            
            # 保存
            try:
                with open(FILE_PARTNER_ALIASES, "w", encoding="utf-8") as f:
                    json.dump(PARTNER_ALIASES, f, ensure_ascii=False, indent=4)
                print(f"✅ 已添加: {alias} -> {real_name}")
            except Exception as e:
                log.error(f"保存失败: {e}")
                
        elif choice == '3':
            alias = input("请输入要删除的别名: ").strip()
            if alias in PARTNER_ALIASES:
                del PARTNER_ALIASES[alias]
                # 保存
                try:
                    with open(FILE_PARTNER_ALIASES, "w", encoding="utf-8") as f:
                        json.dump(PARTNER_ALIASES, f, ensure_ascii=False, indent=4)
                    print(f"✅ 已删除: {alias}")
                except Exception as e:
                    log.error(f"保存失败: {e}")
            else:
                print("❌ 找不到该别名")
                
        elif choice == '4':
            print("\n📥 批量导入别名 (文本粘贴)")
            print("格式：别名 -> 标准名称 (每行一条)")
            print("例如：")
            print("张三 -> A公司")
            print("李四 -> B公司")
            print("-------------------")
            print("请粘贴内容，然后按 Ctrl+Z (Windows) 或 Ctrl+D (Linux/Mac) 结束输入，或者输入 'END' 结束：")
            
            lines = []
            while True:
                try:
                    line = input()
                    if line.strip().upper() == 'END':
                        break
                    lines.append(line)
                except EOFError:
                    break
            
            count = 0
            for line in lines:
                if "->" in line:
                    parts = line.split("->")
                    if len(parts) == 2:
                        alias = parts[0].strip()
                        real = parts[1].strip()
                        if alias and real:
                            PARTNER_ALIASES[alias] = real
                            count += 1
            
            if count > 0:
                # Save
                try:
                    with open(FILE_PARTNER_ALIASES, "w", encoding="utf-8") as f:
                        json.dump(PARTNER_ALIASES, f, ensure_ascii=False, indent=4)
                    print(f"✅ 成功导入 {count} 条别名！")
                except Exception as e:
                    log.error(f"保存失败: {e}")
            else:
                print("⚠️ 未识别到有效数据")

        elif choice == '5':
            print("\n📥 批量导入别名 (Excel文件)")
            print("请准备一个Excel文件，包含两列：【别名】和【标准名称】")
            print("如果没有表头，默认第一列是别名，第二列是标准名称。")
            
            excel_path = select_file_interactively("*.xlsx", "请选择别名映射表")
            if not excel_path:
                print("❌ 未选择文件")
                continue
                
            try:
                df = pd.read_excel(excel_path)
                if df.shape[1] < 2:
                    print("❌ 文件列数不足，至少需要两列")
                    continue
                    
                count = 0
                # 尝试寻找标准列名
                alias_col = None
                real_col = None
                
                for col in df.columns:
                    col_str = str(col)
                    if "别名" in col_str or "户名" in col_str:
                        alias_col = col
                    if "标准" in col_str or "全称" in col_str or "单位" in col_str:
                        real_col = col
                        
                # 如果找不到，就用前两列
                if not alias_col: alias_col = df.columns[0]
                if not real_col: real_col = df.columns[1]
                
                print(f"ℹ️ 使用列: 【{alias_col}】 -> 【{real_col}】")
                
                for _, row in df.iterrows():
                    a = str(row[alias_col]).strip()
                    r = str(row[real_col]).strip()
                    if a and r and a != "nan" and r != "nan":
                        PARTNER_ALIASES[a] = r
                        count += 1
                        
                if count > 0:
                     # Save
                    with open(FILE_PARTNER_ALIASES, "w", encoding="utf-8") as f:
                        json.dump(PARTNER_ALIASES, f, ensure_ascii=False, indent=4)
                    print(f"✅ 成功导入 {count} 条别名！")
                else:
                    print("⚠️ 未找到有效数据")
                    
            except Exception as e:
                log.error(f"导入失败: {e}")

def generate_business_statement(client, app_token):
    """生成往来对账单 (支持加工费+收付款合并对账)"""
    print(f"\n{Color.CYAN}🧾 生成往来对账单 (Statement){Color.ENDC}")
    print("--------------------------------")
    print("功能：合并【加工费明细】(应收/应付) 与 【日常台账】(实收/实付)，生成对账单。")
    
    # 1. 选择客户/供应商
    pf_table_id = get_table_id_by_name(client, app_token, "加工费明细表")
    ledger_table_id = get_table_id_by_name(client, app_token, "日常台账表")
    
    if not pf_table_id or not ledger_table_id: return
    
    print("⏳ 正在获取往来单位列表...")
    # 从加工费表获取最近活跃的单位
    now = datetime.now()
    start_ts_preview = int((now - timedelta(days=60)).timestamp() * 1000)
    filter_preview = f'CurrentValue.[日期]>={start_ts_preview}'
    recs = get_all_records(client, app_token, pf_table_id, filter_info=filter_preview)
    
    partners = set()
    for r in recs:
        p = r.fields.get("往来单位", "").strip()
        if p: partners.add(p)
    
    sorted_partners = sorted(list(partners))
    
    if not sorted_partners:
        print("❌ 无近期往来记录")
        # 允许手动输入
    
    print("\n📋 最近往来单位:")
    for i, p in enumerate(sorted_partners):
        print(f"  {i+1}. {p}")
        
    p_choice = input("\n👉 请选择单位序号 (或直接输入名称): ").strip()
    target_partner = ""
    if p_choice.isdigit() and 1 <= int(p_choice) <= len(sorted_partners):
        target_partner = sorted_partners[int(p_choice)-1]
    else:
        target_partner = p_choice
        
    if not target_partner: return
    
    # 2. 日期范围
    print("\n📅 选择对账期间:")
    start_date_str = input("   起始日期 (YYYY-MM-DD) [默认本月1号]: ").strip()
    end_date_str = input("   结束日期 (YYYY-MM-DD) [默认今天]: ").strip()
    
    if not start_date_str:
        start_date_str = datetime.now().strftime("%Y-%m-01")
    if not end_date_str:
        end_date_str = datetime.now().strftime("%Y-%m-%d")
        
    try:
        s_dt = datetime.strptime(start_date_str, "%Y-%m-%d")
        e_dt = datetime.strptime(end_date_str, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
        start_ts = int(s_dt.timestamp() * 1000)
        end_ts = int(e_dt.timestamp() * 1000)
    except:
        print("❌ 日期格式错误")
        return
        
    print(f"\n🔍 正在拉取【{target_partner}】的全量记录以计算期初余额...")
    
    # 3. 拉取数据 (全量以计算期初)
    # 3.1 加工费 (Charges)
    filter_pf = f'CurrentValue.[往来单位]="{target_partner}"'
    pf_recs = get_all_records(client, app_token, pf_table_id, filter_info=filter_pf)
    
    # 3.2 台账 (Payments)
    filter_lg = f'CurrentValue.[往来单位费用]="{target_partner}"'
    lg_recs = get_all_records(client, app_token, ledger_table_id, filter_info=filter_lg)
    
    # 4. 合并数据 & 计算
    # 定义统一结构: {date, type, desc, qty, unit, price, charge, payment, balance}
    # 规则:
    #   加工费(收入-加工服务) -> Charge (+)
    #   加工费(支出-外协加工) -> Charge (-) [如果是供应商对账，这是"应付"，显示为正向的债务增加?]
    #   
    #   让我们统一逻辑：
    #   【客户对账单】 (We are Seller)
    #      Charge (借方/应收): 增加 (Positive)
    #      Payment (贷方/已收): 减少 (Negative)
    #      Balance > 0: 客户欠我们要付钱
    #
    #   【供应商对账单】 (We are Buyer)
    #      Charge (贷方/应付): 增加 (Positive) [Purchase]
    #      Payment (借方/已付): 减少 (Negative)
    #      Balance > 0: 我们欠供应商要付钱
    
    # 自动判定角色: 根据加工费类型
    # 如果大部分是 "收入-加工服务"，则是客户。
    # 如果大部分是 "支出-外协加工"，则是供应商。
    
    income_count = sum(1 for r in pf_recs if r.fields.get("类型") == "收入-加工服务")
    outcome_count = sum(1 for r in pf_recs if r.fields.get("类型") == "支出-外协加工")
    
    is_supplier = False
    if outcome_count > income_count:
        is_supplier = True
        role_str = "供应商"
    else:
        role_str = "客户"
        
    print(f"ℹ️ 识别为: {role_str} (收入记录: {income_count}, 外协记录: {outcome_count})")
    
    all_txns = []
    
    # 处理加工费
    for r in pf_recs:
        f = r.fields
        d = f.get("日期", 0)
        typ = f.get("类型", "")
        amt = float(f.get("总金额", 0))
        item = f.get("品名", "")
        spec = f.get("规格", "")
        qty = f.get("数量", 0)
        unit = f.get("单位", "")
        price = f.get("单价", 0)
        rem = f.get("备注", "")
        
        charge = 0.0
        payment = 0.0
        
        if not is_supplier: # 客户模式
            if typ == "收入-加工服务":
                charge = amt
            elif typ == "支出-外协加工": 
                # 罕见：客户同时也做外协? 忽略或作为抵扣?
                # 假设作为"应付"，即减少应收 -> payment
                payment = amt 
        else: # 供应商模式
            if typ == "支出-外协加工":
                charge = amt # 应付增加
            elif typ == "收入-加工服务":
                payment = amt # 抵扣?
                
        if abs(charge) < 0.01 and abs(payment) < 0.01: continue
        
        all_txns.append({
            "ts": d,
            "date": datetime.fromtimestamp(d/1000).strftime("%Y-%m-%d"),
            "desc": f"{item} {spec}",
            "qty": qty,
            "unit": unit,
            "price": price,
            "charge": charge,
            "payment": payment,
            "remark": rem,
            "source": "PF"
        })
        
    # 处理台账 (收付款)
    for r in lg_recs:
        f = r.fields
        d = f.get("记账日期", 0)
        b_type = f.get("业务类型", "")
        amt = float(f.get("实际收付金额", 0))
        rem = f.get("备注", "")
        summary = f.get("摘要", "")
        
        charge = 0.0
        payment = 0.0
        
        if not is_supplier: # 客户模式
            if b_type == "收款":
                payment = amt # 客户还款，应收减少
            elif b_type == "付款":
                # 退款给客户?
                charge = amt # 应收增加? 或者作为负的 Payment
                payment = -amt
        else: # 供应商模式
            if b_type == "付款":
                payment = amt # 我们付款，应付减少
            elif b_type == "收款":
                payment = -amt # 退款?
                
        if abs(charge) < 0.01 and abs(payment) < 0.01: continue

        all_txns.append({
            "ts": d,
            "date": datetime.fromtimestamp(d/1000).strftime("%Y-%m-%d"),
            "desc": f"【财务】{summary}",
            "qty": "",
            "unit": "",
            "price": "",
            "charge": charge,
            "payment": payment,
            "remark": rem,
            "source": "LG"
        })
        
    # 按日期排序
    all_txns.sort(key=lambda x: x["ts"])
    
    # 计算期初 & 筛选期间数据
    opening_balance = 0.0
    period_txns = []
    
    for txn in all_txns:
        if txn["ts"] < start_ts:
            opening_balance += (txn["charge"] - txn["payment"])
        elif txn["ts"] <= end_ts:
            # 期间内
            period_txns.append(txn)
            
    # 计算行余额
    running_balance = opening_balance
    total_charge = 0.0
    total_payment = 0.0
    
    html_rows = ""
    
    for txn in period_txns:
        c = txn["charge"]
        p = txn["payment"]
        running_balance += (c - p)
        
        total_charge += c
        total_payment += p
        
        c_str = f"{c:,.2f}" if c != 0 else ""
        p_str = f"{p:,.2f}" if p != 0 else ""
        
        bg = "#fff"
        if txn["source"] == "LG": bg = "#f0f8ff" # 财务记录淡蓝背景
        
        html_rows += f"""
        <tr style="background-color:{bg}">
            <td>{txn['date']}</td>
            <td>{txn['desc']}</td>
            <td style="text-align:right">{txn['qty']}</td>
            <td style="text-align:center">{txn['unit']}</td>
            <td style="text-align:right">{txn['price']}</td>
            <td style="text-align:right; color:#d9534f">{c_str}</td>
            <td style="text-align:right; color:#5cb85c">{p_str}</td>
            <td style="text-align:right; font-weight:bold">{running_balance:,.2f}</td>
            <td style="font-size:12px; color:#666">{txn['remark']}</td>
        </tr>
        """
        
    closing_balance = running_balance
    
    # 5. 生成 HTML
    title_str = "客户对账单" if not is_supplier else "供应商对账单"
    
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <title>{target_partner} - {title_str}</title>
        <style>
            body {{ font-family: 'SimHei', 'Microsoft YaHei', sans-serif; max-width: 900px; margin: 0 auto; padding: 20px; }}
            .header {{ text-align: center; margin-bottom: 20px; border-bottom: 2px solid #000; padding-bottom: 10px; }}
            .title {{ font-size: 24px; font-weight: bold; letter-spacing: 2px; }}
            .info-row {{ display: flex; justify-content: space-between; margin-bottom: 10px; font-size: 14px; }}
            table {{ width: 100%; border-collapse: collapse; margin-bottom: 20px; border: 1px solid #ccc; }}
            th, td {{ border: 1px solid #ccc; padding: 8px; font-size: 13px; }}
            th {{ background-color: #eee; text-align: center; }}
            .summary-box {{ background-color: #f9f9f9; padding: 15px; border: 1px solid #ddd; margin-bottom: 20px; }}
            .footer {{ margin-top: 40px; display: flex; justify-content: space-between; font-size: 14px; }}
            .sign {{ border-top: 1px solid #000; width: 150px; display: inline-block; margin-left: 10px; }}
        </style>
    </head>
    <body>
        <div class="header">
            <div style="font-size:18px; font-weight:bold">五金氧化加工中心</div>
            <div class="title">{title_str}</div>
        </div>
        
        <div class="info-row">
            <div>往来单位: <b>{target_partner}</b></div>
            <div>对账期间: {start_date_str} 至 {end_date_str}</div>
            <div>打印日期: {datetime.now().strftime('%Y-%m-%d')}</div>
        </div>
        
        <div class="summary-box">
            <table style="width:100%; border:none; margin:0;">
                <tr style="background:none;">
                    <td style="border:none"><b>期初余额:</b> {opening_balance:,.2f}</td>
                    <td style="border:none"><b>本期发生(应收/付):</b> {total_charge:,.2f}</td>
                    <td style="border:none"><b>本期已结(实收/付):</b> {total_payment:,.2f}</td>
                    <td style="border:none; font-size:16px"><b>期末应结:</b> <span style="color:red">{closing_balance:,.2f}</span></td>
                </tr>
            </table>
        </div>
        
        <table>
            <thead>
                <tr>
                    <th width="12%">日期</th>
                    <th width="25%">品名/摘要</th>
                    <th width="8%">数量</th>
                    <th width="5%">单位</th>
                    <th width="8%">单价</th>
                    <th width="10%">应收/应付</th>
                    <th width="10%">实收/实付</th>
                    <th width="10%">结余</th>
                    <th width="12%">备注</th>
                </tr>
            </thead>
            <tbody>
                {html_rows}
            </tbody>
        </table>
        
        <div class="footer">
            <div>制单人: 财务部</div>
            <div>
                确认签字: <span class="sign"></span>
                <br><br>日期: ________________
            </div>
        </div>
    </body>
    </html>
    """
    
    save_dir = os.path.join(DATA_ROOT, "往来对账单")
    if not os.path.exists(save_dir): os.makedirs(save_dir)
    
    fname = os.path.join(save_dir, f"对账单_{target_partner}_{start_date_str}_{end_date_str}.html")
    with open(fname, "w", encoding="utf-8") as f:
        f.write(html)
        
    print(f"✅ 对账单已生成: {Color.UNDERLINE}{fname}{Color.ENDC}")
    try: os.startfile(fname)
    except: pass





            


# -------------------------------------------------------------------------
# 新增功能：固定资产折旧
# -------------------------------------------------------------------------

def calculate_depreciation(client, app_token, auto_run=False, target_year=None, target_month=None):
    """一键计提折旧 (生成折旧凭证)"""
    log.info("📉 正在计算固定资产折旧...", extra={"solution": "无"})
    
    asset_table_id = get_table_id_by_name(client, app_token, "固定资产表")
    ledger_table_id = get_table_id_by_name(client, app_token, "日常台账表")
    
    if not asset_table_id or not ledger_table_id:
        log.error("❌ 未找到表格，请先初始化", extra={"solution": "运行 --create-table"})
        return

    # 0. 确定计提月份
    now = datetime.now()
    if target_year and target_month:
        current_month_str = f"{target_year}-{target_month:02d}"
        start_dt = datetime(target_year, target_month, 1)
        if target_month == 12:
            end_dt = datetime(target_year + 1, 1, 1)
        else:
            end_dt = datetime(target_year, target_month + 1, 1)
    else:
        # 默认当前月份
        current_month_str = now.strftime('%Y-%m')
        start_dt = datetime(now.year, now.month, 1)
        if now.month == 12:
            end_dt = datetime(now.year + 1, 1, 1)
        else:
            end_dt = datetime(now.year, now.month + 1, 1)
    
    start_ts = int(start_dt.timestamp() * 1000)
    end_ts = int(end_dt.timestamp() * 1000)
    
    # 使用筛选器查询，避免拉取全部数据
    filter_cmd = f'CurrentValue.[记账日期]>={start_ts}&&CurrentValue.[记账日期]<{end_ts}&&CurrentValue.[费用归类]="折旧摊销"'
    
    # 使用缓存读取 (假设频繁操作)
    existing_deps = get_all_records(client, app_token, ledger_table_id, filter_info=filter_cmd, use_cache=True)
    if existing_deps:
        print(f"{Color.WARNING}⚠️ 检测到本月 ({current_month_str}) 已有 {len(existing_deps)} 条折旧记录！{Color.ENDC}")
        if not auto_run:
            if input("❓ 是否继续计提 (可能导致重复)? (y/n) [n]: ").strip().lower() != 'y':
                return
        else:
            log.info("⚠️ 自动模式下跳过重复计提", extra={"solution": "手动强制执行"})
            return

    # 1. 获取所有使用中的资产 (使用缓存)
    assets = get_all_records(client, app_token, asset_table_id, use_cache=True)
    
    depreciation_entries = []
    total_depreciation = 0.0
    
    if not auto_run:
        print(f"\n📋 资产折旧预览 ({current_month_str}):")
        print("-" * 60)
        print(f"{'资产名称':<20} | {'原值':<10} | {'残值%':<5} | {'月折旧额':<10}")
        print("-" * 60)
    
    for asset in assets:
        f = asset.fields
        status = f.get("状态", "")
        # 优化：如果是补提旧月份，可能资产状态现在是'已报废'，但当时是'使用中'？
        # 暂时只支持对当前'使用中'的资产计提，或者假设资产状态维护得当
        if status != "使用中":
            continue
            
        name = f.get("资产名称", "未知资产")
        original_val = float(f.get("原值", 0))
        years = float(f.get("使用年限(年)", 3)) # 默认3年
        
        if years <= 0: continue
        
        # 简单直线法：月折旧 = 原值 * (1 - 残值率) / (年限 * 12)
        # 尝试获取残值率，默认 0%
        salvage_rate = 0.0
        if "残值率(%)" in f:
            try:
                salvage_rate = float(f["残值率(%)"]) / 100.0
            except:
                salvage_rate = 0.0
        
        monthly_dep = (original_val * (1 - salvage_rate)) / (years * 12)
        monthly_dep = round(monthly_dep, 2)
        
        if monthly_dep > 0:
            if not auto_run:
                print(f"{name:<20} | {original_val:<10.2f} | {salvage_rate*100:<5.0f}% | {monthly_dep:<10.2f}")
            
            # 记账日期设为该月最后一天 (或当前时间)
            # 如果是补提，设为该月最后一天中午12点
            entry_ts = int((end_dt - timedelta(hours=12)).timestamp() * 1000)
            
            depreciation_entries.append({
                "记账日期": entry_ts,
                "业务类型": "费用",
                "费用归类": "折旧摊销", # 自动归类
                "往来单位费用": "内部计提",
                "实际收付金额": monthly_dep, 
                "备注": f"{current_month_str} 折旧计提 - {name}",
                "是否现金": "否", 
                "是否有票": "无票",
                "待补票标记": "无"
            })
            total_depreciation += monthly_dep
            
    if not auto_run:
        print("-" * 60)
        print(f"💰 本月折旧总额: {total_depreciation:.2f}")
    
    if total_depreciation == 0:
        if not auto_run: print("⚠️ 没有需要折旧的资产。")
        return
        
    confirm = 'y'
    if not auto_run:
        confirm = input("\n❓ 确认生成以上折旧凭证吗？(y/n): ").strip().lower()
        
    if confirm == 'y':
        # 批量写入
        batch_size = 100
        for i in range(0, len(depreciation_entries), batch_size):
            batch = depreciation_entries[i:i+batch_size]
            
            # Convert dicts to AppTableRecord
            record_objects = [AppTableRecord.builder().fields(entry).build() for entry in batch]
            
            req = BatchCreateAppTableRecordRequest.builder() \
                .app_token(app_token) \
                .table_id(ledger_table_id) \
                .request_body(BatchCreateAppTableRecordRequestBody.builder().records(record_objects).build()) \
                .build()
            resp = client.bitable.v1.app_table_record.batch_create(req)
            if not resp.success():
                log.error(f"❌ 折旧凭证写入失败: {resp.msg}", extra={"solution": "检查网络"})
            
        print("✅ 折旧凭证已生成！")
        send_bot_message(f"✅ 完成 {current_month_str} 折旧计提，总额: {total_depreciation}元", "accountant")
    else:
        print("❌ 已取消")

def year_end_closing(client, app_token):
    """一键年结：备份 -> 归档 -> 初始化"""
    print(f"\n{Color.HEADER}📅 一键年结向导 (Year End Closing){Color.ENDC}")
    print(f"{Color.WARNING}⚠️  警告：此操作将执行以下流程，不可逆！{Color.ENDC}")
    print("1. 完整备份系统数据 (云端+本地)")
    print("2. 导出全年的标准凭证 Excel")
    print("3. 生成年度财务报表 (HTML)")
    print("4. [可选] 清空云端账目表，准备新的一年 (Reset)")
    print("-" * 50)
    
    confirm = input("👉 请输入 'CONFIRM' 确认执行年结: ").strip()
    if confirm != 'CONFIRM':
        print("❌ 操作已取消")
        return

    year = datetime.now().year
    prev_year = year - 1
    
    # 1. 备份
    print(f"\n{Color.CYAN}Step 1: 系统备份{Color.ENDC}")
    if not backup_system_data(client, app_token):
        print("❌ 备份失败，终止年结")
        return
        
    # 2. 导出凭证
    print(f"\n{Color.CYAN}Step 2: 导出全年凭证{Color.ENDC}")
    # 假设现在是2026年1月，要结2025年的账；或者2026年12月结2026的账
    # 让用户选择年份
    target_year_str = input(f"请输入结账年份 (默认 {prev_year}): ").strip()
    if not target_year_str: target_year = prev_year
    else: target_year = int(target_year_str)
    
    export_standard_voucher(client, app_token, target_year=target_year) 
    
    # 3. 年度报表
    print(f"\n{Color.CYAN}Step 3: 生成年度报表{Color.ENDC}")
    # 需要修改该函数支持年份参数
    generate_annual_report_html(client, app_token, target_year=target_year) 
    
    # 4. 重置 (危险操作)
    print(f"\n{Color.CYAN}Step 4: 数据重置 (可选){Color.ENDC}")
    print("如果您希望清空【日常台账表】以开始新的一年，请选择重置。")
    print("注意：基础信息、固定资产、往来单位表【不会】被清空。")
    if input(f"⚠️ 是否清空 {target_year} 年以前的旧数据? (y/n) [n]: ").strip().lower() == 'y':
        # 这里实现删除逻辑比较复杂，需要遍历删除
        # 为了安全，暂不实现自动删除，只提示
        print("💡 提示: 为数据安全，建议手动在飞书多维表格中新建一个 '202X年账本' 视图，而不是物理删除数据。")
        print("✅ 系统已完成备份和归档，您可以放心地开始新一年的记账了！")
    else:
        print("✅ 数据保持不变。")
        
    print(f"\n{Color.OKGREEN}🎉 年结流程结束！{Color.ENDC}")

def export_standard_voucher(client, app_token, target_year=None, target_month=None):
    """导出标准凭证格式 (对接财务软件)"""
    if target_year and target_month:
        log.info(f"📑 正在生成 {target_year}年{target_month}月 标准凭证导出文件...", extra={"solution": "请稍候"})
        filename_prefix = f"标准凭证导出_{target_year}{target_month:02d}"
    else:
        log.info("📑 正在生成标准凭证导出文件...", extra={"solution": "请稍候"})
        filename_prefix = f"标准凭证导出_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

    table_id = get_table_id_by_name(client, app_token, "日常台账表")
    if not table_id: return
    
    # 构建过滤条件
    filter_str = None
    if target_year:
        try:
            if target_month:
                start_dt = datetime(target_year, target_month, 1)
                if target_month == 12:
                    end_dt = datetime(target_year + 1, 1, 1)
                else:
                    end_dt = datetime(target_year, target_month + 1, 1)
            else:
                start_dt = datetime(target_year, 1, 1)
                end_dt = datetime(target_year + 1, 1, 1)
            
            start_ts = int(start_dt.timestamp() * 1000)
            end_ts = int(end_dt.timestamp() * 1000)
            filter_str = f'AND(CurrentValue.[记账日期]>={start_ts}, CurrentValue.[记账日期]<{end_ts})'
        except Exception as e:
            log.error(f"日期计算错误: {e}")
            return
            
    # Get all records
    print("正在拉取凭证数据...")
    records = get_all_records(client, app_token, table_id, filter_info=filter_str)
    if not records:
        print("⚠️ 无数据")
        return

    # Sort records by date first
    records.sort(key=lambda r: r.fields.get("记账日期", 0))

    voucher_rows = []
    
    # Voucher Numbering
    current_month = ""
    month_seq = 0
    
    for r in records:
        f = r.fields
        date_ts = f.get("记账日期", 0)
        if not date_ts: continue
        
        dt = datetime.fromtimestamp(date_ts/1000)
        date_str = dt.strftime('%Y-%m-%d')
        month_str = dt.strftime('%Y%m')
        
        # Reset sequence for new month
        if month_str != current_month:
            current_month = month_str
            month_seq = 0
        
        month_seq += 1
        voucher_id = f"{month_str}-{month_seq:04d}" # e.g. 202310-0001
        
        amt = float(f.get("实际收付金额", 0))
        if amt == 0: continue
        
        b_type = f.get("业务类型", "")
        summary = f.get("备注", "")
        partner = f.get("往来单位费用", "")
        category = f.get("费用归类", "")

        # 如果备注为空，使用往来单位或费用类型作为摘要
        if not summary:
             summary = f"{partner} {category}".strip()
        
        # 借贷逻辑
        bank_acc = f.get("交易银行", "银行存款")
        
        # 优先使用费用归类作为科目，如果为空则使用往来单位
        # 对于费用类支出，通常科目为费用归类；对于往来款，科目为往来单位
        subject = category if category and category != "其他" and category != "nan" else partner
        if not subject or subject == "nan":
            subject = "暂无分类"
        
        # 简单会计分录逻辑
        if b_type == "收款":
            # 借：银行 (Asset Increase)
            voucher_rows.append({
                "日期": date_str,
                "凭证号": voucher_id,
                "摘要": summary,
                "科目名称": bank_acc, # 借方
                "借方金额": amt,
                "贷方金额": 0
            })
            # 贷：收入/往来 (Revenue/Liability Increase)
            voucher_rows.append({
                "日期": date_str,
                "凭证号": voucher_id,
                "摘要": summary,
                "科目名称": subject, # 贷方
                "借方金额": 0,
                "贷方金额": amt
            })
        elif b_type in ["付款", "费用"]:
            # 借：费用/往来 (Expense Increase)
            voucher_rows.append({
                "日期": date_str,
                "凭证号": voucher_id,
                "摘要": summary,
                "科目名称": subject, # 借方
                "借方金额": amt,
                "贷方金额": 0
            })
            # 贷：银行 (Asset Decrease)
            voucher_rows.append({
                "日期": date_str,
                "凭证号": voucher_id,
                "摘要": summary,
                "科目名称": bank_acc, # 贷方
                "借方金额": 0,
                "贷方金额": amt
            })
            
    if not voucher_rows:
        print("⚠️ 没有生成任何凭证分录")
        return

    df = pd.DataFrame(voucher_rows)
    # 按日期排序
    df = df.sort_values(by="日期")
    
    filename = f"{filename_prefix}.xlsx"
    df.to_excel(filename, index=False)
    
    log.info(f"✅ 导出完成: {filename}", extra={"solution": "可直接导入金蝶/用友等财务软件"})
    try:
        os.startfile(filename)
    except:
        pass

# -------------------------------------------------------------------------
# 新增功能：交互式主菜单 (Python版)
# -------------------------------------------------------------------------

def backup_system_data(client=None, app_token=None):
    """备份系统关键配置和数据"""
    print(f"{Color.CYAN}💾 正在进行系统备份...{Color.ENDC}")
    
    backup_root = "backup"
    if not os.path.exists(backup_root):
        os.makedirs(backup_root)
        
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    target_dir = os.path.join(backup_root, timestamp)
    os.makedirs(target_dir)
    
    # 1. 备份配置文件
    files_to_backup = [
        "partner_aliases.json",
        "category_rules.json",
        ".env",
        "run.bat",
        "使用手册_小白必读.txt"
    ]
    
    for f in files_to_backup:
        if os.path.exists(f):
            try:
                shutil.copy(f, target_dir)
                print(f"  - 已备份: {f}")
            except Exception as e:
                print(f"{Color.FAIL}  - 备份失败 {f}: {e}{Color.ENDC}")
    
    # 2. 备份 Excel 文件 (如果存在)
    excel_files = [f for f in os.listdir('.') if f.endswith('.xlsx') and not f.startswith('~$')]
    for f in excel_files:
        try:
            shutil.copy(f, target_dir)
            print(f"  - 已备份: {f}")
        except:
            pass
            
    # 3. [新增] 备份云端数据 (如果提供了client)
    if client and app_token:
        print("  - 正在导出云端数据...")
        export_to_excel(client, app_token, target_path=target_dir)

    # 4. 压缩备份文件夹
    try:
        shutil.make_archive(target_dir, 'zip', target_dir)
        print(f"📦 已创建压缩包: {target_dir}.zip")
        
        # 5. 清理旧备份 (保留最近 30 天)
        try:
            backup_root = BACKUP_DIR
            now = time.time()
            retention_days = 30
            deleted_count = 0
            
            for f in os.listdir(backup_root):
                f_path = os.path.join(backup_root, f)
                # 检查 zip 文件
                if os.path.isfile(f_path) and f.endswith('.zip'):
                    mtime = os.path.getmtime(f_path)
                    if (now - mtime) > (retention_days * 86400):
                        os.remove(f_path)
                        deleted_count += 1
                # 检查文件夹 (如果之前没压缩或者解压了)
                elif os.path.isdir(f_path):
                     mtime = os.path.getmtime(f_path)
                     if (now - mtime) > (retention_days * 86400):
                        shutil.rmtree(f_path)
                        deleted_count += 1
                        
            if deleted_count > 0:
                print(f"🧹 已自动清理 {deleted_count} 个过期备份 (保留最近{retention_days}天)")
        except Exception as e:
            print(f"⚠️ 清理旧备份失败: {e}")
            
    except Exception as e:
        print(f"⚠️ 压缩失败: {e}")

    print(f"{Color.GREEN}✅ 备份完成！保存路径: {target_dir}{Color.ENDC}")

def move_to_archive(filename):
    """归档文件"""
    target_dir = ARCHIVE_DIR
    if not os.path.exists(target_dir):
        os.makedirs(target_dir)
    try:
        basename = os.path.basename(filename)
        shutil.move(filename, os.path.join(target_dir, basename))
        print(f"   ✅ 文件已归档至 {target_dir}")
    except Exception as e:
        print(f"   ❌ 归档失败: {e}")

def send_notification(title, message):
    """发送 Windows 桌面通知 (使用 PowerShell)"""
    if os.name != 'nt': return
    
    try:
        # PowerShell 脚本: 加载 Windows.Forms 和 Drawing，使用 NotifyIcon
        ps_script = f"""
        [void] [System.Reflection.Assembly]::LoadWithPartialName("System.Windows.Forms")
        [void] [System.Reflection.Assembly]::LoadWithPartialName("System.Drawing")
        $notify = New-Object System.Windows.Forms.NotifyIcon
        $notify.Icon = [System.Drawing.SystemIcons]::Information
        $notify.Visible = $True
        $notify.ShowBalloonTip(0, "{title}", "{message}", [System.Windows.Forms.ToolTipIcon]::Info)
        Start-Sleep -s 3
        $notify.Dispose()
        """
        # 转换命令以避免引号问题
        cmd = ["powershell", "-Command", ps_script]
        # 异步执行，不阻塞
        import subprocess
        subprocess.Popen(cmd, shell=True)
    except Exception as e:
        print(f"⚠️ 通知发送失败: {e}")

def move_to_error(filename, error_msg=""):
    """移动到错误文件夹"""
    target_dir = "待处理单据/处理失败"
    if not os.path.exists(target_dir):
        os.makedirs(target_dir)
    try:
        basename = os.path.basename(filename)
        shutil.move(filename, os.path.join(target_dir, basename))
        print(f"   ❌ 文件已移至 {target_dir}")
        # 写个错误日志
        with open(os.path.join(target_dir, f"{basename}.log"), "w", encoding="utf-8") as f:
            f.write(f"Error Time: {datetime.now()}\n")
            f.write(f"Error: {error_msg}\n")
    except Exception as e:
        print(f"   ❌ 移动失败: {e}")

def monitor_folder_mode(client, app_token):
    """自动监听文件夹模式"""
    watch_dir = PENDING_DIR
    watch_dir_abs = os.path.abspath(watch_dir)
    if not os.path.exists(watch_dir):
        os.makedirs(watch_dir)
        
    print(f"\n{Color.HEADER}📡 已启动【文件夹监听模式】 (挂机中...){Color.ENDC}")
    print(f"📂 监听目录: {watch_dir_abs}")
    print(f"💡 说明: 请将 Excel 银行流水/账单放入该文件夹，系统将自动识别并处理。")
    print(f"🔔 处理结果将通过桌面通知反馈。")
    print(f"🛑 按 Ctrl+C 停止监听并返回菜单。\n")
    
    send_notification("飞书财务助手", "挂机模式已启动，正在监听文件夹...")
    
    print("👀 正在等待新文件...")
    
    try:
        while True:
            # 扫描文件
            if os.path.exists(watch_dir):
                files = [f for f in os.listdir(watch_dir) if f.lower().endswith(('.xlsx', '.xls', '.png', '.jpg', '.jpeg', '.bmp')) and not f.startswith('~$')]
                
                if files:
                    print(f"\n⚡ 发现 {len(files)} 个新文件！开始处理...")
                    for filename in files:
                        full_path = os.path.join(watch_dir, filename)
                        
                        # 等待文件写入完成 (简单等待)
                        time.sleep(2)
                        
                        print(f"▶️ 正在处理: {filename}")
                        try:
                            # 图片处理
                            if filename.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp')):
                                print(f"   📸 识别为图片，启动 AI 记账...")
                                smart_image_entry(client, app_token, file_path=full_path, auto_confirm=True)
                                msg = f"图片 {filename} AI 记账完成！"
                            else:
                                # Excel 处理
                                # 默认作为数据导入
                                is_bank_flow = False
                                if "流水" in filename or "对账" in filename or "bank" in filename.lower():
                                    is_bank_flow = True
                                    
                                if is_bank_flow:
                                    print(f"   🏦 识别为银行流水，启动对账模式...")
                                    reconcile_bank_flow(client, app_token, full_path)
                                    msg = f"银行流水 {filename} 对账完成！"
                                else:
                                    print(f"   📥 识别为业务数据，启动导入模式...")
                                    import_from_excel(client, app_token, excel_path=full_path)
                                    msg = f"业务数据 {filename} 导入成功！"
                                
                            # 归档
                            move_to_archive(full_path)
                            send_notification("处理成功", msg)
                            
                        except Exception as e:
                             print(f"❌ 处理出错: {e}")
                             send_notification("处理失败", f"文件 {filename} 处理出错，已移至失败文件夹。")
                             move_to_error(full_path, str(e))
                             
                    print("\n👀 处理完毕，继续等待新文件...")
                
            time.sleep(5)
    except KeyboardInterrupt:
        print("\n🛑 停止监听。")
        return
        
def auto_fix_missing_categories(client, app_token, target_year=None):
    """自动修复缺失的费用归类"""
    if target_year:
        log.info(f"🔧 正在检查并修复 {target_year}年度 缺失的费用归类...", extra={"solution": "自动修复"})
        year = target_year
    else:
        log.info("🔧 正在检查并修复缺失的费用归类...", extra={"solution": "自动修复"})
        year = datetime.now().year
    
    # [V9.6优化] 确保加载历史知识和AI缓存
    if not HISTORY_CATEGORY_MAP or not AI_CACHE_MAP:
        load_history_knowledge(client, app_token)
        
    table_id = get_table_id_by_name(client, app_token, "日常台账表")
    if not table_id: return
    
    # 获取指定年度数据 (减少处理量)
    start_ts = int(datetime(year, 1, 1).timestamp() * 1000)
    end_ts = int(datetime(year + 1, 1, 1).timestamp() * 1000)
    filter_str = f'AND(CurrentValue.[记账日期]>={start_ts}, CurrentValue.[记账日期]<{end_ts})'
    
    records = get_all_records(client, app_token, table_id, filter_info=filter_str)
    
    updates = []
    
    for r in records:
        f = r.fields
        # 仅处理 费用/付款 类型
        if f.get("业务类型") not in ["费用", "付款"]:
            continue
            
        cat = f.get("费用归类", "")
        desc = f.get("备注", "")
        partner = f.get("往来单位费用", "")
        
        # 如果归类为空 或 为默认值 "其他"
        if not cat or cat in ["", "nan", "其他", "未知"]:
            # 尝试自动分类
            new_cat = auto_categorize(desc, "其他", partner_name=partner)
            
            # 如果自动分类找到了非默认值，且与原值不同
            if new_cat != "其他" and new_cat != cat:
                print(f"   🔧 自动修复: {partner} | {desc} -> {new_cat}")
                updates.append(AppTableRecord.builder().record_id(r.record_id).fields({"费用归类": new_cat}).build())
                
    if updates:
        print(f"   📋 发现 {len(updates)} 条记录待修复，正在批量更新...")
        # [V9.6优化] 批量更新 (Batch Update)
        batch_size = 100
        total_success = 0
        
        for i in range(0, len(updates), batch_size):
            batch = updates[i:i+batch_size]
            try:
                req = BatchUpdateAppTableRecordRequest.builder() \
                    .app_token(app_token) \
                    .table_id(table_id) \
                    .request_body(BatchUpdateAppTableRecordRequestBody.builder().records(batch).build()) \
                    .build()
                resp = client.bitable.v1.app_table_record.batch_update(req)
                if resp.success():
                    total_success += len(batch)
                    print(f"      ✅ 已更新批次 {i//batch_size + 1} ({len(batch)}条)")
                else:
                    log.error(f"❌ 批次更新失败: {resp.msg}")
            except Exception as e:
                log.error(f"❌ 批次更新异常: {e}")
                
        print(f"   ✅ 成功修复 {total_success} 条记录")
    else:
        print("   ✅ 费用归类数据完整，无需修复")

def one_click_daily_closing(client, app_token):
    """一键日结：自动处理单据 -> 计提折旧 -> 税务测算 -> 缺票检查 -> 结账报告 -> 备份"""
    print(f"\n{Color.HEADER}🚀 启动一键日结流程 (Daily Closing)...{Color.ENDC}")
    print(f"{Color.CYAN}💡 提示: 系统将自动处理 '待处理单据' 中的文件并归档{Color.ENDC}")
    
    # 询问是否启用全自动静默模式
    auto_mode = False
    if input("\n👉 是否启用全自动静默处理 (自动确认所有操作)? (y/n) [n]: ").strip().lower() == 'y':
        auto_mode = True
        print(f"{Color.OKGREEN}⚡ 全自动模式已开启，请坐和放宽...{Color.ENDC}")
    
    summary = []
    daily_log = [] # 报告详情
    
    # 1. 扫描当前目录下的 Excel 和 图片 文件
    import glob
    # 修改：扫描 PENDING_DIR 目录
    search_path = PENDING_DIR
    if not os.path.exists(search_path):
        os.makedirs(search_path)
        
    excel_files = []
    image_files = []
    
    # 扫描 PENDING_DIR
    excel_files.extend([os.path.join(search_path, f) for f in os.listdir(search_path) if f.lower().endswith(('.xlsx', '.xls')) and not f.startswith("~$")])
    image_files.extend([os.path.join(search_path, f) for f in os.listdir(search_path) if f.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp'))])
    
    # 兼容根目录（为了方便用户过渡，也扫描根目录，但建议用户用新文件夹）
    root_excels = [f for f in glob.glob("*.xlsx") if not f.startswith("~$") and not f.startswith("待补录") and not f.startswith("往来对账单") and not f.startswith("日结报告")]
    root_images = [f for f in glob.glob("*.*") if f.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp'))]
    
    # 如果根目录有文件，提示用户
    if root_excels or root_images:
        print(f"{Color.WARNING}⚠️  提示：建议将文件放入 '{os.path.basename(PENDING_DIR)}' 文件夹中，系统管理更规范。{Color.ENDC}")
        excel_files.extend(root_excels)
        image_files.extend(root_images)

    all_files = excel_files + image_files
    
    if not all_files:
        print(f"{Color.WARNING}⚠️  当前目录下没有找到待处理文件。{Color.ENDC}")
        summary.append("❌ 未发现新文件")
    else:
        print(f"📂 发现 {len(all_files)} 个待处理文件，开始处理...")
        
        # 总体进度条
        total_files = len(all_files)
        
        for idx, f in enumerate(all_files):
            # show_progress_bar(idx, total_files, prefix='总体进度', suffix=f'处理: {os.path.basename(f)}', length=20)
            print(f"\n📄 [{idx+1}/{total_files}] 正在处理文件: {Color.BOLD}{os.path.basename(f)}{Color.ENDC}")
            
            # 图片处理
            if f.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp')):
                print(f"   📸 识别为图片，建议进行 AI 记账")
                
                do_process = auto_mode
                if not auto_mode:
                    if input("   ❓ 是否处理此图片? (y/n) [y]: ").strip().lower() != 'n':
                        do_process = True
                        
                if do_process:
                    smart_image_entry(client, app_token, file_path=f, auto_confirm=True)
                    summary.append(f"✅ 图片记账: {f}")
                    
                    do_archive = auto_mode
                    if not auto_mode:
                        if input("   ❓ 是否归档? (y/n) [y]: ").strip().lower() != 'n':
                            do_archive = True
                            
                    if do_archive:
                        move_to_archive(f)
                else:
                    summary.append(f"⏩ 跳过图片: {f}")
                continue

            # Excel 处理 (保留原有逻辑)
            # 智能判断建议
            suggestion = "3" # 默认跳过
            f_lower = f.lower()
            if "流水" in f or "对账单" in f or "bank" in f_lower:
                suggestion = "2" # 银行对账
                action_str = "银行对账"
            elif "账单" in f or "import" in f_lower or "数据" in f:
                suggestion = "1" # 数据导入
                action_str = "数据导入"
            else:
                suggestion = "3"
                action_str = "跳过"
            
            print(f"   建议操作: {Color.CYAN}{action_str}{Color.ENDC}")
            
            choice = suggestion
            if not auto_mode:
                print("   1. 作为【业务数据】导入 (Upload)")
                print("   2. 作为【银行流水】对账 (Compare)")
                print("   3. 跳过")
                user_choice = input(f"👉 请选择 (1/2/3) [默认{suggestion}]: ").strip()
                if user_choice: choice = user_choice
            
            if choice == '1':
                import_from_excel(client, app_token, f)
                summary.append(f"✅ 导入: {f}")
                
                do_archive = auto_mode
                if not auto_mode:
                    if input("   ❓ 是否将文件移入 '已处理归档' 文件夹? (y/n) [y]: ").strip().lower() != 'n':
                        do_archive = True
                        
                if do_archive:
                    move_to_archive(f)
            elif choice == '2':
                reconcile_bank_flow(client, app_token, f)
                summary.append(f"✅ 对账: {f}")
                
                do_archive = auto_mode
                if not auto_mode:
                    if input("   ❓ 是否将文件移入 '已处理归档' 文件夹? (y/n) [y]: ").strip().lower() != 'n':
                        do_archive = True
                        
                if do_archive:
                    move_to_archive(f)
            else:
                print("   ⏩ 已跳过")
                summary.append(f"⏩ 跳过: {f}")

    # 1.5 自动计提折旧
    print(f"\n{Color.HEADER}📉 检查固定资产折旧...{Color.ENDC}")
    calculate_depreciation(client, app_token, auto_run=True)

    # 1.6 自动修复缺失分类 (New)
    print(f"\n{Color.HEADER}🔧 检查并修复缺失分类...{Color.ENDC}")
    auto_fix_missing_categories(client, app_token)

    # 2. 税务测算 (New)
    print(f"\n{Color.HEADER}🧮 正在进行税务测算...{Color.ENDC}")
    tax_msg = calculate_tax(client, app_token)
    daily_log.append("\n【税务风险测算】\n" + str(tax_msg))

    # 3. 缺票检查 (New)
    print(f"\n{Color.HEADER}🎫 正在检查待补票据...{Color.ENDC}")
    missing_count = export_missing_tickets(client, app_token, silent=True)
    if missing_count > 0:
        summary.append(f"⚠️ 发现 {missing_count} 笔待补票记录")
        daily_log.append(f"\n【待补票据】\n发现 {missing_count} 笔支出未收发票，请及时催收！")
    else:
        summary.append("✅ 票据状态良好")
        daily_log.append("\n【待补票据】\n目前没有待补票记录，非常棒！")

    # 4. 财务体检
    print(f"\n{Color.HEADER}🏥 开始财务健康体检...{Color.ENDC}")
    financial_health_check(client, app_token)
    
    # 5. 系统备份
    print(f"\n{Color.HEADER}💾 开始系统自动备份...{Color.ENDC}")
    backup_system_data(client, app_token)

    # 6. 发送每日简报
    print(f"\n{Color.HEADER}📢 发送每日经营简报...{Color.ENDC}")
    daily_briefing(client, app_token)
    
    # 6. 生成日结报告 (HTML)
    print(f"\n{Color.HEADER}📊 生成每日结账报告...{Color.ENDC}")
    combined_log = []
    if summary:
        combined_log.append("【处理摘要】")
        combined_log.extend(summary)
    if daily_log:
        combined_log.append("\n【详细日志】")
        combined_log.extend(daily_log)
        
    report_file = generate_daily_html_report(client, app_token, summary_log=combined_log)
    
    if report_file:
        print(f"\n{Color.GREEN}========================================{Color.ENDC}")
        print(f"{Color.GREEN}🎉 日结完成！报告已生成: {report_file}{Color.ENDC}")
        print(f"{Color.GREEN}========================================{Color.ENDC}")
        try:
            os.startfile(report_file)
        except:
            pass
    else:
        log.error("生成报告失败")
    
    print(f"\n{Color.GREEN}✅ 一键流程全部完成！{Color.ENDC}")

# -------------------------------------------------------------------------
# 实用小工具
# -------------------------------------------------------------------------

def parse_date_smart(date_str):
    """
    智能日期解析 (支持自然语言)
    输入: 'zuo', 'qian', '-1', '2.5', '2023.1.1', '今天'
    输出: 'YYYY-MM-DD' 或 None
    """
    date_str = date_str.strip().lower()
    if not date_str: return None
    
    today = datetime.now()
    
    # 快捷指令
    if date_str in ['t', 'j', 'today', 'jin', '今天']:
        return today.strftime("%Y-%m-%d")
    elif date_str in ['y', 'z', 'zuo', 'yesterday', '昨天', '-1']:
        return (today - timedelta(days=1)).strftime("%Y-%m-%d")
    elif date_str in ['by', 'q', 'qian', 'before', '前天', '-2']:
        return (today - timedelta(days=2)).strftime("%Y-%m-%d")
        
    # 简写日期 (如 2.5, 2-5, 2/5 -> 当年-02-05)
    # 正则匹配 M.D 或 M-D 或 M/D
    import re
    match_short = re.match(r'^(\d{1,2})[.\-/](\d{1,2})$', date_str)
    if match_short:
        m, d = int(match_short.group(1)), int(match_short.group(2))
        try:
            # 默认为当年
            dt = datetime(today.year, m, d)
            # 如果是未来日期（比如现在是1月，输入12.5），可能是去年？
            # 暂不自作聪明，按当年算
            return dt.strftime("%Y-%m-%d")
        except: pass
        
    # 标准尝试
    for fmt in ["%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%Y%m%d"]:
        try:
            return datetime.strptime(date_str, fmt).strftime("%Y-%m-%d")
        except: pass
        
    return None

def number_to_chinese(n):
    """
    人民币数字转大写 (简化版，支持万亿级别)
    """
    if not isinstance(n, (int, float)):
        try: n = float(n)
        except: return "输入无效"
        
    if n > 999999999999.99: return "金额过大"
    
    fractions = ['角', '分']
    digit = ['零', '壹', '贰', '叁', '肆', '伍', '陆', '柒', '捌', '玖']
    unit = [['元', '万', '亿'], ['', '拾', '佰', '仟']]
    
    n_str = "{:.2f}".format(n)
    left, right = n_str.split('.')
    
    # 小数部分
    res = []
    jiao = int(right[0])
    fen = int(right[1])
    
    if jiao > 0:
        res.append(digit[jiao] + fractions[0])
    elif fen > 0 and int(left) > 0:
        res.append("零")
        
    if fen > 0:
        res.append(digit[fen] + fractions[1])
            
    if not res:
        suffix = "整"
    else:
        suffix = ""
        
    if int(left) == 0:
        if not res: return "零元整"
        return "".join(res)
        
    # 整数部分
    s = ""
    left_str = str(int(left))
    length = len(left_str)
    
    for i in range(length):
        j = length - i - 1
        d = int(left_str[i])
        
        # 单位索引
        u_idx = j % 4
        # 大单位索引
        b_idx = j // 4
        
        if d != 0:
            s += digit[d] + unit[1][u_idx]
        else:
            # 处理零
            if s and s[-1] != digit[0]:
                s += digit[0]
                
        # 添加大单位
        if u_idx == 0:
            if s and s[-1] == digit[0]: s = s[:-1]
            if b_idx < len(unit[0]):
                s += unit[0][b_idx]
                
    # 修复多余的零
    s = s.replace("零万", "万").replace("零亿", "亿").replace("亿万", "亿").replace("零元", "元")
    
    return s + ("".join(res) or "整")



def draw_dashboard_ui():
    """绘制字符画仪表盘"""
    # 0. 获取数据
    inc, exp, net = 0, 0, 0
    cur_month = datetime.now().strftime("%Y-%m")
    try:
        if os.path.exists(FILE_DASHBOARD_CACHE):
            with open(FILE_DASHBOARD_CACHE, "r", encoding="utf-8") as f:
                data = json.load(f)
                if data.get("month") == cur_month:
                    inc = data.get("income", 0)
                    exp = data.get("expense", 0)
                    net = data.get("net", 0)
    except: pass
    
    # 1. 待处理文件
    watch_dir = PENDING_DIR
    pending_count = 0
    if os.path.exists(watch_dir):
        pending_count = len([f for f in os.listdir(watch_dir) if f.lower().endswith(('.xlsx', '.xls', '.csv', '.jpg', '.png'))])
        
    # 2. 最近备份
    backup_dir = BACKUP_DIR
    last_backup = "无"
    if os.path.exists(backup_dir):
        try:
            items = [os.path.join(backup_dir, d) for d in os.listdir(backup_dir)]
            valid_backups = [f for f in items if os.path.isdir(f) or f.lower().endswith(('.xlsx', '.zip'))]
            if valid_backups:
                latest = max(valid_backups, key=os.path.getmtime)
                last_backup = datetime.fromtimestamp(os.path.getmtime(latest)).strftime("%H:%M")
        except: pass

    # 2.5 库存预警
    inv_alert = ""
    try:
        inv_file = os.path.join(DATA_ROOT, "cache", "inventory_alert.json")
        if os.path.exists(inv_file):
            with open(inv_file, "r") as f:
                alerts = json.load(f)
                if alerts:
                    inv_alert = f"⚠️ 库存告急: {len(alerts)}项"
    except: pass

    # 3. 颜色
    c_inc = Color.GREEN
    c_exp = Color.FAIL
    c_net = Color.OKBLUE if net >= 0 else Color.FAIL
    c_rst = Color.ENDC
    c_bld = Color.BOLD
    
    # 4. 绘制
    lines = []
    lines.append(f"{c_bld}╔════════════════════════════════════════════╗{c_rst}")
    lines.append(f"{c_bld}║ 📊 {cur_month} 财务概览                        ║{c_rst}")
    lines.append(f"{c_bld}╠════════════════════════════════════════════╣{c_rst}")
    
    # Income
    s_inc = f"💰 收入: {inc:,.0f}"
    lines.append(f"║ {c_inc}{s_inc:<39}{c_rst}║")
    
    # Expense
    s_exp = f"💸 支出: {exp:,.0f}"
    lines.append(f"║ {c_exp}{s_exp:<39}{c_rst}║")
    
    # Net
    s_net = f"💴 净额: {net:+,.0f}"
    lines.append(f"║ {c_net}{s_net:<39}{c_rst}║")
    
    lines.append(f"{c_bld}╠════════════════════════════════════════════╣{c_rst}")
    
    # [New] Production Stats
    prod_stats = ""
    try:
        if os.path.exists(FILE_DASHBOARD_CACHE):
            with open(FILE_DASHBOARD_CACHE, "r", encoding="utf-8") as f:
                d = json.load(f)
                p = d.get("production", {})
                if p:
                    kg = p.get("kg", 0)
                    area = p.get("area", 0)
                    cnt = p.get("count", 0)
                    if kg > 0 or area > 0:
                         prod_stats = f"🏭 产量: {int(kg)}kg / {int(area)}m² ({cnt}笔)"
    except: pass
    
    if prod_stats:
        lines.append(f"║ {Color.OKBLUE}{prod_stats:<39}{c_rst}║")
        lines.append(f"{c_bld}╠════════════════════════════════════════════╣{c_rst}")
    
    # Pending & Backup & Alert
    p_color = Color.FAIL if pending_count > 0 else Color.OKGREEN
    s_pend = f"🔔 待办: {pending_count}"
    s_back = f"💾 备份: {last_backup}"
    
    lines.append(f"║ {p_color}{s_pend:<16}{c_rst}    {Color.OKBLUE}{s_back:<16}{c_rst} ║")
    
    if inv_alert:
        lines.append(f"║ {Color.FAIL}{inv_alert:<39}{c_rst}║")
    
    lines.append(f"{c_bld}╚════════════════════════════════════════════╝{c_rst}")
    
    return "\n".join(lines)

def update_inventory_alert_cache(client, app_token):
    """更新库存预警缓存"""
    try:
        t_id = get_table_id_by_name(client, app_token, "库存管理表")
        if not t_id: return
        recs = get_all_records(client, app_token, t_id)
        low_stock = []
        if recs:
            for r in recs:
                c = float(r.fields.get("当前库存", 0))
                s = float(r.fields.get("安全库存", 0))
                if s > 0 and c < s:
                    low_stock.append(r.fields.get("物品名称"))
        
        cache_dir = os.path.join(DATA_ROOT, "cache")
        if not os.path.exists(cache_dir): os.makedirs(cache_dir)
        
        with open(os.path.join(cache_dir, "inventory_alert.json"), "w") as f:
            json.dump(low_stock, f)
            
    except: pass



def generate_monthly_expenses(client, app_token):
    """生成每月固定支出"""
    print(f"\n{Color.UNDERLINE}📅 生成每月固定支出 (Fixed Expenses){Color.ENDC}")
    
    config_path = os.path.join(DATA_ROOT, "monthly_expenses.json")
    
    # 1. Check/Create Config
    if not os.path.exists(config_path):
        sample = [
            {"name": "房租", "amount": 5000, "category": "房租物业", "partner": "房东", "type": "费用"},
            {"name": "宽带费", "amount": 199, "category": "办公费", "partner": "电信", "type": "费用"},
            {"name": "保洁费", "amount": 800, "category": "服务费", "partner": "保洁公司", "type": "费用"}
        ]
        try:
            with open(config_path, "w", encoding="utf-8") as f:
                json.dump(sample, f, ensure_ascii=False, indent=2)
            print(f"✅ 已创建示例配置文件: {config_path}")
            print("👉 请修改该文件后重试。")
        except:
            print("❌ 创建配置文件失败")
        return

    try:
        with open(config_path, "r", encoding="utf-8") as f:
            items = json.load(f)
    except Exception as e:
        print(f"❌ 读取配置文件失败: {e}")
        return

    if not items:
        print("❌ 配置列表为空")
        return

    # 2. Confirm Month
    cur_month = datetime.now().strftime("%Y-%m")
    month_input = input(f"\n请输入入账月份 (YYYY-MM) [{cur_month}]: ").strip()
    if not month_input: month_input = cur_month
    
    try:
        # Check format YYYY-MM
        datetime.strptime(month_input, "%Y-%m")
        target_date = f"{month_input}-01"
        ts = int(datetime.strptime(target_date, "%Y-%m-%d").timestamp() * 1000)
    except:
        print("❌ 日期格式错误")
        return

    # 3. Preview
    print(f"\n即将生成以下 {len(items)} 笔支出 ({target_date}):")
    total_amt = 0
    for i in items:
        print(f"  - {i.get('name')}: {i.get('amount')}元 ({i.get('partner')})")
        total_amt += float(i.get('amount', 0))
    
    print(f"  💰 总金额: {total_amt:,.2f} 元")
    
    if input("\n👉 确认生成? (y/n): ").strip().lower() != 'y': return
    
    # 4. Batch Create
    table_id = get_table_id_by_name(client, app_token, "日常台账表")
    if not table_id: return
    
    records = []
    for item in items:
        fields = {
            "记账日期": ts,
            "业务类型": item.get("type", "费用"),
            "费用归类": item.get("category", "未分类"),
            "往来单位费用": item.get("partner", "散户"),
            "实际收付金额": float(item.get("amount", 0)),
            "备注": f"{month_input} {item.get('name')}",
            "是否有票": item.get("has_invoice", "无票"),
            "是否现金": item.get("is_cash", "否"),
            "操作人": "系统自动"
        }
        records.append(AppTableRecord.builder().fields(fields).build())
        
    # Call batch create
    batch_size = 100
    for i in range(0, len(records), batch_size):
        batch = records[i:i+batch_size]
        req = BatchCreateAppTableRecordRequest.builder() \
            .app_token(app_token) \
            .table_id(table_id) \
            .request_body(BatchCreateAppTableRecordRequestBody.builder().records(batch).build()) \
            .build()
        resp = client.bitable.v1.app_table_record.batch_create(req)
        if not resp.success():
             print(f"❌ 生成失败: {resp.msg}")
        else:
             print(f"✅ 第 {i//batch_size + 1} 批生成成功")

    # Update cache
    try:
        update_dashboard_cache_silent(client, app_token)
        update_inventory_alert_cache(client, app_token)
    except: pass

def reconcile_bank_account(client, app_token):
    """资金账户对账"""
    # 1. Load Initial Balance
    config_file = os.path.join(DATA_ROOT, "config", "capital_account.json")
    if not os.path.exists(os.path.dirname(config_file)):
        os.makedirs(os.path.dirname(config_file), exist_ok=True)
        
    data = {}
    if os.path.exists(config_file):
        try:
            with open(config_file, "r", encoding="utf-8") as f:
                data = json.load(f)
        except: pass
        
    print(f"\n{Color.HEADER}💰 资金账户对账 (Reconciliation){Color.ENDC}")
    print("--------------------------------")
    print("功能说明: 核对【系统账面余额】与【实际资金余额】是否一致。")
    
    init_date = data.get("init_date", "2024-01-01")
    init_balance = data.get("init_balance", 0.0)
    
    print(f"\n⚙️  当前期初设置: {init_date} 余额: {init_balance:,.2f}")
    if input("👉 是否修改期初余额? (y/n) [n]: ").strip().lower() == 'y':
        d = input("请输入期初日期 (YYYY-MM-DD): ").strip()
        if d: init_date = d
        b = input("请输入期初余额: ").strip()
        if b: 
            try: init_balance = float(b)
            except: pass
        
        data["init_date"] = init_date
        data["init_balance"] = init_balance
        try:
            with open(config_file, "w", encoding="utf-8") as f:
                json.dump(data, f)
            print("✅ 设置已保存")
        except Exception as e:
            print(f"❌ 保存失败: {e}")
            
    # Calculate System Balance
    print("\n⏳ 正在计算系统余额 (请稍候)...")
    
    t_id = get_table_id_by_name(client, app_token, "日常台账表")
    if not t_id: return
    
    # 既然是小会计，数据量应该不会特别大，直接全量拉取比较稳
    # 如果有缓存用缓存，但为了对账准确，建议刷新
    recs = get_all_records(client, app_token, t_id)
    
    total_in = 0.0
    total_out = 0.0
    
    try:
        init_ts = int(datetime.strptime(init_date, "%Y-%m-%d").timestamp() * 1000)
    except:
        print("❌ 日期格式错误")
        return
    
    valid_count = 0
    for r in recs:
        d = r.fields.get("记账日期", 0)
        if d < init_ts: continue
        
        t = r.fields.get("业务类型", "")
        amt = float(r.fields.get("实际收付金额", 0))
        
        valid_count += 1
        if t == "收款": total_in += amt
        elif t in ["付款", "费用"]: total_out += amt
        
    sys_balance = init_balance + total_in - total_out
    
    print(f"\n📊 系统账面计算 ({init_date} 至今, 共 {valid_count} 笔):")
    print(f"   ➕ 期间收入: {Color.GREEN}{total_in:,.2f}{Color.ENDC}")
    print(f"   ➖ 期间支出: {Color.FAIL}{total_out:,.2f}{Color.ENDC}")
    print(f"   💰 理论余额: {Color.BOLD}{sys_balance:,.2f}{Color.ENDC}")
    
    # User Input
    print(f"\n{Color.CYAN}💳 请输入您手头所有资金的总和 (银行卡+微信+支付宝+现金){Color.ENDC}")
    real_str = input("👉 实际余额: ").strip()
    if not real_str: return
    
    try:
        real_balance = float(real_str)
        diff = real_balance - sys_balance
        
        print("-" * 30)
        if abs(diff) < 1.0: # 允许1元误差
            print(f"✅ {Color.OKGREEN}完美！账实相符！(差异 {diff:.2f}){Color.ENDC}")
            print("🎉 您今天的账记得很棒！")
        else:
            print(f"❌ {Color.FAIL}对账不平！差异: {diff:,.2f}{Color.ENDC}")
            if diff > 0:
                print(f"   🤔 实际比账面【多】了 {abs(diff):,.2f}")
                print("   可能原因: 1. 有收款忘了记  2. 支出记多了  3. 期初余额偏低")
            else:
                print(f"   🤔 实际比账面【少】了 {abs(diff):,.2f}")
                print("   可能原因: 1. 有支出忘了记  2. 收款记多了  3. 期初余额偏高")
                
            if input("\n🔎 是否需要列出今日收支以供核对? (y/n): ").strip().lower() == 'y':
                # Show today's entries
                today_start = int(datetime(datetime.now().year, datetime.now().month, datetime.now().day).timestamp() * 1000)
                print(f"\n📅 今日 ({datetime.now().strftime('%Y-%m-%d')}) 记录:")
                for r in recs:
                    if r.fields.get("记账日期", 0) >= today_start:
                         print(f"   {r.fields.get('业务类型')} | {r.fields.get('实际收付金额')} | {r.fields.get('往来单位费用')} | {r.fields.get('备注')}")
    except:
        print("❌ 金额无效")
    
    input("\n按回车键返回...")

def create_inventory_table(client, app_token):
    """创建库存管理表 (Consumables Inventory)"""
    table_name = "库存管理表"
    table_id = get_table_id_by_name(client, app_token, table_name)
    
    if table_id: return table_id
    
    print(f"⏳ 正在初始化 {table_name} ...")
    
    req = CreateAppTableRequest.builder() \
        .app_token(app_token) \
        .request_body(CreateAppTableRequestBody.builder()
            .table(AppTable.builder()
                .name(table_name)
                .default_view_name("默认视图")
                .fields([
                    AppTableCreateHeader.builder().field_name("物品名称").type(1).build(), # Text
                    AppTableCreateHeader.builder().field_name("规格型号").type(1).build(), # Text
                    AppTableCreateHeader.builder().field_name("当前库存").type(2).build(), # Number
                    AppTableCreateHeader.builder().field_name("单位").type(1).build(),     # Text
                    AppTableCreateHeader.builder().field_name("安全库存").type(2).build(), # Number (Alert Level)
                    AppTableCreateHeader.builder().field_name("最后变动时间").type(5).build(), # Date
                    AppTableCreateHeader.builder().field_name("备注").type(1).build(),
                ])
                .build())
            .build()) \
        .build()
        
    resp = client.bitable.v1.app_table.create(req)
    if resp.success():
        print(f"✅ {table_name} 创建成功")
        return resp.data.table_id
    else:
        print(f"❌ 创建失败: {resp.msg}")
        return None

def manage_inventory(client, app_token):
    """车间耗材库存管理"""
    table_id = create_inventory_table(client, app_token)
    if not table_id: return
    
    while True:
        print(f"\n{Color.HEADER}📦 车间耗材库存管理 (Inventory){Color.ENDC}")
        print("--------------------------------")
        
        # 1. Show Dashboard (Low Stock)
        recs = get_all_records(client, app_token, table_id)
        low_stock = []
        total_items = 0
        
        if recs:
            total_items = len(recs)
            for r in recs:
                curr = float(r.fields.get("当前库存", 0))
                safe = float(r.fields.get("安全库存", 0))
                if safe > 0 and curr < safe:
                    low_stock.append(r)
        
        if low_stock:
            print(f"{Color.FAIL}⚠️  库存预警: {len(low_stock)} 种物品低于安全库存!{Color.ENDC}")
            for r in low_stock[:3]:
                print(f"   - {r.fields.get('物品名称')} (余 {r.fields.get('当前库存')}{r.fields.get('单位')})")
        else:
            print(f"{Color.OKGREEN}✅ 库存状态良好 (共 {total_items} 种物品){Color.ENDC}")
            
        print("\n1. 📋 查看所有库存")
        print("2. 📥 采购入库 (Stock In)")
        print("3. 📤 领料出库 (Stock Out)")
        print("4. 🔄 库存盘点 (Stock Take)")
        print("0. 返回")
        
        choice = input(f"\n👉 请选择: ").strip()
        
        if choice == '0': break
        
        elif choice == '1':
            print(f"\n📋 库存列表:")
            print(f"{'序号':<4} | {'物品名称':<15} | {'规格':<10} | {'当前库存':<10} | {'状态'}")
            print("-" * 60)
            
            # Sort by name
            recs.sort(key=lambda x: x.fields.get("物品名称", ""))
            
            for i, r in enumerate(recs):
                f = r.fields
                curr = float(f.get("当前库存", 0))
                safe = float(f.get("安全库存", 0))
                unit = f.get("单位", "")
                status = "✅"
                if safe > 0 and curr < safe: status = f"{Color.FAIL}⚠️ 补货{Color.ENDC}"
                
                print(f"{i+1:<4} | {f.get('物品名称'):<15} | {f.get('规格型号',''):<10} | {curr:<6}{unit} | {status}")
            input("\n按回车继续...")
            
        elif choice == '2': # 入库
            print(f"\n{Color.CYAN}📥 采购入库{Color.ENDC}")
            name = input("物品名称 (如 '片碱'): ").strip()
            if not name: continue
            
            # Check existing
            target_rec = None
            for r in recs:
                if r.fields.get("物品名称") == name:
                    target_rec = r
                    break
            
            curr_qty = 0
            unit = "kg"
            
            if target_rec:
                print(f"✅ 找到已有物品: {name} (当前: {target_rec.fields.get('当前库存')})")
                curr_qty = float(target_rec.fields.get("当前库存", 0))
                unit = target_rec.fields.get("单位", "kg")
            else:
                print("🆕 新物品登记")
                unit = input("单位 (默认 kg): ").strip()
                if not unit: unit = "kg"
                
            qty_in = float(input(f"入库数量 ({unit}): ").strip())
            
            new_qty = curr_qty + qty_in
            
            # Save
            fields = {
                "物品名称": name,
                "当前库存": new_qty,
                "最后变动时间": int(datetime.now().timestamp() * 1000),
                "单位": unit
            }
            
            if target_rec:
                # Update
                client.bitable.v1.app_table_record.update(
                    UpdateAppTableRecordRequest.builder().app_token(app_token).table_id(table_id).record_id(target_rec.record_id)
                    .request_body(AppTableRecord.builder().fields(fields).build()).build()
                )
            else:
                # Create
                # Ask for safety stock
                s_stock = input("设置安全库存 (默认 0): ").strip()
                if s_stock: fields["安全库存"] = float(s_stock)
                
                client.bitable.v1.app_table_record.create(
                    CreateAppTableRecordRequest.builder().app_token(app_token).table_id(table_id)
                    .request_body(AppTableRecord.builder().fields(fields).build()).build()
                )
            
            print(f"✅ 入库完成！当前库存: {new_qty} {unit}")
            
            # Link to Expense
            if input("💰 是否同时记录一笔【采购支出】? (y/n) [y]: ").strip().lower() != 'n':
                amt = float(input("采购金额 (元): ").strip())
                remark = f"采购 {name} {qty_in}{unit}"
                
                # Call register logic directly
                l_id = get_table_id_by_name(client, app_token, "日常台账表")
                if l_id:
                     ef = {
                         "记账日期": int(datetime.now().timestamp() * 1000),
                         "业务类型": "费用",
                         "费用归类": "原材料-三酸/片碱/色粉", # Default
                         "实际收付金额": amt,
                         "备注": remark,
                         "往来单位费用": "供应商"
                     }
                     client.bitable.v1.app_table_record.create(
                        CreateAppTableRecordRequest.builder().app_token(app_token).table_id(l_id)
                        .request_body(AppTableRecord.builder().fields(ef).build()).build()
                     )
                     print("✅ 支出已记录")

        elif choice == '3': # 出库
            print(f"\n{Color.CYAN}📤 领料出库{Color.ENDC}")
            name = input("物品名称: ").strip()
            target_rec = None
            for r in recs:
                if r.fields.get("物品名称") == name:
                    target_rec = r
                    break
            
            if not target_rec:
                print("❌ 物品不存在")
                continue
                
            curr = float(target_rec.fields.get("当前库存", 0))
            print(f"当前库存: {curr} {target_rec.fields.get('单位')}")
            
            qty_out = float(input("领用数量: ").strip())
            if qty_out > curr:
                print(f"⚠️ 库存不足! (缺 {qty_out - curr})")
                if input("是否强制出库? (y/n): ").strip().lower() != 'y':
                    continue
            
            new_qty = curr - qty_out
            client.bitable.v1.app_table_record.update(
                UpdateAppTableRecordRequest.builder().app_token(app_token).table_id(table_id).record_id(target_rec.record_id)
                .request_body(AppTableRecord.builder().fields({"当前库存": new_qty, "最后变动时间": int(datetime.now().timestamp() * 1000)}).build()).build()
            )
            print(f"✅ 出库完成！剩余: {new_qty}")
            
        elif choice == '4': # 盘点
             print(f"\n{Color.CYAN}🔄 库存盘点{Color.ENDC}")
             name = input("物品名称: ").strip()
             target_rec = None
             for r in recs:
                 if r.fields.get("物品名称") == name:
                     target_rec = r
                     break
             if not target_rec:
                 print("❌ 物品不存在")
                 continue
                 
             print(f"系统库存: {target_rec.fields.get('当前库存')}")
             real_qty = float(input("实际盘点数量: ").strip())
             
             client.bitable.v1.app_table_record.update(
                 UpdateAppTableRecordRequest.builder().app_token(app_token).table_id(table_id).record_id(target_rec.record_id)
                 .request_body(AppTableRecord.builder().fields({"当前库存": real_qty, "最后变动时间": int(datetime.now().timestamp() * 1000)}).build()).build()
             )
             print(f"✅ 盘点已更新")

# 智能回款/付款核销助手
def smart_payment_matcher(client, app_token):
    """智能凑单工具：查找哪几笔账单凑成了这笔款项 (支持回款和付款)"""
    print(f"\n{Color.HEADER}🧩 智能核销助手 (Smart Matcher){Color.ENDC}")
    print("功能: 输入金额，自动查找对应的未结账单组合。")
    print("--------------------------------")
    print("1. 客户回款凑单 (查加工费收入)")
    print("2. 供应商付款凑单 (查外协/材料支出)")
    print("0. 返回")
    
    mode = input(f"\n👉 请选择模式 (1/2): ").strip()
    if mode == '0': return
    
    pf_table_id = get_table_id_by_name(client, app_token, "加工费明细表")
    if not pf_table_id: return

    if mode == '1':
        target_type = "收入-加工服务"
        name_prompt = "客户名称"
        amt_prompt = "回款金额"
    else:
        target_type = "支出-外协加工" # 简单起见先查外协，如果需要查材料可能要去日常台账
        # But wait, user said "Outsourced processing" is a major cost.
        # "Acid/Soda" might be in inventory or daily ledger.
        # Let's check "加工费明细表" first for outsourcing.
        name_prompt = "供应商名称"
        amt_prompt = "付款金额"

    name = input(f"👉 请输入{name_prompt} (支持模糊): ").strip()
    if not name: return
    
    # 1. Fetch Records
    print(f"⏳ 正在查询 '{name}' 的相关记录...")
    
    # Fetch based on mode
    target_recs = []
    
    if mode == '1' or mode == '2':
        # Check Processing Fee Table (covers Income and Outsourcing)
        filter_p = f'CurrentValue.[类型]="{target_type}"'
        records = get_all_records(client, app_token, pf_table_id, filter_info=filter_p)
        
        for r in records:
            p = r.fields.get("往来单位", "")
            if name in p:
                try:
                    amt = float(r.fields.get("总金额", 0))
                except: amt = 0
                
                target_recs.append({
                    "amt": amt,
                    "date": r.fields.get("日期", 0),
                    "item": r.fields.get("物品名称", "未知"),
                    "spec": r.fields.get("规格", ""),
                    "id": r.record_id,
                    "source": "加工费表"
                })
                
    # If mode 2, also check Daily Ledger for "Material" costs if not found enough?
    # For now, keep it simple.
            
    if not target_recs:
        print(f"❌ 未找到包含 '{name}' 的相关未结记录")
        return
        
    print(f"✅ 找到 {len(target_recs)} 条相关记录")
    
    # 2. Input Amount
    try:
        target_amt = float(input(f"👉 请输入{amt_prompt} (实际发生额): ").strip())
    except:
        print("❌ 金额无效")
        return
        
    print(f"⏳ 正在计算凑单组合 (目标: {target_amt})...")
    
    # 3. Find Subset
    import itertools
    
    found = False
    
    # Limit records to last 50 to avoid explosion
    target_recs.sort(key=lambda x: x["date"], reverse=True)
    working_recs = target_recs[:50] 
    
    for r in range(1, 6): # Try 1 to 5 bills
        for combo in itertools.combinations(working_recs, r):
            s = sum(c["amt"] for c in combo)
            if abs(s - target_amt) < 1.0: # Tolerance 1 yuan
                print(f"\n{Color.OKGREEN}🎉 找到匹配组合! (误差: {s - target_amt:.2f}){Color.ENDC}")
                for c in combo:
                    d_str = datetime.fromtimestamp(c["date"]/1000).strftime("%Y-%m-%d")
                    print(f" - {d_str} | {c['item']} {c['spec']} | ¥ {c['amt']}")
                found = True
                break
        if found: break
        
    if not found:
        print(f"\n{Color.WARNING}⚠️ 未找到精确匹配的组合 (尝试了最近50笔中的1-5笔组合){Color.ENDC}")
        print("建议：手动核对或检查是否有抹零/扣款。")
        
    input("\n按回车继续...")

# 老板查账/业务速查
def boss_quick_search(client, app_token):
    """老板查账：快速查询客户/供应商/库存/资金"""
    while True:
        print(f"\n{Color.HEADER}🔎 老板查账 (Quick Search){Color.ENDC}")
        print("--------------------------------")
        print("1. 👤 查客户 (欠款/最近交易)")
        print("2. 🏭 查供应商 (应付/最近采购)")
        print("3. 📦 查库存 (数量/价格)")
        print("4. 💰 查资金 (账户余额)")
        print("5. 🏷️ 查历史单价 (Price History)")
        print("0. 返回")
        
        c = input(f"\n👉 请选择: ").strip()
        if c == '0': break
        
        if c == '1': # Customer
            name = input("请输入客户名称 (模糊): ").strip()
            if not name: continue
            
            # Fetch Processing Fee (Income)
            pf_id = get_table_id_by_name(client, app_token, "加工费明细表")
            if not pf_id: continue
            
            print(f"⏳ 正在查询 '{name}' ...")
            filter_p = f'CurrentValue.[类型]="收入-加工服务"'
            recs = get_all_records(client, app_token, pf_id, filter_info=filter_p)
            
            found_recs = [r for r in recs if name in r.fields.get("往来单位", "")]
            if not found_recs:
                print("❌ 未找到记录")
                continue
            
            # Calc Stats
            total_amt = sum([float(r.fields.get("总金额", 0)) for r in found_recs])
            unpaid = 0.0
            for r in found_recs:
                # 简单判断：如果状态不是"已结清" (需确保字段存在)
                status = r.fields.get("状态", "未结") 
                if status != "已结清":
                    unpaid += float(r.fields.get("总金额", 0))
            
            found_recs.sort(key=lambda x: x.fields.get("日期", 0), reverse=True)
            last_3 = found_recs[:3]
            
            print(f"\n📊 {name} 数据概览:")
            print(f"   💰 累计加工费: {total_amt:,.2f}")
            print(f"   ⚠️ 当前未结清: {Color.FAIL}{unpaid:,.2f}{Color.ENDC}")
            print(f"   📝 最近3笔交易:")
            for r in last_3:
                d = datetime.fromtimestamp(r.fields.get("日期",0)/1000).strftime("%Y-%m-%d")
                item = r.fields.get("物品名称","")
                spec = r.fields.get("规格","")
                amt = float(r.fields.get("总金额", 0))
                print(f"     - {d} | {item} {spec} | ¥ {amt}")
                
        elif c == '2': # Supplier
             name = input("请输入供应商名称 (模糊): ").strip()
             pf_id = get_table_id_by_name(client, app_token, "加工费明细表")
             print(f"⏳ 正在查询 '{name}' ...")
             
             # Check Outsourcing
             filter_p = f'CurrentValue.[类型]="支出-外协加工"'
             recs = get_all_records(client, app_token, pf_id, filter_info=filter_p)
             found = [r for r in recs if name in r.fields.get("往来单位", "")]
             
             total = sum([float(r.fields.get("总金额", 0)) for r in found])
             unpaid = sum([float(r.fields.get("总金额", 0)) for r in found if r.fields.get("状态") != "已结清"])
             
             print(f"\n📊 {name} (外协) 数据概览:")
             print(f"   💰 累计外协费: {total:,.2f}")
             print(f"   ⚠️ 当前未付:   {Color.FAIL}{unpaid:,.2f}{Color.ENDC}")
             if found:
                 found.sort(key=lambda x: x.fields.get("日期", 0), reverse=True)
                 print(f"   📝 最近3笔:")
                 for r in found[:3]:
                     d = datetime.fromtimestamp(r.fields.get("日期",0)/1000).strftime("%Y-%m-%d")
                     print(f"     - {d} | {r.fields.get('物品名称')} | ¥ {r.fields.get('总金额')}")
                     
        elif c == '3': # Stock
             name = input("请输入物品名称 (模糊): ").strip()
             inv_id = get_table_id_by_name(client, app_token, "车间耗材库存表")
             if not inv_id: 
                 print("❌ 未启用库存表")
                 continue
             recs = get_all_records(client, app_token, inv_id)
             found = [r for r in recs if name in r.fields.get("物品名称", "")]
             
             if not found:
                 print("❌ 未找到物品")
             else:
                 print(f"\n📦 库存查询结果:")
                 for r in found:
                     n = r.fields.get("物品名称")
                     q = r.fields.get("当前库存")
                     u = r.fields.get("单位")
                     safe = r.fields.get("安全库存", 0)
                     print(f"   - {n}: {Color.OKGREEN}{q} {u}{Color.ENDC} (安全线: {safe})")
        
        elif c == '4': # Cash
             l_id = get_table_id_by_name(client, app_token, "日常台账表")
             recs = get_all_records(client, app_token, l_id)
             
             total_inc = sum([float(r.fields.get("实际收付金额",0)) for r in recs if r.fields.get("业务类型")=="收款"])
             total_exp = sum([float(r.fields.get("实际收付金额",0)) for r in recs if r.fields.get("业务类型") in ["付款","费用"]])
             
             print(f"\n💰 资金概览 (基于流水计算):")
             print(f"   总收入: {total_inc:,.2f}")
             print(f"   总支出: {total_exp:,.2f}")
             print(f"   💵 结余: {Color.OKGREEN}{total_inc - total_exp:,.2f}{Color.ENDC}")

        elif c == '5': # Price History
             p_name = input("请输入物品名称 (模糊): ").strip()
             c_name = input("请输入客户名称 (可选, 回车跳过): ").strip()
             
             pf_id = get_table_id_by_name(client, app_token, "加工费明细表")
             if pf_id:
                 print(f"⏳ 正在查询 '{p_name}' ...")
                 # Filter: Type="收入-加工服务"
                 filter_p = f'CurrentValue.[类型]="收入-加工服务"'
                 recs = get_all_records(client, app_token, pf_id, filter_info=filter_p)
                 
                 found = []
                 for r in recs:
                     item = r.fields.get("品名", "")
                     spec = r.fields.get("规格", "")
                     cust = r.fields.get("往来单位", "")
                     
                     if p_name in item or p_name in spec:
                         if c_name and c_name not in cust:
                             continue
                         found.append(r)
                 
                 if not found:
                     print("❌ 未找到记录")
                 else:
                     # Sort by date desc
                     found.sort(key=lambda x: x.fields.get("日期", 0), reverse=True)
                     
                     print(f"\n🏷️ 历史单价查询结果 (最近 10 笔):")
                     print(f"{'日期':<10} | {'客户':<10} | {'品名/规格':<20} | {'数量':<8} | {'单价':<8} | {'备注'}")
                     print("-" * 80)
                     
                     for r in found[:10]:
                         d = datetime.fromtimestamp(r.fields.get("日期",0)/1000).strftime("%Y-%m-%d")
                         cust = r.fields.get("往来单位", "")[:10]
                         desc = f"{r.fields.get('品名','')} {r.fields.get('规格','')}"[:20]
                         qty = f"{r.fields.get('数量',0)}{r.fields.get('单位','')}"
                         price = f"{r.fields.get('单价',0)}"
                         rem = r.fields.get("备注", "")
                         
                         print(f"{d:<10} | {cust:<10} | {desc:<20} | {qty:<8} | {Color.OKGREEN}{price:<8}{Color.ENDC} | {rem}")
             
        input("\n按回车继续...")

def calculate_piecework_salary():
    """简易计件工资计算器"""
    print(f"\n{Color.HEADER}👷 简易计件工资计算器{Color.ENDC}")
    print("--------------------------------")
    
    entries = []
    while True:
        print(f"\n📝 录入第 {len(entries)+1} 项 (输入 0 结束录入, c 清空):")
        process = input("   工序/产品 (如 '挂具', 'A款灯杯'): ").strip()
        if process == '0': break
        if process.lower() == 'c':
            entries = []
            print("已清空")
            continue
            
        try:
            qty_str = input("   数量 (个/扎): ").strip()
            price_str = input("   单价 (元): ").strip()
            
            qty = float(qty_str)
            price = float(price_str)
            total = qty * price
            
            entries.append({
                "process": process,
                "qty": qty,
                "price": price,
                "total": total
            })
            
            print(f"   ✅ 已记: {qty} * {price} = {total:.2f}")
            
        except:
            print("❌ 输入格式错误，请重试")
            
    if not entries: return
    
    # 汇总
    print(f"\n{Color.OKGREEN}📊 工资汇总单{Color.ENDC}")
    print("-" * 50)
    print(f"{'工序/产品':<15} | {'数量':<10} | {'单价':<8} | {'金额':<10}")
    print("-" * 50)
    
    grand_total = 0.0
    for e in entries:
        print(f"{e['process']:<15} | {e['qty']:<10.1f} | {e['price']:<8.3f} | {e['total']:<10.2f}")
        grand_total += e['total']
        
    print("-" * 50)
    print(f"{'总计':<37} | {Color.FAIL}¥ {grand_total:,.2f}{Color.ENDC}")
    
    # 复制提示
    print("\n💡 提示: 您可以截图或复制上方表格到微信发给员工核对。")
    input("按回车返回...")

def anodizing_price_calculator():
    """氧化计价助手 (面积/重量/价格计算)"""
    import math
    
    print(f"\n{Color.HEADER}📐 氧化计价助手 (报价神器){Color.ENDC}")
    print("--------------------------------")
    
    # 铝密度 (kg/m3)
    DENSITY_AL = 2700.0 
    
    while True:
        print("\n请选择计算模式:")
        print("1. 🟢 管/棒材 (输入: 直径x壁厚x长度)")
        print("2. 🟨 板/片材 (输入: 长x宽x厚)")
        print("3. 🔷 型材 (输入: 截面周长x米重x长度)")
        print("0. 返回")
        
        mode = input("👉 请选择: ").strip()
        if mode == '0': break
        
        area_total = 0.0 # 平方米
        weight_total = 0.0 # kg
        count = 0
        desc = ""
        
        try:
            if mode == '1': # 管/棒
                d_mm = float(input("   外径 (mm): "))
                wall_str = input("   壁厚 (mm) [实心棒填0]: ").strip()
                wall_mm = float(wall_str) if wall_str else 0
                l_m = float(input("   长度 (m): "))
                qty = int(input("   数量 (支): "))
                
                # Area = pi * d * L * qty (外表面积)
                area_one = math.pi * (d_mm / 1000.0) * l_m
                area_total = area_one * qty
                
                # Weight
                if wall_mm > 0:
                    # Tube: pi * (R^2 - r^2) * L * density
                    R = d_mm / 2.0 / 1000.0
                    r = (d_mm - 2*wall_mm) / 2.0 / 1000.0
                    vol_one = math.pi * (R**2 - r**2) * l_m
                    desc = f"管材 (Φ{d_mm}*{wall_mm}mm * {l_m}m)"
                else:
                    # Rod: pi * R^2 * L * density
                    R = d_mm / 2.0 / 1000.0
                    vol_one = math.pi * (R**2) * l_m
                    desc = f"棒材 (Φ{d_mm}mm * {l_m}m)"
                    
                weight_total = vol_one * DENSITY_AL * qty
                count = qty
                
            elif mode == '2': # 板
                l_mm = float(input("   长 (mm): "))
                w_mm = float(input("   宽 (mm): "))
                h_str = input("   厚 (mm) [用于算重/侧边]: ").strip()
                h_mm = float(h_str) if h_str else 0
                
                is_double = input("   是否双面氧化? (y/n) [y]: ").strip().lower() != 'n'
                qty = int(input("   数量 (片): "))
                
                # Area (Main face)
                area_one = (l_mm / 1000.0) * (w_mm / 1000.0)
                if is_double: area_one *= 2
                
                # Add side area if thickness provided
                if h_mm > 0:
                    perim = 2 * (l_mm + w_mm) / 1000.0
                    area_one += perim * (h_mm / 1000.0)
                
                area_total = area_one * qty
                
                # Weight
                if h_mm > 0:
                    vol_one = (l_mm/1000.0) * (w_mm/1000.0) * (h_mm/1000.0)
                    weight_total = vol_one * DENSITY_AL * qty
                    
                desc = f"板材 ({l_mm}*{w_mm}*{h_mm}mm)"
                count = qty
                
            elif mode == '3': # 型材
                p_mm = float(input("   截面周长 (mm): "))
                w_str = input("   米重 (kg/m) [用于算重, 不知可空]: ").strip()
                w_per_m = float(w_str) if w_str else 0
                
                l_m = float(input("   长度 (m): "))
                qty = int(input("   数量 (支): "))
                
                area_total = (p_mm / 1000.0) * l_m * qty
                
                if w_per_m > 0:
                    weight_total = w_per_m * l_m * qty
                    
                desc = f"型材 (周长{p_mm}mm * {l_m}m)"
                count = qty
                
            else:
                continue
                
            print(f"\n📊 计算结果: {desc}")
            print(f"   数量: {count}")
            print(f"   总面积: {Color.OKGREEN}{area_total:.4f} m²{Color.ENDC}")
            if weight_total > 0:
                print(f"   总重量: {Color.OKBLUE}{weight_total:.3f} kg{Color.ENDC} (估算)")
            
            # 算钱
            print("\n💰 计价方式:")
            print("   1. 按面积 (元/m²)")
            if weight_total > 0:
                print("   2. 按重量 (元/kg)")
            print("   3. 按数量 (元/件)")
            
            p_mode = input("👉 请选择计价方式 [1]: ").strip()
            
            total_amt = 0.0
            price_unit = ""
            
            if p_mode == '2' and weight_total > 0:
                price = float(input("   请输入单价 (元/kg): "))
                total_amt = weight_total * price
                price_unit = "元/kg"
            elif p_mode == '3':
                price = float(input("   请输入单价 (元/件): "))
                total_amt = count * price
                price_unit = "元/件"
            else:
                # Default Area
                price = float(input("   请输入单价 (元/m²): "))
                total_amt = area_total * price
                price_unit = "元/m²"
                
            price_per_item = total_amt / count if count > 0 else 0
            
            print(f"   --------------------")
            print(f"   总金额: {Color.FAIL}¥ {total_amt:,.2f}{Color.ENDC}")
            print(f"   折合单价: ¥ {price_per_item:.4f} /支(片)")
            
            input("\n按回车继续...")
            
        except Exception as e:
            print(f"❌ 输入错误: {e}")

def manage_small_tools(client, app_token):
    while True:
        print(f"\n{Color.BOLD}🧰 会计实用工具箱{Color.ENDC}")
        print(f"{Color.CYAN}--- 常用计算器 ---{Color.ENDC}")
        print("  1. 🔢 金额转大写")
        print("  2. 🧮 税额计算器 (含税/不含税)")
        print("  3. 📅 日期计算器")
        print("  7. 💸 贷款计算器 (等额本息)")
        print("  9. 🧾 增值税估算器 (进项抵扣)")
        
        print(f"{Color.CYAN}--- 数据维护 ---{Color.ENDC}")
        print("  4. 📥 生成 Excel 导入模板")
        print("  5. 📤 导出最新备份到桌面")
        print("  6. ♻️ 从回收站还原数据")
        print("  8. 📅 生成每月固定支出 (房租等)")
        
        print(f"{Color.CYAN}--- 行业与管理工具 ---{Color.ENDC}")
        print(" 10. 🏭 氧化厂模拟数据生成")
        print(" 11. 💰 资金账户对账 (余额核对)")
        print(" 12. 📦 车间耗材库存管理")
        print(" 13. 📢 应收账款催收助手")
        print(" 14. 📊 月度经营分析报告 (Visual)")
        print(" 15. 💴 薪酬个税管理")
        print(" 16. 🧩 智能回款核销助手")
        print(" 17. 🔎 老板查账 (Quick Search)")
        print(" 18. 👷 简易计件工资计算器")
        print(" 19. 📐 氧化计价助手 (报价神器)")
        
        print("  0. 返回主菜单")
        
        choice = input(f"👉 {Color.BOLD}请选择: {Color.ENDC}").strip()
        if choice == '0': break

        if choice == '19':
            anodizing_price_calculator()

        if choice == '18':
            calculate_piecework_salary()

        if choice == '17':
            boss_quick_search(client, app_token)
        
        if choice == '16':
            smart_payment_matcher(client, app_token)
            
        if choice == '15':
            manage_salary_flow(client, app_token)

        if choice == '13':
            debt_collection_assistant(client, app_token)

        if choice == '14':
            generate_monthly_visual_report(client, app_token)

        if choice == '12':
            manage_inventory(client, app_token)
            
        if choice == '11':
            reconcile_bank_account(client, app_token)
        
        if choice == '1':
            print(f"\n{Color.UNDERLINE}🔢 金额转大写{Color.ENDC}")
            while True:
                s = input("请输入金额 (输入 0 返回): ").strip()
                if s == '0': break
                try:
                    val = float(s)
                    cn = number_to_chinese(val)
                    print(f"👉 大写: {Color.OKGREEN}{cn}{Color.ENDC}")
                    # 同时也显示复制提示
                    print(f"   (已生成，可直接选中复制)")
                except:
                    print("❌ 无效数字")
                    
        elif choice == '2':
            print(f"\n{Color.UNDERLINE}🧮 税额计算器{Color.ENDC}")
            print("💡 提示: 输入 '100' 代表不含税，输入 'h 113' 代表含税")
            while True:
                s = input("\n请输入金额 (0 返回): ").strip()
                if s == '0': break
                
                is_inclusive = False
                val = 0.0
                if s.lower().startswith('h') or s.startswith('含'):
                    is_inclusive = True
                    try: val = float(s.lstrip('h含 '))
                    except: pass
                else:
                    try: val = float(s)
                    except: pass
                    
                if val == 0: continue
                
                r_str = input("税率% [13]: ").strip()
                if not r_str: r_str = "13"
                try:
                    rate = float(r_str) / 100.0
                except:
                    print("❌ 税率无效")
                    continue
                
                if is_inclusive:
                    amt = val / (1 + rate)
                    tax = val - amt
                    print(f"📉 [含税 {val:,.2f}] (税率 {int(rate*100)}%)")
                    print(f"   ✅ 不含税金额: {Color.OKGREEN}{amt:,.2f}{Color.ENDC}")
                    print(f"   ✅ 税额:       {Color.OKGREEN}{tax:,.2f}{Color.ENDC}")
                else:
                    tax = val * rate
                    total = val + tax
                    print(f"📈 [不含税 {val:,.2f}] (税率 {int(rate*100)}%)")
                    print(f"   ✅ 税额:       {Color.OKGREEN}{tax:,.2f}{Color.ENDC}")
                    print(f"   ✅ 价税合计:   {Color.OKGREEN}{total:,.2f}{Color.ENDC}")

        elif choice == '3':
            print(f"\n{Color.UNDERLINE}📅 日期计算器{Color.ENDC}")
            print("💡 示例: 输入 '30' (30天后) 或 '-7' (7天前)")
            while True:
                s = input("\n请输入天数 (0 返回): ").strip()
                if s == '0': break
                
                try:
                    days = int(s)
                    target_date = datetime.now() + timedelta(days=days)
                    desc = "后" if days > 0 else "前"
                    print(f"👉 {abs(days)}天{desc}: {Color.OKGREEN}{target_date.strftime('%Y-%m-%d')} ({target_date.strftime('%A')}){Color.ENDC}")
                except:
                    print("❌ 无效天数")

        elif choice == '4':
            generate_excel_template()

        elif choice == '5':
            print(f"\n{Color.UNDERLINE}📤 导出最新备份{Color.ENDC}")
            backup_root = BACKUP_DIR
            if not os.path.exists(backup_root):
                print("❌ 没有找到备份记录")
                continue
                
            # Find latest
            try:
                items = [os.path.join(backup_root, d) for d in os.listdir(backup_root)]
                valid = [d for d in items if os.path.isdir(d) or d.endswith('.zip')]
                if not valid:
                    print("❌ 没有找到有效备份")
                    continue
                    
                latest = max(valid, key=os.path.getmtime)
                fname = os.path.basename(latest)
                
                # Desktop path
                desktop = os.path.join(os.path.expanduser("~"), "Desktop")
                target = os.path.join(desktop, fname)
                
                import shutil
                if os.path.isdir(latest):
                    if os.path.exists(target):
                        shutil.rmtree(target)
                    shutil.copytree(latest, target)
                else:
                    shutil.copy2(latest, target)
                    
                print(f"✅ 已导出到桌面: {Color.GREEN}{target}{Color.ENDC}")
            except Exception as e:
                print(f"❌ 导出失败: {e}")

        elif choice == '6':
            # Restore from Recycle Bin
            print(f"\n{Color.CYAN}♻️ 数据还原向导{Color.ENDC}")
            recycle_log = os.path.join(DATA_ROOT, "系统日志", "recycle_bin.jsonl")
            if not os.path.exists(recycle_log):
                print("❌ 回收站为空")
                continue
                
            entries = []
            try:
                with open(recycle_log, "r", encoding="utf-8") as f:
                    for line in f:
                        if line.strip():
                            entries.append(json.loads(line))
            except: pass
            
            if not entries:
                print("❌ 回收站为空")
                continue
                
            # Show last 10 deleted
            print(f"\n最近删除记录 (共 {len(entries)} 条):")
            print("-" * 60)
            print(f"{'序号':<4} | {'删除时间':<20} | {'表名':<15} | {'内容摘要'}")
            print("-" * 60)
            
            last_10 = entries[-10:]
            for i, e in enumerate(reversed(last_10)):
                idx = len(entries) - i
                data_summary = str(e.get('data', {}))[:30] + "..."
                print(f"{idx:<4} | {e.get('deleted_at'):<20} | {e.get('table'):<15} | {data_summary}")
                
            print("-" * 60)
            print("💡 提示: 暂不支持直接一键还原，请根据上述信息手动补录。")
            print("   (完整日志请查看: 财务数据/系统日志/recycle_bin.jsonl)")
            input("\n按回车返回...")

        elif choice == '7':
            print(f"\n{Color.UNDERLINE}💸 贷款计算器 (等额本息){Color.ENDC}")
            while True:
                p_str = input("\n请输入贷款金额 (万元) [0返回]: ").strip()
                if p_str == '0': break
                
                try:
                    principal = float(p_str) * 10000
                    rate_str = input("请输入年利率% (如 3.85): ").strip()
                    years_str = input("请输入贷款年限 (年): ").strip()
                    
                    rate = float(rate_str) / 100.0
                    years = int(years_str)
                    months = years * 12
                    month_rate = rate / 12
                    
                    # 等额本息公式: PMT = P * i * (1+i)^n / ((1+i)^n - 1)
                    if month_rate == 0:
                        pmt = principal / months
                        total_interest = 0
                    else:
                        pmt = principal * month_rate * pow(1 + month_rate, months) / (pow(1 + month_rate, months) - 1)
                        total_interest = (pmt * months) - principal
                        
                    print(f"\n📊 计算结果:")
                    print(f"   💰 贷款总额: {principal/10000:.2f} 万元")
                    print(f"   📅 贷款期限: {years} 年 ({months} 期)")
                    print(f"   📉 年利率:   {rate*100:.2f}%")
                    print("-" * 30)
                    print(f"   ✅ 每月还款: {Color.OKGREEN}{pmt:.2f}{Color.ENDC} 元")
                    print(f"   ✅ 总支付利息: {total_interest/10000:.2f} 万元")
                    print(f"   ✅ 本息合计:   {(principal + total_interest)/10000:.2f} 万元")
                    
                except Exception as e:
                    print(f"❌ 输入错误: {e}")

        elif choice == '8':
            generate_monthly_expenses(client, app_token)

        elif choice == '9':
            estimate_vat_payable(client, app_token)
 
        elif choice == '10':
            try:
                # 延迟导入，避免顶层依赖问题
                from lark_oapi.api.bitable.v1.model import BatchCreateAppTableRecordRequest, BatchCreateAppTableRecordRequestBody, AppTableRecord
            except Exception:
                pass
            try:
                import simulate_factory_data as sfd
                sfd.update_rules()
                sfd.generate_excel()
                now = datetime.now()
                ym = f"{now.year}{now.month}"
                base_dir = os.path.dirname(os.path.abspath(__file__))
                g_path = os.path.join(base_dir, "财务数据", "待处理单据", f"模拟_G银行_对公流水_{ym}.xlsx")
                n_path = os.path.join(base_dir, "财务数据", "待处理单据", f"模拟_N银行_微信流水_{ym}.xlsx")
                for p in [g_path, n_path]:
                    if os.path.exists(p):
                        print(f"📥 正在导入: {p}")
                        import_from_excel(client, app_token, p)
                    else:
                        print(f"⚠️ 未找到文件: {p}")
                # 加工费示例记录
                pf_table_id = create_processing_fee_table(client, app_token)
                create_processing_price_table(client, app_token)
                if pf_table_id:
                    demo_date = int(datetime(now.year, now.month, 15).timestamp() * 1000)
                    recs = []
                    def add(fields):
                        recs.append(AppTableRecord.builder().fields(fields).build())
                    add({"日期": demo_date, "往来单位": "A灯饰厂", "品名": "铝型材", "规格": "20x30",
                         "类型": "收入-加工服务", "计价方式": "按件/个", "数量": 500, "单位": "件",
                         "单价": 1.200, "总金额": 600.00, "备注": "常规氧化-亮银"})
                    add({"日期": demo_date, "往来单位": "B五金制品", "品名": "铝条", "规格": "L=2m",
                         "类型": "收入-加工服务", "计价方式": "按米长", "数量": 800, "单位": "米",
                         "单价": 0.800, "总金额": 640.00, "备注": "拉丝后氧化"})
                    add({"日期": demo_date, "往来单位": "C电子科技", "品名": "散热片", "规格": "米重=150g",
                         "类型": "收入-加工服务", "计价方式": "按重量", "数量": 120.0, "单位": "kg",
                         "单价": 6.500, "总金额": 780.00, "备注": "按米重折算"})
                    add({"日期": demo_date, "往来单位": "D铝业", "品名": "铝板", "规格": "展开500mm",
                         "类型": "收入-加工服务", "计价方式": "按平方", "数量": 300.0, "单位": "m²",
                         "单价": 2.200, "总金额": 660.00, "备注": "按展开周长折算面积"})
                    add({"日期": demo_date, "往来单位": "精艺抛光厂", "品名": "抛光服务", "规格": "来料铝件",
                         "类型": "支出-外协加工", "计价方式": "按件/个", "数量": 500, "单位": "件",
                         "单价": 0.500, "总金额": 250.00, "备注": "外发抛光"})
                    add({"日期": demo_date, "往来单位": "锐砂喷砂", "品名": "喷砂服务", "规格": "铝型材",
                         "类型": "支出-外协加工", "计价方式": "按米长", "数量": 800, "单位": "米",
                         "单价": 0.300, "总金额": 240.00, "备注": "外发喷砂"})
                    try:
                        req = BatchCreateAppTableRecordRequest.builder() \
                            .app_token(app_token) \
                            .table_id(pf_table_id) \
                            .request_body(BatchCreateAppTableRecordRequestBody.builder().records(recs).build()) \
                            .build()
                        resp = client.bitable.v1.app_table_record.batch_create(req)
                        if resp.success():
                            print("✅ 已插入加工费示例记录 (收入/外协)")
                        else:
                            print(f"❌ 插入加工费记录失败: {resp.msg}")
                    except Exception as e:
                        print(f"❌ 批量写入失败: {e}")
                print(f"\n{Color.OKGREEN}🎉 氧化厂模拟数据已导入完毕！{Color.ENDC}")
                print("下一步建议：")
                print("  - 输入 22 运行【财务体检】，查看无票/大额现金与经营风险")
                print("  - 输入 21 运行【快速查账】，搜索 '外协加工费' 或 '氧化加工费'")
                print("  - 输入 23 打开【工具箱】，可用 9 估算本月增值税")
            except Exception as e:
                print(f"❌ 执行失败: {e}")

def estimate_vat_payable(client, app_token):
    """简易增值税估算器"""
    print(f"\n{Color.HEADER}📊 增值税估算器 (VAT Estimator){Color.ENDC}")
    print("----------------------------------------")
    print("本工具用于估算本期应交增值税额 (仅供参考，以税务申报为准)")
    
    # 1. 选择纳税人类型
    print("\n请选择纳税人类型:")
    print("  1. 小规模纳税人 (1% 征收率)")
    print("  2. 小规模纳税人 (3% 征收率)")
    print("  3. 一般纳税人 (13% / 6%) - 支持进项抵扣")
    
    t_choice = input("👉 请选择 [1]: ").strip()
    if not t_choice: t_choice = '1'
    
    tax_rate = 0.01
    input_deductible = False
    
    if t_choice == '2': 
        tax_rate = 0.03
    elif t_choice == '3':
        tax_rate = 0.13 # 默认一般税率
        input_deductible = True
        
    # 2. 选择期间
    cur_month = datetime.now().strftime("%Y-%m")
    month_input = input(f"\n请输入估算月份 (YYYY-MM) [{cur_month}]: ").strip()
    if not month_input: month_input = cur_month
    
    try:
        start_dt = datetime.strptime(month_input, "%Y-%m")
        if start_dt.month == 12:
            end_dt = datetime(start_dt.year + 1, 1, 1)
        else:
            end_dt = datetime(start_dt.year, start_dt.month + 1, 1)
            
        start_ts = int(start_dt.timestamp() * 1000)
        end_ts = int(end_dt.timestamp() * 1000)
    except:
        print("❌ 日期格式错误")
        return

    # 3. 拉取数据
    table_id = get_table_id_by_name(client, app_token, "日常台账表")
    if not table_id: return
    
    print(f"正在拉取 {month_input} 的账目数据...")
    filter_cmd = f'AND(CurrentValue.[记账日期]>={start_ts}, CurrentValue.[记账日期]<{end_ts})'
    records = get_all_records(client, app_token, table_id, filter_info=filter_cmd)
    
    # 4. 计算
    total_sales_inc = 0.0 # 含税销售额
    total_input_inc = 0.0 # 含税进项额 (有票)
    
    detail_lines = []
    
    for r in records:
        f = r.fields
        amt = float(f.get("实际收付金额", 0))
        b_type = f.get("业务类型", "")
        has_ticket = f.get("是否有票") == "有票"
        
        # 销项: 只要是“收款”且业务类型暗示是收入
        # 简单起见，所有“收款”视为收入 (需剔除往来? 暂时无法区分)
        # 优化: 排除备注含 "借款", "退回" 等
        memo = str(f.get("备注", ""))
        if b_type == "收款":
            if "借款" not in memo and "退款" not in memo:
                total_sales_inc += amt
        
        # 进项: “付款”或“费用” 且 “有票”
        if (b_type in ["付款", "费用"]) and has_ticket:
            total_input_inc += amt
            
    # 5. 估算税额
    if not input_deductible:
        # 小规模: 销售额 / (1+征收率) * 征收率
        sales_ex = total_sales_inc / (1 + tax_rate)
        vat_out = sales_ex * tax_rate
        vat_in = 0.0
        vat_payable = vat_out
        
        print(f"\n🧾 估算结果 ({month_input}):")
        print(f"   💰 含税销售额: {total_sales_inc:,.2f}")
        print(f"   📉 不含税销售: {sales_ex:,.2f}")
        print(f"   应交增值税:   {Color.FAIL}{vat_payable:,.2f}{Color.ENDC} (按 {int(tax_rate*100)}% 简易征收)")
        
    else:
        # 一般纳税人: (销项 - 进项)
        # 假设销项税率 13%, 进项税率 13% (简化)
        sales_ex = total_sales_inc / (1.13)
        vat_out = sales_ex * 0.13
        
        # 进项倒推 (假设都是专用发票)
        input_ex = total_input_inc / (1.13)
        vat_in = input_ex * 0.13
        
        vat_payable = vat_out - vat_in
        
        print(f"\n🧾 估算结果 ({month_input}) [一般纳税人模式]:")
        print(f"   💰 销项 (估):   {vat_out:,.2f} (基于含税收入 {total_sales_inc:,.0f})")
        print(f"   🎫 进项 (估):   {vat_in:,.2f} (基于有票支出 {total_input_inc:,.0f})")
        
        c = Color.FAIL if vat_payable > 0 else Color.GREEN
        print(f"   应交增值税:   {c}{vat_payable:,.2f}{Color.ENDC}")
        
        if vat_payable > 0:
            print(f"   💡 提示: 您还需要 {vat_payable/0.13:,.0f} 元的进项发票来抵扣税款。")
            
    print("\n⚠️ 注意: 此功能仅作资金预算参考，实际申报请咨询专业会计。")
    input("\n按回车返回...")
 

def generate_excel_template():
    """生成 Excel 导入模板"""
    print(f"\n{Color.UNDERLINE}📥 生成 Excel 导入模板{Color.ENDC}")
    try:
        import pandas as pd
        
        # 1. 定义标准列名
        columns = ["记账日期", "业务类型", "费用归类", "实际收付金额", "往来单位费用", "备注", "是否有票", "是否现金"]
        
        # 2. 创建示例数据
        data = [
            ["2024-01-01", "收款", "", 10000, "客户A", "货款", "有票", "否"],
            ["2024-01-02", "付款", "原材料采购", 5000, "供应商B", "采购材料", "有票", "否"],
            ["2024-01-03", "费用", "办公费", 200, "京东", "买纸笔", "有票", "否"],
            ["2024-01-04", "费用", "差旅费-交通", 50, "滴滴", "打车去税务局", "无票", "否"]
        ]
        
        df = pd.DataFrame(data, columns=columns)
        
        # 3. 保存文件
        if not os.path.exists(TEMPLATE_DIR):
            os.makedirs(TEMPLATE_DIR)
            
        fname = os.path.join(TEMPLATE_DIR, f"导入模板_{datetime.now().strftime('%Y%m%d')}.xlsx")
        
        # 使用 ExcelWriter 设置列宽
        with pd.ExcelWriter(fname, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="日常台账表")
            
        print(f"✅ 模板已生成: {Color.OKGREEN}{fname}{Color.ENDC}")
        # 尝试打开文件夹
        try:
            os.startfile(TEMPLATE_DIR)
        except:
            pass
            
            # 尝试调整列宽
            worksheet = writer.sheets['日常台账表']
            for i, col in enumerate(columns):
                worksheet.column_dimensions[chr(65+i)].width = 15
                
        print(f"✅ 模板已生成: {Color.GREEN}{fname}{Color.ENDC}")
        print("💡 提示: 请在模板中填入数据，然后使用 '3. 从 Excel 导入数据' 功能。")
        
        # 尝试打开文件夹
        try:
            os.startfile(os.getcwd())
        except: pass
        
    except Exception as e:
        print(f"❌ 生成失败: {e}")


def load_voucher_templates():
    """加载凭证模板"""
    if os.path.exists(FILE_VOUCHER_TEMPLATES):
        try:
            with open(FILE_VOUCHER_TEMPLATES, "r", encoding="utf-8") as f:
                return json.load(f)
        except: pass
    return {}

def save_voucher_templates(templates):
    """保存凭证模板"""
    try:
        with open(FILE_VOUCHER_TEMPLATES, "w", encoding="utf-8") as f:
            json.dump(templates, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"❌ 保存模板失败: {e}")

def update_dashboard_cache_silent(client, app_token):
    """静默更新仪表盘缓存 (不打印日志)"""
    try:
        table_id = get_table_id_by_name(client, app_token, "日常台账表")
        if not table_id: return

        # 当月时间范围过滤，减少数据拉取量
        now = datetime.now()
        cur_month = now.strftime("%Y-%m")
        start_dt = datetime(now.year, now.month, 1)
        if now.month == 12:
            end_dt = datetime(now.year + 1, 1, 1)
        else:
            end_dt = datetime(now.year, now.month + 1, 1)
        start_ts = int(start_dt.timestamp() * 1000)
        end_ts = int(end_dt.timestamp() * 1000)
        filter_cmd = f'CurrentValue.[记账日期]>={start_ts}&&CurrentValue.[记账日期]<{end_ts}'
        
        records = get_all_records(client, app_token, table_id, filter_info=filter_cmd, field_names=["记账日期", "实际收付金额", "业务类型"])
        
        inc = 0.0
        exp = 0.0
        
        for r in records:
            f = r.fields
            try:
                ts = f.get("记账日期", 0)
                val = float(f.get("实际收付金额", 0))
                b_type = f.get("业务类型", "")
                
                if b_type == "收款":
                    inc += val
                elif b_type in ["付款", "费用"]:
                    exp += val
            except: pass
            
        net = inc - exp
        
        data = {
            "month": cur_month,
            "income": inc,
            "expense": exp,
            "net": net,
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        
        with open(FILE_DASHBOARD_CACHE, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=4)
            
    except Exception:
        pass # Silent

def show_progress_bar(current, total, prefix='', suffix='', decimals=1, length=30, fill='█'):
    """
    终端进度条生成器
    [██████████] 100.0% Complete
    """
    percent = ("{0:." + str(decimals) + "f}").format(100 * (current / float(total)))
    filled_length = int(length * current // total)
    bar = fill * filled_length + '-' * (length - filled_length)
    print(f'\r{prefix} |{bar}| {percent}% {suffix}', end='\r')
    # Print New Line on Complete
    if current == total: 
        print()

def register_voucher(client, app_token):
    """手工录入凭证 (CLI Wizard) - 支持模板"""
    print(f"\n{Color.HEADER}📝 手工录入凭证 (Voucher Entry){Color.ENDC}")
    print("-----------------------------------------------")
    
    # 0. Load Template Option
    templates = load_voucher_templates()
    template_data = None
    
    if templates:
        print(f"{Color.CYAN}📋 可用模板:{Color.ENDC}")
        t_keys = list(templates.keys())
        for idx, k in enumerate(t_keys):
            t = templates[k]
            print(f"  {idx+1}. {k} ({t.get('type', '')} {t.get('amount', '')})")
        print("  0. 不使用模板")
        print("  -1. 📋 智能粘贴录入 (New!)")
        
        t_choice = input(f"\n👉 选择模板 (0-{len(t_keys)}): ").strip()
        
        if t_choice == '-1':
             print(f"\n{Color.OKBLUE}📋 请粘贴整段文本 (例如: '昨天付给张三货款5000元'){Color.ENDC}")
             raw_text = input("👉 文本内容: ").strip()
             if raw_text:
                 smart_data = parse_smart_text(raw_text)
                 if smart_data:
                     print(f"✅ 智能解析成功! (已自动填入相关字段)")
                     template_data = smart_data
        elif t_choice.isdigit() and 1 <= int(t_choice) <= len(t_keys):
            key = t_keys[int(t_choice)-1]
            template_data = templates[key]
            print(f"✅ 已加载模板: {key}")
    
    # 1. Date
    default_date = datetime.now().strftime("%Y-%m-%d")
    if template_data and template_data.get("date"):
        default_date = template_data.get("date")
        
    date_input = input(f"\n1. 📅 日期 [{default_date}] (支持 '昨天', '2.5'): ").strip()
    
    if not date_input: 
        date_str = default_date
    else:
        # 使用智能解析
        parsed = parse_date_smart(date_input)
        if parsed:
            date_str = parsed
            print(f"   ✅ 识别为: {date_str}")
        else:
            date_str = default_date
            print(f"   ⚠️ 无法识别，使用默认: {date_str}")

    try:
        ts = int(datetime.strptime(date_str, "%Y-%m-%d").timestamp() * 1000)
    except:
        print("❌ 日期格式错误")
        return

    # 2. Type
    default_type = template_data.get('type', '费用') if template_data else '费用'
    print(f"\n2. 🏷️ 业务类型 (当前默认: {default_type})")
    print("  1. 收款 (+)")
    print("  2. 付款 (-)")
    print("  3. 费用 (-)")
    t_map = {"1": "收款", "2": "付款", "3": "费用"}
    
    t_input = input("👉 选择 (1-3) 或直接回车: ").strip()
    if t_input in t_map:
        biz_type = t_map[t_input]
    else:
        biz_type = default_type

    # 3. Amount
    default_amt = str(template_data.get('amount', '')) if template_data else ''
    amt_prompt = f"[{default_amt}]" if default_amt else ""
    amt_str = input(f"\n3. 💰 金额 {amt_prompt}: ").strip()
    if not amt_str and default_amt: amt_str = default_amt
    
    try:
        amount = float(eval(amt_str, {"__builtins__": None}, {}))
    except:
        print("❌ 金额错误")
        return

    # 4. Partner
    default_partner = template_data.get('partner', '') if template_data else ''
    partner_prompt = f"[{default_partner}]" if default_partner else ""
    partner = input(f"\n4. 👤 往来单位 {partner_prompt}: ").strip()
    if not partner: partner = default_partner if default_partner else "散户"

    # 5. Category
    default_cat = template_data.get('category', '') if template_data else ''
    cat_prompt = f"[{default_cat}]" if default_cat else ""
    category = input(f"\n5. 📂 费用归类 {cat_prompt}: ").strip()
    if not category: category = default_cat if default_cat else "未分类"

    # 6. Remark
    default_remark = template_data.get('remark', '') if template_data else ''
    remark_prompt = f"[{default_remark}]" if default_remark else ""
    remark = input(f"\n6. 📝 备注摘要 {remark_prompt}: ").strip()
    if not remark: remark = default_remark

    # 7. Invoice
    has_invoice = "无票"
    if input("\n7. 🧾 是否有票? (y/n) [n]: ").strip().lower() == 'y':
        has_invoice = "有票"

    # Review
    print(f"\n{Color.BOLD}👀 确认信息:{Color.ENDC}")
    print(f"  📅 日期: {date_str}")
    print(f"  🏷️ 类型: {biz_type}")
    print(f"  💰 金额: {amount:,.2f}")
    print(f"  🏢 单位: {partner}")
    print(f"  📂 分类: {category}")
    print(f"  📝 备注: {remark}")
    print(f"  🧾 发票: {has_invoice}")
    
    if input("\n确认保存吗? (y/n): ").strip().lower() != 'y': return

    # Save to Feishu
    table_id = get_table_id_by_name(client, app_token, "日常台账表")
    if not table_id: return
    
    fields = {
        "记账日期": ts,
        "业务类型": biz_type,
        "费用归类": category,
        "往来单位费用": partner,
        "实际收付金额": amount,
        "备注": remark,
        "是否有票": has_invoice,
        "是否现金": "否"
    }
    
    try:
        req = CreateAppTableRecordRequest.builder() \
            .app_token(app_token) \
            .table_id(table_id) \
            .request_body(AppTableRecord.builder().fields(fields).build()) \
            .build()
            
        resp = client.bitable.v1.app_table_record.create(req)
        if resp.success():
            new_record_id = resp.data.record_id
            print(f"\n✅ {Color.GREEN}凭证保存成功！{Color.ENDC}")
            
            # Undo Logic
            print(f"{Color.WARNING}👉 如需撤销，请在 3 秒内输入 'u' 并回车...{Color.ENDC}")
            # 这里不能 sleep 否则会卡住，只能是普通提示
            # 或者直接问
            if input("↩️ 输入 'u' 撤销录入，或直接回车继续: ").strip().lower() == 'u':
                try:
                     req_del = DeleteAppTableRecordRequest.builder() \
                        .app_token(app_token) \
                        .table_id(table_id) \
                        .record_id(new_record_id) \
                        .build()
                     if client.bitable.v1.app_table_record.delete(req_del).success():
                         print(f"🗑️ {Color.OKGREEN}已撤销上一条录入。{Color.ENDC}")
                         # 软删除日志
                         try:
                             recycle_log = os.path.join(DATA_ROOT, "系统日志", "recycle_bin.jsonl")
                             with open(recycle_log, "a", encoding="utf-8") as f:
                                 log_entry = {
                                     "deleted_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                     "table": "日常台账表 (Undo)",
                                     "record_id": new_record_id,
                                     "data": fields
                                 }
                                 f.write(json.dumps(log_entry, ensure_ascii=False) + "\n")
                         except: pass
                         return # 撤销后直接返回
                     else:
                         print("❌ 撤销失败")
                except Exception as e:
                     print(f"❌ 撤销异常: {e}")
            
            # Silent Update Dashboard (Only if not undone)
            print("⏳ 正在更新仪表盘...", end="", flush=True)
            update_dashboard_cache_silent(client, app_token)
            print("\r" + " " * 30 + "\r", end="", flush=True)
            
            # Save as Template Option
            if input("\n💾 是否保存为常用模板? (y/n): ").strip().lower() == 'y':
                t_name = input("请输入模板名称 (如 '每月房租'): ").strip()
                if t_name:
                    new_t = {
                        "type": biz_type,
                        "category": category,
                        "amount": amount,
                        "partner": partner,
                        "remark": remark
                    }
                    templates = load_voucher_templates()
                    templates[t_name] = new_t
                    save_voucher_templates(templates)
                    print(f"✅ 模板 '{t_name}' 已保存")
                
        
    except Exception as e:
        log.error(f"保存异常: {e}")


def manage_category_rules():
    """管理自动分类规则"""
    global AUTO_CATEGORY_RULES
    
    while True:
        print(f"\n{Color.HEADER}🏷️ 自动分类规则管理{Color.ENDC}")
        print("-------------------")
        print("1. 查看当前规则")
        print("2. 添加新规则")
        print("3. 删除规则")
        print("0. 返回")
        print("-------------------")
        
        choice = input("请选择 (0-3): ").strip()
        
        if choice == '0':
            break
            
        elif choice == '1':
            print(f"\n📋 当前规则 ({len(AUTO_CATEGORY_RULES)}条):")
            print(f"{'关键词':<20} -> {'分类'}")
            print("-" * 40)
            if not AUTO_CATEGORY_RULES:
                print("(暂无规则)")
            else:
                # 只显示前50条，避免太长
                count = 0
                for k, v in AUTO_CATEGORY_RULES.items():
                    print(f"{k:<20} -> {v}")
                    count += 1
                    if count >= 50:
                        print(f"... (还有 {len(AUTO_CATEGORY_RULES)-50} 条)")
                        break
            input("\n按回车继续...")
            
        elif choice == '2':
            print("\n➕ 添加新规则")
            key = input("请输入关键词 (如 '滴滴'): ").strip()
            if not key: continue
            
            cat = input(f"请输入 '{key}' 对应的分类 (如 '差旅费-交通'): ").strip()
            if not cat: continue
            
            AUTO_CATEGORY_RULES[key] = cat
            
            try:
                with open(FILE_CATEGORY_RULES, "w", encoding="utf-8") as f:
                    json.dump(AUTO_CATEGORY_RULES, f, ensure_ascii=False, indent=4)
                print(f"✅ 已添加: {key} -> {cat}")
            except Exception as e:
                log.error(f"保存失败: {e}")
                
        elif choice == '3':
            key = input("请输入要删除的关键词: ").strip()
            if key in AUTO_CATEGORY_RULES:
                del AUTO_CATEGORY_RULES[key]
                try:
                    with open(FILE_CATEGORY_RULES, "w", encoding="utf-8") as f:
                        json.dump(AUTO_CATEGORY_RULES, f, ensure_ascii=False, indent=4)
                    print(f"✅ 已删除: {key}")
                except Exception as e:
                    log.error(f"保存失败: {e}")
            else:
                print("❌ 找不到该规则")

def manage_config_menu():
    """配置管理菜单 (别名/规则)"""
    while True:
        print(f"\n{Color.HEADER}⚙️ 系统配置管理{Color.ENDC}")
        print("--------------------------------")
        print("  1. 往来单位别名管理 (Partner Aliases)")
        print("  2. 自动分类规则管理 (Category Rules)")
        print("  0. 返回主菜单")
        
        choice = input(f"\n👉 请选择: ").strip()
        if choice == '0': break
        
        if choice == '1':
            manage_partner_aliases()
        elif choice == '2':
            manage_category_rules()

def manage_partner_aliases():
    """往来单位别名管理 (CRUD)"""
    json_file = FILE_PARTNER_ALIASES
    
    while True:
        # Load latest
        aliases = {}
        if os.path.exists(json_file):
            try:
                with open(json_file, "r", encoding="utf-8") as f:
                    aliases = json.load(f)
            except: pass
            
        print(f"\n{Color.UNDERLINE}👤 往来单位别名管理 ({len(aliases)} 条){Color.ENDC}")
        print("  1. 📋 查看所有别名")
        print("  2. ➕ 添加/修改别名")
        print("  3. ❌ 删除别名")
        print("  0. 返回")
        
        c = input("👉 请选择: ").strip()
        if c == '0': break
        
        if c == '1':
            print("\n--------------------------------")
            print(f"{'关键词 (Excel)':<15} -> {'标准名称 (飞书)':<15}")
            print("--------------------------------")
            for k, v in aliases.items():
                print(f"{k:<15} -> {Color.OKGREEN}{v}{Color.ENDC}")
            print("--------------------------------")
            input("按回车继续...")
            
        elif c == '2':
            print("\n💡 提示: 输入 Excel 里的名字 (如 '张三') 和 飞书里的标准名 (如 'A客户')")
            k = input("🔑 关键词 (Excel出现的名字): ").strip()
            if not k: continue
            v = input(f"🏷️ 标准名 (飞书里的名字) [当前: {aliases.get(k, '无')}]: ").strip()
            if not v: continue
            
            aliases[k] = v
            with open(json_file, "w", encoding="utf-8") as f:
                json.dump(aliases, f, indent=4, ensure_ascii=False)
            print(f"✅ 已保存: {k} -> {v}")
            
        elif c == '3':
            k = input("🗑️ 请输入要删除的关键词: ").strip()
            if k in aliases:
                del aliases[k]
                with open(json_file, "w", encoding="utf-8") as f:
                    json.dump(aliases, f, indent=4, ensure_ascii=False)
                print(f"✅ 已删除: {k}")
            else:
                print("❌ 找不到该关键词")

def interactive_menu():
    """Python版交互主菜单 (重构：按频率分组)"""
    # 启用 Windows ANSI 支持 (如果是 Windows)
    if os.name == 'nt':
        os.system('color')
        
    while True:
        # 清屏 (兼容 Windows/Linux)
        os.system('cls' if os.name == 'nt' else 'clear')
        
        print(f"{Color.HEADER}===============================================")
        print(f"       🚀 飞书财务小助手 V9.8 - 旗舰版")
        print(f"==============================================={Color.ENDC}")
        
        # 显示仪表盘状态
        print(f"\n{draw_dashboard_ui()}")
        
        print(f"\n{Color.OKGREEN}☀️ 日常高频 (Daily){Color.ENDC}")
        print("  00. 🚀 一键日结 (自动扫描+处理+备份) [推荐]")
        print("  1.  📝 凭证登记 (手工/模板)")
        print("  2.  🏭 加工费登记 (工厂专用)")
        print("  3.  📥 导入 Excel (流水/单据)")
        print("  4.  🏦 银行对账 (自动勾兑)")
        print("  5.  📊 每日简报 (老板看板)")
        
        print(f"\n{Color.CYAN}🌙 月末结账 (Monthly){Color.ENDC}")
        print("  11. 📉 计提折旧 (固定资产)")
        print("  12. 💰 薪酬工资 (个税/社保)")
        print("  13. 🧾 发票管理 (进项/销项)")
        print("  14. 🤝 往来对账中心 (New!)")
        print("  15. 🗓️ 月度结账 (利润表/归档)")
        
        print(f"\n{Color.OKBLUE}🔧 实用工具 (Tools){Color.ENDC}")
        print("  21. 🔍 快速查账 (搜索/导出)")
        print("  22. 🏥 财务体检 (风险扫描)")
        print("  23. 🧰 会计工具箱 (税额/大写/模板)")
        print("  97. ⚙️ 系统配置 (分类规则/别名)")
        print("  98. 🤖 AI 助手 (自然语言问答)")
        print("  99. ❌ 退出系统")
        
        # 兼容旧代码的输入处理，映射新菜单到旧逻辑
        # 我们需要保留原有的 choice 处理逻辑，但界面上只显示精简的
        # 这里做一个映射表，将新菜单号映射到实际执行的功能号
        # 或者直接修改下面的 if-elif 逻辑，但这改动太大
        # 方案：保持 input 接收，如果用户输入了旧代码也能用，但界面引导用新的
        
        print("\n👉 请输入功能编号 (支持搜索，如 '折旧'): ")
        choice = input("   您的选择: ").strip()
        
        # 模糊搜索支持
        if not choice.isdigit():
            # 关键词映射表
            keywords = {
                "日结": "00", "一键": "00",
                "凭证": "27", "手工": "27",
                "加工": "26", "工厂": "26",
                "导入": "3", "excel": "3",
                "对账": "4", "银行": "4",
                "简报": "8", "日报": "8",
                "折旧": "12", "固定资产": "12",
                "工资": "24", "薪酬": "24",
                "发票": "25", "税务": "13",
                "往来": "5",
                "月结": "11", "结账": "11",
                "查账": "18", "搜索": "18",
                "体检": "9", "检查": "9",
                "工具": "28", "大写": "28",
                "配置": "97", "设置": "97",
                "ai": "10", "助手": "10",
                "退出": "99"
            }
            # 简单匹配
            match = None
            for k, v in keywords.items():
                if k in choice:
                    match = v
                    break
            
            if match:
                print(f"🔍 已识别指令: {choice} -> {match}")
                choice = match
                time.sleep(0.5)
            else:
                print("❌ 未识别指令，请重试")
                time.sleep(1)
                continue

        # 菜单路由映射 (New UI -> Old Logic)
        # 1 -> 27 (凭证)
        # 2 -> 26 (加工费)
        # 3 -> 3
        # 4 -> 4
        # 5 -> 8 (简报)
        # 11 -> 12 (折旧)
        # 12 -> 24 (薪酬)
        # 13 -> 25 (发票)
        # 14 -> 5 (往来)
        # 15 -> 11 (月结)
        # 21 -> 18 (查账)
        # 22 -> 9 (体检)
        # 23 -> 28 (工具箱)
        # 97 -> 97 (配置)
        # 98 -> 10 (AI)
        
        real_choice = choice
        if choice == '1': real_choice = '27'
        elif choice == '2': real_choice = '26'
        elif choice == '5': real_choice = '8'
        elif choice == '11': real_choice = '12'
        elif choice == '12': real_choice = '24'
        elif choice == '13': real_choice = '25'
        elif choice == '14': real_choice = '5'
        elif choice == '15': real_choice = '11'
        elif choice == '21': real_choice = '18'
        elif choice == '22': real_choice = '9'
        elif choice == '23': real_choice = '28'
        elif choice == '97': real_choice = '97'
        elif choice == '98': real_choice = '10'
        
        choice = real_choice # 传递给后续逻辑
        
        if choice == '00':
            one_click_daily_closing(client, app_token)
        elif choice == '1': # 保留旧的截图记账入口，但在UI上隐藏了
            smart_image_entry(client, app_token)
        elif choice == '2': # 保留旧的文本记账入口
            smart_text_entry(client, app_token)
        elif choice == '27':
            register_voucher(client, app_token)
        elif choice == '3':
            import_from_excel(client, app_token)
        elif choice == '4':
            reconcile_bank_flow(client, app_token, None)
        elif choice == '5':
            reconciliation_hub(client, app_token)
        elif choice == '6':
            export_missing_tickets(client, app_token)
        elif choice == '7':
            generate_daily_html_report(client, app_token)
        elif choice == '8':
            daily_briefing(client, app_token)
        elif choice == '9':
            financial_health_check(client, app_token)
        elif choice == '10':
            smart_query_assistant(client, app_token)
        elif choice == '11':
            monthly_closing(client, app_token)
        elif choice == '22': # 一键年结
            annual_closing(client, app_token)
        elif choice == '12':
            calculate_depreciation(client, app_token)
        elif choice == '13':
            calculate_tax(client, app_token)
        elif choice == '14': # 旧的别名管理，现在合并到97
            manage_config_menu()
        elif choice == '97':
            manage_config_menu()
        elif choice == '16':
            export_vouchers(client, app_token)
        elif choice == '17':
            smart_learning_mode(client, app_token)
        elif choice == '18':
            quick_search_ledger(client, app_token)
        elif choice == '19':
            backup_system_data(client, app_token)
        elif choice == '20':
            folder_monitor_mode(client, app_token)
        elif choice == '21':
            generate_annual_report_html(client, app_token)
        elif choice == '23':
            reconcile_external_bill(client, app_token)
        elif choice == '24':
            manage_salary_flow(client, app_token)
        elif choice == '25':
            manage_invoice_flow(client, app_token)
        elif choice == '26':
            manage_processing_fee_flow(client, app_token)
        elif choice == '28':
            manage_small_tools(client, app_token)
        elif choice == '95':
            setup_auto_run_task()
        elif choice == '96':
            restore_from_backup(client, app_token)
        elif choice == '99':
            print("👋 再见！")
            break
        else:
            print("❌ 无效选项，请重试")
            time.sleep(1)
        print("  16. 导出标准凭证 (财务软件用) [新]")
        print("  17. 智能学习分类规则 (越用越聪明) [新]")
        print("  18. 万能查账 (金额/日期/关键词) [新]")
        print("  19. 导出云端数据到 Excel [备份]")
        print("  20. 启动文件夹监听模式 (支持Excel/图片) [新]")

        print(f"\n{Color.CYAN}🧰 实用工具{Color.ENDC}")
        print("  28. 会计常用工具箱 (大写/税额) [新]")
        
        print(f"\n{Color.CYAN}🛠️ 系统工具{Color.ENDC}")
        print("  95. 设置每日自动运行 (Windows任务) [新]")
        print("  96. 从备份恢复数据 [新]")
        print("  97. 初始化飞书表格 (首次)")
        print("  98. 备份系统数据 (配置+Excel)")
        print("  99. 显示云端后台链接")
        print("  0. 退出")
        print(f"{Color.HEADER}==============================================={Color.ENDC}")
        
        choice = input(f"\n👉 {Color.BOLD}请选择功能 (输入数字): {Color.ENDC}").strip()
        
        # 处理选择
        if choice == '0':
            print("👋 再见！")
            sys.exit(0)
        
        elif choice == '29':
            manage_config_menu()
            continue
            
        # [无需联网的功能优先处理]
        if choice == '95': 
            setup_auto_task()
            input("\n✅ 操作完成，按回车返回...")
            continue
        elif choice == '96': 
            restore_from_backup()
            input("\n✅ 操作完成，按回车返回...")
            continue
        elif choice == '98': 
            backup_system_data()
            input("\n✅ 操作完成，按回车返回...")
            continue
            
        # 懒加载 client，避免启动太慢
        # global client, APP_TOKEN # Remove syntax error
        
        current_client = None
        current_token = None
        
        if 'client' in globals() and client:
             current_client = client
        else:
             print(f"{Color.WARNING}🔄 正在连接飞书云端...{Color.ENDC}")
             current_client = init_clients()
             if not current_client: 
                 input(f"{Color.FAIL}❌ 初始化失败，按回车退出...{Color.ENDC}")
                 sys.exit(1)
             # Update global
             client = current_client
             
        current_token = APP_TOKEN
                 
        if choice == '00': one_click_daily_closing(current_client, current_token)
        elif choice == '1': smart_image_entry(current_client, current_token)
        elif choice == '2': smart_text_entry(current_client, current_token)
        elif choice == '27': register_voucher(current_client, current_token)
        elif choice == '3': 
             import_from_excel(current_client, current_token, None)
             
        elif choice == '4': 
             reconcile_bank_flow(current_client, current_token, None)
             
        elif choice == '5': generate_business_statement(current_client, current_token)
        elif choice == '23': reconcile_external_bill(current_client, current_token)
        elif choice == '24': manage_salary_flow(current_client, current_token)
        elif choice == '25': manage_invoice_flow(current_client, current_token)
        elif choice == '26': manage_processing_fee_flow(current_client, current_token)
        elif choice == '28': manage_small_tools(current_client, current_token)
        elif choice == '6': export_missing_tickets(current_client, current_token)
        
        elif choice == '7': generate_daily_html_report(current_client, current_token)
        elif choice == '8': daily_briefing(current_client, current_token)
        elif choice == '9': financial_health_check(current_client, current_token)
        elif choice == '10': smart_query_assistant(current_client, current_token)
        
        elif choice == '11': monthly_closing(current_client, current_token)
        elif choice == '22': annual_closing(current_client, current_token)
        elif choice == '12': calculate_depreciation(current_client, current_token)
        elif choice == '13': calculate_tax(current_client, current_token)
        elif choice == '14': manage_partner_aliases()
        elif choice == '15': manage_config_menu() # 原settings_menu改名或移除
        elif choice == '16': export_vouchers(current_client, current_token)
        elif choice == '17': smart_learning_mode(current_client, current_token)
        elif choice == '18': quick_search_ledger(current_client, current_token)
        elif choice == '19': backup_system_data(current_client, current_token)
        elif choice == '20': folder_monitor_mode(current_client, current_token)
        elif choice == '21': generate_annual_report_html(current_client, current_token)
        
        elif choice == '97': 
             manage_config_menu()
             
        elif choice == '99': 
             print("👋 再见！")
             break
        
        else:
            print(f"{Color.FAIL}❌ 无效选项{Color.ENDC}")
            
        input("\n✅ 操作完成，按回车返回主菜单...")

def create_processing_fee_table(client, app_token):
    """创建加工费明细表"""
    table_id = get_table_id_by_name(client, app_token, "加工费明细表")
    if table_id:
        log.info("✅ 加工费明细表已存在", extra={"solution": "无需创建"})
        return table_id

    log.info("🔨 正在创建加工费明细表...", extra={"solution": "请稍候"})
    
    req = CreateAppTableRequest.builder() \
        .app_token(app_token) \
        .request_body(CreateAppTableRequestBody.builder()
            .table(AppTableCreateHeader.builder()
                .name("加工费明细表")
                .fields([
                    AppTableField.builder().field_name("日期").type(FT.DATE).build(),
                    AppTableField.builder().field_name("往来单位").type(FT.TEXT).build(),
                    AppTableField.builder().field_name("品名").type(FT.TEXT).build(),
                    AppTableField.builder().field_name("规格").type(FT.TEXT).build(),
                    AppTableField.builder().field_name("类型").type(FT.SELECT).property(
                        AppTableFieldProperty.builder().options([
                            AppTableFieldPropertyOption.builder().name("支出-外协加工").build(),
                            AppTableFieldPropertyOption.builder().name("收入-加工服务").build()
                        ]).build()
                    ).build(),
                    AppTableField.builder().field_name("计价方式").type(FT.SELECT).property(
                        AppTableFieldProperty.builder().options([
                            AppTableFieldPropertyOption.builder().name("按件/个").build(),
                            AppTableFieldPropertyOption.builder().name("按米长").build(),
                            AppTableFieldPropertyOption.builder().name("按重量").build(),
                            AppTableFieldPropertyOption.builder().name("按平方").build()
                        ]).build()
                    ).build(),
                    AppTableField.builder().field_name("数量").type(FT.NUMBER).build(),
                    AppTableField.builder().field_name("单位").type(FT.TEXT).build(),
                    AppTableField.builder().field_name("单价").type(FT.NUMBER).property(
                        AppTableFieldProperty.builder().formatter("0.000").build() # 单价保留3位小数
                    ).build(),
                    AppTableField.builder().field_name("总金额").type(FT.NUMBER).property(
                        AppTableFieldProperty.builder().formatter("0.00").build()
                    ).build(),
                    AppTableField.builder().field_name("备注").type(FT.TEXT).build()
                ])
                .build())
            .build()) \
        .build()

    resp = client.bitable.v1.app_table.create(req)
    if resp.success():
        log.info("✅ 加工费明细表创建成功", extra={"solution": "无"})
        return resp.data.table_id
    else:
        log.error(f"❌ 创建失败: {resp.msg}", extra={"solution": "检查权限"})
        return None

def create_processing_price_table(client, app_token):
    """创建加工费价目表 (Price List)"""
    table_id = get_table_id_by_name(client, app_token, "加工费价目表")
    if table_id: return table_id

    log.info("🔨 正在创建加工费价目表...", extra={"solution": "请稍候"})
    
    req = CreateAppTableRequest.builder() \
        .app_token(app_token) \
        .request_body(CreateAppTableRequestBody.builder()
            .table(AppTableCreateHeader.builder()
                .name("加工费价目表")
                .fields([
                    AppTableField.builder().field_name("品名").type(FT.TEXT).build(),
                    AppTableField.builder().field_name("规格").type(FT.TEXT).build(),
                    AppTableField.builder().field_name("单位").type(FT.TEXT).build(),
                    AppTableField.builder().field_name("单价").type(FT.NUMBER).property(
                        AppTableFieldProperty.builder().formatter("0.000").build()
                    ).build(),
                    AppTableField.builder().field_name("备注").type(FT.TEXT).build()
                ])
                .build())
            .build()) \
        .build()

    resp = client.bitable.v1.app_table.create(req)
    if resp.success():
        log.info("✅ 加工费价目表创建成功", extra={"solution": "无"})
        return resp.data.table_id
    else:
        log.error(f"❌ 创建价目表失败: {resp.msg}", extra={"solution": "检查权限"})
        return None

def manage_price_list(client, app_token):
    """维护加工费价目表"""
    table_id = create_processing_price_table(client, app_token)
    if not table_id: return

    while True:
        print(f"\n{Color.CYAN}📋 加工费价目表管理{Color.ENDC}")
        print("1. 查看/搜索价目")
        print("2. 新增单价 (逐条)")
        print("3. 修改/删除单价")
        print("4. Excel 批量导入 (高效)")
        print("0. 返回")
        
        choice = input("👉 请选择: ").strip()
        
        if choice == '0': break
        
        if choice == '1':
            records = get_all_records(client, app_token, table_id)
            if not records:
                print("📭 暂无价目")
            else:
                print(f"\n{'品名':<15} | {'规格':<15} | {'单位':<6} | {'单价':<8} | {'备注'}")
                print("-" * 70)
                filter_kw = input("🔍 搜索关键词 (回车显示全部): ").strip()
                count = 0
                for r in records:
                    f = r.fields
                    if filter_kw and (filter_kw not in f.get('品名','') and filter_kw not in f.get('规格','')):
                        continue
                    print(f"{f.get('品名',''):<15} | {f.get('规格',''):<15} | {f.get('单位',''):<6} | {f.get('单价',0):<8} | {f.get('备注','')}")
                    count += 1
                print(f"共找到 {count} 条记录")
        
        elif choice == '2':
            print("\n➕ 新增价目")
            name = input("品名 (如: 铝型材/螺丝): ").strip()
            spec = input("规格 (如: 黑色氧化/周长20cm): ").strip()
            unit = input("单位 (如: 米, kg, 件): ").strip()
            try:
                price = float(input("单价 (元): ").strip())
            except:
                print("❌ 单价无效")
                continue
            remark = input("备注 (选填): ").strip()
            
            fields = {
                "品名": name,
                "规格": spec,
                "单位": unit,
                "单价": price,
                "备注": remark
            }
            
            req = CreateAppTableRecordRequest.builder() \
                .app_token(app_token) \
                .table_id(table_id) \
                .request_body(AppTableRecord.builder().fields(fields).build()) \
                .build()
                
            if client.bitable.v1.app_table_record.create(req).success():
                print("✅ 价目已保存")
            else:
                print("❌ 保存失败")

        elif choice == '3':
            # 修改/删除
            print("\n🛠️ 修改/删除单价")
            kw = input("🔍 请输入品名/规格关键词搜索: ").strip()
            if not kw: continue
            
            records = get_all_records(client, app_token, table_id)
            candidates = []
            for r in records:
                f = r.fields
                if kw in f.get('品名','') or kw in f.get('规格',''):
                    candidates.append(r)
            
            if not candidates:
                print("❌ 未找到相关记录")
                continue
                
            print(f"\n{'序号':<4} | {'品名':<15} | {'规格':<15} | {'单价':<8}")
            for i, r in enumerate(candidates):
                f = r.fields
                print(f"{i+1:<4} | {f.get('品名',''):<15} | {f.get('规格',''):<15} | {f.get('单价',0):<8}")
                
            sel = input("\n👉 请输入序号 (0取消): ").strip()
            if not sel or sel == '0': continue
            
            try:
                idx = int(sel) - 1
                if 0 <= idx < len(candidates):
                    target = candidates[idx]
                    print(f"\n已选中: {target.fields.get('品名')} {target.fields.get('规格')} (当前单价: {target.fields.get('单价')})")
                    action = input("👉 请选择操作: 1.修改单价  2.删除记录  0.取消: ").strip()
                    
                    if action == '1':
                        new_price = float(input("请输入新单价: ").strip())
                        
                        req = UpdateAppTableRecordRequest.builder() \
                            .app_token(app_token) \
                            .table_id(table_id) \
                            .record_id(target.record_id) \
                            .request_body(AppTableRecord.builder().fields({"单价": new_price}).build()) \
                            .build()
                            
                        if client.bitable.v1.app_table_record.update(req).success():
                            print("✅ 修改成功")
                        else:
                            print("❌ 修改失败")
                            
                    elif action == '2':
                        if input("⚠️ 确认删除吗? (y/n): ").lower() == 'y':
                            # 软删除日志记录 (模拟回收站)
                            try:
                                recycle_log = os.path.join(DATA_ROOT, "系统日志", "recycle_bin.jsonl")
                                with open(recycle_log, "a", encoding="utf-8") as f:
                                    log_entry = {
                                        "deleted_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                        "table": "加工费价目表",
                                        "record_id": target.record_id,
                                        "data": target.fields
                                    }
                                    f.write(json.dumps(log_entry, ensure_ascii=False) + "\n")
                                log.info(f"🗑️ 已移入回收站: {target.record_id}", extra={"solution": "查看 recycle_bin.jsonl"})
                            except: pass

                            req = DeleteAppTableRecordRequest.builder() \
                                .app_token(app_token) \
                                .table_id(table_id) \
                                .record_id(target.record_id) \
                                .build()
                            if client.bitable.v1.app_table_record.delete(req).success():
                                print("✅ 删除成功")
                            else:
                                print("❌ 删除失败")
            except Exception as e:
                print(f"❌ 操作错误: {e}")

        elif choice == '4':
            # Excel 导入
            print(f"\n{Color.OKBLUE}📂 Excel 批量导入价目表{Color.ENDC}")
            print("请准备 Excel 文件，包含以下列: [品名, 规格, 单位, 单价, 备注(选填)]")
            path = input("请输入文件路径 (直接回车扫描当前目录): ").strip()
            
            if not path:
                cands = [f for f in os.listdir('.') if '价目' in f and f.endswith('.xlsx')]
                if cands:
                    path = cands[0]
                    print(f"🔍 自动找到: {path}")
                else:
                    print("❌ 未找到文件")
                    continue
                    
            if not os.path.exists(path):
                print("❌ 文件不存在")
                continue
                
            try:
                df = pd.read_excel(path)
                required = ['品名', '规格', '单位', '单价']
                if not all(c in df.columns for c in required):
                    print(f"❌ 缺少必要列: {required}")
                    continue
                
                print(f"📄 读取到 {len(df)} 条记录，正在导入...")
                
                # 获取现有记录以进行排重/更新 (可选，这里简化为追加模式，或者先查后插)
                # 为了效率，暂用追加模式，用户需自行管理重复
                # 高级版：构建 map (name+spec -> record_id) 进行 upsert
                
                existing_map = {}
                print("🔄 正在同步现有数据以支持更新...")
                all_recs = get_all_records(client, app_token, table_id)
                if all_recs:
                    for r in all_recs:
                        key = f"{r.fields.get('品名')}_{r.fields.get('规格')}"
                        existing_map[key] = r.record_id
                
                batch_add = []
                update_count = 0
                
                for _, row in df.iterrows():
                    name = str(row['品名']).strip()
                    spec = str(row['规格']).strip()
                    key = f"{name}_{spec}"
                    
                    fields = {
                        "品名": name,
                        "规格": spec,
                        "单位": str(row['单位']).strip(),
                        "单价": float(row['单价']),
                        "备注": str(row.get('备注', ''))
                    }
                    
                    if key in existing_map:
                        # Update
                        rid = existing_map[key]
                        req = UpdateAppTableRecordRequest.builder() \
                            .app_token(app_token) \
                            .table_id(table_id) \
                            .record_id(rid) \
                            .request_body(AppTableRecord.builder().fields(fields).build()) \
                            .build()
                        client.bitable.v1.app_table_record.update(req)
                        update_count += 1
                        print(f"   🔄 更新: {name} {spec}")
                    else:
                        # Add
                        batch_add.append(AppTableRecord.builder().fields(fields).build())
                
                # Execute Batch Add
                if batch_add:
                    batch_size = 100
                    for i in range(0, len(batch_add), batch_size):
                        batch = batch_add[i:i+batch_size]
                        req = BatchCreateAppTableRecordRequest.builder() \
                            .app_token(app_token) \
                            .table_id(table_id) \
                            .request_body(BatchCreateAppTableRecordRequestBody.builder().records(batch).build()) \
                            .build()
                        client.bitable.v1.app_table_record.batch_create(req)
                        
                print(f"✅ 导入完成! 新增 {len(batch_add)} 条, 更新 {update_count} 条")
                
            except Exception as e:
                print(f"❌ 导入出错: {e}")



def archive_report(file_path):
    """自动归档报表"""
    try:
        if not os.path.exists(file_path): return
        
        # 归档路径: 财务数据/报表存档/YYYY年/MM月/
        now = datetime.now()
        archive_dir = os.path.join(DATA_ROOT, "报表存档", f"{now.year}年", f"{now.month:02d}月")
        if not os.path.exists(archive_dir):
            os.makedirs(archive_dir)
            
        fname = os.path.basename(file_path)
        dest = os.path.join(archive_dir, fname)
        
        # 复制而非移动，保留根目录副本方便用户立即查看 (或者移动并创建快捷方式? 还是复制简单)
        # 为了保持根目录整洁，建议移动。但用户刚生成可能想打开。
        # 策略：复制一份到归档，根目录保留。用户如果不清理，那是用户的事。
        # 或者：移动到归档，并在根目录打印路径。
        # 既然是 "存档"，应该是移动。但为了用户体验，保留一份在根目录 (或 "查询报告" 目录) 更好。
        # 现有的 "查询报告" 目录 (DATA_ROOT/查询报告) 似乎没怎么用。
        # 让我们把文件移动到 `财务数据/查询报告` 并在 `报表存档` 留底。
        # 简单点：只复制到 `报表存档`，根目录保留。
        
        import shutil
        shutil.copy2(file_path, dest)
        # print(f"💾 报表已归档至: {dest}")
    except: pass

def generate_customer_processing_report(client, app_token):
    """生成客户分品类加工费月报 (增强版: 拆分/标准价/归档)"""
    print(f"\n{Color.CYAN}📊 正在生成客户加工费月报...{Color.ENDC}")
    
    table_id = get_table_id_by_name(client, app_token, "加工费明细表")
    if not table_id:
        print("❌ 未找到加工费明细表")
        return

    # 预加载价目表 (Standard Price)
    price_map = {}
    try:
        pt_id = create_processing_price_table(client, app_token)
        if pt_id:
            p_recs = get_all_records(client, app_token, pt_id)
            if p_recs:
                for r in p_recs:
                    # Key: (品名, 规格) -> 单价
                    k = (r.fields.get('品名', '').strip(), r.fields.get('规格', '').strip())
                    v = float(r.fields.get('单价', 0))
                    price_map[k] = v
    except: pass

    # 选择月份 (智能默认: 1-10号默认上月, 否则本月)
    now = datetime.now()
    default_input = now.strftime("%Y-%m")
    if now.day <= 10:
        last_month_dt = now.replace(day=1) - timedelta(days=1)
        default_input = last_month_dt.strftime("%Y-%m")
        
    user_input = input(f"请输入查询月份 (YYYY-MM) 或年份 (YYYY) [{default_input}]: ").strip()
    if not user_input: user_input = default_input
    
    is_annual = False
    try:
        if len(user_input) == 4 and user_input.isdigit():
            is_annual = True
            year = int(user_input)
            start_ts = int(datetime(year, 1, 1).timestamp() * 1000)
            end_ts = int(datetime(year + 1, 1, 1).timestamp() * 1000)
            report_name = f"{year}年度客户加工费总表"
        else:
            start_dt = datetime.strptime(user_input, "%Y-%m")
            if start_dt.month == 12:
                end_dt = datetime(start_dt.year + 1, 1, 1)
            else:
                end_dt = datetime(start_dt.year, start_dt.month + 1, 1)
            
            start_ts = int(start_dt.timestamp() * 1000)
            end_ts = int(end_dt.timestamp() * 1000)
            report_name = f"客户加工费月报_{user_input}"
    except:
        print("❌ 日期格式错误")
        return

    # 加载别名映射
    aliases = {}
    if os.path.exists(FILE_PARTNER_ALIASES):
        try:
            with open(FILE_PARTNER_ALIASES, "r", encoding="utf-8") as f:
                aliases = json.load(f)
        except: pass

    # 拉取数据
    filter_cmd = f'AND(CurrentValue.[日期]>={start_ts}, CurrentValue.[日期]<{end_ts}, CurrentValue.[类型]="收入-加工服务")'
    records = get_all_records(client, app_token, table_id, filter_info=filter_cmd)
    
    if not records:
        print(f"📭 {user_input} 无加工费收入记录")
        return

    # 聚合数据
    # Key: (客户, 品名, 规格, 计价方式, 单位)
    data_map = {}
    
    # 额外聚合用于图表
    chart_cust_stats = {}
    chart_prod_stats = {}
    
    # 年度趋势数据 (仅年度模式用)
    # Key: 客户 -> {Month: Amount}
    annual_trend = {}
    
    for r in records:
        f = r.fields
        raw_cust = f.get("往来单位", "未知客户")
        # 别名清洗
        customer = aliases.get(raw_cust, raw_cust)
        
        product = f.get("品名", "未知品名")
        spec = f.get("规格", "-")
        unit = f.get("单位", "")
        pricing = f.get("计价方式", "按件/个")
        
        qty = float(f.get("数量", 0))
        amt = float(f.get("总金额", 0))
        ts = f.get("日期", 0)
        
        key = (customer, product, spec, pricing, unit)
        
        if key not in data_map:
            data_map[key] = {"qty": 0.0, "amt": 0.0}
        
        data_map[key]["qty"] += qty
        data_map[key]["amt"] += amt
        
        # 图表聚合
        chart_cust_stats[customer] = chart_cust_stats.get(customer, 0) + amt
        chart_prod_stats[product] = chart_prod_stats.get(product, 0) + amt
        
        if is_annual:
            month_str = datetime.fromtimestamp(ts/1000).strftime("%m月")
            if customer not in annual_trend: annual_trend[customer] = {}
            annual_trend[customer][month_str] = annual_trend[customer].get(month_str, 0) + amt

    # 生成报表数据
    report_data = []
    for k, v in data_map.items():
        avg_price = v["amt"] / v["qty"] if v["qty"] != 0 else 0
        
        # 标准价对比
        std_price = price_map.get((k[1], k[2]), 0) # (品名, 规格)
        diff_pct = 0.0
        if std_price > 0:
            diff_pct = (avg_price - std_price) / std_price
            
        report_data.append({
            "客户": k[0],
            "品名": k[1],
            "规格": k[2],
            "计价方式": k[3],
            "单位": k[4],
            "总数量": v["qty"],
            "平均单价": avg_price,
            "标准单价": std_price,
            "偏差%": diff_pct,
            "总金额": v["amt"]
        })
    
    # 转 DataFrame 并排序
    df = pd.DataFrame(report_data)
    df.sort_values(by=["客户", "总金额"], ascending=[True, False], inplace=True)
    
    # 导出
    fname = f"{report_name}_{datetime.now().strftime('%H%M')}.xlsx"
    
    # 计算合计行 (用于明细表)
    total_row = pd.DataFrame([{
        "客户": "合计",
        "品名": "-",
        "规格": "-",
        "计价方式": "-",
        "单位": "-",
        "总数量": df["总数量"].sum(),
        "平均单价": 0,
        "标准单价": 0,
        "偏差%": 0,
        "总金额": df["总金额"].sum()
    }])
    df_with_total = pd.concat([df, total_row], ignore_index=True)
    
    # 询问是否拆分文件
    split_files = False
    if not is_annual: # 年度报表通常不需要拆分月度明细，或者也可以拆
        split_files = input("📂 是否为每个客户生成独立文件? (y/n) [n]: ").strip().lower() == 'y'
    
    # 使用 openpyxl 进行美化导出
    with pd.ExcelWriter(fname, engine='openpyxl') as writer:
        from openpyxl.chart import BarChart, Reference, PieChart, LineChart
        
        # 0. 汇总看板 (Dashboard)
        # 准备数据
        dash_cust_df = pd.DataFrame(list(chart_cust_stats.items()), columns=["客户", "总金额"]).sort_values("总金额", ascending=False).head(10)
        dash_prod_df = pd.DataFrame(list(chart_prod_stats.items()), columns=["品名", "总金额"]).sort_values("总金额", ascending=False)
        
        dash_cust_df.to_excel(writer, sheet_name='汇总看板', index=False, startrow=0, startcol=0)
        dash_prod_df.to_excel(writer, sheet_name='汇总看板', index=False, startrow=0, startcol=4)
        
        ws_dash = writer.sheets['汇总看板']
        apply_excel_styles(ws_dash)
        
        # 插入图表 - 客户排行
        chart1 = BarChart()
        chart1.title = f"{user_input} 客户加工费 Top 10"
        chart1.y_axis.title = "金额 (元)"
        data1 = Reference(ws_dash, min_col=2, min_row=1, max_row=len(dash_cust_df)+1)
        cats1 = Reference(ws_dash, min_col=1, min_row=2, max_row=len(dash_cust_df)+1)
        chart1.add_data(data1, titles_from_data=True)
        chart1.set_categories(cats1)
        ws_dash.add_chart(chart1, "A15")
        
        # 插入图表 - 品类分布
        chart2 = PieChart()
        chart2.title = f"{user_input} 加工品类分布"
        data2 = Reference(ws_dash, min_col=6, min_row=1, max_row=len(dash_prod_df)+1)
        cats2 = Reference(ws_dash, min_col=5, min_row=2, max_row=len(dash_prod_df)+1)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats2)
        ws_dash.add_chart(chart2, "E15")
        
        # 年度模式额外图表：月度趋势
        if is_annual:
            # 构建透视数据: Top 5 客户的月度趋势
            top_customers = dash_cust_df["客户"].tolist()[:5]
            months = [f"{i:02d}月" for i in range(1, 13)]
            trend_data = []
            for cust in top_customers:
                row = {"客户": cust}
                for m in months:
                    row[m] = annual_trend.get(cust, {}).get(m, 0)
                trend_data.append(row)
            
            trend_df = pd.DataFrame(trend_data)
            trend_df.to_excel(writer, sheet_name='汇总看板', index=False, startrow=0, startcol=8) # 放在右边
            
            # 折线图
            chart3 = LineChart()
            chart3.title = "Top 5 客户年度月度走势"
            chart3.y_axis.title = "金额"
            chart3.x_axis.title = "月份"
            
            # 数据引用 (假设最多5个客户，12个月)
            # Row 1 is header. Data starts row 2. Cols I(9) to U(21)
            # Reference logic: min_col=9 (I), min_row=1 (Header), max_col=21 (U), max_row=len(trend_df)+1
            # But Series are rows.
            data3 = Reference(ws_dash, min_col=9, min_row=1, max_col=21, max_row=len(trend_df)+1)
            # cats3 should be the header row? No, cats are columns (Jan-Dec).
            # LineChart expects columns as categories usually if plotting rows as series.
            # Let's double check openpyxl behavior.
            # Usually: add_data(data, titles_from_data=True, from_rows=True)
            
            chart3.add_data(data3, titles_from_data=True, from_rows=True)
            # Categories are the first row (headers) from col 2 to end?
            cats3 = Reference(ws_dash, min_col=10, min_row=1, max_col=21, max_row=1)
            chart3.set_categories(cats3)
            
            ws_dash.add_chart(chart3, "A32")
            
            # Apply styles again to cover new data
            apply_excel_styles(ws_dash)

        # 1. 明细 Sheet (包含所有)
        sheet_title = '年度明细' if is_annual else '月报明细'
        df_with_total.to_excel(writer, index=False, sheet_name=sheet_title)
        apply_excel_styles(writer.sheets[sheet_title])
        
        # 格式化偏差列
        ws_detail = writer.sheets[sheet_title]
        for row in ws_detail.iter_rows(min_row=2, max_row=ws_detail.max_row, min_col=9, max_col=9): # I列是偏差%
            for cell in row:
                cell.number_format = '0.0%'
                if cell.value and (cell.value > 0.1 or cell.value < -0.1): # 偏差超过10%
                     cell.font = Font(color="9C0006")
                     cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # 2. 分客户 Sheet (Top 20 + Others)
        if not is_annual: # 年度报表数据量大，分Sheet可能太慢，且通常用于总览。如果需要可以开启。
            # 获取所有唯一客户
            all_customers = df["客户"].unique()
            
            # 简单策略：每个客户一个 Sheet
            for cust in all_customers:
                cust_df = df[df["客户"] == cust].copy()
                if cust_df.empty: continue
                
                # 添加该客户的合计
                total_qty = cust_df["总数量"].sum()
                total_amt = cust_df["总金额"].sum()
                
                cust_total_row = pd.DataFrame([{
                    "客户": "合计",
                    "品名": "-",
                    "规格": "-",
                    "计价方式": "-",
                    "单位": "-",
                    "总数量": total_qty,
                    "平均单价": 0,
                    "标准单价": 0,
                    "偏差%": 0,
                    "总金额": total_amt
                }])
                cust_df = pd.concat([cust_df, cust_total_row], ignore_index=True)
                
                # Sheet 名称清洗
                sheet_name = str(cust)[:30].replace(":", "").replace("\\", "").replace("/", "").replace("?", "").replace("*", "").replace("[", "").replace("]", "")
                
                cust_df.to_excel(writer, index=False, sheet_name=sheet_name)
                ws = writer.sheets[sheet_name]
                apply_excel_styles(ws)
                
                # 偏差高亮
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=9, max_col=9):
                    for cell in row:
                        cell.number_format = '0.0%'
                        if cell.value and abs(cell.value) > 0.1:
                             cell.font = Font(color="9C0006")
                             cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    print(f"✅ 报表已生成: {fname}")
    try: os.startfile(fname)
    except: pass
    
    # 拆分文件逻辑
    if split_files:
        split_dir = os.path.join(DATA_ROOT, "报表存档", f"{now.year}年", f"{now.month:02d}月", f"客户分表_{user_input}")
        if not os.path.exists(split_dir): os.makedirs(split_dir)
        
        print(f"📂 正在生成分客户独立文件 (保存至: {split_dir})...")
        for cust in all_customers:
            cust_df = df[df["客户"] == cust].copy()
            if cust_df.empty: continue
            
            # 合计
            total_qty = cust_df["总数量"].sum()
            total_amt = cust_df["总金额"].sum()
            cust_total_row = pd.DataFrame([{
                "客户": "合计",
                "品名": "-",
                "规格": "-",
                "计价方式": "-",
                "单位": "-",
                "总数量": total_qty,
                "平均单价": 0,
                "标准单价": 0,
                "偏差%": 0,
                "总金额": total_amt
            }])
            cust_df = pd.concat([cust_df, cust_total_row], ignore_index=True)
            
            # 保存
            safe_cust = str(cust).replace("/", "_").replace("\\", "_")
            c_fname = os.path.join(split_dir, f"{safe_cust}_{user_input}.xlsx")
            with pd.ExcelWriter(c_fname, engine='openpyxl') as c_writer:
                cust_df.to_excel(c_writer, index=False, sheet_name="对账单")
                apply_excel_styles(c_writer.sheets["对账单"])
        print(f"✅ 已生成 {len(all_customers)} 个独立对账单")
        try: os.startfile(split_dir)
        except: pass

    # 归档
    archive_report(fname)
    
    # 尝试发送到飞书
    print(f"{Color.CYAN}📤 正在推送报表到飞书...{Color.ENDC}")
    send_bot_message(fname, msg_type="file")

def generate_outsourcing_analysis_report(client, app_token):
    """生成外协费用分析表"""
    print(f"\n{Color.CYAN}📊 正在生成外协费用分析表...{Color.ENDC}")
    
    table_id = get_table_id_by_name(client, app_token, "加工费明细表")
    if not table_id: return

    # 选择年份 (分析通常按年或月)
    cur_year = datetime.now().year
    year_input = input(f"请输入查询年份 (YYYY) [{cur_year}]: ").strip()
    if not year_input: year_input = str(cur_year)
    
    try:
        year = int(year_input)
        start_ts = int(datetime(year, 1, 1).timestamp() * 1000)
        end_ts = int(datetime(year + 1, 1, 1).timestamp() * 1000)
    except:
        print("❌ 年份格式错误")
        return

    # 拉取数据
    filter_cmd = f'AND(CurrentValue.[日期]>={start_ts}, CurrentValue.[日期]<{end_ts}, CurrentValue.[类型]="支出-外协加工")'
    records = get_all_records(client, app_token, table_id, filter_info=filter_cmd)
    
    if not records:
        print(f"📭 {year}年 无外协加工记录")
        return

    # 聚合 1: 按供应商汇总
    supplier_stats = {}
    # 聚合 2: 按工艺(品名)汇总
    process_stats = {}
    # 聚合 3: 月度趋势
    monthly_stats = {}
    
    total_cost = 0.0
    
    for r in records:
        f = r.fields
        amt = float(f.get("总金额", 0))
        supplier = f.get("往来单位", "未知供应商")
        process = f.get("品名", "未知工艺") # 假设品名即工艺，如"喷砂"
        
        ts = f.get("日期", 0)
        month_str = datetime.fromtimestamp(ts/1000).strftime("%Y-%m")
        
        total_cost += amt
        
        supplier_stats[supplier] = supplier_stats.get(supplier, 0) + amt
        process_stats[process] = process_stats.get(process, 0) + amt
        monthly_stats[month_str] = monthly_stats.get(month_str, 0) + amt

    # 打印概览
    print(f"\n💰 {year}年 外协总费用: {total_cost:,.2f} 元")
    
    print(f"\n🏆 Top 5 外协供应商:")
    sorted_supp = sorted(supplier_stats.items(), key=lambda x: x[1], reverse=True)[:5]
    for k, v in sorted_supp:
        print(f"   - {k}: {v:,.2f} ({v/total_cost*100:.1f}%)")
        
    print(f"\n🔨 工艺分布:")
    sorted_proc = sorted(process_stats.items(), key=lambda x: x[1], reverse=True)
    for k, v in sorted_proc:
        print(f"   - {k}: {v:,.2f}")

    # 导出详细 Excel
    with pd.ExcelWriter(f"外协费用分析_{year}.xlsx", engine='openpyxl') as writer:
        from openpyxl.chart import BarChart, Reference, PieChart, LineChart
        
        # Sheet 1: 供应商汇总
        s_df = pd.DataFrame(list(supplier_stats.items()), columns=["供应商", "总金额"])
        s_df.sort_values("总金额", ascending=False, inplace=True)
        s_df.to_excel(writer, sheet_name="供应商排行", index=False)
        
        # Sheet 1 Chart
        ws1 = writer.sheets["供应商排行"]
        chart1 = BarChart()
        chart1.title = "供应商费用排行"
        chart1.y_axis.title = "金额 (元)"
        chart1.x_axis.title = "供应商"
        data1 = Reference(ws1, min_col=2, min_row=1, max_row=len(s_df)+1)
        cats1 = Reference(ws1, min_col=1, min_row=2, max_row=len(s_df)+1)
        chart1.add_data(data1, titles_from_data=True)
        chart1.set_categories(cats1)
        ws1.add_chart(chart1, "D2")
        
        # Sheet 2: 工艺汇总
        p_df = pd.DataFrame(list(process_stats.items()), columns=["工艺类型", "总金额"])
        p_df.sort_values("总金额", ascending=False, inplace=True)
        p_df.to_excel(writer, sheet_name="工艺分布", index=False)
        
        # Sheet 2 Chart
        ws2 = writer.sheets["工艺分布"]
        chart2 = PieChart()
        chart2.title = "外协工艺费用分布"
        data2 = Reference(ws2, min_col=2, min_row=1, max_row=len(p_df)+1)
        cats2 = Reference(ws2, min_col=1, min_row=2, max_row=len(p_df)+1)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats2)
        ws2.add_chart(chart2, "D2")
        
        # Sheet 3: 月度趋势
        m_df = pd.DataFrame(list(monthly_stats.items()), columns=["月份", "总金额"])
        m_df.sort_values("月份", inplace=True)
        m_df.to_excel(writer, sheet_name="月度趋势", index=False)
        
        # Sheet 3 Chart
        ws3 = writer.sheets["月度趋势"]
        chart3 = LineChart()
        chart3.title = "月度费用趋势"
        chart3.y_axis.title = "金额"
        chart3.x_axis.title = "月份"
        data3 = Reference(ws3, min_col=2, min_row=1, max_row=len(m_df)+1)
        cats3 = Reference(ws3, min_col=1, min_row=2, max_row=len(m_df)+1)
        chart3.add_data(data3, titles_from_data=True)
        chart3.set_categories(cats3)
        ws3.add_chart(chart3, "D2")
        
        # Sheet 4: 原始明细
        raw_data = []
        for r in records:
            f = r.fields
            raw_data.append({
                "日期": datetime.fromtimestamp(f.get("日期")/1000).strftime("%Y-%m-%d"),
                "供应商": f.get("往来单位"),
                "工艺": f.get("品名"),
                "规格": f.get("规格"),
                "数量": f.get("数量"),
                "单位": f.get("单位"),
                "单价": f.get("单价"),
                "总金额": f.get("总金额"),
                "备注": f.get("备注")
            })
        pd.DataFrame(raw_data).to_excel(writer, sheet_name="原始明细", index=False)
        
        # 统一美化所有Sheet
        for sheet_name in writer.sheets:
            apply_excel_styles(writer.sheets[sheet_name])

    print(f"✅ 详细分析表已生成: 外协费用分析_{year}.xlsx")
    try: os.startfile(f"外协费用分析_{year}.xlsx")
    except: pass
    
    # 归档
    archive_report(f"外协费用分析_{year}.xlsx")

    # 尝试发送到飞书
    print(f"{Color.CYAN}📤 正在推送报表到飞书...{Color.ENDC}")
    send_bot_message(f"外协费用分析_{year}.xlsx", msg_type="file")

def manage_settlement(client, app_token):
    """结算状态管理 (Mark as Paid)"""
    print(f"\n{Color.CYAN}💰 结算管理 (AR/AP){Color.ENDC}")
    table_id = get_table_id_by_name(client, app_token, "加工费明细表")
    if not table_id:
        print("❌ 未找到加工费明细表")
        return

    # 1. 统计未结算金额
    print("🔄 正在统计未结算金额...")
    
    # Filter: Status != "已结算"
    records = get_all_records(client, app_token, table_id)
    if not records:
        print("📭 暂无记录")
        return
        
    unpaid_map = {} # Cust -> Amount
    unpaid_records = []
    
    for r in records:
        f = r.fields
        status = f.get("结算状态", "未结算")
        if status == "已结算": continue
        
        cust = f.get("往来单位", "未知")
        amt = float(f.get("总金额", 0))
        # Filter out 0 amount
        if amt == 0: continue
        
        unpaid_map[cust] = unpaid_map.get(cust, 0) + amt
        unpaid_records.append(r)
        
    if not unpaid_map:
        print("✅ 所有账单均已结算！")
        return
        
    # Show Top 10
    sorted_cust = sorted(unpaid_map.items(), key=lambda x: x[1], reverse=True)
    
    print("\n📊 欠款排行榜 (Top 10):")
    for i, (c, amt) in enumerate(sorted_cust[:10]):
        print(f"   {i+1}. {c}: {Color.FAIL}{amt:,.2f} 元{Color.ENDC}")
        
    print(f"\n   >> 总未结算金额: {sum(unpaid_map.values()):,.2f} 元")
    
    # Actions
    print("\n操作选项:")
    print("1. 按客户批量结算 (Mark Customer as Paid)")
    print("2. 按月份批量结算 (Mark Month as Paid)")
    print("0. 返回")
    
    op = input("👉 请选择: ").strip()
    
    if op == '1':
        target = input("请输入客户名 (关键词): ").strip()
        if not target: return
        
        # Filter
        matches = [c for c in unpaid_map.keys() if target in c]
        if not matches:
            print("❌ 未找到匹配客户")
            return
            
        if len(matches) > 1:
            print(f"🔍 匹配到多个客户: {matches}")
            target = input("👉 请输入完整客户名确认: ").strip()
            if target not in matches: return
        else:
            target = matches[0]
            
        # Confirm
        total = unpaid_map[target]
        print(f"\n准备将 {Color.BOLD}{target}{Color.ENDC} 的 {len([r for r in unpaid_records if r.fields.get('往来单位')==target])} 笔记录标记为已结算。")
        print(f"涉及金额: {total:,.2f} 元")
        
        if input("❓ 确认执行? (y/n): ").strip().lower() == 'y':
            # Batch Update
            batch_recs = []
            for r in unpaid_records:
                if r.fields.get("往来单位") == target:
                    batch_recs.append(AppTableRecord.builder().record_id(r.record_id).fields({"结算状态": "已结算"}).build())
            
            # Execute Batch
            # Split into 100
            count = 0
            for i in range(0, len(batch_recs), 100):
                batch = batch_recs[i:i+100]
                req = BatchUpdateAppTableRecordRequest.builder() \
                    .app_token(app_token) \
                    .table_id(table_id) \
                    .request_body(BatchUpdateAppTableRecordRequestBody.builder().records(batch).build()) \
                    .build()
                resp = client.bitable.v1.app_table_record.batch_update(req)
                if resp.success():
                    count += len(batch)
            print(f"✅ 成功结算 {count} 笔记录")

    elif op == '2':
        month_str = input("请输入月份 (YYYY-MM): ").strip()
        try:
            target_dt = datetime.strptime(month_str, "%Y-%m")
            # Filter
            batch_recs = []
            total_amt = 0
            for r in unpaid_records:
                ts = r.fields.get("日期", 0)
                rdt = datetime.fromtimestamp(ts/1000)
                if rdt.year == target_dt.year and rdt.month == target_dt.month:
                    batch_recs.append(AppTableRecord.builder().record_id(r.record_id).fields({"结算状态": "已结算"}).build())
                    total_amt += float(r.fields.get("总金额", 0))
            
            if not batch_recs:
                print("❌ 该月份无未结算记录")
                return
                
            print(f"\n准备将 {month_str} 的 {len(batch_recs)} 笔记录标记为已结算。")
            print(f"涉及金额: {total_amt:,.2f} 元")
            
            if input("❓ 确认执行? (y/n): ").strip().lower() == 'y':
                 # Execute Batch
                count = 0
                for i in range(0, len(batch_recs), 100):
                    batch = batch_recs[i:i+100]
                    req = BatchUpdateAppTableRecordRequest.builder() \
                        .app_token(app_token) \
                        .table_id(table_id) \
                        .request_body(BatchUpdateAppTableRecordRequestBody.builder().records(batch).build()) \
                        .build()
                    resp = client.bitable.v1.app_table_record.batch_update(req)
                    if resp.success():
                        count += len(batch)
                print(f"✅ 成功结算 {count} 笔记录")
                
        except:
            print("❌ 日期格式错误")

def import_processing_records_from_excel(client, app_token):
    """批量导入加工费记录 (从 Excel)"""
    table_id = get_table_id_by_name(client, app_token, "加工费明细表")
    if not table_id: 
        print("❌ 未找到加工费明细表")
        return

    print(f"\n{Color.CYAN}📥 批量导入加工费记录{Color.ENDC}")
    print("💡 请准备 Excel 文件，包含以下列 (表头名称必须包含关键词):")
    print("   - 日期 (默认当天)")
    print("   - 客户/供应商 (关键词: 客户, 单位, 往来)")
    print("   - 品名 (关键词: 品名, 产品, 工艺)")
    print("   - 规格 (关键词: 规格, 尺寸)")
    print("   - 数量 (关键词: 数量, 件数, 重量)")
    print("   - 单价 (关键词: 单价, 价格)")
    print("   - 类型 (关键词: 类型 -> 收入/支出, 可选)")
    
    file_path = select_file_interactively("*.xlsx", "请选择加工单 Excel")
    if not file_path: return

    try:
        df = read_excel_smart(file_path) # Use existing smart reader
        if df.empty:
            print("❌ 文件为空或无法读取")
            return
            
        print(f"📄 读取到 {len(df)} 条数据，准备导入...")
        
        # 字段映射
        records = []
        success_count = 0
        
        # 预加载别名
        aliases = {}
        if os.path.exists(FILE_PARTNER_ALIASES):
            try:
                with open(FILE_PARTNER_ALIASES, "r", encoding="utf-8") as f:
                    aliases = json.load(f)
            except: pass

        # 预加载价目表 (用于自动填充单价)
        price_map = {}
        try:
            pt_id = create_processing_price_table(client, app_token)
            if pt_id:
                p_recs = get_all_records(client, app_token, pt_id)
                for r in p_recs:
                    pk = (r.fields.get('品名', '').strip(), r.fields.get('规格', '').strip())
                    price_map[pk] = float(r.fields.get('单价', 0))
        except: pass

        for idx, row in df.iterrows():
            # 智能提取字段
            date_val = None
            for col in df.columns:
                if "日期" in str(col) or "时间" in str(col):
                    try:
                        date_val = pd.to_datetime(row[col])
                        break
                    except: pass
            if not date_val: date_val = datetime.now()
            
            # 客户
            partner = ""
            for col in df.columns:
                if any(k in str(col) for k in ["客户", "单位", "往来", "供应商"]):
                    partner = str(row[col]).strip()
                    break
            # 应用别名
            partner = aliases.get(partner, partner)
            
            # 品名
            product = ""
            for col in df.columns:
                if any(k in str(col) for k in ["品名", "产品", "工艺", "名称"]):
                    product = str(row[col]).strip()
                    break
                    
            # 规格
            spec = "-"
            for col in df.columns:
                if "规格" in str(col) or "尺寸" in str(col):
                    spec = str(row[col]).strip()
                    break
            
            # 数量
            qty = 0.0
            for col in df.columns:
                if any(k in str(col) for k in ["数量", "件数", "重量"]):
                    try: qty = float(row[col])
                    except: pass
                    break
            
            # 单价 (如果Excel里没有，尝试从价目表获取)
            price = 0.0
            found_price = False
            for col in df.columns:
                if "单价" in str(col) or "价格" in str(col):
                    try: 
                        val = float(row[col])
                        if val > 0:
                            price = val
                            found_price = True
                    except: pass
                    break
            
            # 自动补全单价与异常检测
            std_price = 0.0
            price_remark = ""
            
            # 1. 尝试获取标准价
            std_price = price_map.get((product, spec), 0.0)
            if std_price == 0: std_price = price_map.get((product, ""), 0.0)
            
            # 2. 补全单价
            if not found_price:
                price = std_price
                if price > 0: price_remark = " (自动匹配单价)"
            
            # 3. 异常检测 (如果有输入单价且与标准价偏差大)
            elif std_price > 0:
                diff_pct = abs(price - std_price) / std_price
                if diff_pct > 0.2: # 偏差超过 20%
                    price_remark = f" (⚠️ 价格异常: {price} vs 标准{std_price})"

            # 计价方式与单位自动推断 (新增)
            pricing_mode = "按件/只/个" # 默认
            unit = "件"
            
            # 尝试从Excel列读取单位
            for col in df.columns:
                 if "单位" in str(col):
                     val = str(row[col]).strip()
                     if val: unit = val
                     break
            
            # 根据单位推断计价方式
            if unit in ['kg', '公斤', '吨', 'g']:
                pricing_mode = "按重量"
            elif unit in ['m', '米', 'cm']:
                pricing_mode = "按米长"
            elif unit in ['m2', 'm²', '平方', '平米']:
                pricing_mode = "按平方"
            
            # 类型 (默认收入)
            record_type = "收入-加工服务"
            for col in df.columns:
                if "类型" in str(col):
                    val = str(row[col]).strip()
                    if "支出" in val or "外协" in val:
                        record_type = "支出-外协加工"
                    break
            
            if not partner or not product:
                continue # Skip invalid
                
            fields = {
                "日期": int(date_val.timestamp() * 1000),
                "往来单位": partner,
                "品名": product,
                "规格": spec,
                "类型": record_type,
                "计价方式": pricing_mode, # New
                "单位": unit, # New
                "数量": qty,
                "单价": price,
                "总金额": round(qty * price, 2),
                "结算状态": "未结算", # 默认为未结算
                "开票状态": "未开票",
                "备注": "批量导入" + price_remark
            }
            records.append(AppTableRecord.builder().fields(fields).build())
            
        # 预览前5条 (包含异常提示)
        if records:
            print(f"\n{Color.CYAN}👀 导入预览 (前5条):{Color.ENDC}")
            for i, r in enumerate(records[:5]):
                f = r.fields
                d_str = datetime.fromtimestamp(f["日期"]/1000).strftime("%Y-%m-%d")
                
                # 高亮异常备注
                remark = f['备注']
                if "⚠️" in remark:
                    remark = f"{Color.FAIL}{remark}{Color.ENDC}"
                elif "自动" in remark:
                    remark = f"{Color.OKGREEN}{remark}{Color.ENDC}"
                    
                print(f"   {i+1}. {d_str} | {f['往来单位']} | {f['品名']} | {f['单价']} | {remark}")
            
            # 统计异常数量
            abnormal_count = sum(1 for r in records if "⚠️" in r.fields.get("备注", ""))
            if abnormal_count > 0:
                print(f"\n{Color.WARNING}⚠️ 检测到 {abnormal_count} 条记录价格异常 (偏差 > 20%){Color.ENDC}")
            
            if input(f"\n❓ 确认导入共 {len(records)} 条数据? (y/n): ").strip().lower() != 'y':
                print("❌ 已取消导入")
                return

        # 批量写入
        if records:
            batch_size = 100
            for i in range(0, len(records), batch_size):
                batch = records[i:i+batch_size]
                req = BatchCreateAppTableRecordRequest.builder() \
                    .app_token(app_token) \
                    .table_id(table_id) \
                    .request_body(BatchCreateAppTableRecordRequestBody.builder().records(batch).build()) \
                    .build()
                resp = client.bitable.v1.app_table_record.batch_create(req)
                if resp.success():
                    success_count += len(batch)
                    print(f"✅ 已导入 {success_count}/{len(records)} 条")
                else:
                    print(f"❌ 导入失败: {resp.msg}")
                    
            # 导入后，询问是否学习新价格
            learn_prices = input("🎓 是否将导入的新品名/价格自动学习到【价目表】? (y/n) [n]: ").strip().lower() == 'y'
            if learn_prices:
                learn_new_prices(client, app_token, records)
        else:
            print("⚠️ 未解析到有效数据")

    except Exception as e:
        print(f"❌ 导入异常: {e}")

def learn_new_prices(client, app_token, records):
    """自动学习新价格"""
    pt_id = create_processing_price_table(client, app_token)
    if not pt_id: return
    
    # 获取现有价格
    existing_map = {} # (name, spec) -> price
    p_recs = get_all_records(client, app_token, pt_id)
    if p_recs:
        for r in p_recs:
            k = (r.fields.get('品名', '').strip(), r.fields.get('规格', '').strip())
            existing_map[k] = float(r.fields.get('单价', 0))
            
    # 分析新记录
    new_prices = {} # (name, spec) -> price
    for r in records:
        f = r.fields
        name = f.get('品名', '').strip()
        spec = f.get('规格', '').strip()
        price = float(f.get('单价', 0))
        
        if not name or price <= 0: continue
        
        k = (name, spec)
        if k not in existing_map:
            # 简单的策略：直接取最新的价格
            new_prices[k] = price
            
    if not new_prices:
        print("✅ 没有发现新品名或规格")
        return
        
    print(f"🔍 发现 {len(new_prices)} 个新价格组合，正在学习...")
    
    # 批量添加
    batch_recs = []
    for (name, spec), price in new_prices.items():
        fields = {
            "品名": name,
            "规格": spec,
            "单位": "件", # 默认
            "单价": price,
            "备注": f"自动学习 ({datetime.now().strftime('%Y-%m-%d')})"
        }
        batch_recs.append(AppTableRecord.builder().fields(fields).build())
        
    # Execute Batch
    count = 0
    for i in range(0, len(batch_recs), 100):
        batch = batch_recs[i:i+100]
        req = BatchCreateAppTableRecordRequest.builder() \
            .app_token(app_token) \
            .table_id(pt_id) \
            .request_body(BatchCreateAppTableRecordRequestBody.builder().records(batch).build()) \
            .build()
        resp = client.bitable.v1.app_table_record.batch_create(req)
        if resp.success():
            count += len(batch)
            
    print(f"✅ 已自动添加 {count} 条新价格记录到价目表")

def generate_delivery_note(client, app_token):
    """生成送货单 (Delivery Note)"""
    print(f"\n{Color.CYAN}🚚 生成送货单 (Delivery Note){Color.ENDC}")
    print("--------------------------------")
    print("功能：选择未打印的加工单，生成送货单供司机送货和客户签收。")
    
    table_id = get_table_id_by_name(client, app_token, "加工费明细表")
    if not table_id: return
    
    # 1. 选择客户
    # 获取最近 30 天有过记录的客户供选择
    print("⏳ 正在获取最近客户列表...")
    now = datetime.now()
    start_ts = int((now - timedelta(days=30)).timestamp() * 1000)
    filter_cmd = f'AND(CurrentValue.[日期]>={start_ts}, CurrentValue.[类型]="收入-加工服务")'
    
    # 只取部分字段提高速度
    recs = get_all_records(client, app_token, table_id, filter_info=filter_cmd)
    
    partners = set()
    for r in recs:
        p = r.fields.get("往来单位", "").strip()
        if p: partners.add(p)
        
    sorted_partners = sorted(list(partners))
    
    if not sorted_partners:
        print("❌ 最近无加工记录")
        return
        
    print("\n📋 最近往来单位:")
    for i, p in enumerate(sorted_partners):
        print(f"  {i+1}. {p}")
        
    p_choice = input("\n👉 请选择客户序号 (或输入名称): ").strip()
    target_partner = ""
    if p_choice.isdigit() and 1 <= int(p_choice) <= len(sorted_partners):
        target_partner = sorted_partners[int(p_choice)-1]
    else:
        target_partner = p_choice
        
    if not target_partner: return
    
    # 2. 拉取该客户未打印送货单的记录
    days_str = input("查询最近多少天记录 (默认 30): ").strip()
    days = 30
    if days_str.isdigit(): days = int(days_str)
    
    print(f"\n🔍 正在查询 【{target_partner}】 最近 {days} 天的加工记录...")
    
    start_ts = int((now - timedelta(days=days)).timestamp() * 1000)
    filter_p = f'AND(CurrentValue.[往来单位]="{target_partner}", CurrentValue.[日期]>={start_ts}, CurrentValue.[类型]="收入-加工服务")'
    p_recs = get_all_records(client, app_token, table_id, filter_info=filter_p)
    
    if not p_recs:
        print("📭 无近期记录")
        return
        
    # 按日期倒序
    p_recs.sort(key=lambda x: x.fields.get("日期", 0), reverse=True)
    
    selected_recs = []
    
    while True:
        print(f"\n📋 可选记录 (共 {len(p_recs)} 条):")
        print(f"{'序号':<4} | {'日期':<10} | {'品名/规格':<20} | {'数量':<8} | {'金额':<10} | {'备注'}")
        print("-" * 80)
        
        for i, r in enumerate(p_recs):
            f = r.fields
            d_str = datetime.fromtimestamp(f.get("日期", 0)/1000).strftime("%m-%d")
            desc = f"{f.get('品名','')} {f.get('规格','')}"
            qty = f"{f.get('数量',0)}{f.get('单位','')}"
            amt = f"{f.get('总金额',0):.2f}"
            rem = f.get("备注", "")
            
            # Check mark
            mark = "[ ]"
            if r in selected_recs: mark = "[x]"
            
            print(f"{i+1:<4} {mark} | {d_str:<10} | {desc:<20} | {qty:<8} | {amt:<10} | {rem}")
            
        print("-" * 80)
        print("操作: 输入序号选择/取消 (如 '1 3 5')，输入 'a' 全选，输入 'ok' 生成")
        
        op = input("👉 请输入: ").strip().lower()
        
        if op == 'ok':
            if not selected_recs:
                print("❌ 未选择任何记录")
                continue
            break
        elif op == 'a':
            if len(selected_recs) == len(p_recs):
                selected_recs = [] # 全取消
            else:
                selected_recs = list(p_recs) # 全选
        else:
            # Parse numbers
            try:
                idxs = [int(x) for x in op.split()]
                for idx in idxs:
                    if 1 <= idx <= len(p_recs):
                        target = p_recs[idx-1]
                        if target in selected_recs:
                            selected_recs.remove(target)
                        else:
                            selected_recs.append(target)
            except:
                pass

    # 补充送货信息
    driver_info = input("🚚 送货司机/车牌号 (选填): ").strip()
    contact_info = input("📞 联系人/电话 (选填): ").strip()

    # 3. 生成送货单 HTML
    print("\n📄 正在生成送货单...")
    delivery_no = f"DN{datetime.now().strftime('%Y%m%d%H%M')}"
    
    total_qty = 0
    total_amt = 0.0
    items_html = ""
    
    # 统计不同单位的数量
    unit_totals = {}
    
    # Sort selected by date
    selected_recs.sort(key=lambda x: x.fields.get("日期", 0))
    
    for idx, r in enumerate(selected_recs):
        f = r.fields
        d_str = datetime.fromtimestamp(f.get("日期", 0)/1000).strftime("%Y-%m-%d")
        q = float(f.get("数量", 0))
        u = f.get("单位", "")
        a = float(f.get("总金额", 0))
        
        total_qty += q
        total_amt += a
        
        if u not in unit_totals: unit_totals[u] = 0
        unit_totals[u] += q
        
        bg = "#f9f9f9" if idx % 2 == 0 else "#fff"
        
        items_html += f"""
        <tr style="background-color:{bg}">
            <td>{idx+1}</td>
            <td>{f.get('品名','')}</td>
            <td>{f.get('规格','')}</td>
            <td style="text-align:right">{q}</td>
            <td style="text-align:center">{u}</td>
            <td style="text-align:right">{a:.2f}</td>
            <td>{f.get('备注','')}</td>
        </tr>
        """
        
    # 生成合计字符串
    total_desc_parts = []
    for u, q in unit_totals.items():
        total_desc_parts.append(f"{q:.2f} {u}")
    total_desc = " + ".join(total_desc_parts)
        
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <title>送货单 {delivery_no}</title>
        <style>
            body {{ font-family: 'SimHei', 'Microsoft YaHei', sans-serif; max-width: 800px; margin: 0 auto; padding: 20px; }}
            .header {{ text-align: center; margin-bottom: 20px; border-bottom: 2px solid #000; padding-bottom: 10px; }}
            .title {{ font-size: 24px; font-weight: bold; letter-spacing: 5px; }}
            .sub-title {{ margin-top: 5px; font-size: 14px; }}
            .company-name {{ font-size: 18px; margin-bottom: 5px; font-weight: bold; }}
            .info-row {{ display: flex; justify-content: space-between; margin-bottom: 15px; font-size: 14px; }}
            table {{ width: 100%; border-collapse: collapse; margin-bottom: 20px; border: 1px solid #000; }}
            th, td {{ border: 1px solid #000; padding: 8px; font-size: 14px; }}
            th {{ background-color: #eee; text-align: center; }}
            .footer {{ margin-top: 40px; display: flex; justify-content: space-between; font-size: 14px; }}
            .sign {{ border-top: 1px solid #000; width: 150px; display: inline-block; margin-left: 10px; }}
        </style>
    </head>
    <body>
        <div class="header">
            <div class="company-name">五金氧化加工中心</div>
            <div class="title">送 货 单</div>
            <div class="sub-title">Delivery Note</div>
        </div>
        
        <div class="info-row">
            <div style="flex: 1">客户名称: <b>{target_partner}</b></div>
            <div style="flex: 1">单号: {delivery_no}</div>
        </div>
        <div class="info-row">
            <div style="flex: 1">送货日期: {datetime.now().strftime('%Y-%m-%d')}</div>
            <div style="flex: 1">司机/车牌: {driver_info}</div>
        </div>
        
        <table>
            <thead>
                <tr>
                    <th width="5%">序号</th>
                    <th width="25%">品名</th>
                    <th width="20%">规格</th>
                    <th width="10%">数量</th>
                    <th width="10%">单位</th>
                    <th width="15%">金额</th>
                    <th width="15%">备注</th>
                </tr>
            </thead>
            <tbody>
                {items_html}
                <tr style="font-weight:bold; background-color:#eee">
                    <td colspan="3" style="text-align:center">合计</td>
                    <td colspan="2" style="text-align:center">{total_desc}</td>
                    <td style="text-align:right">{total_amt:.2f}</td>
                    <td></td>
                </tr>
            </tbody>
        </table>
        
        <div class="footer">
            <div>
                送货人签字: <span class="sign"></span>
            </div>
            <div>
                客户签收: <span class="sign"></span>
            </div>
        </div>
        
        <div style="margin-top: 20px; font-size: 12px; color: #666; text-align: center;">
            * 请核对数量及规格，如有异议请当面提出。白联:存根 红联:客户 黄联:回单
            <br>{f"联系方式: {contact_info}" if contact_info else ""}
        </div>
    </body>
    </html>
    """
    
    save_dir = os.path.join(DATA_ROOT, "送货单")
    if not os.path.exists(save_dir): os.makedirs(save_dir)
    
    fname = os.path.join(save_dir, f"送货单_{target_partner}_{delivery_no}.html")
    with open(fname, "w", encoding="utf-8") as f:
        f.write(html)
        
    print(f"✅ 送货单已生成: {Color.UNDERLINE}{fname}{Color.ENDC}")
    try: os.startfile(fname)
    except: pass
    
    # 4. 可选：回写备注 (标记已送货)
    if input("👉 是否在备注中标记 '已出单'? (y/n) [y]: ").strip().lower() != 'n':
        print("⏳ 正在更新记录...")
        batch_updates = []
        for r in selected_recs:
            old_rem = r.fields.get("备注", "")
            if "已出单" not in old_rem:
                new_rem = f"{old_rem} [已出单{delivery_no}]".strip()
                batch_updates.append(AppTableRecord.builder().record_id(r.record_id).fields({"备注": new_rem}).build())
        
        if batch_updates:
            # Batch update logic
             for i in range(0, len(batch_updates), 100):
                 req = BatchUpdateAppTableRecordRequest.builder().app_token(app_token).table_id(table_id).request_body(BatchUpdateAppTableRecordRequestBody.builder().records(batch_updates[i:i+100]).build()).build()
                 client.bitable.v1.app_table_record.batch_update(req)
             print("✅ 已标记完成")

def manage_processing_fee_flow(client, app_token):
    """加工费管理 (Menu 26)"""
    # 概览数据
    print(f"\n{Color.CYAN}🔧 加工费管理{Color.ENDC}")
    
    # 尝试加载本月数据概览
    try:
        table_id = get_table_id_by_name(client, app_token, "加工费明细表")
        if table_id:
            now = datetime.now()
            start_ts = int(datetime(now.year, now.month, 1).timestamp() * 1000)
            filter_cmd = f'AND(CurrentValue.[日期]>={start_ts})'
            recs = get_all_records(client, app_token, table_id, filter_info=filter_cmd)
            
            income = 0.0
            expense = 0.0
            unpaid = 0.0
            if recs:
                for r in recs:
                    f = r.fields
                    amt = float(f.get("总金额", 0))
                    if "收入" in f.get("类型", ""): income += amt
                    else: expense += amt
                    
                    if f.get("结算状态") != "已结算":
                        unpaid += amt
            
            print(f"{Color.BOLD}📊 本月概览 ({now.month}月):{Color.ENDC}")
            print(f"   💰 收入: {Color.OKGREEN}{income:,.2f}{Color.ENDC} | 💸 支出: {Color.FAIL}{expense:,.2f}{Color.ENDC} | 🧾 待结算: {Color.WARNING}{unpaid:,.2f}{Color.ENDC}")
    except: pass

    print("-----------------------------------")
    print("1. 批量导入加工单 (Excel)")
    print("2. 导出加工费明细 (Excel)")
    print("3. 维护价目表 (Price List)")
    print(f"{Color.OKGREEN}4. 生成客户加工费月报{Color.ENDC}")
    print(f"{Color.OKGREEN}5. 外协费用分析表{Color.ENDC}")
    print(f"{Color.OKGREEN}6. 结算管理 (AR/AP) [新]{Color.ENDC}")
    print(f"{Color.OKBLUE}8. 同步到总账 (月末汇总) [新]{Color.ENDC}")
    print(f"{Color.OKBLUE}9. 开票管理 (Mark as Invoiced) [新]{Color.ENDC}")
    print(f"{Color.OKGREEN}10. 批量生成客户对账单 (明细版) [新]{Color.ENDC}")
    print(f"{Color.OKBLUE}11. 客户收款登记 (按实际发生) [新]{Color.ENDC}")
    print(f"{Color.OKBLUE}12. 供应商付款登记 (按实际发生) [新]{Color.ENDC}")
    print(f"{Color.CYAN}13. 生成送货单 (Delivery Note) [新]{Color.ENDC}")
    print("7. 登记加工费 (手动)")
    print("0. 返回")
    
    choice = input("\n👉 请选择 (0-13): ").strip()
    
    if choice == '0': return
    
    if choice == '13':
        generate_delivery_note(client, app_token)
        return
    
    if choice == '1':
        import_processing_records_from_excel(client, app_token)
        return

    if choice == '3':
        manage_price_list(client, app_token)
        return
        
    if choice == '4':
        generate_customer_processing_report(client, app_token)
        return

    if choice == '5':
        generate_outsourcing_analysis_report(client, app_token)
        return

    if choice == '6':
        manage_settlement(client, app_token)
        return

    if choice == '8':
        sync_processing_fee_to_ledger(client, app_token)
        return
        
    if choice == '9':
        manage_invoice_status(client, app_token)
        return
        
    if choice == '10':
        batch_generate_customer_statements(client, app_token)
        return
        
    if choice == '11':
        manage_processing_payment(client, app_token)
        return
        
    if choice == '12':
        manage_supplier_payment(client, app_token)
        return

    table_id = create_processing_fee_table(client, app_token) # 确保表存在
    if not table_id: return
    
    # 确保字段存在 (Migration)
    ensure_processing_fee_fields(client, app_token, table_id)
    
    if choice == '2':
        # 导出逻辑
        records = get_all_records(client, app_token, table_id)
        if not records:
            print("❌ 暂无数据")
            return
            
        data = []
        for r in records:
            f = r.fields
            ts = f.get("日期", 0)
            d_str = datetime.fromtimestamp(ts/1000).strftime("%Y-%m-%d") if ts else "-"
            data.append({
                "日期": d_str,
                "往来单位": f.get("往来单位", ""),
                "品名": f.get("品名", ""),
                "规格": f.get("规格", ""),
                "类型": f.get("类型", ""),
                "计价方式": f.get("计价方式", ""),
                "数量": f.get("数量", 0),
                "单位": f.get("单位", ""),
                "单价": f.get("单价", 0),
                "总金额": f.get("总金额", 0),
                "结算状态": f.get("结算状态", "未结算"),
                "开票状态": f.get("开票状态", "未开票"),
                "备注": f.get("备注", "")
            })
        
        df = pd.DataFrame(data)
        fname = f"加工费明细_{datetime.now().strftime('%Y%m%d%H%M')}.xlsx"
        df.to_excel(fname, index=False)
        print(f"✅ 已导出: {fname}")
        try: os.startfile(fname)
        except: pass
        return

    if choice == '7':
        # 登记逻辑
        # 预加载价目表以支持智能学习
        print("🔄 正在加载价目表以支持智能学习...")
        pt_id = create_processing_price_table(client, app_token)
        price_list_map = {} # (name, spec) -> record
        if pt_id:
            p_recs = get_all_records(client, app_token, pt_id)
            if p_recs:
                for r in p_recs:
                    key = (r.fields.get('品名', '').strip(), r.fields.get('规格', '').strip())
                    price_list_map[key] = r
        
        # 记忆变量，用于批量录入时的默认值
        last_date = datetime.now().strftime('%Y-%m-%d')
        last_partner = ""
        last_type_choice = "1"
        
        # [New] 构建历史单价缓存 (Smart Price History) - 优化版 (使用文件缓存)
        print("⏳ 正在加载历史单价缓存...")
        history_price_map = {} # (partner, name, spec) -> {price, unit, date}
        
        cache_file = os.path.join(DATA_ROOT, "cache", "price_history.json")
        last_cache_ts = 0
        
        # 1. Load from file
        if os.path.exists(cache_file):
            try:
                with open(cache_file, "r", encoding="utf-8") as f:
                    cached_data = json.load(f)
                    # Convert list keys back to tuple if needed, but JSON keys are strings
                    # We store as list of dicts or dict with string keys
                    # Let's store as list: [{"key": [p,n,s], "val": {...}}]
                    for item in cached_data:
                        k = tuple(item["key"])
                        history_price_map[k] = item["val"]
                    
                    # Get max timestamp
                    for v in history_price_map.values():
                        if v['date'] > last_cache_ts: last_cache_ts = v['date']
            except: pass
            
        # 2. Fetch incremental updates
        # Filter: Date > last_cache_ts
        filter_cmd = None
        if last_cache_ts > 0:
            filter_cmd = f'CurrentValue.[日期]>{last_cache_ts}'
            
        # Only fetch necessary fields to speed up
        # Unfortunately get_all_records fetches all fields by default unless optimized client used
        # But filter helps.
        new_recs = get_all_records(client, app_token, table_id, filter_info=filter_cmd)
        
        if new_recs:
            print(f"📥 同步了 {len(new_recs)} 条新记录")
            updated = False
            for r in new_recs:
                f = r.fields
                p = f.get("往来单位", "").strip()
                n = f.get("品名", "").strip()
                s = f.get("规格", "").strip()
                pr = float(f.get("单价", 0))
                u = f.get("单位", "")
                d = f.get("日期", 0)
                
                if p and n and pr > 0:
                    key = (p, n, s)
                    # Update if newer
                    if key not in history_price_map or d > history_price_map[key]['date']:
                        history_price_map[key] = {
                            'price': pr,
                            'unit': u,
                            'date': d,
                            'd_str': datetime.fromtimestamp(d/1000).strftime("%Y-%m-%d") if d else ""
                        }
                        updated = True
            
            # 3. Save back to cache if updated
            if updated:
                try:
                    if not os.path.exists(os.path.dirname(cache_file)):
                        os.makedirs(os.path.dirname(cache_file))
                    
                    # Convert to serializable format
                    to_save = []
                    for k, v in history_price_map.items():
                        to_save.append({"key": list(k), "val": v})
                        
                    with open(cache_file, "w", encoding="utf-8") as f:
                        json.dump(to_save, f)
                except: pass
        
        print(f"✅ 历史单价准备就绪 (共 {len(history_price_map)} 条条目)")

        # 批次累计变量
        batch_total_amount = 0.0
        batch_count = 0
        batch_mode = False

        while True:
            # 显示批次累计信息
            if batch_count > 0:
                print(f"\n{Color.HEADER}📊 当前批次累计: {batch_total_amount:,.2f} 元 (共 {batch_count} 笔){Color.ENDC}")
            
            # 定义类型映射
            type_map = {'1': '支出-外协加工', '2': '收入-加工服务'}
            
            if batch_mode:
                print(f"\n{Color.OKBLUE}🔒 批量录入模式 (输入 0 退出当前模式){Color.ENDC}")
                print(f"   📅 {last_date} | 🏢 {last_partner} | 🔖 {type_map.get(last_type_choice)}")
                date_str = last_date
                partner = last_partner
                p_type = type_map.get(last_type_choice, "支出-外协加工")
            else:
                print(f"\n{Color.BOLD}📝 新增加工费记录 (输入 0 退出){Color.ENDC}")
                
                # 日期
                date_str = input(f"日期 (默认 {last_date}): ").strip()
                if date_str == '0': break
                if not date_str: date_str = last_date
                else: last_date = date_str
                
                # 往来单位
                p_prompt = f"往来单位 (默认 '{last_partner}'): " if last_partner else "往来单位: "
                partner = input(p_prompt).strip()
                if not partner and last_partner:
                    partner = last_partner
                
                if not partner: 
                    print("❌ 必须输入单位")
                    continue
                last_partner = partner
                    
                # 类型
                def_type_name = type_map.get(last_type_choice, "支出-外协加工")
                print(f"类型: 1. 支出-外协加工  2. 收入-加工服务 (默认 {last_type_choice}.{def_type_name})")
                t_choice = input("👉 请选择 (1/2): ").strip()
                if not t_choice: t_choice = last_type_choice
                else: last_type_choice = t_choice
                
                p_type = type_map.get(t_choice, "支出-外协加工")
                
                # 询问是否进入批量模式
                if input("⚡ 是否锁定表头进入批量极速模式? (y/n) [n]: ").strip().lower() == 'y':
                    batch_mode = True
                    print(f"{Color.OKGREEN}✅ 已进入批量模式 (锁定: {last_date} | {partner} | {p_type}){Color.ENDC}")
                    print(f"{Color.CYAN}💡 提示: 输入 '0' 可退出批量模式{Color.ENDC}")
            
            # 统一录入/搜索逻辑
            selected_record = None
            price = 0.0
            calc_remark = ""
            product_name = ""
            product_spec = ""
            
            # 批量模式下显示锁定状态
            if batch_mode:
                 print(f"\n{Color.CYAN}🔒 [批量] {last_date} | {partner} | {p_type}{Color.ENDC}")
            
            print(f"{Color.CYAN}🔍 品名录入 (支持关键词搜索，输入 0 返回):{Color.ENDC}")
            p_input = input("👉 品名/关键词: ").strip()
            
            if p_input == '0':
                if batch_mode:
                    batch_mode = False
                    print("🔓 已退出批量模式")
                    continue
                else:
                    break
            
            if not p_input:
                print("❌ 品名不能为空")
                continue

            # 智能搜索
            matches = []
            if price_list_map:
                for r in price_list_map.values():
                    # 简单匹配: 输入包含在品名或规格中，或者品名包含输入
                    p_val = r.fields.get('品名', '')
                    s_val = r.fields.get('规格', '')
                    if p_input in p_val or p_input in s_val:
                        matches.append(r)
            
            # 显示匹配项
            if matches:
                print(f"💡 找到 {len(matches)} 个匹配项:")
                # 按匹配度排序 (完全匹配优先)
                matches.sort(key=lambda x: 0 if x.fields.get('品名') == p_input else 1)
                
                for i, m in enumerate(matches[:5]): # 最多显示5个
                    f = m.fields
                    print(f"   {i+1}. {f.get('品名')} {f.get('规格')} @ {f.get('单价')}元/{f.get('单位')}")
                
                sel = input("👉 选择序号 (回车跳过，直接使用输入值): ").strip()
                if sel.isdigit() and 0 < int(sel) <= len(matches):
                    selected_record = matches[int(sel)-1]
                    print(f"✅ 已选择: {selected_record.fields.get('品名')}")
                else:
                    product_name = p_input # 用户坚持使用输入值
            else:
                product_name = p_input

            if selected_record:
                # 自动推断模式
                f = selected_record.fields
                product_name = f.get('品名', '')
                product_spec = f.get('规格', '')
                unit = f.get('单位', '件')
                
                if unit in ['米', 'm']: m_choice = '2'
                elif unit in ['kg', 'kg', '公斤']: m_choice = '3'
                elif unit in ['平方', 'm²', 'm2']: m_choice = '4'
                else: m_choice = '1'
                
                base_unit = unit
                price = float(f.get('单价', 0))
                
                # 优先使用历史单价 (如果存在)
                hist_key = (partner, product_name, product_spec)
                if hist_key in history_price_map:
                    hist = history_price_map[hist_key]
                    print(f"💡 发现历史成交价: {Color.OKGREEN}{hist['price']}元/{hist['unit']}{Color.ENDC} ({hist['d_str']})")
                    # 如果历史单位和当前推断单位一致，使用历史价格
                    if hist['unit'] == unit:
                        price = hist['price']
                        print(f"   已自动采用历史价格")
                
                calc_remark = f"[价目] {product_name} {product_spec}"
                
                try:
                    qty_str = input(f"数量 ({base_unit}) [支持算式]: ").strip()
                    if '*' in qty_str or '+' in qty_str:
                        try:
                            qty = float(eval(qty_str, {"__builtins__": None}, {}))
                            print(f"   🧮 计算结果: {qty}")
                        except:
                            print("❌ 算式无效")
                            continue
                    else:
                        qty = float(qty_str)
                except:
                    print("❌ 数量无效")
                    continue
            else:
                # 手动模式 - 询问品名/规格 (Smart Learning Key)
                # product_name 已经在上面赋值了
                product_spec = input("规格 (Spec): ").strip()
                
                # 计价方式
                print("计价方式:")
                print("1. 按件/只/个 (Quantity)")
                print("2. 按米长 (Length)")
                print("3. 按重量 (Weight)")
                print("4. 按平方 (Area)")
                
                m_choice = input("👉 请选择 (1-4): ").strip()
                modes = {'1': '按件/只/个', '2': '按米长', '3': '按重量', '4': '按平方'}
                mode_name = modes.get(m_choice, '按件/只/个')
                
                units = {'1': '件', '2': '米', '3': 'kg', '4': 'm²'}
                base_unit = units.get(m_choice, '单位')
                
                # 数量
                qty_val = 0.0
                
                # 数量助手 (针对氧化厂特殊场景)
                if m_choice == '3': # 按重量
                    print(f"\n{Color.CYAN}⚖️ 重量计算助手:{Color.ENDC}")
                    print("   A. 直接输入重量 (kg)")
                    print("   B. 通过【总长 x 米重】计算 (理论重)")
                    q_choice = input("   👉 请选择 (A/B) [默认A]: ").strip().upper()
                    
                    if q_choice == 'B':
                        try:
                            l_val = float(eval(input("   请输入总长度 (米) [支持算式]: ").strip(), {"__builtins__": None}, {}))
                            w_val = float(input("   请输入米重 (kg/m): ").strip())
                            qty_val = round(l_val * w_val, 3)
                            print(f"   ⚖️ 计算重量: {l_val}m * {w_val}kg/m = {qty_val}kg")
                            calc_remark += f" [理论重: {l_val}m*{w_val}]"
                        except:
                             print("❌ 计算错误")
                             continue
                    else:
                         try: qty_val = float(eval(input(f"数量 ({base_unit}) [支持算式]: ").strip(), {"__builtins__": None}, {}))
                         except: continue

                elif m_choice == '2': # 按米长
                     try: qty_val = float(eval(input(f"数量 ({base_unit}) [支持算式]: ").strip(), {"__builtins__": None}, {}))
                     except: continue
                
                elif m_choice == '4': # 按平方
                    print(f"\n{Color.CYAN}📐 面积计算助手:{Color.ENDC}")
                    print("   A. 直接输入面积 (m²)")
                    print("   B. 通过【长 x 宽 x 数量】计算")
                    q_choice = input("   👉 请选择 (A/B) [默认A]: ").strip().upper()
                    
                    if q_choice == 'B':
                         try:
                             l = float(input("   长 (mm): "))
                             w = float(input("   宽 (mm): "))
                             n = float(input("   数量 (件): "))
                             area = (l * w * n) / 1000000.0 # mm^2 to m^2
                             qty_val = round(area, 3)
                             print(f"   📐 计算面积: {qty_val} m²")
                             calc_remark += f" [尺寸: {l}x{w}mm * {n}件]"
                         except: continue
                    else:
                         try: qty_val = float(eval(input(f"数量 ({base_unit}) [支持算式]: ").strip(), {"__builtins__": None}, {}))
                         except: continue
                else:
                    # 默认按件
                    try: qty_val = float(eval(input(f"数量 ({base_unit}) [支持算式]: ").strip(), {"__builtins__": None}, {}))
                    except: continue

                qty = qty_val
                
                # 单价
                # 尝试从历史记录获取默认单价
                def_price = 0.0
                hist_key = (partner, product_name, product_spec)
                if hist_key in history_price_map:
                    hist = history_price_map[hist_key]
                    print(f"💡 发现历史成交价: {Color.OKGREEN}{hist['price']}元/{hist['unit']}{Color.ENDC} ({hist['d_str']})")
                    if hist['unit'] == base_unit:
                        def_price = hist['price']
                
                try:
                    p_in = input(f"单价 (元/{base_unit}) [默认 {def_price}]: ").strip()
                    if not p_in:
                        price = def_price
                    else:
                        price = float(p_in)
                except:
                    print("❌ 单价无效")
                    continue
            
            total = round(qty * price, 2)
            print(f"🧮 自动计算总额: {qty} * {price} = {total}")
            
            # 重新获取 mode_name 如果是从价目表选择的 (因为 mode_name 之前可能没设置)
            modes = {'1': '按件/只/个', '2': '按米长', '3': '按重量', '4': '按平方'}
            mode_name = modes.get(m_choice, '按件/只/个')

            print(f"💰 总金额: {total:,.2f} 元")
            
            r_prompt = f"备注 (默认 '{calc_remark}'): " if calc_remark else "备注: "
            remark = input(r_prompt).strip()
            if not remark and calc_remark:
                remark = calc_remark
            elif remark and calc_remark:
                remark = f"{calc_remark} {remark}"
            
            # 确认保存
            print(f"\n即将保存: [{date_str}] {partner} - {product_name} {product_spec} - {mode_name} {qty}{base_unit} * {price} = {total}")
            if input("确认保存? (y/n): ").strip().lower() == 'y':
                fields = {
                    "日期": int(pd.to_datetime(date_str).timestamp() * 1000),
                    "往来单位": partner,
                    "品名": product_name,
                    "规格": product_spec,
                    "类型": p_type,
                    "计价方式": mode_name,
                    "数量": qty,
                    "单位": base_unit,
                    "单价": price,
                    "总金额": total,
                    "备注": remark
                }
                
                req = CreateAppTableRecordRequest.builder() \
                    .app_token(app_token) \
                    .table_id(table_id) \
                    .request_body(AppTableRecord.builder().fields(fields).build()) \
                    .build()
                    
                resp = client.bitable.v1.app_table_record.create(req)
                if resp.success():
                    print("✅ 保存成功！")
                    batch_total_amount += total
                    batch_count += 1
                    
                    # --- 智能学习逻辑 (Smart Learning) ---
                    if product_name:
                        key = (product_name, product_spec)
                        existing_rec = price_list_map.get(key)
                        
                        if existing_rec:
                            # 存在，检查价格差异
                            old_price = float(existing_rec.fields.get('单价', 0))
                            if abs(old_price - price) > 0.0001:
                                print(f"\n{Color.WARNING}💡 价格变动提醒: 价目表单价为 {old_price}，本次录入 {price}{Color.ENDC}")
                                if input("   👉 是否更新价目表? (y/n) [n]: ").strip().lower() == 'y':
                                    req = UpdateAppTableRecordRequest.builder() \
                                        .app_token(app_token) \
                                        .table_id(pt_id) \
                                        .record_id(existing_rec.record_id) \
                                        .request_body(AppTableRecord.builder().fields({"单价": price}).build()) \
                                        .build()
                                    if client.bitable.v1.app_table_record.update(req).success():
                                        print("   ✅ 价目表已更新")
                                        # Update local cache
                                        existing_rec.fields['单价'] = price
                                        price_list_map[key] = existing_rec
                        else:
                            # 不存在，提示新增
                            print(f"\n{Color.OKGREEN}💡 发现新项目: {product_name} {product_spec} @ {price}{Color.ENDC}")
                            if input("   👉 是否添加到价目表? (y/n) [y]: ").strip().lower() != 'n':
                                fields = {
                                    "品名": product_name,
                                    "规格": product_spec,
                                    "单位": base_unit,
                                    "单价": price,
                                    "备注": "自动学习"
                                }
                                req = CreateAppTableRecordRequest.builder() \
                                    .app_token(app_token) \
                                    .table_id(pt_id) \
                                    .request_body(AppTableRecord.builder().fields(fields).build()) \
                                    .build()
                                resp = client.bitable.v1.app_table_record.create(req)
                                if resp.success():
                                    print("   ✅ 已添加到价目表")
                                    # Update local cache
                                    price_list_map[key] = resp.data.record

                else:
                    print(f"❌ 保存失败: {resp.msg}")



# 发票管理流程 (新)
def debt_collection_assistant(client, app_token):
    """应收账款催收助手 (Debt Collection Assistant)"""
    print(f"\n{Color.FAIL}📢 应收账款催收助手 (Debt Collection){Color.ENDC}")
    print("--------------------------------")
    print("功能: 扫描所有客户的欠款情况，进行账龄分析 (0-30/30-60/60-90/>90天)，并生成催款话术。")
    
    table_id = get_table_id_by_name(client, app_token, "日常台账表")
    if not table_id: return
    
    print("⏳ 正在计算全量客户余额 (可能需要一点时间)...")
    
    # 1. Calculate All Balances
    recs = get_all_records(client, app_token, table_id)
    
    cust_receipts = {} # Customer -> Total Receipt
    last_pay_date = {} # Customer -> Timestamp
    
    # 2. Ledger Receipts (Payment)
    for r in recs:
        t = r.fields.get("业务类型", "")
        if t == "收款":
            p = r.fields.get("往来单位费用", "").strip()
            amt = float(r.fields.get("实际收付金额", 0))
            d = r.fields.get("记账日期", 0)
            if p:
                if p not in cust_receipts: cust_receipts[p] = 0.0
                cust_receipts[p] += amt
                
                if p not in last_pay_date or d > last_pay_date[p]:
                    last_pay_date[p] = d
            
    # 3. Processing Fees (Debt)
    pf_table_id = get_table_id_by_name(client, app_token, "加工费明细表")
    if not pf_table_id: return
    
    print("⏳ 正在拉取加工费记录并分析账龄...")
    pf_recs = get_all_records(client, app_token, pf_table_id)
    
    cust_debts = {} # Customer -> Total Fee
    last_biz_date = {} # Customer -> Timestamp
    partner_records = {} # Customer -> List of records
    
    all_partners = set()
    
    for r in pf_recs:
        t = r.fields.get("类型", "")
        if t == "收入-加工服务":
            p = r.fields.get("往来单位", "").strip()
            amt = float(r.fields.get("总金额", 0))
            d = r.fields.get("日期", 0)
            if p:
                if p not in cust_debts: cust_debts[p] = 0.0
                cust_debts[p] += amt
                all_partners.add(p)
                
                if p not in last_biz_date or d > last_biz_date[p]:
                    last_biz_date[p] = d
                    
                if p not in partner_records: partner_records[p] = []
                partner_records[p].append({
                    "amt": amt,
                    "date": d
                })

    # 4. Calculate Aging
    final_list = []
    now_ts = int(datetime.now().timestamp() * 1000)
    
    print(f"\n📋 欠款客户清单 (按欠款金额排序):")
    # Header with Aging
    print(f"{'排名':<4} | {'客户名称':<10} | {'欠款余额':<10} | {'0-30天':<8} | {'30-60天':<8} | {'60-90天':<8} | {'>90天':<8}")
    print("-" * 90)
    
    for p in all_partners:
        debt = cust_debts.get(p, 0)
        paid = cust_receipts.get(p, 0)
        balance = debt - paid
        
        if balance > 10: # Ignore small change
            # Aging Logic
            aging = {"0-30": 0.0, "30-60": 0.0, "60-90": 0.0, "90+": 0.0}
            
            # Sort records Newest -> Oldest
            p_recs = partner_records.get(p, [])
            p_recs.sort(key=lambda x: x["date"], reverse=True)
            
            remaining_bal = balance
            
            for r in p_recs:
                if remaining_bal <= 0.01: break
                
                amt = r["amt"]
                # allocate
                this_amt = min(remaining_bal, amt)
                remaining_bal -= this_amt
                
                # check age
                r_date = r["date"]
                days_diff = (now_ts - r_date) / (1000 * 3600 * 24)
                
                if days_diff <= 30: aging["0-30"] += this_amt
                elif days_diff <= 60: aging["30-60"] += this_amt
                elif days_diff <= 90: aging["60-90"] += this_amt
                else: aging["90+"] += this_amt
            
            # Handle edge case: if balance remains (maybe from opening balance not in records)
            if remaining_bal > 0.01:
                aging["90+"] += remaining_bal # Assume very old
            
            l_biz = last_biz_date.get(p, 0)
            l_pay = last_pay_date.get(p, 0)
            
            final_list.append({
                "name": p,
                "balance": balance,
                "aging": aging,
                "last_pay": datetime.fromtimestamp(l_pay/1000).strftime("%Y-%m-%d") if l_pay else "-"
            })
            
    final_list.sort(key=lambda x: x["balance"], reverse=True)
    
    for i, item in enumerate(final_list):
        a = item["aging"]
        print(f"{i+1:<4} | {item['name']:<10} | {Color.FAIL}{item['balance']:<10,.0f}{Color.ENDC} | "
              f"{a['0-30']:<8,.0f} | {a['30-60']:<8,.0f} | {a['60-90']:<8,.0f} | {Color.FAIL}{a['90+']:<8,.0f}{Color.ENDC}")
        
    print("-" * 90)
    print(f"💰 总欠款金额: {sum(x['balance'] for x in final_list):,.2f}")
    
    # 5. Generate Reminder
    while True:
        print(f"\n{Color.OKBLUE}功能操作:{Color.ENDC}")
        print(" - 输入序号 (如 1): 生成微信催款话术")
        print(" - 输入 h+序号 (如 h1): 生成HTML正式对账单 (发给客户)")
        print(" - 输入 0: 返回")
        
        idx_str = input("👉 请选择: ").strip().lower()
        if idx_str == '0': break
        
        is_html = False
        if idx_str.startswith('h'):
            is_html = True
            idx_str = idx_str[1:]
        
        try:
            idx = int(idx_str) - 1
            if 0 <= idx < len(final_list):
                target = final_list[idx]
                name = target["name"]
                bal = target["balance"]
                l_pay = target["last_pay"]
                ag = target["aging"]
                
                if is_html:
                    # Generate HTML Statement
                    recs = partner_records.get(name, [])
                    recs.sort(key=lambda x: x["date"], reverse=True)
                    
                    html = f"""
                    <!DOCTYPE html>
                    <html>
                    <head>
                        <meta charset="utf-8">
                        <title>{name} - 对账单</title>
                        <style>
                            body {{ font-family: 'Segoe UI', sans-serif; padding: 40px; max-width: 800px; margin: 0 auto; background: #fff; }}
                            .header {{ text-align: center; border-bottom: 2px solid #333; padding-bottom: 20px; margin-bottom: 30px; }}
                            .info {{ display: flex; justify-content: space-between; margin-bottom: 20px; }}
                            table {{ width: 100%; border-collapse: collapse; margin-bottom: 30px; }}
                            th, td {{ border: 1px solid #ddd; padding: 12px; text-align: left; }}
                            th {{ background-color: #f8f9fa; }}
                            .total {{ text-align: right; font-size: 20px; font-weight: bold; color: #c0392b; }}
                            .footer {{ margin-top: 50px; text-align: center; color: #7f8c8d; font-size: 14px; }}
                            @media print {{ body {{ padding: 0; }} }}
                        </style>
                    </head>
                    <body>
                        <div class="header">
                            <h1>对 账 单 (Statement)</h1>
                            <p>日期: {datetime.now().strftime('%Y-%m-%d')}</p>
                        </div>
                        <div class="info">
                            <div><strong>客户名称:</strong> {name}</div>
                            <div><strong>截止日期:</strong> {datetime.now().strftime('%Y-%m-%d')}</div>
                        </div>
                        
                        <table>
                            <thead>
                                <tr>
                                    <th>日期</th>
                                    <th>摘要/业务</th>
                                    <th style="text-align:right">金额 (元)</th>
                                </tr>
                            </thead>
                            <tbody>
                    """
                    
                    # Show recent 20 records
                    for r in recs[:20]:
                        d_str = datetime.fromtimestamp(r["date"]/1000).strftime("%Y-%m-%d")
                        html += f"""
                                <tr>
                                    <td>{d_str}</td>
                                    <td>加工费</td>
                                    <td style="text-align:right">{r['amt']:,.2f}</td>
                                </tr>
                        """
                        
                    html += f"""
                            </tbody>
                        </table>
                        
                        <div class="total">
                            当前欠款余额: ¥ {bal:,.2f}
                        </div>
                        
                        <div style="margin-top: 20px; border: 1px dashed #ccc; padding: 15px; background: #fffcf5;">
                            <strong>账龄分析:</strong><br>
                            0-30天: {ag['0-30']:,.2f} | 30-60天: {ag['30-60']:,.2f} | 60-90天: {ag['60-90']:,.2f} | >90天: {ag['90+']:,.2f}
                        </div>

                        <div class="footer">
                            <p>请核对上述账单，如有疑问请及时联系。</p>
                            <p>谢谢您的支持！</p>
                        </div>
                    </body>
                    </html>
                    """
                    
                    save_path = os.path.join(DATA_ROOT, f"对账单_{name}_{datetime.now().strftime('%Y%m%d')}.html")
                    with open(save_path, "w", encoding="utf-8") as f:
                        f.write(html)
                    print(f"✅ 对账单已生成: {Color.UNDERLINE}{save_path}{Color.ENDC}")
                    try: os.startfile(save_path)
                    except: pass
                
                else:
                    print(f"\n📱 {name} 催款微信模板:")
                    print("--------------------------------")
                    print(f"{name}老板您好，")
                    print(f"打扰了，这边核对了一下账单，截止到今天，贵司还有 {bal:,.2f} 元加工费未结。")
                    
                    # Add aging detail if long overdue
                    long_overdue = ag["60-90"] + ag["90+"]
                    if long_overdue > 0:
                        print(f"其中 {long_overdue:,.0f} 元已超过2个月，请重点关注一下。")
                    
                    if l_pay != "-":
                        print(f"(上次回款日期: {l_pay})")
                    print(f"麻烦您抽空安排一下，谢谢支持！🙏")
                    print("--------------------------------")
                    print("💡 提示: 选中上方文字 -> 右键复制 -> 发送微信")
            else:
                print("❌ 序号无效")
        except:
            print("❌ 输入无效")

def generate_monthly_visual_report(client, app_token):
    """生成月度经营分析图表报告 (Visual Report)"""
    print(f"\n{Color.OKBLUE}📊 生成月度经营分析报告 (Visual){Color.ENDC}")
    
    month_str = input("👉 请输入月份 (YYYY-MM) [默认本月]: ").strip()
    if not month_str: month_str = datetime.now().strftime("%Y-%m")
    
    try:
        start_dt = datetime.strptime(month_str, "%Y-%m")
        if start_dt.month == 12:
            end_dt = datetime(start_dt.year + 1, 1, 1)
        else:
            end_dt = datetime(start_dt.year, start_dt.month + 1, 1)
        
        start_ts = int(start_dt.timestamp() * 1000)
        end_ts = int(end_dt.timestamp() * 1000)
    except:
        print("❌ 日期格式错误")
        return
        
    # Fetch Data
    # 1. Ledger (Income/Expense/Cost)
    l_id = get_table_id_by_name(client, app_token, "日常台账表")
    if not l_id: return
    
    print("⏳ 正在分析财务数据...")
    filter_l = f'AND(CurrentValue.[记账日期]>={start_ts}, CurrentValue.[记账日期]<{end_ts})'
    l_recs = get_all_records(client, app_token, l_id, filter_info=filter_l)
    
    total_inc = 0.0
    total_exp = 0.0
    exp_cats = {} # Category -> Amount
    
    # Energy Cost Analysis
    energy_cost = 0.0
    outsourced_cost = 0.0
    
    for r in l_recs:
        t = r.fields.get("业务类型", "")
        amt = float(r.fields.get("实际收付金额", 0))
        cat = r.fields.get("费用归类", "其他")
        
        if t == "收款":
            total_inc += amt
        elif t in ["付款", "费用"]:
            total_exp += amt
            if cat not in exp_cats: exp_cats[cat] = 0.0
            exp_cats[cat] += amt
            
            # Identify Energy Costs
            if "电" in cat or "水" in cat or "气" in cat:
                energy_cost += amt
            
            # Identify Outsourced Costs
            if "外协" in cat:
                outsourced_cost += amt
            
    # 2. Production (Processing Fee)
    pf_id = get_table_id_by_name(client, app_token, "加工费明细表")
    print("⏳ 正在分析生产数据...")
    filter_p = f'AND(CurrentValue.[日期]>={start_ts}, CurrentValue.[日期]<{end_ts}, CurrentValue.[类型]="收入-加工服务")'
    p_recs = get_all_records(client, app_token, pf_id, filter_info=filter_p)
    
    prod_qty = 0.0
    cust_sales = {} # Customer -> Amount
    cust_qtys = {} # Customer -> Quantity
    
    for r in p_recs:
        q = float(r.fields.get("数量", 0))
        amt = float(r.fields.get("总金额", 0))
        cust = r.fields.get("往来单位", "散客")
        
        prod_qty += q 
        if cust not in cust_sales: 
            cust_sales[cust] = 0.0
            cust_qtys[cust] = 0.0
            
        cust_sales[cust] += amt
        cust_qtys[cust] += q
    
    # Calculate Financial Ratios
    cost_rate = (total_exp / total_inc * 100) if total_inc > 0 else 0
    energy_rate = (energy_cost / total_inc * 100) if total_inc > 0 else 0
    outsourced_rate = (outsourced_cost / total_inc * 100) if total_inc > 0 else 0
    
    # Keep Unit Cost for reference
    unit_cost = total_exp / prod_qty if prod_qty > 0 else 0
        
    # Sort Data
    top_cust = sorted(cust_sales.items(), key=lambda x: x[1], reverse=True)[:5]
    top_exp = sorted(exp_cats.items(), key=lambda x: x[1], reverse=True)[:5]
    
    # Customer Value Analysis (Avg Price per Unit)
    cust_value = []
    for c, s_amt in cust_sales.items():
        qty = cust_qtys.get(c, 0)
        avg_p = s_amt / qty if qty > 0 else 0
        cust_value.append((c, avg_p, s_amt))
    
    # Sort by Avg Price (find high value customers)
    cust_value.sort(key=lambda x: x[1], reverse=True)
    top_value_cust = cust_value[:5] # Highest price per unit
    
    # Generate HTML with Chart.js
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <title>月度经营分析报告 {month_str}</title>
        <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
        <style>
            body {{ font-family: 'Segoe UI', sans-serif; background: #f4f6f9; padding: 20px; max-width: 1000px; margin: 0 auto; }}
            .card {{ background: white; padding: 20px; border-radius: 8px; box-shadow: 0 2px 5px rgba(0,0,0,0.05); margin-bottom: 20px; }}
            .row {{ display: flex; gap: 20px; flex-wrap: wrap; }}
            .col {{ flex: 1; min-width: 300px; }}
            h2 {{ color: #2c3e50; border-left: 5px solid #3498db; padding-left: 10px; }}
            .kpi-box {{ display: flex; justify-content: space-around; text-align: center; }}
            .kpi {{ padding: 10px; }}
            .kpi-val {{ font-size: 24px; font-weight: bold; }}
            .green {{ color: #27ae60; }} .red {{ color: #c0392b; }} .blue {{ color: #2980b9; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
            th, td {{ padding: 8px; text-align: left; border-bottom: 1px solid #ddd; }}
            th {{ background-color: #f8f9fa; color: #7f8c8d; font-size: 12px; }}
        </style>
    </head>
    <body>
        <div class="card">
            <h1 style="text-align:center">📊 {month_str} 月度经营分析报告</h1>
            <div class="kpi-box">
                <div class="kpi">
                    <div style="color:#7f8c8d">总收入 (Cash)</div>
                    <div class="kpi-val green">¥ {total_inc:,.0f}</div>
                </div>
                <div class="kpi">
                    <div style="color:#7f8c8d">总支出 (Cash)</div>
                    <div class="kpi-val red">¥ {total_exp:,.0f}</div>
                </div>
                <div class="kpi">
                    <div style="color:#7f8c8d">净现金流</div>
                    <div class="kpi-val blue">¥ {total_inc - total_exp:+,.0f}</div>
                </div>
                <div class="kpi">
                    <div style="color:#7f8c8d">产值 (Production)</div>
                    <div class="kpi-val" style="color:#e67e22">¥ {sum(cust_sales.values()):,.0f}</div>
                </div>
            </div>
             <div style="text-align:center; margin-top: 15px; font-size: 14px; color: #7f8c8d; border-top: 1px solid #eee; padding-top: 10px;">
                🏭 本月总产量: {prod_qty:,.0f} | 📉 综合成本率: {cost_rate:.1f}% (能耗{energy_rate:.1f}%/外协{outsourced_rate:.1f}%) | 💰 单位成本: ¥ {unit_cost:.2f}
            </div>
        </div>
        
        <div class="row">
            <div class="col card">
                <h2>🏆 客户产值贡献 TOP5</h2>
                <canvas id="custChart"></canvas>
            </div>
            <div class="col card">
                <h2>💸 支出构成 TOP5</h2>
                <canvas id="expChart"></canvas>
            </div>
        </div>
        
        <div class="card">
            <h2>💎 客户单价价值分析 (TOP 5 High Value)</h2>
            <p style="color: #7f8c8d; font-size: 12px;">* 单价 = 总加工费 / 总数量 (反映客户利润空间)</p>
            <table>
                <thead>
                    <tr>
                        <th>客户名称</th>
                        <th>平均单价 (元/单位)</th>
                        <th>总产值贡献</th>
                    </tr>
                </thead>
                <tbody>
                    {''.join([f"<tr><td>{c}</td><td>¥ {p:.2f}</td><td>¥ {v:,.0f}</td></tr>" for c, p, v in top_value_cust])}
                </tbody>
            </table>
        </div>
        
        <script>
            // Customer Chart
            new Chart(document.getElementById('custChart'), {{
                type: 'bar',
                data: {{
                    labels: {json.dumps([x[0] for x in top_cust], ensure_ascii=False)},
                    datasets: [{{
                        label: '加工费产值',
                        data: {json.dumps([x[1] for x in top_cust])},
                        backgroundColor: 'rgba(52, 152, 219, 0.6)',
                        borderColor: 'rgba(52, 152, 219, 1)',
                        borderWidth: 1
                    }}]
                }},
                options: {{ indexAxis: 'y' }}
            }});
            
            // Expense Chart
            new Chart(document.getElementById('expChart'), {{
                type: 'doughnut',
                data: {{
                    labels: {json.dumps([x[0] for x in top_exp], ensure_ascii=False)},
                    datasets: [{{
                        data: {json.dumps([x[1] for x in top_exp])},
                        backgroundColor: [
                            '#e74c3c', '#e67e22', '#f1c40f', '#2ecc71', '#9b59b6', '#95a5a6'
                        ]
                    }}]
                }}
            }});
        </script>
        
        <div style="text-align:center; color:#999; margin-top:20px;">
            Generated by CWZS System
        </div>
    </body>
    </html>
    """
    
    save_path = os.path.join(DATA_ROOT, f"月度分析_{month_str}.html")
    with open(save_path, "w", encoding="utf-8") as f:
        f.write(html)
        
    print(f"✅ 报告已生成: {Color.UNDERLINE}{save_path}{Color.ENDC}")
    try: os.startfile(save_path)
    except: pass

def manage_invoice_flow(client, app_token):
    """发票管理：录入销项/进项，查看统计"""
    while True:
        print(f"\n{Color.HEADER}🧾 发票管理 (进项/销项){Color.ENDC}")
        print("---------------------------------------")
        print("1. [销项] 登记已开发票 (给客户)")
        print("2. [进项] 登记收到发票 (供应商)")
        print("3. 查看最近发票记录 (20条)")
        print("4. 发票统计 (本月/本年)")
        print("5. 🛡️ 税控额度与税负分析 (Risk Monitor) [新]")
        print("0. 返回主菜单")
        
        choice = input(f"{Color.OKBLUE}请选择功能 (0-5): {Color.ENDC}").strip()
        
        if choice == '0': break
        
        table_id = get_table_id_by_name(client, app_token, "发票管理表")
        if not table_id:
            create_invoice_table(client, app_token)
            table_id = get_table_id_by_name(client, app_token, "发票管理表")
            
        if choice in ['1', '2']:
            is_sales = (choice == '1')
            type_prefix = "销项" if is_sales else "进项"
            
            # 批量模式状态
            batch_mode = False
            b_code = ""
            b_date = ""
            b_type = ""
            b_target = ""
            
            while True:
                print(f"\n{Color.BOLD}➕ 登记{type_prefix}发票{Color.ENDC}")
                if batch_mode:
                    print(f"{Color.OKBLUE}🔒 批量锁定模式 (输入 0 退出当前模式){Color.ENDC}")
                    print(f"   📅 {b_date} | 🏷️ {b_type} | 🏢 {b_target} | 🔢 代码:{b_code}")
                    inv_code = b_code
                    inv_date = b_date
                    inv_type = b_type
                    target = b_target
                else:
                    print("   (输入 0 返回上级菜单)")

                # 1. 发票号码
                inv_no = input("发票号码: ").strip()
                if inv_no == '0':
                    if batch_mode:
                        batch_mode = False
                        print("🔓 已退出批量锁定模式")
                        continue
                    break
                
                if not batch_mode:
                    inv_code = input("发票代码 (选填): ").strip()
                    inv_date = input("开票日期 (YYYY-MM-DD, 回车默认今天): ").strip()
                    if not inv_date: inv_date = datetime.now().strftime("%Y-%m-%d")
                    
                    print(f"发票类型: 1.专票  2.普票")
                    t_choice = input("请选择 (1/2): ").strip()
                    inv_type = f"{type_prefix}{'专票' if t_choice=='1' else '普票'}"
                    
                    target = input(f"{'购买方' if is_sales else '销售方'}名称: ").strip()

                # 金额处理
                amount = 0.0
                tax = 0.0
                auto_calculated = False

                while True:
                    print(f"💰 金额录入 (输入 'h 113' 可按含税价自动拆分，默认税率13%/1%/3%)")
                    amt_input = input("   不含税金额: ").strip()
                    
                    # 检查是否为含税模式
                    if amt_input.lower().startswith('h') or amt_input.startswith('含'):
                        try:
                            # 解析含税总额
                            total_inc = float(amt_input[1:].strip())
                            
                            # 询问税率
                            rate_str = input("   请输入税率% (默认 1, 输入 13/6/3/1): ").strip()
                            if not rate_str: rate_str = "1" # 普票常见1%，专票常见13%
                            rate = float(rate_str) / 100.0
                            
                            amount = total_inc / (1 + rate)
                            tax = total_inc - amount
                            amount = round(amount, 2)
                            tax = round(tax, 2)
                            
                            print(f"   ✨ 自动拆分: 不含税 {amount} + 税额 {tax} = 总额 {total_inc}")
                            auto_calculated = True
                            break
                        except:
                            print("❌ 格式错误，请使用 'h 113'")
                            continue

                    try:
                        amount = float(amt_input or 0)
                        break
                    except:
                        print("❌ 金额无效")
                
                try:
                    if not auto_calculated:
                        # 简易税额计算辅助
                        tax_input = input("税额 (直接回车可按税率估算): ").strip()
                        if not tax_input and amount > 0:
                             pass
                        tax = float(tax_input or 0)
                    
                    total = amount + tax
                    print(f"💰 价税合计: {total:,.2f}")
                except:
                    print("❌ 金额输入错误")
                    continue
                
                remark = f"手动录入 {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                    
                fields = {
                    "发票号码": inv_no,
                    "发票代码": inv_code,
                    "开票日期": int(pd.to_datetime(inv_date).timestamp() * 1000),
                    "类型": inv_type,
                    "购买方/销售方": target,
                    "不含税金额": amount,
                    "税额": tax,
                    "价税合计": total,
                    "状态": "正常",
                    "备注": remark
                }
                
                req = BatchCreateAppTableRecordRequest.builder().app_token(app_token).table_id(table_id).request_body(
                    BatchCreateAppTableRecordRequestBody.builder().records([AppTableRecord.builder().fields(fields).build()]).build()).build()
                
                if client.bitable.v1.app_table_record.batch_create(req).success():
                    print(f"{Color.OKGREEN}✅ {type_prefix}发票已登记: {inv_no}{Color.ENDC}")
                    
                    # 询问进入批量模式
                    if not batch_mode:
                        if input("⚡ 是否锁定表头(代码/日期/类型/对方)进入批量模式? (y/n) [n]: ").strip().lower() == 'y':
                            batch_mode = True
                            b_code = inv_code
                            b_date = inv_date
                            b_type = inv_type
                            b_target = target
                            print(f"{Color.OKGREEN}✅ 已进入批量模式，接下来只需输入号码和金额{Color.ENDC}")
                else:
                    print("❌ 登记失败")
                
        elif choice == '3':
            records = get_all_records(client, app_token, table_id)
            if not records:
                print("📭 暂无发票记录")
            else:
                print(f"\n{Color.UNDERLINE}最近 20 条发票记录:{Color.ENDC}")
                print(f"{'日期':<12} | {'类型':<8} | {'号码':<10} | {'金额':<10} | {'对方名称'}")
                print("-" * 60)
                records.sort(key=lambda x: x.fields.get("开票日期", 0))
                for r in records[-20:]:
                    f = r.fields
                    ts = f.get("开票日期", 0)
                    d_str = datetime.fromtimestamp(ts/1000).strftime("%Y-%m-%d") if ts else "-"
                    print(f"{d_str:<12} | {f.get('类型',''):<8} | {f.get('发票号码',''):<10} | {f.get('价税合计',0):<10.2f} | {f.get('购买方/销售方','')}")

        elif choice == '4':
            # 简单统计
            records = get_all_records(client, app_token, table_id)
            input_tax = 0.0
            output_tax = 0.0
            cur_month = datetime.now().strftime("%Y-%m")
            m_in = 0.0
            m_out = 0.0
            
            for r in records:
                f = r.fields
                tax = float(f.get("税额", 0))
                itype = f.get("类型", "")
                ts = f.get("开票日期", 0)
                d_str = datetime.fromtimestamp(ts/1000).strftime("%Y-%m") if ts else ""
                
                if "进项" in itype:
                    input_tax += tax
                    if d_str == cur_month: m_in += tax
                elif "销项" in itype:
                    output_tax += tax
                    if d_str == cur_month: m_out += tax
                    
            print(f"\n📊 发票统计摘要")
            print(f"本月 ({cur_month}): 进项税 {m_in:,.2f} | 销项税 {m_out:,.2f} | 差额 {m_out - m_in:,.2f}")
            print(f"累计历史: 进项税 {input_tax:,.2f} | 销项税 {output_tax:,.2f}")
            input("\n按回车继续...")

        elif choice == '5':
            # Tax Quota & Burden Analysis
            print(f"\n{Color.HEADER}🛡️ 税控额度与税负分析{Color.ENDC}")
            print("--------------------------------")
            records = get_all_records(client, app_token, table_id)
            
            # Determine Quarter
            now = datetime.now()
            q_start_month = (now.month - 1) // 3 * 3 + 1
            q_start = datetime(now.year, q_start_month, 1)
            if q_start_month + 3 > 12:
                q_end = datetime(now.year + 1, 1, 1)
            else:
                q_end = datetime(now.year, q_start_month + 3, 1)
            
            q_start_ts = int(q_start.timestamp() * 1000)
            q_end_ts = int(q_end.timestamp() * 1000)
            
            # Calc Quarter Sales (No Tax)
            q_sales_no_tax = 0.0
            q_sales_total = 0.0
            
            # Calc Year Totals for Burden
            y_start_ts = int(datetime(now.year, 1, 1).timestamp() * 1000)
            y_out_tax = 0.0
            y_in_tax = 0.0
            y_sales_no_tax = 0.0
            
            for r in records:
                f = r.fields
                ts = f.get("开票日期", 0)
                itype = f.get("类型", "")
                amt = float(f.get("不含税金额", 0))
                tax = float(f.get("税额", 0))
                total = float(f.get("价税合计", 0))
                
                if "销项" in itype:
                    if ts >= q_start_ts and ts < q_end_ts:
                        q_sales_no_tax += amt
                        q_sales_total += total
                    
                    if ts >= y_start_ts:
                        y_out_tax += tax
                        y_sales_no_tax += amt
                        
                elif "进项" in itype:
                    if ts >= y_start_ts:
                        y_in_tax += tax
            
            # 1. Quota Monitor (Small Taxpayer)
            print(f"\n📊 本季度 ({q_start.strftime('%Y-%m')}) 销售概况:")
            print(f"   不含税销售额: {q_sales_no_tax:,.2f}")
            print(f"   价税合计:     {q_sales_total:,.2f}")
            
            limit = 300000 # Default for small taxpayer exemption
            pct = (q_sales_total / limit) * 100 if limit > 0 else 0
            
            bar_len = 30
            filled = int(bar_len * pct / 100)
            filled = min(filled, bar_len)
            bar = "█" * filled + "░" * (bar_len - filled)
            
            color = Color.OKGREEN
            if pct > 80: color = Color.WARNING
            if pct > 100: color = Color.FAIL
            
            print(f"\n📉 小规模免税额度监控 (默认30万/季):")
            print(f"   进度: {color}[{bar}] {pct:.1f}%{Color.ENDC}")
            if pct > 90:
                print(f"   {Color.FAIL}⚠️ 警告: 即将或已经超过免税额度!{Color.ENDC}")
            else:
                print(f"   ✅ 额度充足 (剩余: {limit - q_sales_total:,.2f})")
                
            # 2. Burden Rate (General Taxpayer)
            print(f"\n⚖️  税负率估算 (本年度):")
            net_tax = y_out_tax - y_in_tax
            burden = (net_tax / y_sales_no_tax * 100) if y_sales_no_tax > 0 else 0
            
            print(f"   销项税: {y_out_tax:,.2f}")
            print(f"   进项税: {y_in_tax:,.2f}")
            print(f"   应纳税: {net_tax:,.2f}")
            print(f"   税负率: {burden:.2f}% (应纳税/不含税销售)")
            
            input("\n按回车继续...")

# 菜单：设置
def settings_menu():
    """系统设置菜单"""
    global TAX_RATE, ZHIPU_API_KEY
    
    while True:
        print(f"\n{Color.CYAN}⚙️ 系统设置{Color.ENDC}")
        print(f"  1. 设置税率 (当前: {TAX_RATE}%)")
        print(f"  2. 设置智谱AI Key (当前: {ZHIPU_API_KEY[:6]}******)")
        print("  0. 返回")
        
        choice = input("请选择: ").strip()
        
        if choice == '0':
            break
        elif choice == '1':
            try:
                new_rate = float(input("请输入新税率 (例如 1, 3, 6, 13): "))
                TAX_RATE = new_rate
                # Update .env
                update_env_key("TAX_RATE", str(TAX_RATE))
                print("✅ 税率已更新")
            except:
                print("❌ 输入无效")
        elif choice == '2':
            key = input("请输入新的 API Key: ").strip()
            if key:
                ZHIPU_API_KEY = key
                update_env_key("ZHIPU_API_KEY", key)
                print("✅ Key 已更新")
        elif choice == '14':
            backup_system_data(client, APP_TOKEN)
            
        else:
            print("❌ 无效选项")

def update_env_key(key, value):
    """更新 .env 文件"""
    lines = []
    if os.path.exists(".env"):
        try:
            with open(".env", "r", encoding="utf-8") as f:
                lines = f.readlines()
        except:
            pass
            
    new_lines = []
    found = False
    for line in lines:
        if line.strip().startswith(f"{key}="):
            new_lines.append(f"{key}={value}\n")
            found = True
        else:
            new_lines.append(line)
            
    if not found:
        if new_lines and not new_lines[-1].endswith('\n'):
             new_lines.append('\n')
        new_lines.append(f"{key}={value}\n")
        
    try:
        with open(".env", "w", encoding="utf-8") as f:
            f.writelines(new_lines)
    except Exception as e:
        print(f"❌ 保存配置失败: {e}")

def financial_health_check(client, app_token, target_year=None):
    """一键财务体检：扫描税务风险和数据异常 (生成HTML报告)"""
    if target_year:
        log.info(f"🏥 正在进行 {target_year}年度 财务体检...", extra={"solution": "全面扫描中"})
        year = target_year
    else:
        log.info("🏥 正在进行财务体检...", extra={"solution": "全面扫描中"})
        year = datetime.now().year
    
    table_id = get_table_id_by_name(client, app_token, "日常台账表")
    if not table_id:
        log.error("❌ 找不到日常台账表", extra={"solution": "请先初始化表格"})
        return

    # 优化：只获取指定年度数据
    start_ts = int(datetime(year, 1, 1).timestamp() * 1000)
    end_ts = int(datetime(year + 1, 1, 1).timestamp() * 1000)
    filter_str = f'AND(CurrentValue.[记账日期]>={start_ts}, CurrentValue.[记账日期]<{end_ts})'

    records = get_all_records(client, app_token, table_id, filter_info=filter_str)
    
    risks = []
    stats = {
        "total_count": len(records),
        "cash_txns": 0,
        "no_ticket_amt": 0.0,
        "large_cash": 0,
        "total_income": 0.0,
        "total_expense": 0.0,
        "missing_invoice_high_risk": 0 # >5000 无票
    }
    
    print("\n📋 体检报告")
    print("-" * 40)
    
    # 风险详情列表 (用于生成报告)
    risk_details = []
    seen_txns = {} # 用于查重 (key: date_amt_partner)
    
    # [新增] 可修复的异常
    duplicate_ids = [] # 待删除的重复记录ID
    has_depreciation = False # 本月是否有折旧

    for r in records:
        f = r.fields
        amt = float(f.get("实际收付金额", 0))
        is_cash = f.get("是否现金", "否") == "是"
        has_ticket = f.get("是否有票", "有票") == "无票"
        biz_type = f.get("业务类型", "")
        expense_type = f.get("费用归类", "")
        remark = f.get("备注") or ""
        partner = f.get("往来单位费用") or ""
        date_ts = f.get("记账日期", 0)
        try:
            date_str = datetime.fromtimestamp(date_ts/1000).strftime("%Y-%m-%d")
        except:
            date_str = "未知日期"
            
        # 检查是否已计提折旧
        if "折旧" in expense_type:
            has_depreciation = True
        
        # 统计
        if is_cash: stats["cash_txns"] += 1
        if has_ticket and biz_type in ["付款", "费用"]: stats["no_ticket_amt"] += amt
        
        # 统计收支
        if biz_type == "收入":
            stats["total_income"] += amt
        elif biz_type in ["支出", "费用"]:
            stats["total_expense"] += amt
        
        # 规则 0: 重复录入检测 (新增)
        # 简单指纹: 日期 + 金额 + 对象 (忽略时分秒)
        # 注意: 只有非零金额才查重
        if amt != 0:
            dup_key = f"{date_str}_{amt}_{partner}"
            if dup_key in seen_txns:
                msg = f"⚠️ [重复风险] 疑似重复录入: {date_str} {amt}元 - {partner}"
                risks.append(msg)
                risk_details.append({"date": date_str, "type": "重复录入", "amt": amt, "desc": f"与已有记录重复: {partner}", "level": "高"})
                print(msg)
                
                # 收集重复记录ID (假设是后录入的为重复)
                rid = getattr(r, "record_id", None)
                if rid:
                    duplicate_ids.append(rid)
            else:
                seen_txns[dup_key] = True

        # 规则 1: 大额现金支付 (>5000)
        if is_cash and amt > 5000 and biz_type in ["付款", "费用"]:
            msg = f"⚠️ [高风险] 大额现金支出: {amt}元 ({remark})"
            risks.append(msg)
            risk_details.append({"date": date_str, "type": "大额现金", "amt": amt, "desc": remark, "level": "高"})
            print(msg)
            stats["large_cash"] += 1
            
        # 规则 2: 大额无票费用 (>1000)
        if has_ticket and biz_type == "费用":
            if amt > 5000:
                 msg = f"⚠️ [税务风险] 大额无票费用(>5000): {amt}元 ({remark})"
                 risks.append(msg)
                 risk_details.append({"date": date_str, "type": "税务高危", "amt": amt, "desc": "无票且金额>5000", "level": "高"})
                 print(msg)
                 stats["missing_invoice_high_risk"] += 1
            elif amt > 1000:
                msg = f"⚠️ [税务风险] 大额无票费用: {amt}元 ({remark})"
                risks.append(msg)
                risk_details.append({"date": date_str, "type": "大额无票", "amt": amt, "desc": remark, "level": "中"})
                print(msg)
            
        # 规则 3: 摘要缺失
        if len(remark) < 2:
            print(f"ℹ️ [数据规范] 摘要过短或缺失: {amt}元")
            risk_details.append({"date": date_str, "type": "摘要缺失", "amt": amt, "desc": "摘要为空或过短", "level": "低"})

        # 规则 5: 费用归类缺失 (Daily Closing Validation)
        if biz_type == "费用" and (not expense_type or expense_type in ["", "nan", "未知"]):
            msg = f"⚠️ [数据完善] 费用归类缺失: {amt}元 ({remark})"
            risks.append(msg)
            risk_details.append({"date": date_str, "type": "归类缺失", "amt": amt, "desc": "请补充费用归类", "level": "中"})
            print(msg)

    # 规则 4: 本月折旧未计提
    current_month_str = datetime.now().strftime('%Y-%m')
    check_depreciation = False
    
    # 仅当检查当前年份时，才检查本月折旧
    if year == datetime.now().year:
        check_depreciation = True
        
    has_depreciation = False
    if check_depreciation:
        for r in records:
            f = r.fields
            # 检查是否为本月记录且费用归类为折旧摊销
            r_date = f.get("记账日期", 0)
            try:
                r_month = datetime.fromtimestamp(r_date/1000).strftime('%Y-%m')
            except:
                r_month = ""
                
            if r_month == current_month_str and f.get("费用归类") == "折旧摊销":
                has_depreciation = True
                break
            
        if not has_depreciation:
            msg = f"⚠️ [合规风险] 本月尚未计提固定资产折旧 ({current_month_str})"
            risks.append(msg)
            risk_details.append({"date": datetime.now().strftime("%Y-%m-%d"), "type": "折旧缺失", "amt": 0, "desc": "本月未计提折旧", "level": "中"})
            print(msg)

    # -------------------------------------------------------------------------
    # 新增：加工费明细表体检
    # -------------------------------------------------------------------------
    pf_table_id = get_table_id_by_name(client, app_token, "加工费明细表")
    if pf_table_id:
        log.info("🏥 正在扫描加工费明细表...", extra={"solution": "无"})
        # 针对 "日期" 字段的过滤器
        pf_filter = f'AND(CurrentValue.[日期]>={start_ts}, CurrentValue.[日期]<{end_ts})'
        try:
            pf_records = get_all_records(client, app_token, pf_table_id, filter_info=pf_filter)
            
            for r in pf_records:
                f = r.fields
                date_ts = f.get("日期", 0)
                try:
                    d_str = datetime.fromtimestamp(date_ts/1000).strftime("%Y-%m-%d")
                except:
                    d_str = "未知"
                
                # Check 1: 有数量无金额
                total = float(f.get("总金额", 0))
                qty = float(f.get("数量", 0))
                price = float(f.get("单价", 0))
                
                if qty != 0 and total == 0:
                     msg = f"⚠️ [加工费异常] 有数量无金额: {d_str} {f.get('往来单位','')}"
                     risks.append(msg)
                     risk_details.append({"date": d_str, "type": "加工费异常", "amt": 0, "desc": "有数量无金额", "level": "高"})
                     print(msg)
                
                # Check 2: 单价为0 (提醒)
                if qty != 0 and price == 0:
                     msg = f"ℹ️ [数据提醒] 加工费单价为0: {d_str} {f.get('往来单位','')}"
                     print(msg)

                # Check 3: 往来单位缺失
                if not f.get("往来单位"):
                     msg = f"⚠️ [数据缺失] 加工费未指定往来单位: {d_str}"
                     risks.append(msg)
                     risk_details.append({"date": d_str, "type": "数据缺失", "amt": total, "desc": "未指定往来单位", "level": "中"})
                     print(msg)
        except Exception as e:
            log.warning(f"⚠️ 扫描加工费表失败: {e}")

    # 规则 6: 经营风险预警 (New)
    income = stats["total_income"]
    expense = stats["total_expense"]
    
    if income > 0:
        margin = (income - expense) / income
        if margin < 0:
             msg = f"⚠️ [经营风险] 当前处于亏损状态 (净利润率: {margin*100:.1f}%)"
             risks.append(msg)
             risk_details.append({"date": datetime.now().strftime("%Y-%m-%d"), "type": "亏损预警", "amt": income-expense, "desc": "支出大于收入", "level": "高"})
             print(msg)
        elif margin < 0.1:
             msg = f"⚠️ [经营风险] 利润率过低 ({margin*100:.1f}%)"
             print(msg)
    elif expense > 0 and income == 0:
         msg = f"⚠️ [经营风险] 本期暂无收入"
         risks.append(msg)
         risk_details.append({"date": datetime.now().strftime("%Y-%m-%d"), "type": "亏损预警", "amt": -expense, "desc": "只有支出无收入", "level": "高"})
         print(msg)

    print("-" * 40)
    print(f"扫描完成。共 {len(records)} 条记录。")
    print(f"本期收入: {stats['total_income']:.2f} 元")
    print(f"本期支出: {stats['total_expense']:.2f} 元")
    print(f"大额现金笔数: {stats['large_cash']}")
    print(f"无票金额总计: {stats['no_ticket_amt']:.2f} 元")
    
    ai_advice = "暂无建议"
    
    # 如果有 AI，生成建议
    if zhipu_client and risks:
        try:
            log.info("🤖 AI 正在生成整改建议...", extra={"solution": "请稍候"})
            risk_text = "\n".join(risks[:10]) # 限制长度
            response = zhipu_client.chat.completions.create(
                model="glm-4-flash",
                messages=[
                    {"role": "system", "content": "你是一名资深税务会计。用户进行了一次财务体检，发现了以下风险点。请给出简短、专业的整改建议（3条以内，用HTML列表格式输出）。"},
                    {"role": "user", "content": risk_text}
                ]
            )
            ai_advice = response.choices[0].message.content
            print(f"\n💡 AI 建议生成完毕")
        except:
            pass
            
    # 生成 HTML 报告
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>财务体检报告 - {datetime.now().strftime('%Y-%m-%d')}</title>
        <style>
            body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: #f4f6f9; margin: 0; padding: 20px; }}
            .container {{ max-width: 900px; margin: 0 auto; background: white; padding: 40px; border-radius: 10px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); }}
            h1 {{ color: #2c3e50; border-bottom: 2px solid #e74c3c; padding-bottom: 10px; }}
            .stats {{ display: flex; justify-content: space-between; margin-bottom: 30px; }}
            .stat-box {{ background: #f8f9fa; padding: 20px; border-radius: 8px; text-align: center; width: 30%; }}
            .stat-val {{ font-size: 24px; font-weight: bold; color: #e74c3c; }}
            .risk-table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
            th, td {{ border: 1px solid #ddd; padding: 12px; text-align: left; }}
            th {{ background-color: #e74c3c; color: white; }}
            tr:nth-child(even) {{ background-color: #f2f2f2; }}
            .tag {{ padding: 4px 8px; border-radius: 4px; color: white; font-size: 12px; }}
            .tag-high {{ background-color: #c0392b; }}
            .tag-mid {{ background-color: #e67e22; }}
            .tag-low {{ background-color: #f1c40f; color: #333; }}
            .ai-box {{ background-color: #e8f6f3; padding: 20px; border-left: 5px solid #1abc9c; margin-top: 30px; border-radius: 4px; }}
        </style>
    </head>
    <body>
        <div class="container">
            <h1>🏥 财务健康体检报告</h1>
            <p>生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            
            <div class="stats">
                <div class="stat-box">
                    <div class="stat-val">{stats['large_cash']}</div>
                    <div>大额现金笔数 (>5k)</div>
                </div>
                <div class="stat-box">
                    <div class="stat-val">{stats['no_ticket_amt']:,.2f}</div>
                    <div>无票支出总额</div>
                </div>
                <div class="stat-box">
                    <div class="stat-val">{len(risk_details)}</div>
                    <div>发现风险点总数</div>
                </div>
            </div>

            <div class="ai-box">
                <h3>🤖 AI 整改建议</h3>
                {ai_advice}
            </div>

            <h3>⚠️ 风险详情清单</h3>
            <table class="risk-table">
                <thead>
                    <tr>
                        <th>风险等级</th>
                        <th>日期</th>
                        <th>类型</th>
                        <th>金额</th>
                        <th>说明</th>
                    </tr>
                </thead>
                <tbody>
    """
    
    for r in risk_details:
        tag_cls = "tag-low"
        if r['level'] == "高": tag_cls = "tag-high"
        elif r['level'] == "中": tag_cls = "tag-mid"
        
        html += f"""
        <tr>
            <td><span class="tag {tag_cls}">{r['level']}</span></td>
            <td>{r['date']}</td>
            <td>{r['type']}</td>
            <td>{r['amt']:,.2f}</td>
            <td>{r['desc']}</td>
        </tr>
        """
        
    html += """
                </tbody>
            </table>
        </div>
    </body>
    </html>
    """
    
    report_dir = "财务数据备份"
    if not os.path.exists(report_dir): os.makedirs(report_dir)
    filename = f"{report_dir}/体检报告_{datetime.now().strftime('%Y%m%d%H%M%S')}.html"
    
    with open(filename, "w", encoding="utf-8") as f:
        f.write(html)
        
    log.info(f"📄 体检报告已生成: {filename}", extra={"solution": "浏览器打开查看"})
    try:
        os.startfile(filename)
    except:
        pass
    
    # [新增] 交互式修复逻辑
    if duplicate_ids:
        print(f"\n🔧 [一键修复] 发现 {len(duplicate_ids)} 条重复记录。")
        if input("👉 是否立即删除这些重复项? (y/n): ").strip().lower() == 'y':
            print("🗑️ 正在删除重复记录...")
            try:
                # 批量删除
                req = BatchDeleteAppTableRecordRequest.builder() \
                    .app_token(app_token) \
                    .table_id(table_id) \
                    .request_body(BatchDeleteAppTableRecordRequestBody.builder().records(duplicate_ids).build()) \
                    .build()
                
                resp = client.bitable.v1.app_table_record.batch_delete(req)
                if resp.success():
                    print(f"✅ 已成功删除 {len(duplicate_ids)} 条重复记录！")
                else:
                    print(f"❌ 删除失败: {resp.msg}")
            except Exception as e:
                print(f"❌ 删除出错: {e}")
                
    if not has_depreciation and datetime.now().day > 20:
        print(f"\n🔧 [一键修复] 本月尚未计提折旧 (通常月底计提)。")
        if input("👉 是否立即运行折旧计算? (y/n): ").strip().lower() == 'y':
            calculate_depreciation(client, app_token)

    return True

def generate_daily_html_report(client, app_token, summary_log=None):
    """生成每日结账 HTML 报告"""
    log.info("📊 正在生成今日结账报告...", extra={"solution": "无"})
    
    table_id = get_table_id_by_name(client, app_token, "日常台账表")
    if not table_id: return None
    
    # 1. 获取今日数据
    now = datetime.now()
    today_start = datetime(now.year, now.month, now.day)
    start_ts = int(today_start.timestamp() * 1000)
    filter_str = f'CurrentValue.[记账日期]>={start_ts}'
    
    records = get_all_records(client, app_token, table_id, filter_info=filter_str)
    
    today_income = 0.0
    today_expense = 0.0
    tx_count = len(records)
    details = []
    
    for r in records:
        f = r.fields
        amt = float(f.get("实际收付金额", 0))
        b_type = f.get("业务类型", "")
        desc = f.get("备注") or f.get("往来单位费用", "")
        
        if b_type == "收款":
            today_income += amt
        elif b_type in ["付款", "费用"]:
            today_expense += amt
            
        details.append({
            "type": b_type,
            "amt": amt,
            "desc": desc,
            "partner": f.get("往来单位费用", "-")
        })
        
    # 2. 待办事项 (检查文件夹)
    pending_files = []
    watch_dir = "待处理单据"
    if os.path.exists(watch_dir):
        pending_files = [f for f in os.listdir(watch_dir) if not f.startswith("~$") and f.lower().endswith(('.xlsx', '.png', '.jpg'))]
        
    # 3. 待补票据 (简单查询)
    missing_count = 0
    # 这里为了速度，暂时不全量查，只看传入的 summary_log 是否有提及
    
    # 生成 HTML
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <title>每日结账报告 - {now.strftime('%Y-%m-%d')}</title>
        <style>
            body {{ font-family: 'Segoe UI', sans-serif; background: #f0f2f5; padding: 20px; }}
            .container {{ max-width: 800px; margin: 0 auto; background: white; padding: 30px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
            h1 {{ color: #1a1a1a; border-bottom: 3px solid #3498db; padding-bottom: 10px; }}
            .summary-box {{ display: flex; gap: 20px; margin: 20px 0; }}
            .card {{ flex: 1; background: #f8f9fa; padding: 15px; border-radius: 8px; text-align: center; border: 1px solid #e9ecef; }}
            .num {{ font-size: 24px; font-weight: bold; color: #2c3e50; }}
            .label {{ color: #7f8c8d; font-size: 14px; }}
            .income {{ color: #27ae60; }}
            .expense {{ color: #c0392b; }}
            
            h3 {{ margin-top: 30px; color: #34495e; border-left: 5px solid #3498db; padding-left: 10px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
            th, td {{ padding: 10px; text-align: left; border-bottom: 1px solid #eee; }}
            th {{ background: #f8f9fa; color: #7f8c8d; }}
            
            .log-box {{ background: #2c3e50; color: #ecf0f1; padding: 15px; border-radius: 5px; font-family: monospace; max-height: 200px; overflow-y: auto; }}
            .pending-alert {{ background: #fff3cd; color: #856404; padding: 10px; border-radius: 5px; margin-top: 10px; }}
        </style>
    </head>
    <body>
        <div class="container">
            <h1>📅 每日结账报告 <small style="font-size: 16px; color: #7f8c8d">{now.strftime('%Y-%m-%d %H:%M')}</small></h1>
            
            <div class="summary-box">
                <div class="card">
                    <div class="num income">+{today_income:,.2f}</div>
                    <div class="label">今日收款</div>
                </div>
                <div class="card">
                    <div class="num expense">-{today_expense:,.2f}</div>
                    <div class="label">今日支出</div>
                </div>
                <div class="card">
                    <div class="num">{tx_count}</div>
                    <div class="label">业务笔数</div>
                </div>
                <div class="card">
                    <div class="num" style="color: #2980b9">{today_income - today_expense:,.2f}</div>
                    <div class="label">今日净现金流</div>
                </div>
            </div>
            
            <h3>📝 今日业务明细</h3>
            """
            
    if details:
        html += "<table><thead><tr><th>类型</th><th>金额</th><th>对象</th><th>摘要</th></tr></thead><tbody>"
        for d in details:
            color = "green" if d['type'] == "收款" else "red"
            html += f"<tr><td><span style='color:{color}'>{d['type']}</span></td><td>{d['amt']:,.2f}</td><td>{d['partner']}</td><td>{d['desc']}</td></tr>"
        html += "</tbody></table>"
    else:
        html += "<p style='color:#999; text-align:center'>今日暂无收支记录</p>"
        
    if pending_files:
        html += f"""
        <h3>🔔 待办提醒</h3>
        <div class="pending-alert">
            <strong>发现 {len(pending_files)} 个待处理文件:</strong><br>
            {', '.join(pending_files[:5])} {'...' if len(pending_files)>5 else ''}
        </div>
        """
        
    if summary_log:
        html += """
        <h3>⚙️ 系统处理日志</h3>
        <div class="log-box">
        """
        for line in summary_log:
            html += f"<div>{line}</div>"
        html += "</div>"
        
    html += """
        </div>
    </body>
    </html>
    """
    
    report_dir = "日结报告"
    if not os.path.exists(report_dir): os.makedirs(report_dir)
    filename = f"{report_dir}/日结_{now.strftime('%Y%m%d')}.html"
    
    with open(filename, "w", encoding="utf-8") as f:
        f.write(html)
        
    log.info(f"📄 日结报告已生成: {filename}")
    return filename

def generate_annual_report(client, app_token, year=None):
    """生成年度财务报表"""
    if not year:
        year = datetime.now().year
    
    log.info(f"📊 正在生成 {year} 年度报表...", extra={"solution": "无"})
    
    table_id = get_table_id_by_name(client, app_token, "日常台账表")
    if not table_id: return False
    
    # 获取全年数据
    records = get_all_records(client, app_token, table_id)
    
    monthly_data = {m: {"income": 0.0, "expense": 0.0, "count": 0} for m in range(1, 13)}
    category_summary = {} # {category: amount}
    
    total_income = 0.0
    total_expense = 0.0
    
    for r in records:
        f = r.fields
        ts = f.get("记账日期", 0)
        if not ts: continue
        
        dt = datetime.fromtimestamp(ts / 1000)
        if dt.year != year: continue
        
        amt = float(f.get("实际收付金额", 0))
        b_type = f.get("业务类型", "")
        cat = f.get("费用归类", "未分类")
        
        monthly_data[dt.month]["count"] += 1
        
        if b_type == "收款":
            monthly_data[dt.month]["income"] += amt
            total_income += amt
        elif b_type in ["付款", "费用"]:
            monthly_data[dt.month]["expense"] += amt
            total_expense += amt
            
            # 统计费用分类
            if cat not in category_summary: category_summary[cat] = 0.0
            category_summary[cat] += amt

    # 生成 HTML
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <title>{year} 年度财务报表</title>
        <style>
            body {{ font-family: 'Segoe UI', sans-serif; background: #f0f2f5; padding: 20px; }}
            .container {{ max-width: 1000px; margin: 0 auto; background: white; padding: 40px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
            h1 {{ color: #2c3e50; text-align: center; border-bottom: 2px solid #3498db; padding-bottom: 20px; }}
            .summary-cards {{ display: flex; gap: 20px; margin: 30px 0; }}
            .card {{ flex: 1; background: #f8f9fa; padding: 20px; border-radius: 8px; text-align: center; box-shadow: 0 1px 3px rgba(0,0,0,0.05); }}
            .num {{ font-size: 28px; font-weight: bold; margin: 10px 0; }}
            .income {{ color: #27ae60; }}
            .expense {{ color: #c0392b; }}
            .profit {{ color: #2980b9; }}
            
            table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
            th, td {{ padding: 12px; text-align: center; border-bottom: 1px solid #eee; }}
            th {{ background: #34495e; color: white; }}
            tr:nth-child(even) {{ background: #f9f9f9; }}
            
            .chart-box {{ height: 300px; margin-top: 40px; border: 1px solid #eee; padding: 10px; }}
            h2 {{ color: #34495e; margin-top: 40px; border-left: 5px solid #3498db; padding-left: 10px; }}
        </style>
    </head>
    <body>
        <div class="container">
            <h1>📅 {year} 年度财务报表</h1>
            
            <div class="summary-cards">
                <div class="card">
                    <div class="label">全年总收入</div>
                    <div class="num income">¥{total_income:,.2f}</div>
                </div>
                <div class="card">
                    <div class="label">全年总支出</div>
                    <div class="num expense">¥{total_expense:,.2f}</div>
                </div>
                <div class="card">
                    <div class="label">全年净利润</div>
                    <div class="num profit">¥{total_income - total_expense:,.2f}</div>
                </div>
            </div>
            
            <h2>📈 月度收支明细</h2>
            <table>
                <thead>
                    <tr>
                        <th>月份</th>
                        <th>收入</th>
                        <th>支出</th>
                        <th>结余</th>
                        <th>笔数</th>
                    </tr>
                </thead>
                <tbody>
    """
    
    for m in range(1, 13):
        d = monthly_data[m]
        balance = d["income"] - d["expense"]
        color = "#27ae60" if balance >= 0 else "#c0392b"
        html += f"""
        <tr>
            <td>{m}月</td>
            <td style="color:#27ae60">+{d['income']:,.2f}</td>
            <td style="color:#c0392b">-{d['expense']:,.2f}</td>
            <td style="font-weight:bold; color:{color}">{balance:,.2f}</td>
            <td>{d['count']}</td>
        </tr>
        """
        
    html += """
                </tbody>
            </table>
            
            <h2>📊 费用支出分布</h2>
            <table>
                <thead>
                    <tr>
                        <th>费用类型</th>
                        <th>金额</th>
                        <th>占比</th>
                    </tr>
                </thead>
                <tbody>
    """
    
    sorted_cats = sorted(category_summary.items(), key=lambda x: x[1], reverse=True)
    for cat, amt in sorted_cats:
        percent = (amt / total_expense * 100) if total_expense > 0 else 0
        html += f"""
        <tr>
            <td>{cat}</td>
            <td>{amt:,.2f}</td>
            <td>{percent:.1f}%</td>
        </tr>
        """
        
    html += """
                </tbody>
            </table>
            
            <p style="margin-top: 40px; color: #7f8c8d; text-align: center; font-size: 12px;">生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Powered by 飞书财务小助手</p>
        </div>
    </body>
    </html>
    """
    
    report_dir = "年度报告"
    if not os.path.exists(report_dir): os.makedirs(report_dir)
    filename = f"{report_dir}/{year}年度报表_{datetime.now().strftime('%Y%m%d')}.html"
    
    with open(filename, "w", encoding="utf-8") as f:
        f.write(html)
        
    log.info(f"✅ 年度报表已生成: {filename}")
    os.startfile(filename) # 自动打开
    return True

def backup_system_data(client, app_token):
    """全量数据备份"""
    print(f"\n{Color.CYAN}💾 正在进行全量数据备份...{Color.ENDC}")
    
    backup_dir = os.path.join(DATA_ROOT, "备份", datetime.now().strftime("%Y%m%d_%H%M%S"))
    if not os.path.exists(backup_dir): os.makedirs(backup_dir)
    
    tables = [
        "日常台账表", "加工费明细表", "薪酬管理表", 
        "发票管理表", "固定资产表", "往来单位表", "加工费价目表"
    ]
    
    success_count = 0
    for t_name in tables:
        t_id = get_table_id_by_name(client, app_token, t_name)
        if t_id:
            records = get_all_records(client, app_token, t_id)
            if records:
                data = [r.fields for r in records]
                # Save as JSON
                with open(os.path.join(backup_dir, f"{t_name}.json"), "w", encoding="utf-8") as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                # Save as Excel (optional, but good for user)
                try:
                    df = pd.DataFrame(data)
                    df.to_excel(os.path.join(backup_dir, f"{t_name}.xlsx"), index=False)
                except: pass
                
                print(f"   ✓ {t_name}: {len(records)} 条")
                success_count += 1
            else:
                print(f"   - {t_name}: 无数据")
    
    print(f"✅ 备份完成！路径: {backup_dir}")
    return backup_dir

def reset_system_data(client, app_token):
    """系统初始化/重置 (数据清空)"""
    print(f"\n{Color.FAIL}🛑 危险操作：系统数据重置{Color.ENDC}")
    print("此操作将清空所有业务数据，仅保留表结构。通常用于：")
    print("1. 试用模拟数据后，准备正式启用")
    print("2. 重新开始记账")
    print("注意：操作不可逆！请确保已备份重要数据。")
    
    # 强制备份
    print(f"\n{Color.CYAN}🛡️ 为了安全起见，系统将自动执行一次全量备份...{Color.ENDC}")
    backup_system_data(client, app_token)
    
    confirm = input(f"\n👉 请输入 {Color.BOLD}RESET{Color.ENDC} 确认清空所有数据: ").strip()
    if confirm != "RESET":
        print("❌ 操作已取消")
        return
        
    tables = [
        "日常台账表",
        "加工费明细表",
        "薪酬管理表",
        "固定资产表",
        "发票管理表",
        "加工费价目表", # Optional: maybe keep this?
        "往来单位表"   # Optional: maybe keep this?
    ]
    
    print("\n请选择要清空的范围:")
    print("1. 仅清空业务流水 (保留价目表、客户信息)")
    print("2. 彻底清空所有数据 (包括价目表、客户信息)")
    scope = input("👉 请选择 (1/2): ").strip()
    
    if scope == '1':
        target_tables = ["日常台账表", "加工费明细表", "薪酬管理表", "发票管理表"]
    elif scope == '2':
        target_tables = tables
    else:
        return
        
    print("⏳ 正在清空数据...")
    for t_name in target_tables:
        t_id = get_table_id_by_name(client, app_token, t_name)
        if t_id:
            # Get all records
            recs = get_all_records(client, app_token, t_id)
            if recs:
                print(f"   🗑️ 正在清空 {t_name} ({len(recs)} 条)...")
                # Batch delete
                batch_ids = [r.record_id for r in recs]
                for i in range(0, len(batch_ids), 100):
                    batch = batch_ids[i:i+100]
                    client.bitable.v1.app_table_record.batch_delete(
                        BatchDeleteAppTableRecordRequest.builder()
                        .app_token(app_token).table_id(t_id)
                        .request_body(BatchDeleteAppTableRecordRequestBody.builder().records(batch).build())
                        .build()
                    )
            else:
                print(f"   ✓ {t_name} 已为空")
                
    print(f"\n{Color.OKGREEN}✅ 系统重置完成！您可以开始新的记账了。{Color.ENDC}")

def daily_closing_wizard(client, app_token):
    """每日结单向导 (End of Day)"""
    print(f"\n{Color.HEADER}🌙 每日结单向导 (End of Day){Color.ENDC}")
    print("-----------------------------------")
    print("本向导将协助您完成今日的财务收尾工作，确保数据不遗漏。")
    
    # 1. 检查加工费
    print(f"\n{Color.BOLD}1. 加工费核对{Color.ENDC}")
    if input("👉 今天是否有新的【加工单】需要录入? (y/n) [n]: ").strip().lower() == 'y':
        print("   -> 跳转至批量导入/手动录入...")
        manage_processing_fee_flow(client, app_token)
        
    # 2. 检查收款
    print(f"\n{Color.BOLD}2. 收款核对{Color.ENDC}")
    if input("👉 今天是否收到客户的【货款】? (y/n) [n]: ").strip().lower() == 'y':
        manage_processing_payment(client, app_token)
        
    # 3. 检查付款
    print(f"\n{Color.BOLD}3. 付款核对{Color.ENDC}")
    if input("👉 今天是否支付了【供应商货款】或【外协费】? (y/n) [n]: ").strip().lower() == 'y':
        manage_supplier_payment(client, app_token)
        
    # 4. 日常费用
    print(f"\n{Color.BOLD}4. 日常费用{Color.ENDC}")
    if input("👉 今天是否有【打车/餐饮/买菜】等零星支出? (y/n) [n]: ").strip().lower() == 'y':
        quick_entry(client, app_token)
        
    # 4.5 财务体检
    print(f"\n{Color.BOLD}4.5 财务体检 (自动扫描异常){Color.ENDC}")
    financial_health_check(client, app_token)
        
    # 5. 今日汇总
    print(f"\n{Color.BOLD}5. 今日经营快报{Color.ENDC}")
    now = datetime.now()
    today_start = int(datetime(now.year, now.month, now.day).timestamp() * 1000)
    today_end = int((datetime(now.year, now.month, now.day) + timedelta(days=1)).timestamp() * 1000)
    
    income = 0.0
    expense = 0.0
    
    # Ledger
    t_ledger = get_table_id_by_name(client, app_token, "日常台账表")
    if t_ledger:
        filter_cmd = f'AND(CurrentValue.[记账日期]>={today_start}, CurrentValue.[记账日期]<{today_end})'
        recs = get_all_records(client, app_token, t_ledger, filter_info=filter_cmd)
        if recs:
            for r in recs:
                t = r.fields.get("业务类型", "")
                amt = float(r.fields.get("实际收付金额", 0))
                if t == "收款": income += amt
                elif t in ["付款", "费用"]: expense += amt
                
    print(f"   📅 日期: {now.strftime('%Y-%m-%d')}")
    print(f"   💰 今日实收: {Color.OKGREEN}{income:,.2f}{Color.ENDC}")
    print(f"   💸 今日实付: {Color.FAIL}{expense:,.2f}{Color.ENDC}")
    print(f"   📈 今日净流: {income - expense:,.2f}")
    
    # 5.5 生成日报
    generate_daily_html_report(client, app_token)
    
    # 6. 自动备份
    print(f"\n{Color.BOLD}6. 数据归档{Color.ENDC}")
    print("⏳ 正在执行每日自动备份...")
    backup_system_data(client, app_token)
    
    print(f"\n{Color.OKGREEN}✅ 今日结单完成！辛苦了！{Color.ENDC}")
    input("按回车键返回主菜单...")

def clean_partner_names(client, app_token):
    """客户/供应商名称清洗"""
    print(f"\n{Color.CYAN}🧹 客户/供应商名称清洗{Color.ENDC}")
    print("功能：合并重复的客户名称 (如 '张三' 和 '张三门窗' 合并为 '张三门窗')")
    
    # 1. 收集所有名称
    print("⏳ 正在扫描所有记录...")
    names = {} # Name -> Count
    
    # Scan Processing Fee
    t_pf = get_table_id_by_name(client, app_token, "加工费明细表")
    if t_pf:
        recs = get_all_records(client, app_token, t_pf)
        if recs:
            for r in recs:
                n = r.fields.get("往来单位", "").strip()
                if n: names[n] = names.get(n, 0) + 1
                
    # Scan Ledger
    t_lg = get_table_id_by_name(client, app_token, "日常台账表")
    if t_lg:
        recs = get_all_records(client, app_token, t_lg)
        if recs:
            for r in recs:
                n = r.fields.get("往来单位费用", "").strip()
                if n: names[n] = names.get(n, 0) + 1
                
    sorted_names = sorted(names.items(), key=lambda x: x[1], reverse=True)
    
    print(f"\n📊 发现 {len(sorted_names)} 个独立往来单位:")
    for i, (n, c) in enumerate(sorted_names[:20]):
        print(f"   {i+1}. {n} ({c}次)")
    if len(sorted_names) > 20: print("   ...")
    
    print("\n操作选项:")
    print("1. 手动合并名称 (Merge A into B)")
    print("0. 返回")
    
    op = input("👉 请选择: ").strip()
    
    if op == '1':
        old_name = input("请输入【错误/旧】名称 (将被替换): ").strip()
        if old_name not in names:
            print("❌ 名称不存在")
            return
            
        new_name = input("请输入【正确/新】名称 (目标名称): ").strip()
        if not new_name: return
        
        print(f"\n⚠️  即将把所有 '{old_name}' 修改为 '{new_name}'")
        if input("❓ 确认执行? (y/n): ").strip().lower() == 'y':
            count = 0
            # Update PF
            if t_pf:
                pf_recs = get_all_records(client, app_token, t_pf, filter_info=f'CurrentValue.[往来单位]="{old_name}"')
                if pf_recs:
                    batch = []
                    for r in pf_recs:
                        batch.append(AppTableRecord.builder().record_id(r.record_id).fields({"往来单位": new_name}).build())
                    
                    # Batch Update
                    for i in range(0, len(batch), 100):
                        req = BatchUpdateAppTableRecordRequest.builder().app_token(app_token).table_id(t_pf).request_body(BatchUpdateAppTableRecordRequestBody.builder().records(batch[i:i+100]).build()).build()
                        client.bitable.v1.app_table_record.batch_update(req)
                    count += len(pf_recs)
                    
            # Update Ledger
            if t_lg:
                lg_recs = get_all_records(client, app_token, t_lg, filter_info=f'CurrentValue.[往来单位费用]="{old_name}"')
                if lg_recs:
                    batch = []
                    for r in lg_recs:
                        batch.append(AppTableRecord.builder().record_id(r.record_id).fields({"往来单位费用": new_name}).build())
                        
                    for i in range(0, len(batch), 100):
                        req = BatchUpdateAppTableRecordRequest.builder().app_token(app_token).table_id(t_lg).request_body(BatchUpdateAppTableRecordRequestBody.builder().records(batch[i:i+100]).build()).build()
                        client.bitable.v1.app_table_record.batch_update(req)
                    count += len(lg_recs)
            
            print(f"✅ 已合并 {count} 条记录！")

# 全局台账缓存 (用于快速查账)
GLOBAL_LEDGER_CACHE = None

def quick_search_ledger(client, app_token):
    """快速查账 (优化版：支持金额、日期、关键词智能搜索)"""
    global GLOBAL_LEDGER_CACHE
    
    print(f"\n{Color.HEADER}🔍 万能查账助手{Color.ENDC}")
    print(f"{Color.CYAN}提示：支持输入 金额(100)、日期(2024-01)、关键词(京东){Color.ENDC}")
    
    table_id = get_table_id_by_name(client, app_token, "日常台账表")
    if not table_id: return
    
    # 首次加载或刷新
    if GLOBAL_LEDGER_CACHE is None:
        print("⏳ 正在拉取全量台账数据 (首次加载)...")
        GLOBAL_LEDGER_CACHE = get_all_records(client, app_token, table_id)
        print(f"✅ 已缓存 {len(GLOBAL_LEDGER_CACHE)} 条记录")
    else:
        print(f"⚡ 使用本地缓存 ({len(GLOBAL_LEDGER_CACHE)} 条) - 输入 'reload' 强制刷新")
    
    import re

    while True:
        print("-" * 30)
        query = input("👉 请输入查询内容 (q:退出, reload:刷新): ").strip()
        
        if not query: continue
        if query.lower() == 'q': break
        
        if query.lower() == 'reload':
            print("🔄 正在刷新数据...")
            GLOBAL_LEDGER_CACHE = get_all_records(client, app_token, table_id)
            print(f"✅ 刷新完成: {len(GLOBAL_LEDGER_CACHE)} 条")
            continue
            
        matches = []
        total_income = 0.0
        total_expense = 0.0
        
        # 智能解析查询意图
        target_amt = None
        target_month = None
        target_date = None
        is_fuzzy_text = True
        
        # 1. 尝试解析为金额
        try:
            target_amt = float(query)
            # 即使解析为金额，如果它是整数，也可能是文本的一部分（如单号），所以不完全禁用文本搜索
            # 但为了精准，如果用户输入的是 100.00，那肯定是金额。如果是 2024，可能是年份。
            # 策略：如果匹配到金额，就加入；同时如果文本匹配，也加入。
        except:
            pass
            
        # 2. 尝试解析为日期 (YYYY-MM 或 YYYY-MM-DD)
        if re.match(r"^\d{4}-\d{2}$", query):
            target_month = query
            is_fuzzy_text = False # 明确是月份搜索
        elif re.match(r"^\d{4}-\d{2}-\d{2}$", query):
            target_date = query
            is_fuzzy_text = False # 明确是日期搜索

        print(f"🔎 正在搜索: {query} ...")
        
        for r in GLOBAL_LEDGER_CACHE:
            f = r.fields
            
            # 提取关键字段
            ts = f.get("记账日期", 0)
            date_str = datetime.fromtimestamp(ts/1000).strftime("%Y-%m-%d") if ts else ""
            r_amt = float(f.get("实际收付金额", 0))
            desc = f.get("备注", "")
            partner = f.get("往来单位费用", "")
            cat = f.get("费用归类", "")
            b_type = f.get("业务类型", "")
            
            # 构建全文本索引 (用于模糊搜)
            full_text = f"{date_str} {r_amt} {desc} {partner} {b_type} {cat}".lower()
            
            matched = False
            
            # 逻辑 A: 金额匹配 (绝对值误差 0.01)
            if target_amt is not None:
                if abs(abs(r_amt) - abs(target_amt)) < 0.01:
                    matched = True
            
            # 逻辑 B: 日期匹配
            if target_month and date_str.startswith(target_month):
                matched = True
            if target_date and date_str == target_date:
                matched = True
                    
            # 逻辑 C: 文本匹配 (只要关键词在任意字段中)
            if query.lower() in full_text:
                matched = True
            
            if matched:
                # 注入 record_id 以便后续操作 (如删除)
                item = r.fields.copy()
                if hasattr(r, 'record_id'):
                    item['_record_id'] = r.record_id
                matches.append(item)
                
                if b_type == "收款":
                    total_income += r_amt
                elif b_type in ["付款", "费用"]:
                    total_expense += r_amt

        if matches:
            print(f"\n✅ 找到 {len(matches)} 条记录:")
            print(f"{'日期':<12} | {'类型':<6} | {'金额':<10} | {'往来单位/备注'}")
            print("-" * 65)
            
            # 按日期排序
            matches.sort(key=lambda x: x.get("记账日期", 0), reverse=True)
            
            limit = 20
            for m in matches[:limit]:
                ts = m.get("记账日期", 0)
                d_str = datetime.fromtimestamp(ts/1000).strftime("%Y-%m-%d") if ts else "-"
                amt = float(m.get("实际收付金额", 0))
                b_type = m.get("业务类型", "")
                desc = f"{m.get('往来单位费用', '')} {m.get('备注', '')}".replace('\n', ' ')
                
                # Color
                amt_str = f"{amt:,.2f}"
                row_str = f"{d_str:<12} | {b_type:<6} | {amt_str:<10} | {desc[:30]}"
                if b_type == "收款":
                    print(f"{Color.GREEN}{row_str}{Color.ENDC}")
                else:
                    print(row_str)
                
            if len(matches) > limit:
                print(f"... (还有 {len(matches)-limit} 条，建议导出)")
                
            print("-" * 65)
            net = total_income - total_expense
            print(f"💰 统计结果: 收入 {total_income:,.2f} | 支出 {total_expense:,.2f} | 净额 {net:,.2f}")
            
            # 操作选项
            print(f"{Color.CYAN}操作: [x]Excel [h]HTML报表 [d]删除记录 [n]新搜索{Color.ENDC}")
            opt = input("👉 请输入操作指令 [n]: ").strip().lower()
            
            if opt == 'd':
                try:
                    max_idx = min(len(matches), limit)
                    del_idx = int(input(f"👉 输入要删除的序号 (1-{max_idx}): ")) - 1
                    if 0 <= del_idx < max_idx:
                        target = matches[del_idx]
                        rid = target.get('_record_id')
                        if not rid:
                            print("❌ 无法删除：未找到记录ID")
                            continue
                        
                        ts = target.get('记账日期', 0)
                        d_str = datetime.fromtimestamp(ts/1000).strftime("%Y-%m-%d") if ts else "-"
                        desc_str = f"{d_str} | {target.get('实际收付金额', 0)} | {target.get('备注', '')}"
                        
                        confirm = input(f"{Color.FAIL}⚠️ 确认删除? {desc_str} (y/n): {Color.ENDC}")
                        
                        if confirm.lower() == 'y':
                            req = DeleteAppTableRecordRequest.builder() \
                                .app_token(app_token) \
                                .table_id(table_id) \
                                .record_id(rid) \
                                .build()
                            resp = client.bitable.v1.app_table_record.delete(req)
                            
                            if resp.success():
                                print("✅ 删除成功")
                                # Update cache
                                GLOBAL_LEDGER_CACHE = [r for r in GLOBAL_LEDGER_CACHE if getattr(r, 'record_id', '') != rid]
                                print("🔄 数据已更新")
                            else:
                                print(f"❌ 删除失败: {resp.msg}")
                    else:
                        print("❌ 序号无效")
                except ValueError:
                    print("❌ 输入无效")

            elif opt == 'x':
                try:
                    df = pd.DataFrame(matches)
                    # 简单清洗列
                    cols = ["记账日期", "业务类型", "费用归类", "实际收付金额", "往来单位费用", "备注"]
                    # 确保列存在
                    exist_cols = [c for c in cols if c in df.columns]
                    df = df[exist_cols]
                    # 格式化日期
                    if "记账日期" in df.columns:
                        df["记账日期"] = df["记账日期"].apply(lambda x: datetime.fromtimestamp(x/1000).strftime("%Y-%m-%d") if x else "")
                    
                    fname = f"查询结果_{query.replace(':','-')}_{datetime.now().strftime('%H%M%S')}.xlsx"
                    df.to_excel(fname, index=False)
                    print(f"✅ 已导出: {fname}")
                    os.startfile(fname) # Windows only
                except Exception as e:
                    print(f"❌ 导出失败: {e}")

            elif opt == 'h':
                try:
                    # 准备数据
                    # 1. 支出分类统计
                    cat_stats = {}
                    for m in matches:
                        if m.get("业务类型") in ["付款", "费用"]:
                            c = m.get("费用归类", "未分类")
                            if not c: c = "未分类"
                            amt = float(m.get("实际收付金额", 0))
                            cat_stats[c] = cat_stats.get(c, 0) + amt
                    
                    sorted_cats = sorted(cat_stats.items(), key=lambda x: x[1], reverse=True)[:10] # Top 10
                    
                    # 生成 SVG 柱状图
                    chart_html = ""
                    if sorted_cats:
                        max_val = sorted_cats[0][1] if sorted_cats else 1
                        svg_height = len(sorted_cats) * 40 + 20
                        svg_bars = ""
                        for idx, (k, v) in enumerate(sorted_cats):
                            y = idx * 40
                            w = (v / max_val) * 500
                            color = "#e74c3c"
                            svg_bars += f"""
                            <g transform="translate(0, {y})">
                                <text x="0" y="20" font-size="12" fill="#555">{k}</text>
                                <rect x="100" y="5" width="{w}" height="20" fill="{color}" rx="3" opacity="0.8"/>
                                <text x="{100 + w + 10}" y="20" font-size="12" fill="#333">{v:,.2f}</text>
                            </g>
                            """
                        
                        chart_html = f"""
                        <div class="chart-container">
                            <div class="chart-title">💸 支出分类 TOP10</div>
                            <svg width="100%" height="{svg_height}" viewBox="0 0 800 {svg_height}">
                                <g transform="translate(20, 10)">
                                    {svg_bars}
                                </g>
                            </svg>
                        </div>
                        """
                    
                    # 生成 HTML
                    html = f"""
                    <!DOCTYPE html>
                    <html>
                    <head>
                        <meta charset="utf-8">
                        <title>查账报告 - {query}</title>
                        <style>
                            body {{ font-family: 'Segoe UI', sans-serif; background: #f5f7fa; color: #333; padding: 20px; }}
                            .container {{ max-width: 900px; margin: 0 auto; background: white; padding: 40px; border-radius: 12px; box-shadow: 0 4px 20px rgba(0,0,0,0.05); }}
                            .header {{ border-bottom: 2px solid #eee; padding-bottom: 20px; margin-bottom: 30px; }}
                            .title h1 {{ margin: 0; font-size: 24px; color: #2c3e50; }}
                            .cards {{ display: flex; gap: 20px; margin-bottom: 40px; }}
                            .card {{ flex: 1; background: #f8f9fa; padding: 20px; border-radius: 8px; text-align: center; border: 1px solid #e9ecef; }}
                            .card .val {{ font-size: 24px; font-weight: bold; margin-bottom: 5px; }}
                            .card .lbl {{ font-size: 13px; color: #7f8c8d; }}
                            .c-in {{ color: #27ae60; }}
                            .c-out {{ color: #c0392b; }}
                            .c-net {{ color: #2980b9; }}
                            
                            table {{ width: 100%; border-collapse: collapse; font-size: 14px; }}
                            th {{ text-align: left; padding: 12px; background: #f8f9fa; color: #7f8c8d; border-bottom: 2px solid #eee; }}
                            td {{ padding: 12px; border-bottom: 1px solid #f1f1f1; }}
                            tr:hover {{ background: #fafafa; }}
                            
                            .chart-container {{ margin: 30px 0; padding: 20px; border: 1px solid #eee; border-radius: 8px; background: #fff; }}
                            .chart-title {{ font-size: 14px; font-weight: bold; color: #34495e; margin-bottom: 15px; }}
                        </style>
                    </head>
                    <body>
                        <div class="container">
                            <div class="header">
                                <div class="title">
                                    <h1>🔎 查账报告: {query}</h1>
                                    <p style="color:#999; font-size:12px; margin-top:5px">生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}</p>
                                </div>
                            </div>
                            
                            <div class="cards">
                                <div class="card">
                                    <div class="val c-in">+{total_income:,.2f}</div>
                                    <div class="lbl">总收入</div>
                                </div>
                                <div class="card">
                                    <div class="val c-out">-{total_expense:,.2f}</div>
                                    <div class="lbl">总支出</div>
                                </div>
                                <div class="card">
                                    <div class="val c-net">{net:+,.2f}</div>
                                    <div class="lbl">净额</div>
                                </div>
                                <div class="card">
                                    <div class="val">{len(matches)}</div>
                                    <div class="lbl">记录数</div>
                                </div>
                            </div>
                            
                            {chart_html}
                            
                            <h3>📝 详细记录</h3>
                            <table>
                                <thead>
                                    <tr>
                                        <th>日期</th>
                                        <th>类型</th>
                                        <th>费用归类</th>
                                        <th>往来/备注</th>
                                        <th style="text-align:right">金额</th>
                                    </tr>
                                </thead>
                                <tbody>
                    """
                    
                    for m in matches:
                        ts = m.get("记账日期", 0)
                        d_str = datetime.fromtimestamp(ts/1000).strftime("%Y-%m-%d") if ts else "-"
                        b_type = m.get("业务类型", "")
                        cat = m.get("费用归类", "")
                        desc = f"{m.get('往来单位费用', '')} {m.get('备注', '')}"
                        amt = float(m.get("实际收付金额", 0))
                        
                        color = "#333"
                        if b_type == "收款": color = "#27ae60"
                        elif b_type in ["付款", "费用"]: color = "#c0392b"
                        
                        html += f"""
                        <tr>
                            <td>{d_str}</td>
                            <td><span style="padding:2px 6px; border-radius:4px; font-size:12px; background:{'#e8f5e9' if b_type=='收款' else '#ffebee'}; color:{color}">{b_type}</span></td>
                            <td>{cat}</td>
                            <td>{desc}</td>
                            <td style="text-align:right; font-weight:bold; color:{color}">{amt:,.2f}</td>
                        </tr>
                        """
                        
                    html += """
                                </tbody>
                            </table>
                        </div>
                    </body>
                    </html>
                    """
                    
                    report_dir = REPORT_DIR
                    # if not os.path.exists(report_dir): os.makedirs(report_dir) # Already created
                    fname = os.path.join(report_dir, f"查账_{query.replace(':','-')}_{datetime.now().strftime('%Y%m%d%H%M%S')}.html")
                    
                    with open(fname, "w", encoding="utf-8") as f:
                        f.write(html)
                        
                    print(f"✅ 报表已生成: {fname}")
                    try:
                        os.startfile(os.path.abspath(fname))
                    except:
                        pass
                    
                except Exception as e:
                    print(f"❌ 生成报表失败: {e}")
                    import traceback
                    traceback.print_exc()
                    
        else:
            print("❌ 未找到相关记录")

# 菜单：设置
def settings_menu():
    print("\n⚙️  系统设置 (修改后自动保存到 .env)")
    print("-----------------------------------")
    print(f"1. 增值税率 (当前: {VAT_RATE}%)")
    print(f"2. 对账容差天数 (当前: {TOLERANCE_DAYS}天)")
    print(f"3. 智谱AI Key (当前: {ZHIPUAI_API_KEY[:8]}...)" if ZHIPUAI_API_KEY else "3. 智谱AI Key (当前: 未配置)")
    print("0. 返回主菜单")
    
    choice = input("\n请选择要修改的项 (0-3): ").strip()
    
    if choice == "1":
        val = input("请输入新的税率 (例如 3): ").strip()
        if val.isdigit():
            update_env("VAT_RATE", val)
            print("✅ 税率已更新")
        else:
            print("❌ 输入无效")
            
    elif choice == "2":
        val = input("请输入新的容差天数 (例如 2): ").strip()
        if val.isdigit():
            update_env("TOLERANCE_DAYS", val)
            print("✅ 容差天数已更新")
        else:
            print("❌ 输入无效")
            
    elif choice == "3":
        val = input("请输入智谱AI API Key: ").strip()
        if val:
            update_env("ZHIPUAI_API_KEY", val)
            print("✅ API Key 已更新")
    
    elif choice == "0":
        return

def manage_invoice_status(client, app_token):
    """开票状态管理 (Mark as Invoiced)"""
    print(f"\n{Color.CYAN}🧾 开票管理 (Invoice Status){Color.ENDC}")
    table_id = get_table_id_by_name(client, app_token, "加工费明细表")
    if not table_id:
        print("❌ 未找到加工费明细表")
        return

    # 1. 统计未开票金额
    print("🔄 正在统计未开票金额...")
    
    # Filter: Status != "已开票"
    records = get_all_records(client, app_token, table_id)
    if not records:
        print("📭 暂无记录")
        return
        
    uninvoiced_map = {} # Cust -> Amount
    uninvoiced_records = []
    
    for r in records:
        f = r.fields
        status = f.get("开票状态", "未开票")
        if status == "已开票": continue
        
        # 仅收入才需要开票
        t = f.get("类型", "")
        if "收入" not in t: continue
        
        cust = f.get("往来单位", "未知")
        amt = float(f.get("总金额", 0))
        # Filter out 0 amount
        if amt == 0: continue
        
        uninvoiced_map[cust] = uninvoiced_map.get(cust, 0) + amt
        uninvoiced_records.append(r)
        
    if not uninvoiced_map:
        print("✅ 所有收入均已开票！")
        return
        
    # Show Top 10
    sorted_cust = sorted(uninvoiced_map.items(), key=lambda x: x[1], reverse=True)
    
    print("\n📊 待开票排行榜 (Top 10):")
    for i, (c, amt) in enumerate(sorted_cust[:10]):
        print(f"   {i+1}. {c}: {Color.WARNING}{amt:,.2f} 元{Color.ENDC}")
        
    print(f"\n   >> 总待开票金额: {sum(uninvoiced_map.values()):,.2f} 元")
    
    # Actions
    print("\n操作选项:")
    print("1. 按客户批量开票 (Mark Customer as Invoiced)")
    print("2. 按月份批量开票 (Mark Month as Invoiced)")
    print("0. 返回")
    
    op = input("👉 请选择: ").strip()
    
    if op == '1':
        target = input("请输入客户名 (关键词): ").strip()
        if not target: return
        
        # Filter
        matches = [c for c in uninvoiced_map.keys() if target in c]
        if not matches:
            print("❌ 未找到匹配客户")
            return
            
        if len(matches) > 1:
            print(f"🔍 匹配到多个客户: {matches}")
            target = input("👉 请输入完整客户名确认: ").strip()
            if target not in matches: return
        else:
            target = matches[0]
            
        # Confirm
        total = uninvoiced_map[target]
        print(f"\n准备将 {Color.BOLD}{target}{Color.ENDC} 的 {len([r for r in uninvoiced_records if r.fields.get('往来单位')==target])} 笔记录标记为已开票。")
        print(f"涉及金额: {total:,.2f} 元")
        
        if input("❓ 确认执行? (y/n): ").strip().lower() == 'y':
            # Batch Update
            batch_recs = []
            for r in uninvoiced_records:
                if r.fields.get("往来单位") == target:
                    batch_recs.append(AppTableRecord.builder().record_id(r.record_id).fields({"开票状态": "已开票"}).build())
            
            # Execute Batch
            count = 0
            for i in range(0, len(batch_recs), 100):
                batch = batch_recs[i:i+100]
                req = BatchUpdateAppTableRecordRequest.builder() \
                    .app_token(app_token) \
                    .table_id(table_id) \
                    .request_body(BatchUpdateAppTableRecordRequestBody.builder().records(batch).build()) \
                    .build()
                resp = client.bitable.v1.app_table_record.batch_update(req)
                if resp.success():
                    count += len(batch)
            print(f"✅ 成功标记 {count} 笔记录为已开票")

    elif op == '2':
        month_str = input("请输入月份 (YYYY-MM): ").strip()
        try:
            target_dt = datetime.strptime(month_str, "%Y-%m")
            # Filter
            batch_recs = []
            total_amt = 0
            for r in uninvoiced_records:
                ts = r.fields.get("日期", 0)
                rdt = datetime.fromtimestamp(ts/1000)
                if rdt.year == target_dt.year and rdt.month == target_dt.month:
                    batch_recs.append(AppTableRecord.builder().record_id(r.record_id).fields({"开票状态": "已开票"}).build())
                    total_amt += float(r.fields.get("总金额", 0))
            
            if not batch_recs:
                print("❌ 该月份无待开票记录")
                return
                
            print(f"\n准备将 {month_str} 的 {len(batch_recs)} 笔记录标记为已开票。")
            print(f"涉及金额: {total_amt:,.2f} 元")
            
            if input("❓ 确认执行? (y/n): ").strip().lower() == 'y':
                 # Execute Batch
                count = 0
                for i in range(0, len(batch_recs), 100):
                    batch = batch_recs[i:i+100]
                    req = BatchUpdateAppTableRecordRequest.builder() \
                        .app_token(app_token) \
                        .table_id(table_id) \
                        .request_body(BatchUpdateAppTableRecordRequestBody.builder().records(batch).build()) \
                        .build()
                    resp = client.bitable.v1.app_table_record.batch_update(req)
                    if resp.success():
                        count += len(batch)
                print(f"✅ 成功标记 {count} 笔记录为已开票")
                
        except:
            print("❌ 日期格式错误")



def generate_statement_html(cust_name, month_str, items, total_qty, total_amt, save_dir):
    """生成对账单 HTML 版本"""
    fname = os.path.join(save_dir, f"{str(cust_name).replace('/','_')}_{month_str}_对账单.html")
    
    rows = ""
    for idx, it in enumerate(items):
        bg = "#f9f9f9" if idx % 2 == 0 else "#fff"
        rows += f"""
        <tr style="background-color:{bg}">
            <td>{it['日期']}</td>
            <td>{it['品名']}</td>
            <td>{it['规格']}</td>
            <td style="text-align:right">{it['数量']}</td>
            <td style="text-align:center">{it['单位']}</td>
            <td style="text-align:right">{it['单价']:.2f}</td>
            <td style="text-align:right;font-weight:bold">{it['金额']:.2f}</td>
            <td style="color:#666;font-size:0.8em">{it['备注']}</td>
        </tr>
        """
        
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <title>{cust_name} 对账单 {month_str}</title>
        <style>
            body {{ font-family: 'Segoe UI', 'Microsoft YaHei', sans-serif; max-width: 900px; margin: 0 auto; padding: 30px; color: #333; }}
            .header {{ text-align: center; border-bottom: 3px solid #3498db; padding-bottom: 20px; margin-bottom: 30px; }}
            .title {{ font-size: 28px; font-weight: bold; color: #2c3e50; }}
            .subtitle {{ font-size: 16px; color: #7f8c8d; margin-top: 5px; }}
            .info-box {{ display: flex; justify-content: space-between; margin-bottom: 30px; background: #f8f9fa; padding: 20px; border-radius: 8px; }}
            .info-item {{ font-size: 14px; }}
            .label {{ color: #7f8c8d; font-weight: 600; }}
            
            table {{ width: 100%; border-collapse: collapse; margin-bottom: 30px; }}
            th {{ background: #3498db; color: white; padding: 12px 8px; text-align: left; font-size: 14px; }}
            td {{ padding: 10px 8px; border-bottom: 1px solid #eee; font-size: 14px; }}
            
            .summary {{ display: flex; justify-content: flex-end; margin-top: 20px; }}
            .total-box {{ background: #fff3cd; padding: 15px 30px; border-radius: 8px; border: 1px solid #ffeeba; }}
            .total-line {{ font-size: 16px; margin: 5px 0; text-align: right; }}
            .grand-total {{ font-size: 24px; font-weight: bold; color: #d35400; border-top: 1px solid #e0c49e; padding-top: 10px; margin-top: 5px; }}
            
            .footer {{ margin-top: 50px; border-top: 1px solid #eee; padding-top: 20px; display: flex; justify-content: space-between; font-size: 14px; color: #7f8c8d; }}
            .sign-area {{ width: 200px; height: 80px; border-bottom: 1px solid #333; margin-top: 30px; }}
        </style>
    </head>
    <body>
        <div class="header">
            <div class="title">往来对账单 Statement of Account</div>
            <div class="subtitle">月份 Period: {month_str}</div>
        </div>
        
        <div class="info-box">
            <div>
                <div class="info-item"><span class="label">往来单位 (Partner):</span> {cust_name}</div>
                <div class="info-item"><span class="label">打印日期 (Date):</span> {datetime.now().strftime('%Y-%m-%d')}</div>
            </div>
            <div style="text-align:right">
                <div class="info-item"><span class="label">共计笔数:</span> {len(items)} 笔</div>
            </div>
        </div>
        
        <table>
            <thead>
                <tr>
                    <th width="12%">日期</th>
                    <th width="20%">品名</th>
                    <th width="15%">规格</th>
                    <th width="10%" style="text-align:right">数量</th>
                    <th width="8%" style="text-align:center">单位</th>
                    <th width="10%" style="text-align:right">单价</th>
                    <th width="12%" style="text-align:right">金额</th>
                    <th width="13%">备注</th>
                </tr>
            </thead>
            <tbody>
                {rows}
            </tbody>
        </table>
        
        <div class="summary">
            <div class="total-box">
                <div class="total-line">数量合计: <b>{total_qty:,.2f}</b></div>
                <div class="total-line grand-total">金额合计: ¥ {total_amt:,.2f}</div>
            </div>
        </div>
        
        <div class="footer">
            <div style="text-align:center">
                <div>我方制单 (Prepared By)</div>
                <div class="sign-area"></div>
            </div>
            <div style="text-align:center">
                <div>对方确认 (Confirmed By)</div>
                <div class="sign-area"></div>
                <div>请核对无误后签字盖章回传</div>
            </div>
        </div>
    </body>
    </html>
    """
    
    with open(fname, "w", encoding="utf-8") as f:
        f.write(html)
    return fname

def batch_generate_business_statements(client, app_token, pre_mode=None):
    """批量生成对账单 (支持 客户加工费 / 供应商外协费)"""
    print(f"\n{Color.HEADER}📑 批量生成业务对账单{Color.ENDC}")
    print("--------------------------------")
    
    if pre_mode:
        mode_choice = str(pre_mode)
    else:
        print("1. 客户对账单 (收入-加工服务) - 发给客户")
        print("2. 供应商对账单 (支出-外协加工) - 发给外协厂")
        print("0. 返回")
        mode_choice = input("👉 请选择 (1/2): ").strip()
        
    if mode_choice == '0': return
    
    target_type = "收入-加工服务"
    mode_name = "客户"
    if mode_choice == '2':
        target_type = "支出-外协加工"
        mode_name = "供应商"
        
    table_id = get_table_id_by_name(client, app_token, "加工费明细表")
    if not table_id: return

    # 选择月份
    now = datetime.now()
    default_input = now.strftime("%Y-%m")
    if now.day <= 10:
        last_month_dt = now.replace(day=1) - timedelta(days=1)
        default_input = last_month_dt.strftime("%Y-%m")
        
    user_input = input(f"请输入对账月份 (YYYY-MM) [{default_input}]: ").strip()
    if not user_input: user_input = default_input
    
    try:
        start_dt = datetime.strptime(user_input, "%Y-%m")
        if start_dt.month == 12:
            end_dt = datetime(start_dt.year + 1, 1, 1)
        else:
            end_dt = datetime(start_dt.year, start_dt.month + 1, 1)
        
        start_ts = int(start_dt.timestamp() * 1000)
        end_ts = int(end_dt.timestamp() * 1000)
    except:
        print("❌ 日期格式错误")
        return

    # 加载别名
    aliases = {}
    if os.path.exists(FILE_PARTNER_ALIASES):
        try:
            with open(FILE_PARTNER_ALIASES, "r", encoding="utf-8") as f:
                aliases = json.load(f)
        except: pass

    # 拉取数据
    print(f"⏳ 正在拉取 {mode_name} 数据 ({user_input})...")
    filter_cmd = f'AND(CurrentValue.[日期]>={start_ts}, CurrentValue.[日期]<{end_ts}, CurrentValue.[类型]="{target_type}")'
    records = get_all_records(client, app_token, table_id, filter_info=filter_cmd)
    
    if not records:
        print(f"📭 {user_input} 无{mode_name}记录")
        return

    # 分组数据
    partner_data = {} # Partner -> List of dict
    
    for r in records:
        f = r.fields
        raw_p = f.get("往来单位", "未知单位")
        p = aliases.get(raw_p, raw_p)
        
        ts = f.get("日期", 0)
        d_str = datetime.fromtimestamp(ts/1000).strftime("%Y-%m-%d")
        
        item = {
            "日期": d_str,
            "品名": f.get("品名", ""),
            "规格": f.get("规格", ""),
            "数量": float(f.get("数量", 0)),
            "单位": f.get("单位", "件"),
            "单价": float(f.get("单价", 0)),
            "金额": float(f.get("总金额", 0)),
            "备注": f.get("备注", "")
        }
        if p not in partner_data: partner_data[p] = []
        partner_data[p].append(item)
        
    # 生成文件
    save_dir = os.path.join(DATA_ROOT, f"{mode_name}对账单", user_input)
    if not os.path.exists(save_dir): os.makedirs(save_dir)
    
    print(f"📂 正在生成对账单 (共 {len(partner_data)} 家)...")
    
    for p_name, items in partner_data.items():
        # 按日期排序
        items.sort(key=lambda x: x["日期"])
        
        # 转 DataFrame
        df = pd.DataFrame(items)
        
        # 添加合计行
        total_qty = df["数量"].sum()
        total_amt = df["金额"].sum()
        
        total_row = pd.DataFrame([{
            "日期": "合计",
            "品名": f"{len(items)} 笔",
            "规格": "",
            "数量": total_qty,
            "单位": "",
            "单价": "",
            "金额": total_amt,
            "备注": ""
        }])
        df = pd.concat([df, total_row], ignore_index=True)
        
        safe_name = str(p_name).replace("/", "_").replace("\\", "_")
        
        # 1. Excel
        fname_xlsx = os.path.join(save_dir, f"{safe_name}_{user_input}_对账单.xlsx")
        try:
            with pd.ExcelWriter(fname_xlsx, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name="对账单")
                ws = writer.sheets["对账单"]
                apply_excel_styles(ws)
                # 调整列宽
                ws.column_dimensions['A'].width = 12
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['G'].width = 12
                ws.column_dimensions['H'].width = 20
        except Exception as e:
            print(f"❌ 生成 Excel 失败: {e}")

        # 2. HTML (Visual)
        try:
            generate_statement_html(p_name, user_input, items, total_qty, total_amt, save_dir)
        except Exception as e:
            print(f"❌ 生成 HTML 失败: {e}")
            
    print(f"✅ 全部生成完毕！文件保存在: {Color.UNDERLINE}{save_dir}{Color.ENDC}")
    try: os.startfile(save_dir)
    except: pass

def reconciliation_hub(client, app_token):
    """往来对账中心 (Reconciliation Hub)"""
    while True:
        print(f"\n{Color.HEADER}🤝 往来对账中心 (Reconciliation Center){Color.ENDC}")
        print("-----------------------------------------------")
        print("  1. 📤 批量生成客户对账单 (加工费收入) [Excel/HTML]")
        print("  2. 📤 批量生成供应商对账单 (外协费支出) [Excel/HTML]")
        print("  3. 📊 生成往来单位余额表 (应收应付总览)")
        print("  4. 📥 外部账单智能比对 (Excel vs 系统台账)")
        print("  5. 💰 资金账户对账 (余额核对)")
        print("  0. 返回主菜单")
        
        choice = input(f"\n👉 请选择: ").strip()
        
        if choice == '0': break
        elif choice == '1':
            batch_generate_business_statements(client, app_token, pre_mode=1)
        elif choice == '2':
            batch_generate_business_statements(client, app_token, pre_mode=2)
        elif choice == '3':
            generate_business_statement(client, app_token)
        elif choice == '4':
            reconcile_partner_flow(client, app_token)
        elif choice == '5':
            reconcile_bank_account(client, app_token)

def batch_generate_customer_statements(client, app_token):
    # Deprecated wrapper, redirect to new function
    batch_generate_business_statements(client, app_token)

def manage_supplier_payment(client, app_token):
    """供应商付款登记 (按实际发生)"""
    print(f"\n{Color.CYAN}💸 供应商付款登记 (按实际发生){Color.ENDC}")
    print("说明: 记录付给供应商的实际款项（预付/尾款），并自动同步到【日常台账】。")
    
    # 1. 输入信息
    date_str = input(f"付款日期 (YYYY-MM-DD) [默认今天]: ").strip()
    if not date_str: date_str = datetime.now().strftime("%Y-%m-%d")
    
    partner = input("供应商名称: ").strip()
    if not partner: return
    
    amount = float(input("付款金额 (元): ").strip())
    
    bank_map = {'1': 'G银行基本户(有票)', '2': 'N银行/微信(无票)'}
    print("付款账户:")
    print("1. G银行基本户(有票)")
    print("2. N银行/微信(无票)")
    b_choice = input("👉 请选择 (1/2): ").strip()
    bank = bank_map.get(b_choice, 'G银行基本户(有票)')
    
    remark = input("备注 (如 '1月材料款'): ").strip()
    
    # 2. 写入日常台账 (作为总的付款记录)
    ledger_id = get_table_id_by_name(client, app_token, "日常台账表")
    if not ledger_id:
        print("❌ 未找到日常台账表")
        return
        
    fields = {
        "记账日期": int(pd.to_datetime(date_str).timestamp() * 1000),
        "业务类型": "付款", # 付款
        "费用归类": "外协加工费", # 默认归类，后续可能需要细化
        "往来单位费用": partner,
        "实际收付金额": amount,
        "账面金额": amount,
        "交易银行": bank,
        "是否有票": "有票" if "有票" in bank else "无票",
        "是否现金": "否" if "有票" in bank else "是",
        "备注": f"{remark} (付款登记)"
    }
    
    # 询问费用归类
    print("费用归类:")
    print("1. 原材料-三酸/片碱/色粉")
    print("2. 辅料-挂具/除油剂")
    print("3. 外协加工费")
    print("4. 房租水电")
    print("5. 其他")
    c_choice = input("👉 请选择 (1-5) [默认3]: ").strip()
    cats = {'1': '原材料-三酸/片碱/色粉', '2': '辅料-挂具/除油剂', '3': '外协加工费', '4': '房租水电', '5': '其他'}
    fields["费用归类"] = cats.get(c_choice, '外协加工费')
    
    req = CreateAppTableRecordRequest.builder() \
        .app_token(app_token) \
        .table_id(ledger_id) \
        .request_body(AppTableRecord.builder().fields(fields).build()) \
        .build()
        
    resp = client.bitable.v1.app_table_record.create(req)
    if resp.success():
        print(f"✅ 付款已记录到台账！")
        
        # 3. 智能核销建议 (仅针对外协加工费)
        if fields["费用归类"] == "外协加工费":
            if input("👉 是否要自动核销该供应商的旧欠款(外协费)? (y/n) [y]: ").strip().lower() != 'n':
                 # 查找该供应商未结算的记录 (在加工费明细表中，类型=支出-外协加工)
                 pf_table_id = get_table_id_by_name(client, app_token, "加工费明细表")
                 if pf_table_id:
                     # 获取所有未结算
                     filter_cmd = f'AND(CurrentValue.[往来单位]="{partner}", CurrentValue.[结算状态]!="已结算", CurrentValue.[类型]="支出-外协加工")'
                     unpaid_recs = get_all_records(client, app_token, pf_table_id, filter_info=filter_cmd)
                     
                     if unpaid_recs:
                         # 按日期排序 (FIFO)
                         unpaid_recs.sort(key=lambda x: x.fields.get("日期", 0))
                         
                         to_settle = []
                         remaining = amount
                         
                         for r in unpaid_recs:
                             rec_amt = float(r.fields.get("总金额", 0))
                             if remaining >= rec_amt:
                                 to_settle.append(r)
                                 remaining -= rec_amt
                             else:
                                 break 
                                 
                         if to_settle:
                             print(f"💡 系统建议核销最早的 {len(to_settle)} 笔未结算外协记录 (共 {amount - remaining:,.2f} 元)")
                             if input("❓ 确认核销? (y/n): ").strip().lower() == 'y':
                                 batch_recs = []
                                 for r in to_settle:
                                     batch_recs.append(AppTableRecord.builder().record_id(r.record_id).fields({"结算状态": "已结算"}).build())
                                 
                                 # Execute Batch
                                 for i in range(0, len(batch_recs), 100):
                                     batch = batch_recs[i:i+100]
                                     req_b = BatchUpdateAppTableRecordRequest.builder() \
                                         .app_token(app_token) \
                                         .table_id(pf_table_id) \
                                         .request_body(BatchUpdateAppTableRecordRequestBody.builder().records(batch).build()) \
                                         .build()
                                     client.bitable.v1.app_table_record.batch_update(req_b)
                                 print(f"✅ 已自动核销 {len(to_settle)} 笔记录")
                         else:
                             print("⚠️ 付款金额不足以核销最早的一笔记录，暂不执行核销。")
                     else:
                         print("🎉 该供应商没有未结算的外协记录。")
    else:
        print(f"❌ 记录失败: {resp.msg}")

def manage_processing_payment(client, app_token):
    """客户收款登记 (按实际发生)"""
    print(f"\n{Color.CYAN}💰 客户收款登记 (按实际发生){Color.ENDC}")
    print("说明: 记录客户的实际付款（预收/尾款），并自动同步到【日常台账】。")
    
    # 1. 输入信息
    date_str = input(f"收款日期 (YYYY-MM-DD) [默认今天]: ").strip()
    if not date_str: date_str = datetime.now().strftime("%Y-%m-%d")
    
    partner = input("客户名称: ").strip()
    if not partner: return
    
    amount = float(input("收款金额 (元): ").strip())
    
    bank_map = {'1': 'G银行基本户(有票)', '2': 'N银行/微信(无票)'}
    print("收款账户:")
    print("1. G银行基本户(有票)")
    print("2. N银行/微信(无票)")
    b_choice = input("👉 请选择 (1/2): ").strip()
    bank = bank_map.get(b_choice, 'G银行基本户(有票)')
    
    remark = input("备注 (如 '1月货款'): ").strip()
    
    # 2. 写入日常台账 (作为总的收款记录)
    ledger_id = get_table_id_by_name(client, app_token, "日常台账表")
    if not ledger_id:
        print("❌ 未找到日常台账表")
        return
        
    fields = {
        "记账日期": int(pd.to_datetime(date_str).timestamp() * 1000),
        "业务类型": "收款",
        "费用归类": "加工服务收入", # 或者 "预收账款"
        "往来单位费用": partner,
        "实际收付金额": amount,
        "账面金额": amount,
        "交易银行": bank,
        "是否有票": "有票" if "有票" in bank else "无票",
        "是否现金": "否" if "有票" in bank else "是",
        "备注": f"{remark} (加工费收款)"
    }
    
    req = CreateAppTableRecordRequest.builder() \
        .app_token(app_token) \
        .table_id(ledger_id) \
        .request_body(AppTableRecord.builder().fields(fields).build()) \
        .build()
        
    resp = client.bitable.v1.app_table_record.create(req)
    if resp.success():
        print(f"✅ 收款已记录到台账！")
        
        # 3. 智能核销建议
        # 询问是否要核销旧账单
        if input("👉 是否要自动核销该客户的旧欠款? (y/n) [y]: ").strip().lower() != 'n':
             # 查找该客户未结算的记录
             pf_table_id = get_table_id_by_name(client, app_token, "加工费明细表")
             if pf_table_id:
                 # 获取所有未结算
                 filter_cmd = f'AND(CurrentValue.[往来单位]="{partner}", CurrentValue.[结算状态]!="已结算", CurrentValue.[类型]="收入-加工服务")'
                 unpaid_recs = get_all_records(client, app_token, pf_table_id, filter_info=filter_cmd)
                 
                 if unpaid_recs:
                     # 按日期排序 (FIFO)
                     unpaid_recs.sort(key=lambda x: x.fields.get("日期", 0))
                     
                     to_settle = []
                     remaining = amount
                     
                     for r in unpaid_recs:
                         rec_amt = float(r.fields.get("总金额", 0))
                         if remaining >= rec_amt:
                             to_settle.append(r)
                             remaining -= rec_amt
                         else:
                             break # 钱不够了，剩下的部分不核销（或者部分核销？为了简单，暂时只核销全额匹配的）
                             
                     if to_settle:
                         print(f"💡 系统建议核销最早的 {len(to_settle)} 笔未结算记录 (共 {amount - remaining:,.2f} 元)")
                         if input("❓ 确认核销? (y/n): ").strip().lower() == 'y':
                             batch_recs = []
                             for r in to_settle:
                                 batch_recs.append(AppTableRecord.builder().record_id(r.record_id).fields({"结算状态": "已结算"}).build())
                             
                             # Execute Batch
                             for i in range(0, len(batch_recs), 100):
                                 batch = batch_recs[i:i+100]
                                 req_b = BatchUpdateAppTableRecordRequest.builder() \
                                     .app_token(app_token) \
                                     .table_id(pf_table_id) \
                                     .request_body(BatchUpdateAppTableRecordRequestBody.builder().records(batch).build()) \
                                     .build()
                                 client.bitable.v1.app_table_record.batch_update(req_b)
                             print(f"✅ 已自动核销 {len(to_settle)} 笔记录")
                     else:
                         print("⚠️ 收款金额不足以核销最早的一笔记录，暂不执行核销。")
                 else:
                     print("🎉 该客户没有未结算记录。")
    else:
        print(f"❌ 记录失败: {resp.msg}")

def generate_anodizing_demo_data(client, app_token):
    """生成氧化厂模拟数据 (小白专用)"""
    print(f"\n{Color.HEADER}🏭 正在生成氧化厂模拟数据...{Color.ENDC}")
    print("场景: 小型氧化加工厂，包含加工费收入、原材料采购、水电房租等。")
    
    # 1. 填充价目表 (Price List)
    pt_id = create_processing_price_table(client, app_token)
    if pt_id:
        print("1. 正在生成价目表...")
        prices = [
            {"品名": "铝型材-6063", "规格": "喷砂氧化", "单位": "kg", "单价": 4.5, "备注": "常规料"},
            {"品名": "散热器", "规格": "拉丝黑", "单位": "件", "单价": 2.5, "备注": "精密件"},
            {"品名": "铝板", "规格": "本色氧化", "单位": "m²", "单价": 35.0, "备注": "大板"},
            {"品名": "装饰条", "规格": "抛光金", "单位": "米", "单价": 1.8, "备注": "高光"}
        ]
        batch = []
        for p in prices:
            batch.append(AppTableRecord.builder().fields(p).build())
        client.bitable.v1.app_table_record.batch_create(BatchCreateAppTableRecordRequest.builder().app_token(app_token).table_id(pt_id).request_body(BatchCreateAppTableRecordRequestBody.builder().records(batch).build()).build())

    # 2. 填充加工费记录 (Processing Fees)
    pf_id = create_processing_fee_table(client, app_token)
    if pf_id:
        print("2. 正在生成加工单...")
        now = datetime.now()
        records = []
        # A. 铝型材 (按kg)
        records.append({
            "日期": int((now - timedelta(days=5)).timestamp() * 1000),
            "往来单位": "张三门窗厂",
            "品名": "铝型材-6063",
            "规格": "喷砂氧化",
            "类型": "收入-加工服务",
            "计价方式": "按重量",
            "数量": 500.0,
            "单位": "kg",
            "单价": 4.5,
            "总金额": 2250.0,
            "结算状态": "未结算",
            "开票状态": "未开票",
            "备注": "送货单号: SH20260201"
        })
        # B. 散热器 (按件)
        records.append({
            "日期": int((now - timedelta(days=3)).timestamp() * 1000),
            "往来单位": "李四电子",
            "品名": "散热器",
            "规格": "拉丝黑",
            "类型": "收入-加工服务",
            "计价方式": "按件/只/个",
            "数量": 1000.0,
            "单位": "件",
            "单价": 2.5,
            "总金额": 2500.0,
            "结算状态": "已结算",
            "开票状态": "已开票",
            "备注": "加急"
        })
        # C. 装饰条 (按米)
        records.append({
            "日期": int((now - timedelta(days=1)).timestamp() * 1000),
            "往来单位": "王五装饰",
            "品名": "装饰条",
            "规格": "抛光金",
            "类型": "收入-加工服务",
            "计价方式": "按米长",
            "数量": 2000.0,
            "单位": "米",
            "单价": 1.8,
            "总金额": 3600.0,
            "结算状态": "未结算",
            "开票状态": "未开票",
            "备注": ""
        })
        
        batch = []
        for r in records:
            batch.append(AppTableRecord.builder().fields(r).build())
        client.bitable.v1.app_table_record.batch_create(BatchCreateAppTableRecordRequest.builder().app_token(app_token).table_id(pf_id).request_body(BatchCreateAppTableRecordRequestBody.builder().records(batch).build()).build())

    # 3. 填充日常台账 (Ledger)
    lg_id = create_ledger_table(client, app_token)
    if lg_id:
        print("3. 正在生成日常支出与收款...")
        recs = []
        # 支出：原材料
        recs.append({
            "记账日期": int((now - timedelta(days=10)).timestamp() * 1000),
            "业务类型": "费用",
            "费用归类": "原材料-三酸/片碱/色粉",
            "往来单位费用": "化工原料行",
            "实际收付金额": 5000.0,
            "账面金额": 5000.0,
            "交易银行": "G银行基本户(有票)",
            "是否有票": "有票",
            "是否现金": "否",
            "备注": "采购硫酸、硝酸"
        })
        # 支出：外协
        recs.append({
            "记账日期": int((now - timedelta(days=8)).timestamp() * 1000),
            "业务类型": "费用",
            "费用归类": "外协加工费",
            "往来单位费用": "老王抛光厂",
            "实际收付金额": 1200.0,
            "账面金额": 1200.0,
            "交易银行": "N银行/微信(无票)",
            "是否有票": "无票",
            "是否现金": "是",
            "备注": "支付抛光费"
        })
        # 收入：收款 (对应李四电子)
        recs.append({
            "记账日期": int((now - timedelta(days=2)).timestamp() * 1000),
            "业务类型": "收款",
            "费用归类": "加工服务收入",
            "往来单位费用": "李四电子",
            "实际收付金额": 2500.0,
            "账面金额": 2500.0,
            "交易银行": "G银行基本户(有票)",
            "是否有票": "有票",
            "是否现金": "否",
            "备注": "收2月加工费"
        })
        
        batch = []
        for r in recs:
            batch.append(AppTableRecord.builder().fields(r).build())
        client.bitable.v1.app_table_record.batch_create(BatchCreateAppTableRecordRequest.builder().app_token(app_token).table_id(lg_id).request_body(BatchCreateAppTableRecordRequestBody.builder().records(batch).build()).build())
        
    print(f"{Color.OKGREEN}✅ 模拟数据生成完毕！请进入各个菜单查看效果。{Color.ENDC}")

def update_env(key, value):
    # 读取现有内容
    lines = []
    if os.path.exists(".env"):
        with open(".env", "r", encoding="utf-8") as f:
            lines = f.readlines()
            
    # 更新或添加
    found = False
    new_lines = []
    for line in lines:
        if line.startswith(f"{key}="):
            new_lines.append(f"{key}={value}\n")
            found = True
        else:
            new_lines.append(line)
            
    if not found:
        new_lines.append(f"\n{key}={value}\n")
        
    with open(".env", "w", encoding="utf-8") as f:
        f.writelines(new_lines)
        
    # 重新加载变量 (简单方式: 告诉用户重启，或者尝试热加载)
    # 这里我们只做简单的全局变量更新，为了生效最好重启，但部分变量可以热更新
    global VAT_RATE, TOLERANCE_DAYS, ZHIPUAI_API_KEY
    if key == "VAT_RATE":
        VAT_RATE = float(value)
    elif key == "TOLERANCE_DAYS":
        TOLERANCE_DAYS = int(value)
    elif key == "ZHIPUAI_API_KEY":
        ZHIPUAI_API_KEY = value

# 主函数
def main():
    global ZHIPUAI_API_KEY
    import argparse
    parser = argparse.ArgumentParser(description="飞书财务小助手V8.8 (Lark OAPI V2)")
    parser.add_argument("--create-table", action="store_true", help="创建台账表格+填充测试数据")
    parser.add_argument("--import-excel", type=str, nargs='?', const="", help="从Excel导入数据（路径）")
    parser.add_argument("--export-excel", action="store_true", help="导出台账数据")
    parser.add_argument("--reconcile", type=str, nargs='?', const="", help="银行流水对账（Excel路径）")
    parser.add_argument("--calculate-tax", action="store_true", help="统计税务数据")
    parser.add_argument("--find-missing-ticket", action="store_true", help="查找待补票记录")
    parser.add_argument("--generate-report", action="store_true", help="生成HTML可视化报表")
    parser.add_argument("--monthly-close", action="store_true", help="一键月度结账")
    parser.add_argument("--daily-briefing", action="store_true", help="发送每日经营简报")
    parser.add_argument("--health-check", action="store_true", help="[新] 一键财务体检")
    parser.add_argument("--ai-chat", action="store_true", help="[新] AI 查数助手")
    parser.add_argument("--smart-entry", action="store_true", help="[新] 智能文本录入")
    parser.add_argument("--smart-image", action="store_true", help="[新] 智能截图记账")
    parser.add_argument("--learn-rules", action="store_true", help="[新] 智能学习分类规则")
    parser.add_argument("--partner-statement", action="store_true", help="[新] 生成往来对账单")
    parser.add_argument("--manage-aliases", action="store_true", help="[新] 管理往来单位别名")
    parser.add_argument("--show-urls", action="store_true", help="显示云端后台链接")
    parser.add_argument("--settings", action="store_true", help="进入设置菜单")
    parser.add_argument("--menu", action="store_true", help="进入交互式主菜单")
    parser.add_argument("--auto-run", action="store_true", help="[新] 自动运行每日任务(无交互)")
    parser.add_argument("--reconcile-partner", type=str, nargs='?', const="", help="[新] 往来对账（Excel路径）")
    parser.add_argument("--salary", action="store_true", help="[新] 薪酬管理")
    parser.add_argument("--invoice", action="store_true", help="[新] 发票管理")
    parser.add_argument("--processing-fee", action="store_true", help="[新] 加工费管理")
    parser.add_argument("--generate-demo", action="store_true", help="[新] 生成氧化厂模拟数据 (小白专用)")
    parser.add_argument("--reset-system", action="store_true", help="[新] 系统初始化/重置 (数据清空)")
    parser.add_argument("--backup", action="store_true", help="[新] 全量数据备份")
    
    args = parser.parse_args()

    # 如果没有参数，默认进入交互式菜单
    import sys
    if len(sys.argv) == 1:
        interactive_menu()
        return

    if args.auto_run:
        log.info("🤖 自动运行模式启动...")
        client = init_clients()
        if client:
            one_click_daily_closing(client, APP_TOKEN)
        return

    if args.menu:
        interactive_menu()
        return
        
    if args.settings:
        settings_menu()
        return

    client = init_clients()
    if not client:
        # Try wizard if client init failed (likely due to missing config)
        if not interactive_setup_wizard():
             return

        # Retry init
        client = init_clients()
        if not client: return
    
    # 自动引导配置 (Legacy Check, kept for CLI args)
    if not APP_TOKEN:
        print("\n⚠️  检测到未配置 FEISHU_APP_TOKEN (Base Token)")
        token = input("请输入您的飞书多维表格 Token (通常在URL中): ").strip()
        if token:
            with open(".env", "a", encoding="utf-8") as f:
                f.write(f"\nFEISHU_APP_TOKEN={token}")
            log.info("✅ 配置已保存，请重新运行程序", extra={"solution": "重启"})
            return
        else:
            log.error("❌ 未配置 Token，程序退出", extra={"solution": "请在 .env 中配置"})
            return

    if args.create_table:
        # Create Tables
        log.info("🛠️ 正在初始化飞书多维表格结构...")
        create_basic_info_table(client, APP_TOKEN)
        create_ledger_table(client, APP_TOKEN)
        create_partner_table(client, APP_TOKEN) # 新增
        create_invoice_table(client, APP_TOKEN) # 新增
        create_asset_table(client, APP_TOKEN) # 新增
        create_salary_table(client, APP_TOKEN) # 新增
        create_processing_fee_table(client, APP_TOKEN) # 新增
        print("✅ 所有表格初始化完成！")
        fill_test_data(client, APP_TOKEN)
        send_bot_message(f"✅ 表格初始化完成！Base: {APP_TOKEN}", "accountant")

    if args.import_excel is not None:
        import_from_excel(client, APP_TOKEN, args.import_excel)
        
    if args.reconcile is not None:
        reconcile_bank_flow(client, APP_TOKEN, args.reconcile)
        
    if args.calculate_tax:
        # 如果是新功能，检查API KEY
        if not ZHIPUAI_API_KEY:
            print("\n💡 提示：您尚未配置 ZHIPUAI_API_KEY，无法使用 AI 财务诊断。")
            key = input("请输入智谱AI Key (按回车跳过): ").strip()
            if key:
                with open(".env", "a", encoding="utf-8") as f:
                    f.write(f"\nZHIPUAI_API_KEY={key}")
                # 重新加载
                ZHIPUAI_API_KEY = key
                log.info("✅ AI Key 已保存", extra={"solution": "无"})
        calculate_tax(client, APP_TOKEN)
        
    if args.find_missing_ticket:
        export_missing_tickets(client, APP_TOKEN)
        
    if args.generate_report:
        generate_html_report(client, APP_TOKEN)
        
    if args.monthly_close:
        monthly_close(client, APP_TOKEN)
        
    if args.daily_briefing:
        daily_briefing(client, APP_TOKEN)

    if args.health_check:
        financial_health_check(client, APP_TOKEN)

    if args.ai_chat:
        ai_data_query(client, APP_TOKEN)

    if args.smart_entry:
        smart_text_entry(client, APP_TOKEN)

    if args.smart_image:
        smart_image_entry(client, APP_TOKEN)

    if args.learn_rules:
        learn_category_rules(client, APP_TOKEN)

    if args.partner_statement:
        generate_business_statement(client, APP_TOKEN)

    if args.manage_aliases:
        manage_partners_flow(client, APP_TOKEN)

    if args.show_urls:
        show_cloud_urls(client, APP_TOKEN)

    if args.export_excel:
        export_to_excel(client, APP_TOKEN)

    if args.reconcile_partner:
        reconcile_partner_flow(client, APP_TOKEN, args.reconcile_partner)

    if args.salary:
        manage_salary_flow(client, APP_TOKEN)

    if args.invoice:
        manage_invoice_flow(client, APP_TOKEN)

    if args.processing_fee:
        manage_processing_fee_flow(client, APP_TOKEN)

    if args.generate_demo:
        generate_anodizing_demo_data(client, APP_TOKEN)

    if args.reset_system:
        reset_system_data(client, APP_TOKEN)
        return

    if args.backup:
        backup_system_data(client, APP_TOKEN)
        return

    # 4. 交互式菜单
    if args.menu:
        # 启动时显示看板
        show_ascii_dashboard(client, APP_TOKEN)
        interactive_menu()
        return

def check_for_updates():
    """检查 Git 更新 (仅在有 .git 目录时生效)"""
    if not os.path.exists(os.path.join(ROOT_DIR, ".git")):
        return
        
    try:
        # 使用 git fetch 检查远程状态
        import subprocess
        log.info("🔍 正在检查软件更新...", extra={"solution": "无"})
        
        # 1. Fetch
        subprocess.run(["git", "fetch"], check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        
        # 2. Check status
        status = subprocess.check_output(["git", "status", "-uno"], encoding="utf-8")
        
        if "behind" in status:
            print(f"\n{Color.OKGREEN}🚀 发现新版本！{Color.ENDC}")
            if input("👉 是否立即更新? (y/n) [y]: ").strip().lower() != 'n':
                print("🔄 正在更新代码...")
                subprocess.run(["git", "pull"], check=True)
                print(f"{Color.OKGREEN}✅ 更新完成，请重启程序。{Color.ENDC}")
                sys.exit(0)
        else:
            log.info("✅ 当前已是最新版本", extra={"solution": "无"})
            
    except Exception as e:
        log.warning(f"⚠️ 检查更新失败: {e}", extra={"solution": "请手动 git pull"})

if __name__ == "__main__":
    # 启用 Windows ANSI 支持
    if os.name == 'nt':
        os.system('color')
        
    # 自动检查更新
    check_for_updates()
    
    main()
